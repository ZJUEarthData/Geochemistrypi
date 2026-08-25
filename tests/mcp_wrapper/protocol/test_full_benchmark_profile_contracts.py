import hashlib
from pathlib import Path

import pytest
from geochemistrypi_mcp.api.schemas import DatasetFilterRule, DatasetPreparationContract, EnvironmentProfileContract
from geochemistrypi_mcp.data.inspector import snapshot_dataset
from geochemistrypi_mcp.data.preparation import DatasetPreparationError, prepare_dataset_view
from geochemistrypi_mcp.planning.artifact_mapping import build_adapter_artifact_mappings
from geochemistrypi_mcp.planning.profiles import BenchmarkProfile, BenchmarkProfileNotReadyError, ProfileWorkflow, load_benchmark_profile
from geochemistrypi_mcp.planning.scientific_contract import resolved_environment_profile
from openpyxl import Workbook


def test_range_filter_is_deterministic_and_bound_to_preparation_hash(tmp_path: Path) -> None:
    source = tmp_path / "elements.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["SampleID", "AGE", "MGO", "SIO2"])
    sheet.append(["S1", 10, 8.1, 42])
    sheet.append(["S2", 20, 8.2, 43])
    sheet.append(["S3", 30, 8.3, 47])
    sheet.append(["S4", 40, 8.4, 51])
    sheet.append(["S5", 50, 8.5, 52])
    workbook.save(source)
    snapshot = snapshot_dataset(source, 10 * 1024 * 1024)
    contract = DatasetPreparationContract(
        worksheet="Data",
        selected_columns=("SampleID", "AGE", "MGO", "SIO2"),
        filters=(
            DatasetFilterRule(column="SIO2", operator="not_null"),
            DatasetFilterRule(
                column="SIO2",
                operator="between",
                minimum=43,
                maximum=51,
            ),
        ),
    )

    first = prepare_dataset_view(
        snapshot,
        contract,
        tmp_path / "state",
        10 * 1024 * 1024,
        256,
    )
    second = prepare_dataset_view(
        snapshot,
        contract,
        tmp_path / "state",
        10 * 1024 * 1024,
        256,
    )
    narrower = prepare_dataset_view(
        snapshot,
        contract.model_copy(
            update={
                "filters": (
                    DatasetFilterRule(column="SIO2", operator="not_null"),
                    DatasetFilterRule(
                        column="SIO2",
                        operator="between",
                        minimum=44,
                        maximum=50,
                    ),
                )
            }
        ),
        tmp_path / "state",
        10 * 1024 * 1024,
        256,
    )

    assert first.snapshot.sha256 == second.snapshot.sha256
    assert first.record["contract_hash"] == second.record["contract_hash"]
    assert first.record["contract_hash"] != narrower.record["contract_hash"]
    assert first.snapshot.resolved_path.read_text(encoding="utf-8").splitlines() == [
        "SampleID,AGE,MGO,SIO2",
        "S2,20,8.2,43",
        "S3,30,8.3,47",
        "S4,40,8.4,51",
    ]
    assert first.record["table"]["filters"][1] == {
        "column": "SIO2",
        "operator": "between",
        "minimum": 43,
        "maximum": 51,
    }


def test_compound_headers_flatten_without_losing_source_row_lineage(
    tmp_path: Path,
) -> None:
    source = tmp_path / "paired.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Training"
    sheet.append(["SampleID", "Liquid", "Liquid", "Cpx", "Cpx", "Target"])
    sheet.append([None, "SiO2", "MgO", "SiO2", "MgO", "Pressure"])
    sheet.append(["A", 50.1, 4.2, 49.1, 15.2, 10.5])
    sheet.append(["B", 51.2, 3.9, 50.0, 14.8, 11.2])
    workbook.save(source)
    contract = DatasetPreparationContract(
        worksheet="Training",
        header_row_indices=(0, 1),
        header_join_separator="__",
        selected_columns=(
            "SampleID",
            "Liquid__SiO2",
            "Liquid__MgO",
            "Cpx__SiO2",
            "Cpx__MgO",
            "Target__Pressure",
        ),
    )

    prepared = prepare_dataset_view(
        snapshot_dataset(source, 10 * 1024 * 1024),
        contract,
        tmp_path / "state",
        10 * 1024 * 1024,
        256,
    )

    assert prepared.snapshot.resolved_path.read_text(encoding="utf-8").splitlines() == [
        "SampleID,Liquid__SiO2,Liquid__MgO,Cpx__SiO2,Cpx__MgO,Target__Pressure",
        "A,50.1,4.2,49.1,15.2,10.5",
        "B,51.2,3.9,50,14.8,11.2",
    ]
    assert prepared.record["table"]["header_row_indices"] == [0, 1]
    assert prepared.record["table"]["row_identity"]["ordered_sha256"]
    assert "compose_header_rows" in prepared.record["executed_view_operations"]


def test_explicit_duplicate_header_policy_preserves_selected_column_identity(
    tmp_path: Path,
) -> None:
    source = tmp_path / "duplicate.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["SampleID", "VALUE", "VALUE", "AGE"])
    sheet.append(["S1", 1, 2, 10])
    workbook.save(source)
    contract = DatasetPreparationContract(
        worksheet="Data",
        selected_columns=("SampleID", "AGE"),
        duplicate_header_policy="suffix",
    )

    prepared = prepare_dataset_view(
        snapshot_dataset(source, 10 * 1024 * 1024),
        contract,
        tmp_path / "state",
        10 * 1024 * 1024,
        256,
    )

    assert prepared.snapshot.resolved_path.read_text(encoding="utf-8").splitlines() == [
        "SampleID,AGE",
        "S1,10",
    ]
    assert prepared.record["table"]["duplicate_header_policy"] == "suffix"


def test_selected_projection_ignores_unrelated_duplicates_but_rejects_ambiguous_selected_columns(
    tmp_path: Path,
) -> None:
    source = tmp_path / "projected-duplicates.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["SampleID", "Feature", "Unused", "Unused"])
    sheet.append(["S1", 1.5, 10, 20])
    workbook.save(source)
    snapshot = snapshot_dataset(source, 10 * 1024 * 1024)

    prepared = prepare_dataset_view(
        snapshot,
        DatasetPreparationContract(
            worksheet="Data",
            selected_columns=("SampleID", "Feature"),
        ),
        tmp_path / "state",
        10 * 1024 * 1024,
        256,
    )

    assert prepared.snapshot.resolved_path.read_text(encoding="utf-8").splitlines() == [
        "SampleID,Feature",
        "S1,1.5",
    ]
    with pytest.raises(DatasetPreparationError, match="referenced dataset columns are ambiguous"):
        prepare_dataset_view(
            snapshot,
            DatasetPreparationContract(
                worksheet="Data",
                selected_columns=("SampleID", "Unused"),
            ),
            tmp_path / "ambiguous-state",
            10 * 1024 * 1024,
            256,
        )


def test_explicit_header_whitespace_policy_is_deterministic_and_collision_safe(
    tmp_path: Path,
) -> None:
    source = tmp_path / "whitespace.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["SampleID ", "Feature "])
    sheet.append(["S1", 2.5])
    workbook.save(source)
    snapshot = snapshot_dataset(source, 10 * 1024 * 1024)
    contract = DatasetPreparationContract(
        worksheet="Data",
        header_whitespace_policy="strip",
        selected_columns=("SampleID", "Feature"),
    )

    first = prepare_dataset_view(
        snapshot,
        contract,
        tmp_path / "state",
        10 * 1024 * 1024,
        256,
    )
    second = prepare_dataset_view(
        snapshot,
        contract,
        tmp_path / "state",
        10 * 1024 * 1024,
        256,
    )

    assert first.snapshot.sha256 == second.snapshot.sha256
    assert first.record["contract_hash"] == second.record["contract_hash"]
    assert first.snapshot.resolved_path.read_text(encoding="utf-8").splitlines() == [
        "SampleID,Feature",
        "S1,2.5",
    ]

    colliding = tmp_path / "whitespace-collision.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["SampleID", "Feature", "Feature "])
    sheet.append(["S1", 1, 2])
    workbook.save(colliding)
    with pytest.raises(DatasetPreparationError, match="referenced dataset columns are ambiguous"):
        prepare_dataset_view(
            snapshot_dataset(colliding, 10 * 1024 * 1024),
            contract,
            tmp_path / "collision-state",
            10 * 1024 * 1024,
            256,
        )


def test_explicit_header_bom_policy_propagates_the_normalized_column_identity(
    tmp_path: Path,
) -> None:
    source = tmp_path / "bom-header.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["\ufeffDate", "Ratio"])
    sheet.append(["2024-01-01", 1.25])
    workbook.save(source)
    contract = DatasetPreparationContract(
        worksheet="Data",
        header_bom_policy="strip",
        selected_columns=("Date", "Ratio"),
        row_identity={"strategy": "column_values", "columns": ("Date",)},
    )

    prepared = prepare_dataset_view(
        snapshot_dataset(source, 10 * 1024 * 1024),
        contract,
        tmp_path / "state",
        10 * 1024 * 1024,
        256,
    )

    assert prepared.snapshot.resolved_path.read_text(encoding="utf-8").splitlines() == [
        "Date,Ratio",
        "2024-01-01,1.25",
    ]
    assert prepared.record["table"]["row_identity"]["columns"] == ["Date"]


def test_multi_sheet_row_union_aligns_columns_and_hashes_sheet_row_lineage(
    tmp_path: Path,
) -> None:
    source = tmp_path / "two-populations.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "Population_A"
    first.append(["SampleID", "Group", "F1", "F2"])
    first.append(["A1", "A", 1.0, 2.0])
    first.append(["A2", "A", None, 3.0])
    second = workbook.create_sheet("Population_B")
    second.append(["SampleID", "Group", "Unused", "F2", "F1"])
    second.append(["B1", "B", "x", 5.0, 4.0])
    workbook.save(source)
    contract = DatasetPreparationContract(
        worksheets=("Population_A", "Population_B"),
        union_mode="rows",
        source_sheet_column="__source_sheet__",
        source_row_column="__source_row__",
        selected_columns=(
            "SampleID",
            "Group",
            "F1",
            "F2",
            "__source_sheet__",
            "__source_row__",
        ),
        filters=(DatasetFilterRule(column="F1", operator="not_null"),),
        row_identity={
            "strategy": "column_values",
            "columns": ("__source_sheet__", "__source_row__"),
        },
    )

    prepared = prepare_dataset_view(
        snapshot_dataset(source, 10 * 1024 * 1024),
        contract,
        tmp_path / "state",
        10 * 1024 * 1024,
        256,
    )

    assert prepared.snapshot.resolved_path.read_text(encoding="utf-8").splitlines() == [
        "SampleID,Group,F1,F2,__source_sheet__,__source_row__",
        "A1,A,1,2,Population_A,2",
        "B1,B,4,5,Population_B,2",
    ]
    assert prepared.record["table"]["worksheets"] == [
        "Population_A",
        "Population_B",
    ]
    assert prepared.record["table"]["source_row_count"] == 2
    assert prepared.record["table"]["per_sheet"]["Population_A"] == {
        "input_row_count": 2,
        "source_row_count": 1,
        "filtered_row_count": 1,
    }
    assert prepared.record["table"]["row_identity"]["ordered_sha256"]
    assert "union_worksheets_by_rows" in prepared.record["executed_view_operations"]


def test_named_environment_profile_preserves_mcp_and_geochemistrypi_versions() -> None:
    profile = EnvironmentProfileContract(
        profile_id="scientific-runtime",
        python="3.10.20",
        geochemistrypi="0.8.0",
        mcp="0.2.1",
        package_versions={"scikit-learn": "1.1.3"},
    )

    class Request:
        environment_profile = profile

    resolved = resolved_environment_profile(Request())

    assert resolved["python"] == "3.10.20"
    assert resolved["geochemistrypi"] == "0.8.0"
    assert resolved["mcp"] == "0.2.1"
    assert resolved["dependency_versions"] == {"scikit-learn": "1.1.3"}


def test_profile_reproducibility_is_forwarded_and_unknown_template_cannot_run(
    tmp_path: Path,
) -> None:
    dataset = tmp_path / "rocks.csv"
    dataset.write_text(
        "SampleID,F1,F2,Label\nS1,1,2,A\nS2,2,3,B\n",
        encoding="utf-8",
    )
    ready = tmp_path / "ready.yaml"
    ready.write_text(
        f"""profile_version: 1
benchmark:
  profile_id: generic_ready
  title: Generic ready profile
workflow:
  family: supervised_learning
  mode: classification
  method: logistic_regression
dataset:
  source: path
  path: '{dataset.as_posix()}'
  expected_sha256: {hashlib.sha256(dataset.read_bytes()).hexdigest()}
reproducibility:
  split_seed: 42
  model_seed: 42
  deterministic_policy: fixed_seed_required
parameters:
  experiment_name: Generic
  run_name: Ready
  identifier_column: SampleID
  feature_columns: [F1, F2]
  target_column: Label
  model:
    type: logistic_regression
""",
        encoding="utf-8",
    )
    template = tmp_path / "template.yaml"
    template.write_text(
        """profile_version: 1
benchmark:
  profile_id: generic_template
  title: Generic blocked template
profile_state:
  execution_ready: false
  blocker_category: DATA_OR_PARAMETER_GAP
  evidence_level: input_missing
  unknown_fields: [dataset.path, parameters.feature_columns]
workflow:
  family: dimension_reduction
  mode: embedding
  method: tsne
dataset:
  source: path
  path: UNKNOWN
parameters:
  experiment_name: Generic
  run_name: Template
  identifier_column: UNKNOWN
  feature_columns: [UNKNOWN]
  model:
    type: tsne
expected_artifacts:
  - requirement_id: dimension.coordinates
    scientific_type: embedding_coordinates
    output_role: dimension_reduction.coordinates
""",
        encoding="utf-8",
    )

    ready_profile, _ = load_benchmark_profile(ready)
    request = ready_profile.to_analysis_request()
    blocked_profile, _ = load_benchmark_profile(template)

    assert request.reproducibility.split_seed == 42
    assert request.reproducibility.model_seed == 42
    with pytest.raises(BenchmarkProfileNotReadyError, match="non-executable template"):
        blocked_profile.to_analysis_request()
    plan = blocked_profile.compatibility_plan()
    assert plan.execution_ready is False
    assert plan.public_command == ()
    assert plan.adapter_status == "requirements_unmet"
    assert plan.artifact_mappings[0].availability == "unavailable"
    assert "dataset.path" in " ".join(plan.blocking_issues)


def test_profile_stage_graph_rejects_cycles() -> None:
    with pytest.raises(ValueError, match="dependency cycle"):
        ProfileWorkflow(
            family="clustering",
            mode="pipeline",
            method="agglomerative",
            stages=(
                {
                    "stage_id": "first",
                    "family": "table_transform",
                    "method": "one",
                    "inputs": ["second_output"],
                    "outputs": ["first_output"],
                },
                {
                    "stage_id": "second",
                    "family": "clustering",
                    "method": "two",
                    "inputs": ["first_output"],
                    "outputs": ["second_output"],
                },
            ),
        )


def test_post_run_stages_do_not_block_an_executable_core_workflow(
    tmp_path: Path,
) -> None:
    dataset = tmp_path / "embedding.csv"
    dataset.write_text(
        "SampleID,F1,F2\nS1,1,2\nS2,2,3\nS3,3,4\n",
        encoding="utf-8",
    )
    profile = BenchmarkProfile.model_validate(
        {
            "benchmark": {
                "profile_id": "generic_post_run",
                "title": "Generic post-run evidence profile",
            },
            "profile_state": {
                "execution_ready": True,
                "comparison_ready": False,
                "claim_ready": False,
            },
            "workflow": {
                "family": "dimension_reduction",
                "mode": "embedding",
                "method": "tsne",
                "stages": [
                    {
                        "stage_id": "reference_comparison",
                        "family": "evaluation",
                        "method": "qualitative_neighborhood_recovery",
                        "inputs": ["coordinates", "reference_figure"],
                        "outputs": ["comparison_report"],
                        "execution_phase": "post_run_comparison",
                    }
                ],
            },
            "dataset": {
                "source": "path",
                "path": dataset,
            },
            "parameters": {
                "experiment_name": "Generic",
                "run_name": "Embedding",
                "identifier_column": "SampleID",
                "feature_columns": ["F1", "F2"],
                "model": {"type": "tsne", "perplexity": 1},
            },
        }
    )

    request = profile.to_analysis_request()

    assert request.task == "decomposition"
    assert profile.workflow.stages[0].execution_phase == "post_run_comparison"


def test_method_aware_artifact_mapping_reports_real_cli_limits() -> None:
    decomposition = build_adapter_artifact_mappings(
        "dimension_reduction",
        "embedding",
        ("geopi_output/E/R/artifacts/data/X Reduced.xlsx",),
        "tsne",
    )
    anomaly = build_adapter_artifact_mappings(
        "anomaly_detection",
        "outlier_detection",
        ("geopi_output/E/R/artifacts/data/X Abnormal Detection.xlsx",),
        "isolation_forest",
    )

    assert any(item.availability == "available" and item.output_role == "dimension_reduction.coordinates" for item in decomposition)
    assert any(item.availability == "unavailable" and item.output_role == "dimension_reduction.quality" for item in decomposition)
    assert any(item.availability == "available" and item.output_role == "anomaly_detection.labels" for item in anomaly)
    assert any(item.availability == "unavailable" and item.output_role == "anomaly_detection.scores" for item in anomaly)
