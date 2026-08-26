import hashlib
import json
import struct
import sys
import zlib
from pathlib import Path

import pytest
from geochemistrypi_mcp.api.schemas import (
    ArtifactRequirement,
    ClassificationRequest,
    DatasetFilterRule,
    DatasetPreparationContract,
    EnvironmentProfileContract,
    ExplicitDatasetReference,
    ReproducibilityContract,
    SourceRowIdentityContract,
    XGBoostSettings,
)
from geochemistrypi_mcp.config.constants import CLI_VERSION
from geochemistrypi_mcp.config.settings import McpSettings
from geochemistrypi_mcp.data.inspector import snapshot_dataset
from geochemistrypi_mcp.data.preparation import DatasetPreparationError, prepare_dataset_view
from geochemistrypi_mcp.planning.artifact_mapping import AdapterArtifactMapping
from geochemistrypi_mcp.planning.interaction_plan import AnalysisPlanCompiler
from geochemistrypi_mcp.planning.profiles import attest_profile_plan, load_benchmark_profile
from geochemistrypi_mcp.planning.scientific_contract import assess_scientific_compatibility, planned_artifact_requirements
from geochemistrypi_mcp.runtime.artifacts import discover_artifacts
from geochemistrypi_mcp.runtime.environment import EnvironmentSnapshot
from geochemistrypi_mcp.runtime.runs import InputIntegrityError, RunManager
from openpyxl import Workbook


def _workbook(path: Path, rows: int = 40) -> Path:
    workbook = Workbook()
    notes = workbook.active
    notes.title = "Notes"
    notes.append(["This is not the analysis table"])
    data = workbook.create_sheet("Analysis Data")
    data.append(["provenance note", None, None, None])
    data.append(["SampleID", "F1", "F2", "Label"])
    for index in range(rows):
        data.append([f"S{index:03d}", float(index + 1), float((index % 7) + 0.5), "A" if index % 2 == 0 else "B"])
    workbook.save(path)
    return path


def _environment() -> EnvironmentSnapshot:
    record = {
        "schema_version": 1,
        "cli_executable": {"path": str(Path(sys.executable).resolve()), "sha256": "1" * 64},
        "python": {
            "executable": str(Path(sys.executable).resolve()),
            "executable_sha256": "2" * 64,
            "version": "3.11.0",
            "implementation": "CPython",
        },
        "platform": "test-platform",
        "dependencies": {"geochemistrypi": CLI_VERSION, "xgboost": "1.3.3"},
    }
    return EnvironmentSnapshot(identity_sha256="a" * 64, record=record)


def _write_rgb_png(path: Path, pixels: list[list[tuple[int, int, int]]]) -> None:
    height = len(pixels)
    width = len(pixels[0])
    scanlines = b"".join(b"\x00" + bytes(channel for pixel in row for channel in pixel) for row in pixels)

    def chunk(kind: bytes, payload: bytes) -> bytes:
        return struct.pack(">I", len(payload)) + kind + payload + struct.pack(">I", zlib.crc32(kind + payload) & 0xFFFFFFFF)

    path.write_bytes(b"\x89PNG\r\n\x1a\n" + chunk(b"IHDR", struct.pack(">IIBBBBB", width, height, 8, 2, 0, 0, 0)) + chunk(b"IDAT", zlib.compress(scanlines)) + chunk(b"IEND", b""))


def _classification_request(dataset: Path, preparation: DatasetPreparationContract, **updates) -> ClassificationRequest:
    request = ClassificationRequest(
        training_dataset=ExplicitDatasetReference(
            path=dataset,
            expected_sha256=hashlib.sha256(dataset.read_bytes()).hexdigest(),
            preparation=preparation,
        ),
        experiment_name="R",
        run_name="C",
        identifier_column="SampleID",
        feature_columns=("F1", "F2"),
        target_column="Label",
    )
    return request.model_copy(update=updates)


def test_excel_view_is_deterministic_and_preserves_source_lineage(tmp_path: Path) -> None:
    source = _workbook(tmp_path / "raw.xlsx", rows=3)
    source_mapping = tmp_path / "source-mapping.json"
    source_mapping.write_text('{"S000":1,"S001":2,"S002":3}', encoding="utf-8")
    before = source.read_bytes()
    snapshot = snapshot_dataset(source, 10 * 1024 * 1024)
    contract = DatasetPreparationContract(
        worksheet="Analysis Data",
        header_row_index=1,
        excluded_columns=("F1",),
        row_identity=SourceRowIdentityContract(
            strategy="column_values",
            columns=("SampleID",),
            source_mapping_path=source_mapping,
            source_mapping_sha256=hashlib.sha256(source_mapping.read_bytes()).hexdigest(),
        ),
        operations=("transformation",),
    )

    first = prepare_dataset_view(snapshot, contract, tmp_path / "state", 10 * 1024 * 1024, 256)
    second = prepare_dataset_view(snapshot, contract, tmp_path / "state", 10 * 1024 * 1024, 256)

    assert source.read_bytes() == before
    assert first.snapshot.resolved_path == second.snapshot.resolved_path
    assert first.snapshot.sha256 == second.snapshot.sha256
    assert first.snapshot.resolved_path.read_text(encoding="utf-8").splitlines() == [
        "SampleID,F2,Label",
        "S000,0.5,A",
        "S001,1.5,B",
        "S002,2.5,A",
    ]
    assert first.record["source_file"]["sha256"] == snapshot.sha256
    assert first.record["table"]["worksheet"] == "Analysis Data"
    assert first.record["table"]["header_row_index"] == 1
    assert len(first.record["table"]["row_identity"]["ordered_sha256"]) == 64
    assert first.record["source_mapping"]["sha256"] == hashlib.sha256(source_mapping.read_bytes()).hexdigest()
    assert first.record["provenance"]["preparation_hash"] == first.record["contract_hash"]


def test_not_null_filter_materializes_a_deterministic_hashed_dataset_view(tmp_path: Path) -> None:
    source = _workbook(tmp_path / "raw.xlsx", rows=4)
    workbook = Workbook()
    data = workbook.active
    data.title = "Analysis Data"
    data.append(["SampleID", "F1", "F2", "Label"])
    data.append(["S001", 1.0, 2.0, "A"])
    data.append(["S002", 2.0, 3.0, None])
    data.append(["S003", 3.0, 4.0, " "])
    data.append(["S004", 4.0, 5.0, "B"])
    workbook.save(source)
    snapshot = snapshot_dataset(source, 10 * 1024 * 1024)
    contract = DatasetPreparationContract(
        worksheet="Analysis Data",
        selected_columns=("SampleID", "F1", "F2", "Label"),
        filters=(DatasetFilterRule(column="Label", operator="not_null"),),
    )

    first = prepare_dataset_view(snapshot, contract, tmp_path / "state", 10 * 1024 * 1024, 256)
    second = prepare_dataset_view(snapshot, contract, tmp_path / "state", 10 * 1024 * 1024, 256)

    assert first.snapshot.sha256 == second.snapshot.sha256
    assert first.snapshot.resolved_path.read_text(encoding="utf-8").splitlines() == [
        "SampleID,F1,F2,Label",
        "S001,1,2,A",
        "S004,4,5,B",
    ]
    assert first.record["table"]["input_row_count"] == 4
    assert first.record["table"]["source_row_count"] == 2
    assert first.record["table"]["filtered_row_count"] == 2
    assert first.record["table"]["filters"] == [{"column": "Label", "operator": "not_null"}]
    assert len(first.record["table"]["filter_result_sha256"]) == 64
    assert "filter_rows" in first.record["executed_view_operations"]


def test_dataset_view_rejects_wrong_row_identity_and_records_external_lineage(tmp_path: Path) -> None:
    source = _workbook(tmp_path / "raw.xlsx", rows=3)
    snapshot = snapshot_dataset(source, 10 * 1024 * 1024)
    wrong_identity = DatasetPreparationContract(
        worksheet="Analysis Data",
        header_row_index=1,
        selected_columns=("SampleID", "F1"),
        row_identity=SourceRowIdentityContract(
            strategy="column_values",
            columns=("SampleID",),
            expected_ordered_sha256="0" * 64,
        ),
    )
    external_lineage = DatasetPreparationContract(worksheet="Analysis Data", operations=("filtering",))

    with pytest.raises(DatasetPreparationError, match="expected_ordered_sha256"):
        prepare_dataset_view(snapshot, wrong_identity, tmp_path / "state", 10 * 1024 * 1024, 256)
    with pytest.raises(DatasetPreparationError, match="worksheet must be selected explicitly"):
        prepare_dataset_view(snapshot, DatasetPreparationContract(), tmp_path / "state", 10 * 1024 * 1024, 256)
    prepared = prepare_dataset_view(snapshot, external_lineage, tmp_path / "state", 10 * 1024 * 1024, 256)
    assert prepared.record["declared_operations"] == ["filtering"]
    assert prepared.record["executed_view_operations"] == ["select_worksheet"]


def test_run_validation_compiles_the_prepared_view_and_reports_both_hashes(tmp_path: Path) -> None:
    source = _workbook(tmp_path / "raw.xlsx")
    preparation = DatasetPreparationContract(
        worksheet="Analysis Data",
        header_row_index=1,
        selected_columns=("SampleID", "F1", "F2", "Label"),
        row_identity=SourceRowIdentityContract(strategy="column_values", columns=("SampleID",)),
    )
    request = _classification_request(source, preparation)
    manager = RunManager(
        McpSettings(
            runs_root=tmp_path / "runs",
            cli_executable=Path(sys.executable),
            maximum_dataset_bytes=10 * 1024 * 1024,
        ),
        cli_resolver=lambda: (Path(sys.executable), CLI_VERSION),
        environment_resolver=lambda _: _environment(),
    )
    try:
        preview = manager.validate(request)
    finally:
        manager.close()

    assert preview.execution_ready is True
    assert Path(preview.training_dataset_path).suffix == ".csv"
    assert preview.source_dataset_path == str(source.resolve())
    assert preview.source_dataset_sha256 == hashlib.sha256(source.read_bytes()).hexdigest()
    assert preview.training_sha256 != preview.source_dataset_sha256
    assert preview.columns == ("SampleID", "F1", "F2", "Label")
    assert preview.dataset_preparation["table"]["worksheet"] == "Analysis Data"
    assert preview.environment_identity_sha256 == "a" * 64
    assert preview.environment_status == "UNSPECIFIED"
    assert preview.effective_seeds == {"split": 42, "model": 42}


def test_effective_seeds_environment_and_model_parameters_are_bound_before_execution(tmp_path: Path) -> None:
    source = _workbook(tmp_path / "raw.xlsx")
    preparation = DatasetPreparationContract(
        worksheet="Analysis Data",
        header_row_index=1,
        selected_columns=("SampleID", "F1", "F2", "Label"),
    )
    base = _classification_request(source, preparation)
    prepared = prepare_dataset_view(
        snapshot_dataset(source, 10 * 1024 * 1024),
        preparation,
        tmp_path / "state",
        10 * 1024 * 1024,
        256,
    )
    environment = _environment()

    def compile_request(request: ClassificationRequest):
        return AnalysisPlanCompiler().compile(
            request.model_copy(update={"training_dataset_path": prepared.snapshot.resolved_path, "training_dataset": None}),
            cli_executable=Path(sys.executable),
        )

    exact = base.model_copy(
        update={
            "reproducibility": ReproducibilityContract(
                split_seed=42,
                model_seed=42,
                dependency_constraints={"xgboost": "==1.3.3"},
                environment={"expected_identity_sha256": "a" * 64},
                model_parameter_assertions={"maximum_iterations": 200},
                deterministic_policy="fixed_seed_and_dependency_required",
            )
        }
    )
    plan = compile_request(exact)
    requirements = planned_artifact_requirements(exact, plan)
    accepted = assess_scientific_compatibility(exact, plan, requirements, environment)

    wrong_seed = exact.model_copy(update={"reproducibility": exact.reproducibility.model_copy(update={"split_seed": 800})})
    wrong_seed_plan = compile_request(wrong_seed)
    rejected = assess_scientific_compatibility(
        wrong_seed,
        wrong_seed_plan,
        planned_artifact_requirements(wrong_seed, wrong_seed_plan),
        environment,
    )
    wrong_dependency = exact.model_copy(update={"reproducibility": exact.reproducibility.model_copy(update={"dependency_constraints": {"xgboost": "==9.9.9"}})})
    wrong_dependency_plan = compile_request(wrong_dependency)
    dependency_rejected = assess_scientific_compatibility(
        wrong_dependency,
        wrong_dependency_plan,
        planned_artifact_requirements(wrong_dependency, wrong_dependency_plan),
        environment,
    )

    assert accepted.execution_ready is True
    assert accepted.environment_status == "READY"
    assert {name: json.loads(value) for name, value in plan.effective_model_parameters}["maximum_iterations"] == 200
    assert dict(plan.requested_model_parameters) == dict(plan.effective_model_parameters)
    assert dict(plan.requested_preprocessing_parameters) == dict(plan.effective_preprocessing_parameters)
    assert plan.preprocessing_parameter_binding == "interaction_plan"
    assert rejected.execution_ready is False
    assert rejected.environment_status == "READY"
    assert dict(wrong_seed_plan.requested_seeds)["split"] == 800
    assert dict(wrong_seed_plan.effective_seeds)["split"] == 42
    assert "effective value 42" in " ".join(rejected.blocking_issues)
    assert dependency_rejected.environment_status == "MISMATCH"
    assert dependency_rejected.execution_ready is False


def test_named_environment_profile_matches_or_blocks_before_execution(tmp_path: Path) -> None:
    source = _workbook(tmp_path / "raw.xlsx")
    preparation = DatasetPreparationContract(
        worksheet="Analysis Data",
        header_row_index=1,
        selected_columns=("SampleID", "F1", "F2", "Label"),
    )
    base = _classification_request(source, preparation)
    prepared = prepare_dataset_view(
        snapshot_dataset(source, 10 * 1024 * 1024),
        preparation,
        tmp_path / "state",
        10 * 1024 * 1024,
        256,
    )

    def assess(profile: EnvironmentProfileContract):
        request = base.model_copy(update={"environment_profile": profile})
        plan = AnalysisPlanCompiler().compile(
            request.model_copy(update={"training_dataset_path": prepared.snapshot.resolved_path, "training_dataset": None}),
            cli_executable=Path(sys.executable),
        )
        result = assess_scientific_compatibility(
            request,
            plan,
            planned_artifact_requirements(request, plan),
            _environment(),
        )
        return plan, result

    matching = EnvironmentProfileContract(
        profile_id="classification-py311-xgb133",
        python="3.11.0",
        package_versions={"xgboost": "1.3.3"},
        runtime_constraints={"python_implementation": "CPython", "platform": "test-platform"},
    )
    matching_plan, accepted = assess(matching)
    _, rejected = assess(matching.model_copy(update={"profile_id": "classification-py311-xgb999", "package_versions": {"xgboost": "9.9.9"}}))

    assert accepted.execution_ready is True
    assert accepted.environment_status == "READY"
    assert matching_plan.environment_profile_id == matching.profile_id
    assert len(matching_plan.environment_profile_identity_sha256) == 64
    assert rejected.execution_ready is False
    assert rejected.environment_status == "MISMATCH"


def test_classification_adapter_maps_real_cli_outputs_and_rejects_unavailable_normalization(tmp_path: Path) -> None:
    source = _workbook(tmp_path / "raw.xlsx")
    preparation = DatasetPreparationContract(
        worksheet="Analysis Data",
        header_row_index=1,
        selected_columns=("SampleID", "F1", "F2", "Label"),
    )
    request = _classification_request(source, preparation, model=XGBoostSettings())
    prepared = prepare_dataset_view(
        snapshot_dataset(source, 10 * 1024 * 1024),
        preparation,
        tmp_path / "state",
        10 * 1024 * 1024,
        256,
    )
    plan = AnalysisPlanCompiler().compile(
        request.model_copy(update={"training_dataset_path": prepared.snapshot.resolved_path, "training_dataset": None}),
        cli_executable=Path(sys.executable),
    )
    available = tuple(mapping for mapping in plan.artifact_mappings if mapping.availability == "available")
    declared_paths = {path.replace("\\", "/") for path in plan.expected_output_relative_paths}
    mapped_paths = {mapping.relative_path for mapping in available}
    normalized_requirement = ArtifactRequirement(
        requirement_id="evaluation.confusion.normalized",
        scientific_type="normalized_confusion_matrix_table",
        output_role="evaluation.confusion_matrix.normalized",
    )
    blocked = assess_scientific_compatibility(request, plan, (normalized_requirement,), _environment())

    assert declared_paths <= mapped_paths
    assert {"evaluation.scores", "evaluation.predictions", "model.feature_importance"} <= {mapping.output_role for mapping in available}
    assert blocked.execution_ready is False
    assert blocked.artifact_status == "requirements_unmet"
    assert "raw confusion matrix only" in " ".join(blocked.blocking_issues)


def test_validation_receipt_reports_matching_and_mismatched_environments(tmp_path: Path) -> None:
    source = _workbook(tmp_path / "raw.xlsx")
    preparation = DatasetPreparationContract(
        worksheet="Analysis Data",
        header_row_index=1,
        selected_columns=("SampleID", "F1", "F2", "Label"),
    )
    base = _classification_request(source, preparation)
    settings = McpSettings(
        runs_root=tmp_path / "runs",
        cli_executable=Path(sys.executable),
        maximum_dataset_bytes=10 * 1024 * 1024,
    )
    manager = RunManager(
        settings,
        cli_resolver=lambda: (Path(sys.executable), CLI_VERSION),
        environment_resolver=lambda _: _environment(),
    )
    try:
        ready = manager.validate(
            base.model_copy(
                update={
                    "reproducibility": ReproducibilityContract(
                        environment={"expected_identity_sha256": "a" * 64},
                    )
                }
            )
        )
        mismatch = manager.validate(
            base.model_copy(
                update={
                    "reproducibility": ReproducibilityContract(
                        environment={"expected_identity_sha256": "b" * 64},
                    )
                }
            )
        )
        tampered = base.model_copy(update={"training_dataset": base.training_dataset.model_copy(update={"expected_sha256": "0" * 64})})
        with pytest.raises(InputIntegrityError, match="changed between source resolution and validation"):
            manager.validate(tampered)
    finally:
        manager.close()

    assert ready.environment_status == "READY"
    assert ready.execution_ready is True
    assert mismatch.environment_status == "MISMATCH"
    assert mismatch.execution_ready is False
    assert not list((tmp_path / "runs").glob("run-[0-9a-f]*"))


def test_configuration_only_benchmark_profile_compiles_and_attests_generic_workflow(tmp_path: Path) -> None:
    dataset = tmp_path / "rocks.csv"
    dataset.write_text(
        "SampleID,F1,F2,Label\n" "S1,1,2,A\nS2,2,3,B\nS3,3,4,A\nS4,4,5,B\nS5,5,6,A\n" "S6,6,7,B\nS7,7,8,A\nS8,8,9,B\nS9,9,10,A\nS10,10,11,B\n",
        encoding="utf-8",
    )
    dataset_sha256 = hashlib.sha256(dataset.read_bytes()).hexdigest()
    profile_path = tmp_path / "classification.yaml"
    profile_path.write_text(
        f"""profile_version: 1
benchmark:
  profile_id: generic_classification
  title: Generic classification reproduction
workflow:
  family: supervised_learning
  mode: classification
  method: logistic_regression
dataset:
  source: path
  path: '{dataset.as_posix()}'
  expected_sha256: {dataset_sha256}
environment:
  expected_identity_sha256: {'a' * 64}
parameters:
  experiment_name: Profile
  run_name: Classification
  identifier_column: SampleID
  feature_columns: [F1, F2]
  target_column: Label
  model:
    type: logistic_regression
expected_artifacts: []
acceptance_rules:
  require_execution_ready: true
""",
        encoding="utf-8",
    )
    profile_sha256 = hashlib.sha256(profile_path.read_bytes()).hexdigest()

    profile, observed_sha256 = load_benchmark_profile(profile_path, profile_sha256)
    request = profile.to_analysis_request()
    execution_request = request.model_copy(update={"training_dataset_path": dataset, "training_dataset": None})
    plan = AnalysisPlanCompiler().compile(execution_request, cli_executable=Path(sys.executable))
    plan = AnalysisPlanCompiler.bind_scientific_adapter(plan, request)

    assert observed_sha256 == profile_sha256
    assert request.task == "classification"
    assert attest_profile_plan(profile, plan) == ()
    with pytest.raises(ValueError, match="expected_sha256"):
        load_benchmark_profile(profile_path, "0" * 64)


def test_workflow_aware_artifact_contract_checks_role_cardinality_and_content(tmp_path: Path) -> None:
    output = tmp_path / "output"
    for directory in ("artifacts", "metrics", "parameters", "summary"):
        (output / directory).mkdir(parents=True)
    metric = output / "metrics" / "Model Score - Logistic Regression.txt"
    metric.write_text('{"accuracy":0.75,"f1":0.73}', encoding="utf-8")
    requirement = ArtifactRequirement(
        requirement_id="evaluation.holdout",
        scientific_type="holdout_metrics",
        output_role="evaluation.holdout",
        category="metrics",
        path_pattern="metrics/Model Score*.txt",
        media_types=("application/json",),
        required_json_keys=("accuracy", "f1"),
    )

    complete = discover_artifacts(output, 20, (requirement,))
    missing_key = requirement.model_copy(update={"requirement_id": "evaluation.recall", "required_json_keys": ("recall",)})
    incomplete = discover_artifacts(output, 20, (missing_key,))

    assert complete.missing_requirement_ids == ()
    assert complete.requirement_matches["evaluation.holdout"] == ("metrics/Model Score - Logistic Regression.txt",)
    assert complete.all_index_entries[0]["requirement_ids"] == ["evaluation.holdout"]
    assert incomplete.missing_requirement_ids == ("evaluation.recall",)
    assert "missing required JSON keys" in incomplete.requirement_failures["evaluation.recall"]


def test_observed_predicted_figure_contract_rejects_a_plot_without_visible_data(tmp_path: Path) -> None:
    output = tmp_path / "output"
    for directory in ("artifacts", "metrics", "parameters", "summary"):
        (output / directory).mkdir(parents=True)
    image_directory = output / "artifacts" / "image" / "model_output"
    image_directory.mkdir(parents=True)
    figure = image_directory / "External Predicted vs. Actual - Extra-Trees.png"
    blank_pixels = [[(255, 255, 255) for _ in range(40)] for _ in range(40)]
    for index in range(40):
        blank_pixels[4][index] = (0, 0, 0)
        blank_pixels[35][index] = (0, 0, 0)
    _write_rgb_png(figure, blank_pixels)
    requirement = ArtifactRequirement(
        requirement_id="evaluation.figure",
        scientific_type="observed_predicted_figure",
        output_role="evaluation.figure",
        category="artifacts",
        path_pattern="artifacts/image/model_output/External Predicted vs. Actual*.png",
        media_types=("image/png",),
    )
    mapping = AdapterArtifactMapping(
        mapping_id="regression.external.figure",
        scientific_type="observed_predicted_figure",
        output_role="evaluation.figure",
        relative_path=("artifacts/image/model_output/" "External Predicted vs. Actual - Extra-Trees.png"),
    )

    blank = discover_artifacts(
        output,
        20,
        (requirement,),
        workflow_family="regression",
        artifact_mappings=(mapping,),
    )

    assert blank.missing_requirement_ids == ("evaluation.figure",)
    assert "visible plot data" in blank.requirement_failures["evaluation.figure"]

    plotted_pixels = [[(255, 255, 255) for _ in range(40)] for _ in range(40)]
    for index in range(10, 30):
        plotted_pixels[index][index] = (0, 0, 255)
        plotted_pixels[index][39 - index] = (255, 0, 0)
    _write_rgb_png(figure, plotted_pixels)

    complete = discover_artifacts(
        output,
        20,
        (requirement,),
        workflow_family="regression",
        artifact_mappings=(mapping,),
    )

    assert complete.missing_requirement_ids == ()
    assert complete.requirement_matches["evaluation.figure"] == ("artifacts/image/model_output/External Predicted vs. Actual - Extra-Trees.png",)


def test_artifact_cardinality_counts_unique_content_not_identical_cli_mirrors(
    tmp_path: Path,
) -> None:
    output = tmp_path / "output"
    for directory in ("artifacts", "metrics", "parameters", "summary"):
        (output / directory).mkdir(parents=True)
    artifact = output / "artifacts" / "Subaerial Proportion.pdf"
    summary = output / "summary" / "Subaerial Proportion.pdf"
    artifact.write_bytes(b"same-scientific-figure")
    summary.write_bytes(artifact.read_bytes())
    requirement = ArtifactRequirement(
        requirement_id="time_series.figure",
        scientific_type="time_series_figure",
        output_role="time_series.figure",
        path_pattern="**/Subaerial Proportion.pdf",
        media_types=("application/pdf",),
        minimum_count=1,
        maximum_count=1,
    )

    mirrored = discover_artifacts(output, 20, (requirement,))
    summary.write_bytes(b"different-figure")
    conflicting = discover_artifacts(output, 20, (requirement,))

    assert mirrored.missing_requirement_ids == ()
    assert mirrored.requirement_matches["time_series.figure"] == (
        "artifacts/Subaerial Proportion.pdf",
        "summary/Subaerial Proportion.pdf",
    )
    assert conflicting.missing_requirement_ids == ("time_series.figure",)
    assert "more than maximum_count=1" in conflicting.requirement_failures["time_series.figure"]
