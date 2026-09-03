import hashlib
from pathlib import Path

import geochemistrypi_mcp.data.inspector as inspector_module
import pytest
from geochemistrypi_mcp.api.schemas import DatasetInspectionRequest
from geochemistrypi_mcp.config.settings import McpSettings
from geochemistrypi_mcp.data.inspector import DatasetInspectionError, inspect_dataset, snapshot_dataset
from openpyxl import Workbook
from pydantic import ValidationError


def _settings(tmp_path: Path, maximum_bytes: int = 1024 * 1024) -> McpSettings:
    return McpSettings(runs_root=tmp_path / "runs", cli_executable=None, maximum_dataset_bytes=maximum_bytes)


def test_csv_inspection_is_bounded_typed_and_read_only(tmp_path: Path) -> None:
    dataset = tmp_path / "rocks.csv"
    dataset.write_text(
        "SampleID,SIO2,Valid,Label\n" "A,50.1,true,basalt\n" "B,61.0,false,granite\n" "C,55.2,true,andesite\n",
        encoding="utf-8",
    )
    before = dataset.read_bytes()

    result = inspect_dataset(
        DatasetInspectionRequest(dataset_path=dataset, sample_rows=2, detail="full"),
        _settings(tmp_path),
    )

    assert result.format == "csv"
    assert result.row_count == 3
    assert result.row_count_exact is True
    assert result.column_count == 4
    assert [column.inferred_type for column in result.columns] == ["string", "number", "boolean", "string"]
    assert result.sample_rows == (
        {"SampleID": "A", "SIO2": "50.1", "Valid": "true", "Label": "basalt"},
        {"SampleID": "B", "SIO2": "61.0", "Valid": "false", "Label": "granite"},
    )
    assert result.sample_truncated is True
    assert result.sha256 == hashlib.sha256(before).hexdigest()
    assert dataset.read_bytes() == before


def test_names_only_inspection_returns_all_headers_without_verbose_column_summaries(
    tmp_path: Path,
) -> None:
    dataset = tmp_path / "wide.csv"
    headers = [f"COLUMN_{index:03d}" for index in range(59)]
    dataset.write_text(
        ",".join(headers) + "\n" + ",".join(str(index) for index in range(59)) + "\n",
        encoding="utf-8",
    )

    full = inspect_dataset(
        DatasetInspectionRequest(dataset_path=dataset, sample_rows=0, detail="full"),
        _settings(tmp_path),
    )
    names = inspect_dataset(
        DatasetInspectionRequest(dataset_path=dataset, sample_rows=0, detail="names"),
        _settings(tmp_path),
    )

    assert [column.name for column in full.columns] == headers
    assert full.column_names == ()
    assert names.columns == ()
    assert names.column_names == tuple(headers)
    assert names.column_count == 59
    assert names.sample_rows == ()
    assert len(names.model_dump_json()) < len(full.model_dump_json()) / 2


def test_xlsx_inspection_uses_bounded_rows_and_exact_cli_row_count(tmp_path: Path) -> None:
    dataset = tmp_path / "rocks.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["SampleID", "SIO2", "Label"])
    worksheet.append(["A", 50.1, "basalt"])
    worksheet.append(["B", 61.0, "granite"])
    workbook.save(dataset)
    workbook.close()

    result = inspect_dataset(DatasetInspectionRequest(dataset_path=dataset, sample_rows=1), _settings(tmp_path))

    assert result.format == "xlsx"
    assert result.row_count == 2
    assert result.row_count_exact is True
    assert result.sample_rows == ({"SampleID": "A", "SIO2": 50.1, "Label": "basalt"},)


def test_xlsx_blank_header_matches_pandas_unnamed_column_and_is_read_only(
    tmp_path: Path,
) -> None:
    path = tmp_path / "blank-header.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["SampleID", None, "Value"])
    worksheet.append(["A", "note", 1.0])
    workbook.save(path)
    before = path.read_bytes()

    result = inspect_dataset(
        DatasetInspectionRequest(
            dataset_path=path.resolve(),
            sample_rows=1,
            detail="full",
        ),
        _settings(tmp_path),
    )

    assert [column.name for column in result.columns] == [
        "SampleID",
        "Unnamed: 1",
        "Value",
    ]
    assert result.sample_rows[0]["Unnamed: 1"] == "note"
    assert path.read_bytes() == before


def test_every_bundled_workbook_can_be_inspected_without_modification(
    tmp_path: Path,
) -> None:
    repository = Path(__file__).resolve().parents[3]
    dataset_root = repository / "geochemistrypi" / "data_mining" / "data" / "dataset"
    paths = sorted(dataset_root.glob("*.xlsx"))
    before = {path.name: hashlib.sha256(path.read_bytes()).hexdigest() for path in paths}

    results = [
        inspect_dataset(
            DatasetInspectionRequest(dataset_path=path.resolve(), sample_rows=0),
            _settings(tmp_path, maximum_bytes=20 * 1024 * 1024),
            allow_pandas_duplicate_mangling=True,
        )
        for path in paths
    ]

    assert len(results) == 8
    assert all(result.column_count >= 1 for result in results)
    time_series = next(result for result in results if result.source_path.endswith("Data_Time_Series.xlsx"))
    assert any("FEOT.1" in warning for warning in time_series.header_warnings)
    assert before == {path.name: hashlib.sha256(path.read_bytes()).hexdigest() for path in paths}


@pytest.mark.parametrize(
    ("headers", "message"),
    [
        (["SampleID", "Value", "Value"], "duplicate or colliding"),
        (["SampleID", " Value"], "leading or trailing whitespace"),
        (["SampleID", "x" * 129], "must not exceed 128"),
    ],
)
def test_unsafe_xlsx_headers_fail_deterministically(tmp_path: Path, headers, message: str) -> None:
    path = tmp_path / "unsafe.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(headers)
    worksheet.append([1] * len(headers))
    workbook.save(path)

    with pytest.raises(DatasetInspectionError, match=message):
        inspect_dataset(DatasetInspectionRequest(dataset_path=path.resolve()), _settings(tmp_path))


def test_inspection_resolves_contained_relative_paths_and_rejects_escape_unknown_oversized_and_extra_inputs(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    startup_root = tmp_path / "session"
    startup_root.mkdir()
    dataset = startup_root / "rocks.csv"
    dataset.write_text("A,B\n1,2\n", encoding="utf-8")

    monkeypatch.setattr(
        inspector_module,
        "_MCP_STARTUP_WORKING_DIRECTORY",
        startup_root.resolve(),
    )
    relative_snapshot = snapshot_dataset(Path("rocks.csv"), 100)
    assert relative_snapshot.source_path == dataset.resolve()
    assert relative_snapshot.resolved_path == dataset.resolve()

    outside = tmp_path / "outside.csv"
    outside.write_text("A,B\n1,2\n", encoding="utf-8")
    with pytest.raises(DatasetInspectionError, match="must remain inside"):
        snapshot_dataset(Path("../outside.csv"), 100)
    unknown = tmp_path / "rocks.json"
    unknown.write_text("{}", encoding="utf-8")
    with pytest.raises(DatasetInspectionError, match="Supported dataset formats"):
        snapshot_dataset(unknown, 100)
    with pytest.raises(DatasetInspectionError, match="exceeds"):
        snapshot_dataset(dataset, 1)
    with pytest.raises(ValidationError, match="Extra inputs are not permitted"):
        DatasetInspectionRequest(dataset_path=dataset, unexpected=True)

    long_header = tmp_path / "long-header.csv"
    long_header.write_text(f"{'A' * 129},B\n1,2\n", encoding="utf-8")
    with pytest.raises(DatasetInspectionError, match="must not exceed 128"):
        inspect_dataset(
            DatasetInspectionRequest(dataset_path=long_header),
            _settings(tmp_path),
        )
