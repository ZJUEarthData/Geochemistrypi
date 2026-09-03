import json
import sys
from pathlib import Path

import pytest
from geochemistrypi_mcp.api.schemas import AnomalyDetectionRequest, BuiltInDatasetReference, ClassificationRequest, ClusteringRequest, DecompositionRequest, RegressionRequest, TimeSeriesRequest
from geochemistrypi_mcp.config.constants import CLI_VERSION
from geochemistrypi_mcp.config.settings import McpSettings
from geochemistrypi_mcp.data.catalog import ResolvedDataset
from geochemistrypi_mcp.data.inspector import snapshot_dataset
from geochemistrypi_mcp.data.row_identity import SourceRowIdentityError, build_source_row_lineage
from geochemistrypi_mcp.data.row_pairing import verify_original_row_pairing
from geochemistrypi_mcp.planning.interaction_plan import AnalysisPlanCompiler
from geochemistrypi_mcp.runtime.runs import RunManager
from openpyxl import Workbook


def _mixed_identifier_dataset(tmp_path: Path) -> Path:
    path = tmp_path / "mixed-identifiers.csv"
    rows = ["duplicate,0,1,10,20,30", "duplicate,1,2,11,21,31", ",0,3,12,22,32"]
    rows.extend(f"sample-{index},{index % 2},{index},{10 + index},{20 + index},{30 + index}" for index in range(4, 16))
    path.write_text("SampleID,Label,Target,F1,F2,F3\n" + "\n".join(rows) + "\n", encoding="utf-8")
    return path


@pytest.mark.parametrize(
    "request_factory",
    [
        lambda path: ClassificationRequest(
            training_dataset_path=path,
            experiment_name="Identifier Root Fix",
            run_name="Classification",
            identifier_column="SampleID",
            feature_columns=("F1", "F2", "F3"),
            target_column="Label",
        ),
        lambda path: RegressionRequest(
            training_dataset_path=path,
            experiment_name="Identifier Root Fix",
            run_name="Regression",
            identifier_column="SampleID",
            feature_columns=("F1", "F2", "F3"),
            target_column="Target",
            model={"type": "linear_regression"},
        ),
        lambda path: ClusteringRequest(
            training_dataset_path=path,
            experiment_name="Identifier Root Fix",
            run_name="Clustering",
            identifier_column="SampleID",
            feature_columns=("F1", "F2", "F3"),
            model={"type": "kmeans", "number_of_clusters": 3},
        ),
        lambda path: DecompositionRequest(
            training_dataset_path=path,
            experiment_name="Identifier Root Fix",
            run_name="Decomposition",
            identifier_column="SampleID",
            feature_columns=("F1", "F2", "F3"),
            model={"type": "pca", "number_of_components": 2},
        ),
        lambda path: AnomalyDetectionRequest(
            training_dataset_path=path,
            experiment_name="Identifier Root Fix",
            run_name="Anomaly Detection",
            identifier_column="SampleID",
            feature_columns=("F1", "F2", "F3"),
            model={"type": "isolation_forest", "contamination": 0.1},
        ),
    ],
)
def test_duplicate_and_missing_scientific_identifiers_do_not_change_cli_science(tmp_path: Path, request_factory) -> None:
    dataset = _mixed_identifier_dataset(tmp_path)
    before = dataset.read_bytes()
    request = request_factory(dataset)
    request_before = request.model_dump(mode="json")

    plan = AnalysisPlanCompiler().compile(request, cli_executable=Path(sys.executable))

    serialized_plan = json.dumps(
        {
            "command": plan.public_command,
            "responses": [step.response for step in plan.steps],
        }
    )
    assert request.model_dump(mode="json") == request_before
    assert request.identifier_column not in request.feature_columns
    assert request.identifier_column not in tuple(getattr(request, "resolved_target_columns", ()))
    assert "mcp-row-" not in serialized_plan
    assert plan.requires_source_row_pairing is True
    assert dataset.read_bytes() == before


def test_internal_row_identities_are_unique_stable_and_independent_of_scientific_identifiers(tmp_path: Path) -> None:
    dataset = _mixed_identifier_dataset(tmp_path)
    before = dataset.read_bytes()

    first = snapshot_dataset(dataset, 1024 * 1024)
    second = snapshot_dataset(dataset, 1024 * 1024)

    assert first.row_lineage.source_row_count == 15
    assert len(first.row_lineage.identities) == 15
    assert len(set(first.row_lineage.identities)) == 15
    assert all(identity.startswith("mcp-row-") for identity in first.row_lineage.identities)
    assert first.row_lineage == second.row_lineage
    assert dataset.read_bytes() == before


def test_internal_row_identity_collision_fails_closed(monkeypatch) -> None:
    from geochemistrypi_mcp.data import row_identity

    monkeypatch.setattr(row_identity, "source_row_identity", lambda *_: "mcp-row-collision")
    with pytest.raises(SourceRowIdentityError, match="collision"):
        build_source_row_lineage("0" * 64, 2)


def test_original_result_rows_pair_deterministically_with_duplicate_and_missing_identifiers(tmp_path: Path) -> None:
    source = tmp_path / "source.csv"
    source.write_text("SampleID,F1\nduplicate,1\nduplicate,2\n,3\nsample-4,4\n", encoding="utf-8")
    snapshot = snapshot_dataset(source, 1024 * 1024)
    output = tmp_path / "output"
    data_directory = output / "artifacts" / "data"
    data_directory.mkdir(parents=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["SampleID", "F1"])
    worksheet.append(["duplicate", 1])
    worksheet.append(["duplicate", 2])
    worksheet.append([None, 3])
    worksheet.append(["sample-4", 4])
    workbook.save(data_directory / "Data Original.xlsx")
    workbook.close()

    pairing = verify_original_row_pairing(source, output, "SampleID", snapshot.row_lineage)

    assert pairing["verified"] is True
    assert pairing["source_row_count"] == 4
    assert pairing["scientific_identifier_values_preserved"] is True
    assert len(pairing["ordered_pairing_sha256"]) == 64


def test_original_result_pairing_respects_xlsx_numeric_storage_precision(tmp_path: Path) -> None:
    source = tmp_path / "source.csv"
    source.write_text("SampleID,F1\nA,0.04966603906761905\n", encoding="utf-8")
    snapshot = snapshot_dataset(source, 1024 * 1024)
    output = tmp_path / "output"
    data_directory = output / "artifacts" / "data"
    data_directory.mkdir(parents=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["SampleID", "F1"])
    worksheet.append(["A", 0.049666039067619])
    workbook.save(data_directory / "Data Original.xlsx")
    workbook.close()

    pairing = verify_original_row_pairing(source, output, "SampleID", snapshot.row_lineage)

    assert pairing["verified"] is True
    assert pairing["numeric_comparison_policy"] == "xlsx_relative_1e-14_absolute_1e-15"


def test_original_result_pairing_accepts_small_csv_to_xlsx_round_trip_error(tmp_path: Path) -> None:
    source = tmp_path / "source.csv"
    source.write_text("SampleID,MgO\n43,0.0023160599651085574\n", encoding="utf-8")
    snapshot = snapshot_dataset(source, 1024 * 1024)
    output = tmp_path / "output"
    data_directory = output / "artifacts" / "data"
    data_directory.mkdir(parents=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["SampleID", "MgO"])
    worksheet.append([43, 0.0023160599651085])
    workbook.save(data_directory / "Data Original.xlsx")
    workbook.close()

    pairing = verify_original_row_pairing(source, output, "SampleID", snapshot.row_lineage)

    assert pairing["verified"] is True


def test_original_result_pairing_rejects_meaningful_small_numeric_change(tmp_path: Path) -> None:
    source = tmp_path / "source.csv"
    source.write_text("SampleID,MgO\n43,0.0023160599651085574\n", encoding="utf-8")
    snapshot = snapshot_dataset(source, 1024 * 1024)
    output = tmp_path / "output"
    data_directory = output / "artifacts" / "data"
    data_directory.mkdir(parents=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["SampleID", "MgO"])
    worksheet.append([43, 0.0023160599661085])
    workbook.save(data_directory / "Data Original.xlsx")
    workbook.close()

    with pytest.raises(SourceRowIdentityError, match="changed or reordered source row"):
        verify_original_row_pairing(source, output, "SampleID", snapshot.row_lineage)


def test_original_result_pairing_preserves_numeric_looking_text_from_csv(tmp_path: Path) -> None:
    source = tmp_path / "source.csv"
    source.write_text("SampleID,F1\n110062,1.0386431945040793\n", encoding="utf-8")
    snapshot = snapshot_dataset(source, 1024 * 1024)
    output = tmp_path / "output"
    data_directory = output / "artifacts" / "data"
    data_directory.mkdir(parents=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["SampleID", "F1"])
    worksheet.append(["110062", 1.038643194504079])
    workbook.save(data_directory / "Data Original.xlsx")
    workbook.close()

    pairing = verify_original_row_pairing(source, output, "SampleID", snapshot.row_lineage)

    assert pairing["verified"] is True
    assert pairing["scientific_identifier_values_preserved"] is True


def test_original_result_pairing_still_rejects_numeric_change_above_xlsx_precision(tmp_path: Path) -> None:
    source = tmp_path / "source.csv"
    source.write_text("SampleID,F1\nA,0.04966603906761905\n", encoding="utf-8")
    snapshot = snapshot_dataset(source, 1024 * 1024)
    output = tmp_path / "output"
    data_directory = output / "artifacts" / "data"
    data_directory.mkdir(parents=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["SampleID", "F1"])
    worksheet.append(["A", 0.0496660390677])
    workbook.save(data_directory / "Data Original.xlsx")
    workbook.close()

    with pytest.raises(SourceRowIdentityError, match="changed or reordered source row"):
        verify_original_row_pairing(source, output, "SampleID", snapshot.row_lineage)


def test_original_result_row_count_mismatch_fails_closed(tmp_path: Path) -> None:
    source = tmp_path / "source.csv"
    source.write_text("SampleID,F1\nA,1\nB,2\n", encoding="utf-8")
    snapshot = snapshot_dataset(source, 1024 * 1024)
    output = tmp_path / "output"
    data_directory = output / "artifacts" / "data"
    data_directory.mkdir(parents=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["SampleID", "F1"])
    worksheet.append(["A", 1])
    workbook.save(data_directory / "Data Original.xlsx")
    workbook.close()

    with pytest.raises(SourceRowIdentityError, match="row count"):
        verify_original_row_pairing(source, output, "SampleID", snapshot.row_lineage)


def test_duplicate_identifiers_cannot_hide_result_row_reordering(tmp_path: Path) -> None:
    source = tmp_path / "source.csv"
    source.write_text("SampleID,F1\nduplicate,1\nduplicate,2\n", encoding="utf-8")
    snapshot = snapshot_dataset(source, 1024 * 1024)
    output = tmp_path / "output"
    data_directory = output / "artifacts" / "data"
    data_directory.mkdir(parents=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["SampleID", "F1"])
    worksheet.append(["duplicate", 2])
    worksheet.append(["duplicate", 1])
    workbook.save(data_directory / "Data Original.xlsx")
    workbook.close()

    with pytest.raises(SourceRowIdentityError, match="changed or reordered source row"):
        verify_original_row_pairing(source, output, "SampleID", snapshot.row_lineage)


class _RepositoryBuiltInCatalog:
    def __init__(self, dataset_root: Path):
        self.dataset_root = dataset_root

    def resolve(self, reference, *, task=None, role=None) -> ResolvedDataset:
        names = {
            "builtin:regression": "Data_Regression.xlsx",
            "builtin:classification": "Data_Classification.xlsx",
            "builtin:clustering": "Data_Clustering.xlsx",
            "builtin:decomposition": "Data_Decomposition.xlsx",
            "builtin:anomaly_detection": "Data_AnomalyDetection.xlsx",
            "builtin:time_series": "Data_Time_Series.xlsx",
        }
        return ResolvedDataset(
            path=(self.dataset_root / names[reference.dataset_id]).resolve(),
            expected_sha256=None,
            dataset_id=reference.dataset_id,
            source="builtin",
        )

    def resolve_many(self, requests):
        return tuple(self.resolve(reference, task=task, role=role) for reference, task, role in requests)


@pytest.mark.parametrize(
    ("analysis_request", "expected_rows"),
    [
        (
            RegressionRequest(
                training_dataset=BuiltInDatasetReference(dataset_id="builtin:regression"),
                experiment_name="Identifier Root Fix",
                run_name="Regression Builtin",
                identifier_column="SAMPLE NAME",
                feature_columns=("SIO2(WT%)", "TIO2(WT%)"),
                target_column="MGO(WT%)",
                model={"type": "linear_regression"},
            ),
            109,
        ),
        (
            ClassificationRequest(
                training_dataset=BuiltInDatasetReference(dataset_id="builtin:classification"),
                experiment_name="Identifier Root Fix",
                run_name="Classification Builtin",
                identifier_column="SAMPLE NAME",
                feature_columns=("SIO2(WT%)", "TIO2(WT%)"),
                target_column="Label",
                model={"type": "logistic_regression"},
            ),
            2011,
        ),
        (
            ClusteringRequest(
                training_dataset=BuiltInDatasetReference(dataset_id="builtin:clustering"),
                experiment_name="Identifier Root Fix",
                run_name="Clustering Builtin",
                identifier_column="SAMPLE NAME",
                feature_columns=("SIO2(WT%)", "TIO2(WT%)"),
                model={"type": "kmeans", "number_of_clusters": 3},
            ),
            2011,
        ),
        (
            DecompositionRequest(
                training_dataset=BuiltInDatasetReference(dataset_id="builtin:decomposition"),
                experiment_name="Identifier Root Fix",
                run_name="Decomposition Builtin",
                identifier_column="SAMPLE NAME",
                feature_columns=("SIO2(WT%)", "TIO2(WT%)"),
                model={"type": "pca", "number_of_components": 2},
            ),
            109,
        ),
        (
            AnomalyDetectionRequest(
                training_dataset=BuiltInDatasetReference(dataset_id="builtin:anomaly_detection"),
                experiment_name="Identifier Root Fix",
                run_name="Anomaly Builtin",
                identifier_column="SAMPLE NAME",
                feature_columns=("SIO2(WT%)", "TIO2(WT%)"),
                model={"type": "isolation_forest", "contamination": 0.1},
            ),
            109,
        ),
        (
            TimeSeriesRequest(
                training_dataset=BuiltInDatasetReference(dataset_id="builtin:time_series"),
                experiment_name="Identifier Root Fix",
                run_name="Time Series Builtin",
                identifier_column="ROCK NAME",
                selected_columns=(
                    "LATITUDE",
                    "LONGITUDE",
                    "MIN_AGE",
                    "AGE",
                    "MAX_AGE",
                    "R_MIN_AGE",
                    "R_AGE",
                    "R_MAX_AGE",
                    "Estimated Proportion of Subaerial Basalts",
                ),
                missing_values={"method": "drop_rows", "columns": ()},
                probability_column="Estimated Proportion of Subaerial Basalts",
                bin_width=100,
                iterations=100,
                seed=2025,
                fit_curve=False,
            ),
            22640,
        ),
    ],
)
def test_all_six_builtin_modes_pass_validation_with_internal_row_lineage(tmp_path: Path, analysis_request, expected_rows: int) -> None:
    repository = Path(__file__).resolve().parents[3]
    dataset_root = repository / "geochemistrypi" / "data_mining" / "data" / "dataset"
    runs_root = Path(tmp_path.anchor) / "gpi-id-runs" if sys.platform == "win32" else tmp_path / "runs"
    settings = McpSettings(
        runs_root=runs_root,
        cli_executable=Path(sys.executable),
        maximum_dataset_bytes=20 * 1024 * 1024,
    )
    manager = RunManager(
        settings,
        plan_compiler=AnalysisPlanCompiler(),
        cli_resolver=lambda: (Path(sys.executable), CLI_VERSION),
        dataset_catalog=_RepositoryBuiltInCatalog(dataset_root),
    )
    try:
        validation = manager.validate(analysis_request)
    finally:
        manager.close()

    assert validation.valid is True
    assert validation.analysis_process_started is False
    assert validation.source_row_count == expected_rows
    assert validation.row_identity_scheme == "geochemistrypi-mcp-source-row-v1"
    assert validation.row_identity_sha256 is not None
