import hashlib
import json
import os
import subprocess
import sys
import time
from dataclasses import replace
from pathlib import Path

import psutil
import pytest
from geochemistrypi_mcp import CliInteractionDriver, InteractionPlan, InteractionStep, WorkspacePathError
from geochemistrypi_mcp.api.schemas import (
    ArtifactRequirement,
    BuiltInDatasetReference,
    ClassificationRequest,
    ClusteringRequest,
    DecompositionRequest,
    PendingRunResultResponse,
    RandomForestSettings,
    RegressionRequest,
    TimeSeriesRequest,
)
from geochemistrypi_mcp.api.terminal_receipts import MAX_TERMINAL_ERROR_LENGTH, TerminalRunReceipt, normalize_terminal_error, sanitize_terminal_error
from geochemistrypi_mcp.config.constants import CLI_VERSION
from geochemistrypi_mcp.config.settings import McpSettings
from geochemistrypi_mcp.contracts.classification import MODEL_DISPLAY_NAMES as CLASSIFICATION_MODEL_DISPLAY_NAMES
from geochemistrypi_mcp.contracts.classification import MODEL_ORDER as CLASSIFICATION_MODEL_ORDER
from geochemistrypi_mcp.data.catalog import ResolvedDataset
from geochemistrypi_mcp.planning.interaction_plan import AnalysisPlanCompiler, PlanCompilationError
from geochemistrypi_mcp.runtime.cli_capabilities import CliCapabilityProbe
from geochemistrypi_mcp.runtime.cli_driver import CliDriverError, CliRunResult
from geochemistrypi_mcp.runtime.environment import EnvironmentSnapshot
from geochemistrypi_mcp.runtime.runs import InputIntegrityError, RunManager, RunPaths, RunStateError, _normalize_stored_scientific_identity
from openpyxl import Workbook


def _request(
    dataset: Path,
    run_name: str = "Reference Run",
    application_dataset: Path | None = None,
) -> ClassificationRequest:
    return ClassificationRequest(
        training_dataset_path=dataset,
        experiment_name="MCP Contract",
        run_name=run_name,
        identifier_column="SampleID",
        feature_columns=("SIO2",),
        target_column="Label",
        application_dataset_path=application_dataset,
    )


def _dataset(tmp_path: Path) -> Path:
    path = tmp_path / "rocks.csv"
    path.write_text("SampleID,SIO2,Label\nA,50.1,basalt\nB,61.0,granite\n", encoding="utf-8")
    return path


def test_scientific_identity_fallback_is_limited_to_recognized_legacy_records() -> None:
    assert _normalize_stored_scientific_identity({}, legacy_schema=True, include_execution_bound=True, record_label="legacy result",) == {
        "scientific_contract_id": "scientific-contract-v1/legacy",
        "scientific_execution_contract_bound": False,
    }
    with pytest.raises(RunStateError, match="missing its required scientific identity"):
        _normalize_stored_scientific_identity(
            {},
            legacy_schema=False,
            include_execution_bound=True,
            record_label="current result",
        )
    with pytest.raises(RunStateError, match="incomplete scientific identity"):
        _normalize_stored_scientific_identity(
            {"scientific_contract_id": "scientific-contract-v4/example"},
            legacy_schema=True,
            include_execution_bound=True,
            record_label="legacy result",
        )


def test_validate_prepares_nested_external_evaluation_dataset_independently(
    tmp_path: Path,
) -> None:
    training = tmp_path / "training.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Training"
    sheet.append(["TrainID", "F1", "F2", "Target"])
    for index in range(12):
        sheet.append([f"T-{index}", index + 1, index + 2, index + 3])
    workbook.save(training)
    workbook.close()

    evaluation = tmp_path / "evaluation.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Evaluation"
    sheet.append(["EvalID", "F1", "F2", "Target"])
    for index in range(3):
        sheet.append([f"E-{index}", index + 10, index + 20, index + 30])
    workbook.save(evaluation)
    workbook.close()

    request = RegressionRequest(
        training_dataset={
            "source": "path",
            "path": training,
            "preparation": {
                "worksheet": "Training",
                "selected_columns": ("TrainID", "F1", "F2", "Target"),
            },
        },
        experiment_name="External",
        run_name="Prepared",
        identifier_column="TrainID",
        feature_columns=("F1", "F2"),
        target_column="Target",
        scaling="standardization",
        model={
            "type": "extra_trees",
            "number_of_estimators": 10,
            "maximum_depth": None,
            "maximum_features": 2,
        },
        evaluation={
            "mode": "external_labeled",
            "evaluation_dataset": {
                "source": "path",
                "path": evaluation,
                "preparation": {
                    "worksheet": "Evaluation",
                    "selected_columns": ("EvalID", "F1", "F2", "Target"),
                },
            },
            "external_identifier_column": "EvalID",
        },
        reproducibility={"model_seed": 280},
    )
    manager = RunManager(
        McpSettings(
            runs_root=tmp_path / "runs",
            cli_executable=Path(sys.executable),
            maximum_dataset_bytes=1024 * 1024,
        ),
        cli_resolver=lambda: (Path(sys.executable), CLI_VERSION),
    )
    try:
        preview = manager.validate(request)
    finally:
        manager.close()

    assert preview.execution_ready is False
    assert "option:data-mining:--scientific-config" in " ".join(preview.blocking_issues)
    assert preview.source_row_count == 12
    assert preview.application_source_row_count == 3
    assert preview.application_preparation is not None
    assert preview.application_preparation["contract"]["worksheet"] == "Evaluation"
    assert preview.effective_seeds == {"model": 280}


class ScriptPlanCompiler:
    def __init__(self, script: Path):
        self.script = script

    def compile(
        self,
        request: ClassificationRequest | RegressionRequest | ClusteringRequest | DecompositionRequest | TimeSeriesRequest,
        cli_executable: Path | None = None,
        *,
        dataset_context=None,
    ) -> InteractionPlan:
        return InteractionPlan(
            schema_version=1,
            name="test-cli-plan",
            public_command=(sys.executable, "-u", str(self.script), request.experiment_name, request.run_name),
            steps=(InteractionStep("ready", ("READY>",), "continue"),),
        )


class LegacyLifecycleTestDriver(CliInteractionDriver):
    """Exercise process lifecycle fixtures without impersonating CLI automation."""

    def run(self, plan, **kwargs):
        return super().run(
            replace(plan, scientific_execution_contract_json=None),
            **kwargs,
        )


class IncompleteTimingTraceDriver(LegacyLifecycleTestDriver):
    """Corrupt only the child interval after the real child has exited."""

    def run(self, plan, **kwargs):
        result = super().run(plan, **kwargs)
        trace = json.loads(result.trace_path.read_text(encoding="utf-8"))
        trace.pop("cli_finished_at", None)
        result.trace_path.write_text(json.dumps(trace), encoding="utf-8")
        return result


def _manager(
    tmp_path: Path,
    script: Path,
    maximum_pending_runs: int = 8,
    dataset_catalog=None,
    driver_factory=None,
) -> RunManager:
    settings = McpSettings(
        runs_root=tmp_path / "runs",
        cli_executable=Path(sys.executable),
        maximum_dataset_bytes=1024 * 1024,
        maximum_pending_runs=maximum_pending_runs,
    )
    return RunManager(
        settings,
        plan_compiler=ScriptPlanCompiler(script),
        driver_factory=(
            driver_factory
            or (
                lambda: LegacyLifecycleTestDriver(
                    prompt_timeout_seconds=3,
                    process_timeout_seconds=20,
                )
            )
        ),
        cli_resolver=lambda: (Path(sys.executable), CLI_VERSION),
        dataset_catalog=dataset_catalog,
        capability_resolver=lambda _executable, requirements: CliCapabilityProbe(
            tuple(requirements),
            (),
        ),
    )


class FixedDatasetCatalog:
    def __init__(self, path: Path):
        self.path = path

    def resolve(self, reference, *, task=None, role=None) -> ResolvedDataset:
        assert reference.dataset_id == "builtin:classification"
        assert task == "classification"
        assert role == "training"
        path = self.path.resolve()
        return ResolvedDataset(
            path=path,
            expected_sha256=reference.expected_sha256,
            dataset_id=reference.dataset_id,
            source="builtin",
            file_name=path.name,
            catalog_task=task,
            catalog_role=role,
            root_path=path.parent,
            observed_size_bytes=path.stat().st_size,
            observed_sha256=hashlib.sha256(path.read_bytes()).hexdigest(),
        )

    def resolve_many(self, requests):
        return tuple(self.resolve(reference, task=task, role=role) for reference, task, role in requests)


class FixedTimeSeriesDatasetCatalog:
    def __init__(self, path: Path):
        self.path = path

    def resolve(self, reference, *, task=None, role=None) -> ResolvedDataset:
        assert reference.dataset_id == "builtin:time_series"
        assert task == "time_series"
        assert role == "training"
        path = self.path.resolve()
        return ResolvedDataset(
            path=path,
            expected_sha256=reference.expected_sha256,
            dataset_id=reference.dataset_id,
            source="builtin",
            file_name=path.name,
            catalog_task=task,
            catalog_role=role,
            root_path=path.parent,
            observed_size_bytes=path.stat().st_size,
            observed_sha256=hashlib.sha256(path.read_bytes()).hexdigest(),
        )

    def resolve_many(self, requests):
        return tuple(self.resolve(reference, task=task, role=role) for reference, task, role in requests)


class RecordingClassificationCatalog:
    def __init__(self, training: Path, application: Path):
        self.paths = {
            "builtin:classification": training.resolve(),
            "builtin:application_classification": application.resolve(),
        }
        self.calls = 0
        self.fail_if_called = False

    def resolve_many(self, requests):
        if self.fail_if_called:
            raise AssertionError("start_validated must not query the dataset catalog")
        self.calls += 1
        resolved = []
        for reference, task, role in requests:
            assert task == "classification"
            path = self.paths[reference.dataset_id]
            digest = hashlib.sha256(path.read_bytes()).hexdigest()
            if reference.expected_sha256 is not None and reference.expected_sha256 != digest:
                raise ValueError("expected SHA-256 does not match selected input")
            expected_role = "training" if reference.dataset_id == "builtin:classification" else "application"
            assert role == expected_role
            resolved.append(
                ResolvedDataset(
                    path=path,
                    expected_sha256=reference.expected_sha256,
                    dataset_id=reference.dataset_id,
                    source="builtin",
                    file_name=path.name,
                    catalog_task="classification",
                    catalog_role=expected_role,
                    root_path=path.parent,
                    observed_size_bytes=path.stat().st_size,
                    observed_sha256=digest,
                )
            )
        return tuple(resolved)


def _wait_for_state(manager: RunManager, run_id: str, expected: set[str], timeout: float = 10) -> str:
    deadline = time.monotonic() + timeout
    while time.monotonic() < deadline:
        state = manager.get_status(run_id).state
        if state in expected:
            return state
        time.sleep(0.05)
    raise AssertionError(f"Run {run_id} did not reach {expected}; last state was {manager.get_status(run_id).state}")


def _wait_for_pid(path: Path, timeout: float = 5) -> int:
    deadline = time.monotonic() + timeout
    last_value: str | None = None
    while time.monotonic() < deadline:
        try:
            last_value = path.read_text(encoding="utf-8").strip()
        except FileNotFoundError:
            last_value = None
        if last_value and last_value.isdecimal():
            pid = int(last_value)
            if pid > 0:
                return pid
        time.sleep(0.05)
    raise AssertionError(f"PID file {path} did not contain a valid PID; last value was {last_value!r}")


def test_terminal_error_is_whitespace_normalized_and_bounded() -> None:
    sanitized = sanitize_terminal_error(" first line\nsecond line " + "x" * 2000)

    assert sanitized is not None
    assert sanitized.startswith("first line second line")
    assert "\n" not in sanitized
    assert len(sanitized) == MAX_TERMINAL_ERROR_LENGTH


def test_incomplete_child_timing_seals_a_failed_terminal_receipt(
    tmp_path: Path,
) -> None:
    script = tmp_path / "incomplete-timing-cli.py"
    script.write_text(
        """import sys
from pathlib import Path

print('READY>', flush=True)
input()
root = Path('geopi_output') / sys.argv[1] / sys.argv[2]
for name in ('artifacts', 'metrics', 'parameters', 'summary'):
    (root / name).mkdir(parents=True)
""",
        encoding="utf-8",
    )
    manager = _manager(
        tmp_path,
        script,
        driver_factory=lambda: IncompleteTimingTraceDriver(
            prompt_timeout_seconds=3,
            process_timeout_seconds=20,
        ),
    )
    try:
        started = manager.start(_request(_dataset(tmp_path)))
        assert _wait_for_state(manager, started.run_id, {"failed"}) == "failed"

        terminal = manager.get_result(started.run_id)
        repeated = manager.get_result(started.run_id)
        assert isinstance(terminal, TerminalRunReceipt)
        assert terminal == repeated
        assert terminal.state == "failed"
        assert terminal.result_type == "run_state_invalid"
        assert terminal.retryable is False
        assert terminal.analysis_process_started is True
        assert terminal.cli_exit_code == 0
        assert terminal.cli_started_at is None
        assert terminal.cli_finished_at is None
        assert terminal.cli_execution_duration_seconds is None
        assert terminal.error is not None
        assert "incomplete CLI child-process interval" in terminal.error
        durable = json.loads(RunPaths.create(manager.settings.runs_root, started.run_id).status.read_text(encoding="utf-8"))
        assert durable["state"] == "failed"
    finally:
        manager.close()


@pytest.mark.parametrize(
    ("failure", "expected_result_type"),
    (
        ("cli", "cli_execution_failed"),
        ("internal", "internal_error"),
    ),
)
def test_async_failure_receipt_uses_the_public_recovery_taxonomy(
    tmp_path: Path,
    failure: str,
    expected_result_type: str,
) -> None:
    script = tmp_path / "must-not-start.py"
    script.write_text("raise AssertionError('must not start')\n", encoding="utf-8")

    class FailingDriver:
        def run(self, _plan, *, workspace, **_kwargs):
            if failure == "cli":
                raise CliDriverError("synthetic CLI failure", workspace)
            raise AssertionError("synthetic internal failure")

    manager = _manager(
        tmp_path,
        script,
        driver_factory=FailingDriver,
    )
    try:
        started = manager.start(_request(_dataset(tmp_path)))
        assert _wait_for_state(manager, started.run_id, {"failed"}) == "failed"
        status = manager.get_status(started.run_id)
        terminal = manager.get_result(started.run_id)
        repeated = manager.get_result(started.run_id)

        assert status.result_type == expected_result_type
        assert status.retryable is False
        assert isinstance(terminal, TerminalRunReceipt)
        assert terminal.result_type == expected_result_type
        assert terminal.retryable is False
        assert repeated == terminal
    finally:
        manager.close()


def test_long_wrapper_failure_receipt_is_bounded_and_hash_bound(tmp_path: Path) -> None:
    script = tmp_path / "never-started.py"
    script.write_text("raise AssertionError('must not start')\n", encoding="utf-8")
    manager = _manager(tmp_path, script)
    run_id = "run-0123456789abcdef"
    paths = RunPaths.create(manager.settings.runs_root, run_id)
    paths.wrapper.mkdir(parents=True)
    manager._write_status(
        paths,
        {
            "schema_version": 1,
            "run_id": run_id,
            "state": "running",
            "stage": "running",
            "created_at": "2026-08-29T00:00:00+00:00",
            "started_at": None,
            "finished_at": None,
            "progress_message": "Synthetic wrapper state.",
            "error": None,
        },
    )
    raw_error = " first line\nsecond line " + "错误" * 900
    normalized = normalize_terminal_error(raw_error)
    assert normalized is not None

    try:
        manager._finish(paths, "failed", "The wrapper failed.", raw_error)
        status = manager.get_status(run_id)
        terminal = manager.get_result(run_id)

        assert status.error == normalized[:MAX_TERMINAL_ERROR_LENGTH]
        assert status.error_truncated is True
        assert status.error_total_utf8_bytes == len(normalized.encode("utf-8"))
        assert status.error_sha256 == hashlib.sha256(normalized.encode("utf-8")).hexdigest()
        durable_status = json.loads(paths.status.read_text(encoding="utf-8"))
        assert durable_status["error_truncated"] is True
        assert durable_status["error_total_utf8_bytes"] == status.error_total_utf8_bytes
        assert durable_status["error_sha256"] == status.error_sha256
        assert isinstance(terminal, TerminalRunReceipt)
        assert terminal.result_type == "internal_error"
        assert terminal.retryable is False
        assert terminal.error == normalized[:MAX_TERMINAL_ERROR_LENGTH]
        assert terminal.error_truncated is True
        assert terminal.error_total_utf8_bytes == len(normalized.encode("utf-8"))
        assert terminal.error_sha256 == hashlib.sha256(normalized.encode("utf-8")).hexdigest()
        assert (status.error, status.error_truncated, status.error_sha256, status.error_total_utf8_bytes,) == (
            terminal.error,
            terminal.error_truncated,
            terminal.error_sha256,
            terminal.error_total_utf8_bytes,
        )
        assert len(terminal.error) == MAX_TERMINAL_ERROR_LENGTH
        assert terminal.model_dump_json()
    finally:
        manager.close()


def test_validate_analysis_previews_workload_without_creating_run_or_process(tmp_path: Path) -> None:
    script = tmp_path / "never-started.py"
    script.write_text("raise AssertionError('must not start')\n", encoding="utf-8")
    dataset = _dataset(tmp_path)
    manager = _manager(tmp_path, script)
    try:
        analysis_request = _request(dataset).model_copy(update={"model": RandomForestSettings(number_of_estimators=500)})
        preview = manager.validate(analysis_request)
        assert preview.valid is True
        assert preview.task == "classification"
        assert preview.models == ("random_forest",)
        assert preview.estimated_model_count == 1
        assert preview.columns == ("SampleID", "SIO2", "Label")
        assert preview.identifier_column == "SampleID"
        assert preview.feature_columns == ("SIO2",)
        assert preview.target_column == "Label"
        assert preview.resolved_model_parameters["number_of_estimators"] == 500
        assert preview.execution_decisions["evaluation"] == {
            "requested_mode": "cli_default",
            "effective_mode": "internal_holdout",
            "requested_test_ratio": 0.2,
            "effective_test_ratio": 0.2,
            "requested_split_strategy": "cli_default",
            "effective_split_strategy": "stratified_holdout",
            "requested_cross_validation_folds": None,
            "effective_cross_validation_folds": 10,
            "requested_metrics": (),
            "metric_artifact_bindings": {},
            "required_artifact_ids": (),
            "class_order": (),
            "requested_confusion_matrix_normalization": None,
            "effective_confusion_matrix_normalization": "none",
            "requested_metric_average": "auto",
            "effective_metric_average": "auto",
            "requested_positive_label": None,
            "effective_positive_label": None,
        }
        assert preview.execution_decisions["preprocessing"] == {
            "missing_values": {"method": "error"},
            "scaling": "standardization",
            "feature_selection": {"method": "none"},
            "engineered_features": (),
            "label_customization": {"strategy": "encode_original"},
            "world_map": {"enabled": False},
            "target_transformations": {},
            "sample_balancing": "none",
            "metadata_columns": (),
            "feature_engineering": None,
        }
        assert preview.execution_decisions["application"] == {
            "enabled": False,
            "role": "none",
            "training_identifier_column": "SampleID",
            "secondary_identifier_column": None,
            "target_columns": (),
            "label_used_as_feature": False,
        }
        assert preview.training_sha256
        assert preview.validation_id.startswith("val-")
        assert len(preview.validation_id) == 36
        assert len(preview.request_hash) == 64
        assert preview.validation_expires_at
        assert preview.analysis_process_started is False
        assert "inference will be skipped" in " ".join(preview.warnings)
        assert not list(manager.settings.runs_root.glob("run-*"))
        detail_path = manager.settings.service_state_root / "validations" / f"{preview.validation_id}.detail.json"
        assert detail_path.is_file()
        assert json.loads(detail_path.read_text(encoding="utf-8"))["schema_version"] == 2
        dataset.unlink()
        full_detail = manager.get_validation_detail(
            preview.validation_id,
            preview.request_hash,
        )
        assert full_detail.model_dump(mode="json") == preview.model_dump(mode="json")
        tampered_detail = json.loads(detail_path.read_text(encoding="utf-8"))
        tampered_detail["response"]["warnings"] = ["tampered"]
        detail_path.write_text(json.dumps(tampered_detail), encoding="utf-8")
        with pytest.raises(RunStateError, match="detail integrity"):
            manager.get_validation_detail(
                preview.validation_id,
                preview.request_hash,
            )
        assert not list(manager.settings.runs_root.glob("run-*"))
    finally:
        manager.close()


def test_validation_reference_is_stable_tamper_evident_and_starts_the_exact_request(
    tmp_path: Path,
) -> None:
    script = tmp_path / "validated_cli.py"
    script.write_text(
        """import json
import sys
from pathlib import Path

print("READY>", flush=True)
input()
root = Path("geopi_output") / sys.argv[1] / sys.argv[2]
for name in ("artifacts", "metrics", "parameters", "summary"):
    (root / name).mkdir(parents=True)
(root / "artifacts" / "model.joblib").write_bytes(b"real-cli-model")
(root / "metrics" / "Model Score.txt").write_text(json.dumps({"accuracy": 0.75}), encoding="utf-8")
(root / "parameters" / "Model Parameters.txt").write_text(json.dumps({"solver": "lbfgs"}), encoding="utf-8")
(root / "summary" / "Model Score.txt").write_text(json.dumps({"accuracy": 0.75}), encoding="utf-8")
""",
        encoding="utf-8",
    )
    dataset = _dataset(tmp_path)
    manager = _manager(tmp_path, script)
    request = _request(dataset, run_name="Validated")
    try:
        first = manager.validate(request)
        second = manager.validate(request)
        assert first.validation_id == second.validation_id
        assert first.request_hash == second.request_hash
        assert not list(manager.settings.runs_root.glob("run-*"))

        with pytest.raises(RunStateError, match="request hash"):
            manager.start_validated(first.validation_id, "f" * 64)
        assert not list(manager.settings.runs_root.glob("run-*"))

        acknowledgement = manager.start_validated(first.validation_id, first.request_hash)
        assert acknowledgement.started_from_validation is True
        assert acknowledgement.request_hash == first.request_hash
        result = manager.get_result(acknowledgement.run_id, wait_seconds=5)
        assert result.state == "partial_failure"
        assert result.contract_status == "incomplete"
        request_record = json.loads((manager.settings.runs_root / acknowledgement.run_id / "wrapper" / "request.json").read_text(encoding="utf-8"))
        assert request_record["request"] == request.model_dump(mode="json")
        assert request_record["validation"] == {
            "validation_id": first.validation_id,
            "request_hash": first.request_hash,
            "mode": "validation_reference",
        }

        tampered_preview = manager.validate(_request(dataset, run_name="Tampered"))
        receipt_path = manager.settings.service_state_root / "validations" / f"{tampered_preview.validation_id}.json"
        receipt = json.loads(receipt_path.read_text(encoding="utf-8"))
        receipt["request"]["run_name"] = "Changed after validation"
        receipt_path.write_text(json.dumps(receipt), encoding="utf-8")
        with pytest.raises(RunStateError, match="integrity"):
            manager.start_validated(
                tampered_preview.validation_id,
                tampered_preview.request_hash,
            )
    finally:
        manager.close()


def test_multirow_header_validation_receipt_replays_without_mutually_exclusive_default(
    tmp_path: Path,
) -> None:
    source = tmp_path / "compound.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Training"
    sheet.append(["SampleID", "Liquid", "Target"])
    sheet.append([None, "SiO2", "Class"])
    sheet.append(["A", 50.1, "basalt"])
    sheet.append(["B", 61.0, "granite"])
    workbook.save(source)
    workbook.close()

    script = tmp_path / "compound_cli.py"
    script.write_text(
        """from pathlib import Path
import sys
print('READY>', flush=True)
input()
root = Path('geopi_output') / sys.argv[1] / sys.argv[2]
for name in ('artifacts', 'metrics', 'parameters', 'summary'):
    (root / name).mkdir(parents=True)
""",
        encoding="utf-8",
    )
    request = ClassificationRequest(
        training_dataset={
            "source": "path",
            "path": source,
            "preparation": {
                "worksheet": "Training",
                "header_row_indices": (0, 1),
                "header_join_separator": "__",
                "selected_columns": (
                    "SampleID",
                    "Liquid__SiO2",
                    "Target__Class",
                ),
            },
        },
        experiment_name="Compound",
        run_name="Replay",
        identifier_column="SampleID",
        feature_columns=("Liquid__SiO2",),
        target_column="Target__Class",
    )
    manager = _manager(tmp_path, script)
    try:
        preview = manager.validate(request)
        receipt_path = manager.settings.service_state_root / "validations" / f"{preview.validation_id}.json"
        receipt = json.loads(receipt_path.read_text(encoding="utf-8"))
        preparation = receipt["request"]["training_dataset"]["preparation"]
        assert preparation["header_row_indices"] == [0, 1]
        assert "header_row_index" not in preparation

        acknowledgement = manager.start_validated(
            preview.validation_id,
            preview.request_hash,
        )
        assert _wait_for_state(manager, acknowledgement.run_id, {"partial_failure"}) == "partial_failure"
        result = manager.get_result(acknowledgement.run_id)
        assert result.contract_status == "incomplete"
        request_record = json.loads((manager.settings.runs_root / acknowledgement.run_id / "wrapper" / "request.json").read_text(encoding="utf-8"))
        assert result.request_hash == preview.request_hash
        assert request_record["request_hash"] == preview.request_hash
        assert request_record["request"] == receipt["request"]
    finally:
        manager.close()


def test_validation_fails_closed_when_configured_cli_lacks_required_command(
    tmp_path: Path,
) -> None:
    observations = tmp_path / "reference.csv"
    observations.write_text(
        "when,signal,label\n2024-01-01,1.0,1\n2024-01-02,2.0,0\n",
        encoding="utf-8",
    )
    request = TimeSeriesRequest(
        training_dataset_path=observations,
        mode="reference_anomaly_series",
        experiment_name="Reference",
        run_name="Capability Probe",
        time_column="when",
        signal_columns=("signal",),
        reference_label_column="label",
        reference_positive_values=("1",),
        selected_columns=("when", "signal", "label"),
    )
    environment = EnvironmentSnapshot(
        identity_sha256="a" * 64,
        record={
            "schema_version": 1,
            "geochemistrypi": {"version": CLI_VERSION},
            "mcp": {"version": "0.2.1"},
            "python": {"version": "3.9.19"},
            "platform": "test",
            "runtime": {"kind": "test"},
            "dependencies": {},
        },
    )
    manager = RunManager(
        McpSettings(
            runs_root=tmp_path / "runs",
            cli_executable=Path(sys.executable),
            maximum_dataset_bytes=1024 * 1024,
        ),
        cli_resolver=lambda: (Path(sys.executable), CLI_VERSION),
        environment_resolver=lambda _: environment,
    )
    try:
        preview = manager.validate(request)
    finally:
        manager.close()

    assert preview.execution_ready is False
    assert preview.adapter_status == "unavailable"
    assert "command:reference-anomaly-time-series" in " ".join(preview.blocking_issues)


def test_validation_reference_rejects_dataset_change_before_process_creation(
    tmp_path: Path,
) -> None:
    script = tmp_path / "must-not-start.py"
    script.write_text("raise AssertionError('must not start')\n", encoding="utf-8")
    dataset = _dataset(tmp_path)
    manager = _manager(tmp_path, script)
    try:
        preview = manager.validate(_request(dataset))
        dataset.write_text(
            dataset.read_text(encoding="utf-8") + "C,52.0,basalt\n",
            encoding="utf-8",
        )
        with pytest.raises(InputIntegrityError, match="changed since validation"):
            manager.start_validated(preview.validation_id, preview.request_hash)
        assert not list(manager.settings.runs_root.glob("run-*"))
    finally:
        manager.close()


def test_validated_named_inputs_start_without_catalog_requery_and_ignore_unrelated_changes(
    tmp_path: Path,
) -> None:
    script = tmp_path / "validated_cli.py"
    script.write_text(
        """import sys
from pathlib import Path
print('READY>', flush=True)
input()
root = Path('geopi_output') / sys.argv[1] / sys.argv[2]
for name in ('artifacts', 'metrics', 'parameters', 'summary'):
    (root / name).mkdir(parents=True)
""",
        encoding="utf-8",
    )
    training = _dataset(tmp_path)
    application = tmp_path / "application.csv"
    application.write_text(
        "SampleID,SIO2,Label\nC,52.1,basalt\n",
        encoding="utf-8",
    )
    unrelated = tmp_path / "unrelated.csv"
    unrelated.write_text("id,value\n1,2\n", encoding="utf-8")
    catalog = RecordingClassificationCatalog(training, application)
    manager = _manager(tmp_path, script, dataset_catalog=catalog)
    request = ClassificationRequest(
        training_dataset=BuiltInDatasetReference(dataset_id="builtin:classification"),
        application_dataset=BuiltInDatasetReference(dataset_id="builtin:application_classification"),
        experiment_name="MCP Contract",
        run_name="Receipt Only Start",
        identifier_column="SampleID",
        feature_columns=("SIO2",),
        target_column="Label",
    )
    try:
        preview = manager.validate(request)
        assert catalog.calls == 1
        unrelated.write_text("id,value\n1,changed\n", encoding="utf-8")
        catalog.fail_if_called = True

        started = manager.start_validated(
            preview.validation_id,
            preview.request_hash,
        )

        assert started.run_id.startswith("run-")
        assert catalog.calls == 1
    finally:
        manager.close()


def test_validated_named_input_mutation_fails_without_catalog_requery(
    tmp_path: Path,
) -> None:
    script = tmp_path / "must-not-start.py"
    script.write_text("raise AssertionError('must not start')\n", encoding="utf-8")
    training = _dataset(tmp_path)
    application = tmp_path / "application.csv"
    application.write_text(
        "SampleID,SIO2,Label\nC,52.1,basalt\n",
        encoding="utf-8",
    )
    catalog = RecordingClassificationCatalog(training, application)
    manager = _manager(tmp_path, script, dataset_catalog=catalog)
    request = ClassificationRequest(
        training_dataset=BuiltInDatasetReference(dataset_id="builtin:classification"),
        application_dataset=BuiltInDatasetReference(dataset_id="builtin:application_classification"),
        experiment_name="MCP Contract",
        run_name="Selected Mutation",
        identifier_column="SampleID",
        feature_columns=("SIO2",),
        target_column="Label",
    )
    try:
        preview = manager.validate(request)
        training.write_text(
            training.read_text(encoding="utf-8") + "D,49.0,basalt\n",
            encoding="utf-8",
        )
        catalog.fail_if_called = True

        with pytest.raises(InputIntegrityError, match="changed since validation"):
            manager.start_validated(preview.validation_id, preview.request_hash)

        assert catalog.calls == 1
        assert not list(manager.settings.runs_root.glob("run-*"))
    finally:
        manager.close()


def test_validation_receipt_binds_hmac_role_source_root_and_symlink_state(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    script = tmp_path / "must-not-start.py"
    script.write_text("raise AssertionError('must not start')\n", encoding="utf-8")
    training = _dataset(tmp_path)
    application = tmp_path / "application.csv"
    application.write_text(
        "SampleID,SIO2,Label\nC,52.1,basalt\n",
        encoding="utf-8",
    )
    catalog = RecordingClassificationCatalog(training, application)
    manager = _manager(tmp_path, script, dataset_catalog=catalog)
    request = ClassificationRequest(
        training_dataset=BuiltInDatasetReference(dataset_id="builtin:classification"),
        experiment_name="MCP Contract",
        run_name="Receipt Identity",
        identifier_column="SampleID",
        feature_columns=("SIO2",),
        target_column="Label",
    )
    try:
        preview = manager.validate(request)
        catalog.fail_if_called = True
        receipt_path = manager.settings.service_state_root / "validations" / f"{preview.validation_id}.json"
        original = json.loads(receipt_path.read_text(encoding="utf-8"))

        tampered = json.loads(json.dumps(original))
        tampered["training"]["resolution"]["source"] = "desktop"
        receipt_path.write_text(json.dumps(tampered), encoding="utf-8")
        with pytest.raises(RunStateError, match="integrity check failed"):
            manager.start_validated(preview.validation_id, preview.request_hash)

        for field, value, message in (
            ("requested_role", "application", "role or task"),
            ("source", "desktop", "selector is inconsistent"),
            ("dataset_id", "builtin:regression", "selector is inconsistent"),
        ):
            tampered = json.loads(json.dumps(original))
            tampered["training"]["resolution"][field] = value
            unsigned = {key: item for key, item in tampered.items() if key != "integrity_hmac_sha256"}
            tampered["integrity_hmac_sha256"] = manager._receipt_integrity(unsigned)
            receipt_path.write_text(json.dumps(tampered), encoding="utf-8")
            with pytest.raises(RunStateError, match=message):
                manager.start_validated(preview.validation_id, preview.request_hash)

        outside = tmp_path.parent / f"{tmp_path.name}-outside.csv"
        outside.write_text("SampleID,SIO2,Label\nX,1,basalt\n", encoding="utf-8")
        escaped = json.loads(json.dumps(original))
        escaped["training"]["resolution"]["file_name"] = f"../{outside.name}"
        escaped["training"]["resolution"]["canonical_path"] = str(outside.resolve())
        unsigned = {key: item for key, item in escaped.items() if key != "integrity_hmac_sha256"}
        escaped["integrity_hmac_sha256"] = manager._receipt_integrity(unsigned)
        receipt_path.write_text(json.dumps(escaped), encoding="utf-8")
        with pytest.raises(InputIntegrityError, match="escaped"):
            manager.start_validated(preview.validation_id, preview.request_hash)

        receipt_path.write_text(json.dumps(original), encoding="utf-8")
        original_is_symlink = Path.is_symlink
        selected_text = os.path.normcase(os.path.abspath(str(training)))

        def selected_is_symlink(path: Path) -> bool:
            if os.path.normcase(os.path.abspath(str(path))) == selected_text:
                return True
            return original_is_symlink(path)

        monkeypatch.setattr(Path, "is_symlink", selected_is_symlink)
        with pytest.raises(InputIntegrityError, match="symbolic link"):
            manager.start_validated(preview.validation_id, preview.request_hash)

        assert catalog.calls == 1
        assert not list(manager.settings.runs_root.glob("run-*"))
    finally:
        manager.close()


def test_validation_reference_binds_time_series_event_dataset_identity(
    tmp_path: Path,
) -> None:
    observations = tmp_path / "observations.csv"
    observations.write_text(
        "sample time,signal,reference label\n" "2025-01-01,1.0,1\n" "2025-01-03,2.0,0\n",
        encoding="utf-8",
    )
    events = tmp_path / "events.csv"
    events.write_text(
        "event id,event time,accepted\n" "e1,2025-01-04,true\n",
        encoding="utf-8",
    )
    request = TimeSeriesRequest(
        mode="reference_anomaly_series",
        training_dataset_path=observations,
        experiment_name="Reference Series",
        run_name="Event Identity",
        time_column="sample time",
        signal_columns=("signal",),
        reference_label_column="reference label",
        reference_positive_values=("1",),
        reference_label_provenance="archived_external",
        event_dataset_path=events,
        event_time_column="event time",
        event_identifier_column="event id",
        event_filter_column="accepted",
        event_filter_values=("true",),
        association_window_days=5,
        association_direction="before_event",
    )
    manager = RunManager(
        McpSettings(
            runs_root=tmp_path / "runs",
            cli_executable=Path(sys.executable),
            maximum_dataset_bytes=1024 * 1024,
        ),
        cli_resolver=lambda: (Path(sys.executable), CLI_VERSION),
    )
    try:
        preview = manager.validate(request)
        assert preview.event_dataset_path == str(events.resolve())
        assert preview.event_source_sha256 == hashlib.sha256(events.read_bytes()).hexdigest()
        assert preview.event_size_bytes == events.stat().st_size
        receipt_path = manager.settings.service_state_root / "validations" / f"{preview.validation_id}.json"
        receipt = json.loads(receipt_path.read_text(encoding="utf-8"))
        assert receipt["schema_version"] == 2
        assert receipt["event"] == {
            "resolved_path": str(events.resolve()),
            "size_bytes": events.stat().st_size,
            "sha256": preview.event_source_sha256,
            "format": "csv",
        }

        events.write_text(
            events.read_text(encoding="utf-8") + "e2,2025-01-08,false\n",
            encoding="utf-8",
        )
        with pytest.raises(RunStateError, match="event dataset changed since validation"):
            manager.start_validated(preview.validation_id, preview.request_hash)
        assert not list(manager.settings.runs_root.glob("run-*"))
    finally:
        manager.close()


def test_validate_analysis_reports_resolved_multi_targets_in_dataset_order(tmp_path: Path) -> None:
    script = tmp_path / "never-started.py"
    script.write_text("raise AssertionError('must not start')\n", encoding="utf-8")
    dataset = tmp_path / "multi-target.csv"
    dataset.write_text(
        "SampleID,Target,TargetB,SIO2\nA,1.0,10.0,50.1\nB,2.0,20.0,61.0\n",
        encoding="utf-8",
    )
    manager = _manager(tmp_path, script)
    request = RegressionRequest(
        training_dataset_path=dataset,
        experiment_name="MCP Contract",
        run_name="Multi Target Preview",
        identifier_column="SampleID",
        feature_columns=("SIO2",),
        target_columns=("TargetB", "Target"),
    )
    try:
        preview = manager.validate(request)
        assert preview.target_column is None
        assert preview.target_columns == ("Target", "TargetB")
        assert preview.analysis_process_started is False
        assert not manager.settings.runs_root.exists()
    finally:
        manager.close()


def test_validate_time_series_mangles_duplicate_headers_only_for_trusted_builtin(tmp_path: Path) -> None:
    dataset = tmp_path / "time-series.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(
        [
            "R_AGE",
            "R_MAX_AGE",
            "Estimated Proportion of Subaerial Basalts",
            "LATITUDE",
            "LONGITUDE",
            "FEOT",
            "FEOT",
        ]
    )
    worksheet.append([50, 100, 0.6, 10, 20, 8.1, 8.2])
    workbook.save(dataset)
    workbook.close()

    settings = McpSettings(
        runs_root=tmp_path / "runs",
        cli_executable=Path(sys.executable),
        maximum_dataset_bytes=1024 * 1024,
    )
    request_values = {
        "experiment_name": "MCP Contract",
        "run_name": "Trusted Built-in Time Series",
        "bin_width": 100,
        "probability_column": "Estimated Proportion of Subaerial Basalts",
    }
    external_manager = RunManager(
        settings,
        plan_compiler=AnalysisPlanCompiler(),
        cli_resolver=lambda: (Path(sys.executable), CLI_VERSION),
    )
    try:
        with pytest.raises(PlanCompilationError, match="duplicate or colliding column names"):
            external_manager.validate(TimeSeriesRequest(training_dataset_path=dataset, **request_values))
    finally:
        external_manager.close()

    builtin_manager = RunManager(
        settings,
        plan_compiler=AnalysisPlanCompiler(),
        cli_resolver=lambda: (Path(sys.executable), CLI_VERSION),
        dataset_catalog=FixedTimeSeriesDatasetCatalog(dataset.resolve()),
    )
    try:
        preview = builtin_manager.validate(
            TimeSeriesRequest(
                training_dataset=BuiltInDatasetReference(dataset_id="builtin:time_series"),
                **request_values,
            )
        )
        assert preview.training_source == "builtin"
        assert preview.columns[-2:] == ("FEOT", "FEOT.1")
        assert preview.interaction_plan == "time-series-subaerial-proportion-v1"
    finally:
        builtin_manager.close()


def test_validate_continuous_time_series_reports_the_registered_method(tmp_path: Path) -> None:
    dataset = tmp_path / "continuous-time-series.csv"
    dataset.write_text(
        "AGE,MIN AGE,MAX AGE,MGO,SIO2,LATITUDE,LONGITUDE\n" "10,8,12,8.0,43,-20,100\n" "20,18,22,9.0,51,5,110\n" "115,110,120,6.0,48,30,120\n",
        encoding="utf-8",
    )
    settings = McpSettings(
        runs_root=tmp_path / "runs",
        service_state_root=tmp_path / "service-state",
        cli_executable=Path(sys.executable),
        maximum_dataset_bytes=1024 * 1024,
    )
    manager = RunManager(
        settings,
        plan_compiler=AnalysisPlanCompiler(),
        cli_resolver=lambda: (Path(sys.executable), CLI_VERSION),
    )
    try:
        preview = manager.validate(
            TimeSeriesRequest(
                training_dataset_path=dataset,
                mode="continuous",
                age_column="AGE",
                minimum_age_column="MIN AGE",
                maximum_age_column="MAX AGE",
                value_column="MGO",
                filter_column="SIO2",
                filter_minimum=43,
                filter_maximum=51,
                bin_width=100,
                iterations=8,
                seed=2025,
                relative_value_two_sigma=0.04,
                fit_curve=False,
            )
        )
        assert preview.models == ("spatiotemporal_weighted_continuous_bootstrap",)
        assert preview.workflow_mode == "continuous"
        assert preview.method == "spatiotemporal_weighted_continuous_bootstrap"
    finally:
        manager.close()


@pytest.mark.skipif(os.name != "nt", reason="Windows legacy path budget regression")
def test_validate_and_start_reject_unsafe_output_paths_before_creating_a_run(tmp_path: Path) -> None:
    script = tmp_path / "never-started.py"
    script.write_text("raise AssertionError('must not start')\n", encoding="utf-8")
    dataset = _dataset(tmp_path)
    manager = _manager(tmp_path, script)
    original_compile = manager.plan_compiler.compile

    def compile_with_unsafe_output(request, cli_executable=None, *, dataset_context=None):
        plan = original_compile(
            request,
            cli_executable=cli_executable,
            dataset_context=dataset_context,
        )
        return InteractionPlan(
            schema_version=plan.schema_version,
            name=plan.name,
            public_command=plan.public_command,
            steps=plan.steps,
            expected_output_relative_paths=("x" * 260,),
        )

    manager.plan_compiler.compile = compile_with_unsafe_output
    try:
        with pytest.raises(WorkspacePathError, match="before the CLI starts"):
            manager.validate(_request(dataset))
        with pytest.raises(WorkspacePathError, match="No CLI process was started"):
            manager.start(_request(dataset))
        assert not list(manager.settings.runs_root.glob("run-*"))
    finally:
        manager.close()


def test_run_is_non_blocking_atomic_and_references_only_original_cli_outputs(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    script = tmp_path / "successful_cli.py"
    script.write_text(
        """import json
import sys
import time
from pathlib import Path

print("READY>", flush=True)
input()
time.sleep(0.4)
root = Path("geopi_output") / sys.argv[1] / sys.argv[2]
for name in ("artifacts", "metrics", "parameters", "summary"):
    (root / name).mkdir(parents=True)
(root / "artifacts" / "model.joblib").write_bytes(b"real-cli-model")
(root / "metrics" / "Model Score.txt").write_text(json.dumps({"accuracy": 0.75, "f1": 0.73}), encoding="utf-8")
(root / "parameters" / "Model Parameters.txt").write_text(json.dumps({"solver": "lbfgs"}), encoding="utf-8")
(root / "summary" / "Model Score.txt").write_text(json.dumps({"accuracy": 0.75, "f1": 0.73}), encoding="utf-8")
if sys.argv[2] == "Time Series Run":
    (root / "parameters" / "Time Series Parameters.json").write_text(json.dumps({"preprocessing": {"input_row_count": 2, "analysis_row_count": 2, "dropped_row_count": 0}}), encoding="utf-8")
""",
        encoding="utf-8",
    )
    dataset = _dataset(tmp_path)
    monkeypatch.setattr(
        "geochemistrypi_mcp.runtime.runs.probe_cli_capabilities",
        lambda _executable, requirements: CliCapabilityProbe(tuple(requirements), ()),
    )
    manager = _manager(tmp_path, script)
    started = time.monotonic()
    acknowledgement = manager.start(_request(dataset))
    elapsed = time.monotonic() - started

    try:
        assert elapsed < 0.5
        assert acknowledgement.state == "queued"
        result = manager.get_result(
            acknowledgement.run_id,
            wait_seconds=3,
            artifact_offset=0,
            artifact_limit=2,
            artifact_view="all",
        )
        assert result.state == "partial_failure"
        interaction_trace = json.loads(Path(result.interaction_trace).read_text(encoding="utf-8"))
        assert result.cli_started_at == interaction_trace["cli_started_at"]
        assert result.cli_finished_at == interaction_trace["cli_finished_at"]
        assert result.cli_execution_duration_seconds == interaction_trace["cli_execution_duration_seconds"]
        assert result.cli_execution_duration_seconds is not None
        assert result.cli_execution_duration_seconds >= 0.35
        assert result.contract_status == "incomplete"
        assert result.input_hash_verified is True
        assert result.reported_metrics["Model Score.txt"] == {"accuracy": 0.75, "f1": 0.73}
        assert result.artifact_count == 4
        assert result.artifact_view == "all"
        assert result.artifact_view_count == 4
        assert result.canonical_artifact_count == 3
        assert result.summary_mirror_count == 1
        assert result.artifact_offset == 0
        assert result.returned_artifact_count == 2
        assert result.next_artifact_offset == 2
        assert result.artifacts_truncated is True
        assert result.request_hash and len(result.request_hash) == 64
        assert result.canonical_contract_hash and len(result.canonical_contract_hash) == 64
        assert result.compiled_plan_hash and len(result.compiled_plan_hash) == 64
        assert result.provenance_manifest_path is not None
        assert result.provenance_manifest_sha256 is not None
        assert all(len(item.sha256) == 64 for item in result.artifacts)
        assert all(Path(item.local_path).is_relative_to(Path(result.output_directory)) for item in result.artifacts)
        second_page = manager.get_result(
            acknowledgement.run_id,
            artifact_offset=2,
            artifact_limit=2,
            artifact_view="all",
        )
        assert second_page.returned_artifact_count == 2
        assert second_page.cli_started_at == result.cli_started_at
        assert second_page.cli_finished_at == result.cli_finished_at
        assert second_page.cli_execution_duration_seconds == result.cli_execution_duration_seconds
        assert second_page.next_artifact_offset is None
        assert second_page.artifacts_truncated is False
        assert {item.category for item in (*result.artifacts, *second_page.artifacts)} == {
            "artifacts",
            "metrics",
            "parameters",
            "summary",
        }
        wrapper = tmp_path / "runs" / acknowledgement.run_id / "wrapper"
        canonical = manager.get_result(
            acknowledgement.run_id,
            artifact_limit=10,
            artifact_view="canonical",
        )
        assert canonical.artifact_count == 4
        assert canonical.artifact_view_count == 3
        assert canonical.returned_artifact_count == 3
        assert canonical.next_artifact_offset is None
        assert canonical.artifacts_truncated is False
        assert {item.category for item in canonical.artifacts} == {"artifacts", "metrics", "parameters"}
        assert canonical.result_record_path == str(wrapper / "result.json")
        assert canonical.result_record_sha256 == hashlib.sha256((wrapper / "result.json").read_bytes()).hexdigest()
        assert canonical.artifact_index_path == str(wrapper / "artifact-index.json")
        assert canonical.artifact_index_sha256 == hashlib.sha256((wrapper / "artifact-index.json").read_bytes()).hexdigest()
        assert canonical.artifact_index_sha256 == result.artifact_index_sha256 == second_page.artifact_index_sha256
        assert canonical.result_record_sha256 == result.result_record_sha256 == second_page.result_record_sha256
        durable_status = json.loads((wrapper / "status.json").read_text(encoding="utf-8"))
        assert durable_status["result_record_path"] == str(wrapper / "result.json")
        assert durable_status["result_record_sha256"] == canonical.result_record_sha256
        manifest_path = Path(result.provenance_manifest_path)
        assert manifest_path == wrapper / "scientific-run-manifest.json"
        assert hashlib.sha256(manifest_path.read_bytes()).hexdigest() == result.provenance_manifest_sha256
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        assert manifest["request_identity"]["request_hash"] == result.request_hash
        assert manifest["run_identity"] == {"run_id": acknowledgement.run_id}
        assert manifest["artifacts"]
        assert all(len(item["sha256"]) == 64 for item in manifest["artifacts"])
        assert json.loads((wrapper / "status.json").read_text(encoding="utf-8"))["state"] == "partial_failure"
        assert json.loads((wrapper / "artifact-index.json").read_text(encoding="utf-8"))["artifacts"]
        assert not list(wrapper.glob("*.tmp"))

        clustering = manager.start(
            ClusteringRequest(
                task="clustering",
                training_dataset_path=dataset,
                experiment_name="MCP Contract",
                run_name="Clustering Run",
                identifier_column="SampleID",
                feature_columns=("SIO2",),
            )
        )
        # The synthetic CLI does not fit KMeans or emit the newly required
        # Scientific Execution Attestation.  Keep the product contract strict
        # and assert the truthful terminal state instead of removing that
        # required artifact from the plan.
        assert _wait_for_state(manager, clustering.run_id, {"partial_failure"}) == "partial_failure"
        clustering_result = manager.get_result(clustering.run_id)
        assert clustering_result.contract_status == "incomplete"
        assert clustering_result.task == "clustering"
        assert clustering_result.model == "kmeans"
        assert clustering_result.tuning == "not_applicable"
        assert clustering_result.application_input_sha256 is None

        decomposition = manager.start(
            DecompositionRequest(
                task="decomposition",
                training_dataset_path=dataset,
                experiment_name="MCP Contract",
                run_name="Decomposition Run",
                identifier_column="SampleID",
                feature_columns=("SIO2",),
            )
        )
        # PCA now keeps its public v4 scientific-execution sidecar instead of
        # silently downgrading.  This synthetic CLI does not emit the required
        # attestation, so it must not be published as a complete result.
        assert _wait_for_state(manager, decomposition.run_id, {"partial_failure"}) == "partial_failure"
        decomposition_result = manager.get_result(decomposition.run_id)
        assert decomposition_result.contract_status == "incomplete"
        assert decomposition_result.task == "decomposition"
        assert decomposition_result.model == "pca"
        assert decomposition_result.tuning == "not_applicable"
        assert decomposition_result.reported_metrics["Model Score.txt"] == {
            "accuracy": 0.75,
            "f1": 0.73,
        }

        overlay_coordinates = tmp_path / "overlay-coordinates.csv"
        overlay_coordinates.write_text(
            "SampleID,PC1,PC2\nA,1.0,10.0\nB,2.0,20.0\n",
            encoding="utf-8",
        )
        overlay_labels = tmp_path / "overlay-labels.csv"
        overlay_labels.write_text(
            "RecordID,Anomaly\nA,-1\nB,1\n",
            encoding="utf-8",
        )
        overlay = manager.start(
            DecompositionRequest(
                training_dataset_path=overlay_coordinates,
                application_dataset_path=overlay_labels,
                mode="embedding_label_overlay",
                experiment_name="MCP Contract",
                run_name="Embedding Overlay Run",
                identifier_column="SampleID",
                feature_columns=("PC1", "PC2"),
                scaling="none",
                label_identifier_column="RecordID",
                label_column="Anomaly",
                positive_label_values=("-1",),
            )
        )
        assert _wait_for_state(manager, overlay.run_id, {"succeeded"}) == "succeeded"
        overlay_result = manager.get_result(overlay.run_id)
        assert overlay_result.contract_status == "complete"
        assert overlay_result.task == "decomposition"
        assert overlay_result.model == "embedding_label_overlay"
        assert overlay_result.tuning == "not_applicable"

        time_series = manager.start(
            TimeSeriesRequest(
                training_dataset_path=dataset,
                experiment_name="MCP Contract",
                run_name="Time Series Run",
                bin_width=10,
            )
        )
        assert _wait_for_state(manager, time_series.run_id, {"succeeded"}) == "succeeded"
        time_series_result = manager.get_result(time_series.run_id)
        assert time_series_result.contract_status == "complete"
        assert time_series_result.task == "time_series"
        assert time_series_result.model == "subaerial_proportion_bootstrap"
        assert time_series_result.tuning == "not_applicable"
        assert time_series_result.preprocessing_summary is not None
        assert time_series_result.preprocessing_summary.model_dump() == {
            "input_row_count": 2,
            "analysis_row_count": 2,
            "dropped_row_count": 0,
        }

        (wrapper / "result.json").write_text("{}\n", encoding="utf-8")
        with pytest.raises(RunStateError, match="inconsistent with durable run status"):
            manager.get_result(acknowledgement.run_id)
        status_path = wrapper / "status.json"
        durable_status = json.loads(status_path.read_text(encoding="utf-8"))
        durable_status["result_record_sha256"] = hashlib.sha256((wrapper / "result.json").read_bytes()).hexdigest()
        status_path.write_text(json.dumps(durable_status), encoding="utf-8")
        with pytest.raises(
            RunStateError,
            match="public result contract|required scientific identity",
        ):
            manager.get_result(acknowledgement.run_id)
    finally:
        manager.close()


def test_builtin_reference_is_resolved_before_snapshot_and_recorded(
    tmp_path: Path,
) -> None:
    script = tmp_path / "builtin_cli.py"
    script.write_text(
        """import sys
from pathlib import Path
print('READY>', flush=True)
input()
root = Path('geopi_output') / sys.argv[1] / sys.argv[2]
for name in ('artifacts', 'metrics', 'parameters', 'summary'):
    (root / name).mkdir(parents=True)
""",
        encoding="utf-8",
    )
    dataset = _dataset(tmp_path)
    before = dataset.read_bytes()
    manager = _manager(
        tmp_path,
        script,
        dataset_catalog=FixedDatasetCatalog(dataset.resolve()),
    )
    request = ClassificationRequest(
        training_dataset=BuiltInDatasetReference(dataset_id="builtin:classification"),
        experiment_name="MCP Contract",
        run_name="Built-in Source",
        identifier_column="SampleID",
        feature_columns=("SIO2",),
        target_column="Label",
    )

    started = manager.start(request)
    assert _wait_for_state(manager, started.run_id, {"partial_failure", "failed"}) == "partial_failure"
    assert manager.get_result(started.run_id).contract_status == "incomplete"
    record = json.loads((tmp_path / "runs" / started.run_id / "wrapper" / "request.json").read_text(encoding="utf-8"))

    assert record["request"]["training_dataset"]["dataset_id"] == "builtin:classification"
    assert record["input"]["source"] == "builtin"
    assert record["input"]["dataset_id"] == "builtin:classification"
    assert dataset.read_bytes() == before
    manager.close()


def test_all_models_partial_failure_is_terminal_and_result_remains_available(
    tmp_path: Path,
) -> None:
    models = [CLASSIFICATION_MODEL_DISPLAY_NAMES[model] for model in CLASSIFICATION_MODEL_ORDER]
    script = tmp_path / "aggregate_cli.py"
    source = """import json
import sys
from pathlib import Path

print('READY>', flush=True)
input()
root = Path('geopi_output') / sys.argv[1] / sys.argv[2]
for name in ('artifacts', 'metrics', 'parameters', 'summary'):
    (root / name).mkdir(parents=True)
models = __MODELS__
children = []
for index, model in enumerate(models):
    child = root / model
    child.mkdir()
    failed = index == 1
    children.append({
        'model': model,
        'state': 'failed' if failed else 'succeeded',
        'output_relative_path': model,
        'artifact_count': 0,
        'error': 'bounded child failure' if failed else None,
    })
manifest = {
    'schema_version': 1,
    'task': 'classification',
    'selection_mode': 'all',
    'tuning': 'manual',
    'state': 'partial_failure',
    'expected_model_count': len(models),
    'succeeded_count': len(models) - 1,
    'failed_count': 1,
    'children': children,
}
(root / 'summary' / 'Aggregate Model Results.json').write_text(
    json.dumps(manifest), encoding='utf-8'
)
""".replace(
        "__MODELS__", repr(models)
    )
    script.write_text(source, encoding="utf-8")
    manager = _manager(tmp_path, script)
    request = ClassificationRequest(
        training_dataset_path=_dataset(tmp_path),
        experiment_name="MCP Contract",
        run_name="All Models Partial",
        identifier_column="SampleID",
        feature_columns=("SIO2",),
        target_column="Label",
        model_selection={"mode": "all", "tuning": "manual"},
    )

    started = manager.start(request)
    try:
        assert _wait_for_state(manager, started.run_id, {"partial_failure"}) == "partial_failure"
        result = manager.get_result(started.run_id)
        assert result.state == "partial_failure"
        assert result.model == "all_models"
        assert result.aggregate_state == "partial_failure"
        assert result.aggregate_summary.model_dump() == {
            "expected_model_count": len(CLASSIFICATION_MODEL_ORDER),
            "succeeded_count": len(CLASSIFICATION_MODEL_ORDER) - 1,
            "failed_count": 1,
        }
        assert len(result.children) == len(models)
        assert [child.model for child in result.children] == models
        assert result.children[1].state == "failed"
        assert result.children[1].error == "bounded child failure"
    finally:
        manager.close()


def test_cancellation_terminates_recorded_tree_and_preserves_unrelated_process(tmp_path: Path) -> None:
    script = tmp_path / "long_cli.py"
    script.write_text(
        """import subprocess
import sys
import time
from pathlib import Path

child = subprocess.Popen([sys.executable, "-c", "import time; time.sleep(60)"])
pid_path = Path("child.pid")
pid_path.touch()
time.sleep(0.5)
pid_path.write_text(str(child.pid), encoding="utf-8")
print("READY>", flush=True)
input()
time.sleep(60)
""",
        encoding="utf-8",
    )
    unrelated = subprocess.Popen([sys.executable, "-c", "import time; time.sleep(60)"])
    manager = _manager(tmp_path, script)
    acknowledgement = manager.start(_request(_dataset(tmp_path)))
    workspace = tmp_path / "runs" / acknowledgement.run_id / "workspace"
    try:
        assert _wait_for_state(manager, acknowledgement.run_id, {"running"}) == "running"
        pending = manager.get_result(acknowledgement.run_id)
        assert isinstance(pending, PendingRunResultResponse)
        assert pending.state == "running"
        assert pending.stage == "running_cli"
        assert pending.result_available is False
        assert pending.requery_required is True
        child_pid = _wait_for_pid(workspace / "child.pid")
        response = manager.cancel(acknowledgement.run_id)
        assert response.state == "cancellation_requested"
        assert _wait_for_state(manager, acknowledgement.run_id, {"cancelled"}) == "cancelled"
        terminal = manager.get_result(acknowledgement.run_id)
        assert isinstance(terminal, TerminalRunReceipt)
        assert terminal.state == terminal.stage == "cancelled"
        assert terminal.analysis_process_started is True
        assert isinstance(terminal.cli_exit_code, int)
        assert terminal.cli_started_at is not None
        assert terminal.cli_finished_at is not None
        assert terminal.cli_execution_duration_seconds is not None
        assert terminal.scientific_validity == "not_established"
        assert terminal.artifact_contract_status == "not_evaluated"
        assert terminal.verified_artifact_count == 0
        assert terminal.interaction_trace is not None
        assert terminal.cli_stdout_log is not None
        assert terminal.cli_stderr_log is not None
        assert "artifacts" not in terminal.model_dump(mode="json")
        deadline = time.monotonic() + 5
        while psutil.pid_exists(child_pid) and time.monotonic() < deadline:
            time.sleep(0.05)
        assert not psutil.pid_exists(child_pid)
        assert unrelated.poll() is None
    finally:
        manager.close()
        unrelated.terminate()
        unrelated.wait(timeout=5)


def test_input_change_during_execution_fails_integrity_check(tmp_path: Path) -> None:
    script = tmp_path / "integrity_cli.py"
    script.write_text(
        """import sys
import time
from pathlib import Path

print("READY>", flush=True)
input()
time.sleep(0.5)
root = Path("geopi_output") / sys.argv[1] / sys.argv[2]
for name in ("artifacts", "metrics", "parameters", "summary"):
    (root / name).mkdir(parents=True)
""",
        encoding="utf-8",
    )
    dataset = _dataset(tmp_path)
    manager = _manager(tmp_path, script)
    acknowledgement = manager.start(_request(dataset, run_name="Integrity"))
    try:
        assert _wait_for_state(manager, acknowledgement.run_id, {"running"}) == "running"
        dataset.write_text(dataset.read_text(encoding="utf-8") + "C,55.0,andesite\n", encoding="utf-8")
        assert _wait_for_state(manager, acknowledgement.run_id, {"failed"}) == "failed"
        status = manager.get_status(acknowledgement.run_id)
        assert "changed during CLI execution" in status.error
        assert status.result_type == "input_integrity_changed"
        assert status.retryable is False
        terminal = manager.get_result(acknowledgement.run_id)
        assert isinstance(terminal, TerminalRunReceipt)
        assert terminal.state == terminal.stage == "failed"
        assert terminal.error == status.error
        assert terminal.result_type == "input_integrity_changed"
        assert terminal.retryable is False
        assert terminal.analysis_process_started is True
        assert terminal.cli_exit_code == 0
        trace = json.loads(Path(terminal.interaction_trace.path).read_text(encoding="utf-8"))
        assert terminal.cli_started_at == trace["cli_started_at"]
        assert terminal.cli_finished_at == trace["cli_finished_at"]
        assert terminal.cli_execution_duration_seconds == trace["cli_execution_duration_seconds"]
        assert terminal.cli_execution_duration_seconds is not None
        assert terminal.cli_execution_duration_seconds >= 0.45
        assert terminal.scientific_validity == "not_established"
        assert terminal.artifact_contract_status == "not_evaluated"
        assert terminal.verified_artifact_count == 0
        assert terminal.interaction_trace is not None
        assert terminal.cli_stdout_log is not None
        assert terminal.cli_stderr_log is not None
        result_path = Path(terminal.result_record_path)
        assert terminal.result_record_sha256 == hashlib.sha256(result_path.read_bytes()).hexdigest()
        stored_receipt = json.loads(result_path.read_text(encoding="utf-8"))
        assert "result_record_path" not in stored_receipt
        assert "result_record_sha256" not in stored_receipt
        durable_status = json.loads(result_path.with_name("status.json").read_text(encoding="utf-8"))
        assert durable_status["result_record_path"] == str(result_path)
        assert durable_status["result_record_sha256"] == terminal.result_record_sha256
        with pytest.raises(ValueError, match="Extra inputs are not permitted"):
            TerminalRunReceipt.model_validate({**terminal.model_dump(mode="json"), "artifacts": []})

        original_result = result_path.read_bytes()
        paths = RunPaths.create(manager.settings.runs_root, acknowledgement.run_id)
        with pytest.raises(RunStateError, match="cannot be replaced"):
            manager._finish(paths, "cancelled", "A later transition must not replace the failed receipt.")
        assert result_path.read_bytes() == original_result
        assert manager.get_status(acknowledgement.run_id).state == "failed"
        repeated = manager.get_result(acknowledgement.run_id)
        assert isinstance(repeated, TerminalRunReceipt)
        assert repeated.cli_started_at == terminal.cli_started_at
        assert repeated.cli_finished_at == terminal.cli_finished_at
        assert repeated.cli_execution_duration_seconds == terminal.cli_execution_duration_seconds
        assert repeated.result_type == terminal.result_type
        assert repeated.retryable == terminal.retryable
    finally:
        manager.close()


def test_application_input_change_during_execution_fails_integrity_check(tmp_path: Path) -> None:
    script = tmp_path / "application_integrity_cli.py"
    script.write_text(
        """import sys
import time
from pathlib import Path

print("READY>", flush=True)
input()
time.sleep(0.5)
root = Path("geopi_output") / sys.argv[1] / sys.argv[2]
for name in ("artifacts", "metrics", "parameters", "summary"):
    (root / name).mkdir(parents=True)
""",
        encoding="utf-8",
    )
    training = _dataset(tmp_path)
    application = tmp_path / "application.csv"
    application.write_text("SampleID,SIO2\nP-1,55.0\n", encoding="utf-8")
    manager = _manager(tmp_path, script)
    acknowledgement = manager.start(_request(training, run_name="Application Integrity", application_dataset=application))
    try:
        assert _wait_for_state(manager, acknowledgement.run_id, {"running"}) == "running"
        application.write_text(application.read_text(encoding="utf-8") + "P-2,57.0\n", encoding="utf-8")
        assert _wait_for_state(manager, acknowledgement.run_id, {"failed"}) == "failed"
        status = manager.get_status(acknowledgement.run_id)
        assert "application dataset changed during cli execution" in status.error.lower()
        terminal = manager.get_result(acknowledgement.run_id)
        assert isinstance(terminal, TerminalRunReceipt)
        assert terminal.state == "failed"
        assert terminal.result_type == "input_integrity_changed"
        assert terminal.retryable is False
        assert terminal.error == status.error
        assert terminal.verified_artifact_count == 0
    finally:
        manager.close()


def test_recovery_never_terminates_a_process_from_stale_pid_metadata(tmp_path: Path) -> None:
    runs_root = tmp_path / "runs"
    run_id = "run-0123456789abcdef"
    wrapper = runs_root / run_id / "wrapper"
    wrapper.mkdir(parents=True)
    abandoned_success_record = b'{"response_detail":"full","state":"succeeded"}\n'
    (wrapper / "result.json").write_bytes(abandoned_success_record)
    unrelated = subprocess.Popen([sys.executable, "-c", "import time; time.sleep(60)"])
    (wrapper / "status.json").write_text(
        json.dumps(
            {
                "schema_version": 1,
                "run_id": run_id,
                "state": "running",
                "created_at": "2026-08-02T12:00:00+00:00",
                "started_at": "2026-08-02T12:00:01+00:00",
                "finished_at": None,
                "cli_pid": unrelated.pid,
                "recorded_cli_pid": unrelated.pid,
                "recorded_process_create_time": psutil.Process(unrelated.pid).create_time(),
                "progress_message": "running",
                "error": None,
            }
        ),
        encoding="utf-8",
    )
    settings = McpSettings(runs_root=runs_root, cli_executable=Path(sys.executable))
    manager = RunManager(settings, cli_resolver=lambda: (Path(sys.executable), CLI_VERSION))
    try:
        status = manager.get_status(run_id)
        assert status.state == "failed"
        assert "no stale PID was terminated" in status.error
        assert unrelated.poll() is None
        terminal = manager.get_result(run_id)
        assert isinstance(terminal, TerminalRunReceipt)
        assert terminal.state == terminal.stage == "failed"
        assert terminal.result_type == "internal_error"
        assert terminal.retryable is False
        assert terminal.analysis_process_started is True
        assert terminal.cli_exit_code is None
        assert terminal.interaction_trace is None
        assert terminal.cli_stdout_log is None
        assert terminal.cli_stderr_log is None
        assert Path(terminal.result_record_path) == wrapper / "terminal-result.json"
        assert (wrapper / "result.json").read_bytes() == abandoned_success_record
        with pytest.raises(RunStateError, match="already failed"):
            manager.cancel(run_id)
    finally:
        manager.close()
        unrelated.terminate()
        unrelated.wait(timeout=5)


def test_run_queue_fails_closed_at_capacity_and_reopens_after_cancellation(tmp_path: Path) -> None:
    script = tmp_path / "queue_cli.py"
    script.write_text(
        """import time

print("READY>", flush=True)
input()
time.sleep(60)
""",
        encoding="utf-8",
    )
    manager = _manager(tmp_path, script, maximum_pending_runs=1)
    dataset = _dataset(tmp_path)
    first = manager.start(_request(dataset, run_name="First"))
    try:
        assert _wait_for_state(manager, first.run_id, {"running"}) == "running"
        with pytest.raises(RunStateError, match="run queue is full"):
            manager.start(_request(dataset, run_name="Rejected"))
        assert len(tuple((tmp_path / "runs").glob("run-*"))) == 1

        manager.cancel(first.run_id)
        assert _wait_for_state(manager, first.run_id, {"cancelled"}) == "cancelled"
        replacement = manager.start(_request(dataset, run_name="Replacement"))
        assert replacement.state == "queued"
    finally:
        manager.close()


def test_missing_required_artifact_is_an_explicit_partial_failure(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    script = tmp_path / "missing_required_output_cli.py"
    script.write_text(
        """import json
import sys
from pathlib import Path

print('READY>', flush=True)
input()
root = Path('geopi_output') / sys.argv[1] / sys.argv[2]
for name in ('artifacts', 'metrics', 'parameters', 'summary'):
    (root / name).mkdir(parents=True)
(root / 'metrics' / 'Observed Score.json').write_text(
    json.dumps({'accuracy': 0.75}), encoding='utf-8'
)
""",
        encoding="utf-8",
    )

    class MissingArtifactDriver:
        def run(self, plan, *, workspace, capture_directory, **_kwargs):
            output_root = workspace / "geopi_output"
            root = output_root / "MCP Contract" / "Missing Required Output"
            for name in ("artifacts", "metrics", "parameters", "summary"):
                (root / name).mkdir(parents=True)
            (root / "metrics" / "Observed Score.json").write_text(
                json.dumps({"accuracy": 0.75}),
                encoding="utf-8",
            )
            capture_directory.mkdir(parents=True, exist_ok=True)
            stdout_path = capture_directory / "stdout.log"
            stderr_path = capture_directory / "stderr.log"
            trace_path = capture_directory / "interaction-trace.json"
            stdout_path.write_text("completed\n", encoding="utf-8")
            stderr_path.write_text("", encoding="utf-8")
            trace_path.write_text(
                json.dumps(
                    {
                        "schema_version": 1,
                        "returncode": 0,
                        "cli_started_at": "2026-08-30T00:00:00+00:00",
                        "cli_finished_at": "2026-08-30T00:00:01+00:00",
                        "cli_execution_duration_seconds": 1.0,
                    }
                ),
                encoding="utf-8",
            )
            return CliRunResult(
                workspace=workspace,
                output_root=output_root,
                stdout_path=stdout_path,
                stderr_path=stderr_path,
                trace_path=trace_path,
                returncode=0,
                completed_step_ids=tuple(step.id for step in plan.steps),
            )

    class MissingOutputPlanCompiler(ScriptPlanCompiler):
        def compile(self, *args, **kwargs):
            return replace(
                super().compile(*args, **kwargs),
                expected_output_relative_paths=("artifacts/Required Scientific Output.pdf",),
            )

    settings = McpSettings(
        runs_root=tmp_path / "runs",
        cli_executable=Path(sys.executable),
        maximum_dataset_bytes=1024 * 1024,
    )
    manager = RunManager(
        settings,
        plan_compiler=MissingOutputPlanCompiler(script),
        driver_factory=MissingArtifactDriver,
        cli_resolver=lambda: (Path(sys.executable), CLI_VERSION),
    )
    monkeypatch.setattr(
        "geochemistrypi_mcp.runtime.runs.probe_cli_capabilities",
        lambda _executable, requirements: CliCapabilityProbe(tuple(requirements), ()),
    )
    request = _request(_dataset(tmp_path), run_name="Missing Required Output").model_copy(
        update={
            "artifact_requirements": (
                ArtifactRequirement(
                    requirement_id="required.missing-pdf",
                    scientific_type="scientific_figure",
                    output_role="scientific.output",
                    category="artifacts",
                    expected_relative_path="artifacts/Required Scientific Output.pdf",
                ),
            )
        }
    )

    started = manager.start(request)
    try:
        assert _wait_for_state(manager, started.run_id, {"partial_failure", "failed"}) == "partial_failure"
        status = manager.get_status(started.run_id)
        result = manager.get_result(started.run_id)

        assert status.state == "partial_failure"
        assert "required artifacts are missing" in status.progress_message
        assert result.state == "partial_failure"
        assert result.contract_status == "incomplete"
        artifact_contract = json.loads((manager.settings.runs_root / started.run_id / "wrapper" / "artifact-index.json").read_text(encoding="utf-8"))["artifact_contract"]
        attestation_ids = tuple(requirement["requirement_id"] for requirement in artifact_contract["required"] if requirement["scientific_type"] == "parameter_attestation")
        assert result.missing_artifact_requirement_ids == (
            "required.missing-pdf",
            *attestation_ids,
        )
        assert len(attestation_ids) == 1
        assert result.cli_exit_code == 0
        assert result.artifact_count == 1
        durable = json.loads((tmp_path / "runs" / started.run_id / "wrapper" / "status.json").read_text(encoding="utf-8"))
        assert durable["state"] == "partial_failure"
    finally:
        manager.close()


def test_queued_cancellation_publishes_a_terminal_receipt_without_process_evidence(tmp_path: Path) -> None:
    script = tmp_path / "queued_cli.py"
    script.write_text(
        """import time

print("READY>", flush=True)
input()
time.sleep(60)
""",
        encoding="utf-8",
    )
    manager = _manager(tmp_path, script, maximum_pending_runs=2)
    dataset = _dataset(tmp_path)
    running = manager.start(_request(dataset, run_name="Occupies Worker"))
    queued = None
    try:
        assert _wait_for_state(manager, running.run_id, {"running"}) == "running"
        queued = manager.start(_request(dataset, run_name="Still Queued"))
        assert manager.get_status(queued.run_id).state == "queued"
        pending = manager.get_result(queued.run_id)
        assert isinstance(pending, PendingRunResultResponse)
        assert pending.state == pending.stage == "queued"
        assert pending.result_available is False
        assert pending.requery_required is True

        cancellation = manager.cancel(queued.run_id)
        assert cancellation.state == "cancelled"
        terminal = manager.get_result(queued.run_id)
        assert isinstance(terminal, TerminalRunReceipt)
        assert terminal.state == terminal.stage == "cancelled"
        assert terminal.started_at is None
        assert terminal.analysis_process_started is False
        assert terminal.cli_exit_code is None
        assert terminal.cli_started_at is None
        assert terminal.cli_finished_at is None
        assert terminal.cli_execution_duration_seconds is None
        assert terminal.interaction_trace is None
        assert terminal.cli_stdout_log is None
        assert terminal.cli_stderr_log is None
        assert terminal.verified_artifact_count == 0
    finally:
        if manager.get_status(running.run_id).state not in {"failed", "cancelled", "succeeded", "partial_failure"}:
            manager.cancel(running.run_id)
            _wait_for_state(manager, running.run_id, {"cancelled"})
        manager.close()


def test_default_driver_uses_the_configured_total_process_timeout(tmp_path: Path) -> None:
    settings = McpSettings(
        runs_root=tmp_path / "runs",
        cli_executable=Path(sys.executable),
        maximum_process_seconds=37,
    )
    manager = RunManager(settings, cli_resolver=lambda: (Path(sys.executable), CLI_VERSION))
    try:
        assert manager.driver_factory().process_timeout_seconds == 37
        assert manager.driver_factory().automation_mode is True
    finally:
        manager.close()
