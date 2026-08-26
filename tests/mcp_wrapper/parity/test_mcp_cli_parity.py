import asyncio
import csv
import hashlib
import json
import os
import subprocess
import sys
import tempfile
import time
from dataclasses import replace
from pathlib import Path, PurePosixPath
from typing import Any

import pytest
from geochemistrypi_mcp import (
    AnalysisPlanCompiler,
    AnomalyDetectionPlanCompiler,
    AnomalyDetectionRequest,
    ClassificationPlanCompiler,
    ClassificationRequest,
    ClusteringPlanCompiler,
    ClusteringRequest,
    DecompositionPlanCompiler,
    DecompositionRequest,
    RegressionPlanCompiler,
    RegressionRequest,
    TimeSeriesPlanCompiler,
    TimeSeriesRequest,
)
from geochemistrypi_mcp.config.constants import CLI_VERSION, ISOLATED_CLI_ENVIRONMENT_VARIABLES, SERVER_VERSION
from geochemistrypi_mcp.config.settings import resolve_cli_interpreter
from mcp import Client, StdioServerParameters
from mcp.client.stdio import stdio_client
from openpyxl import load_workbook

REPOSITORY_ROOT = Path(__file__).resolve().parents[3]
FIXTURE_DIRECTORY = REPOSITORY_ROOT / "tests" / "cli_contract" / "fixtures"
DATASET_PATH = FIXTURE_DIRECTORY / "classification_baseline.csv"
INTERACTION_PATH = FIXTURE_DIRECTORY / "classification_interaction_v1.json"
MANIFEST_PATH = FIXTURE_DIRECTORY / "classification_output_manifest_v1.json"
REGRESSION_DATASET_PATH = FIXTURE_DIRECTORY / "regression_baseline.csv"
CLUSTERING_DATASET_PATH = FIXTURE_DIRECTORY / "clustering_baseline.csv"
DECOMPOSITION_DATASET_PATH = FIXTURE_DIRECTORY / "decomposition_baseline.csv"
ANOMALY_DETECTION_DATASET_PATH = FIXTURE_DIRECTORY / "anomaly_detection_baseline.csv"
CLUSTERING_METRIC_RELATIVE_TOLERANCE = 1e-4
CLUSTERING_METRIC_ABSOLUTE_TOLERANCE = 1e-8


@pytest.fixture
def anyio_backend() -> str:
    return "asyncio"


def _load_json(path: Path) -> dict[str, Any]:
    with path.open(encoding="utf-8") as stream:
        return json.load(stream)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _expected_output_files(manifest: dict[str, Any]) -> list[str]:
    primary_files = [f"artifacts/{path}" for path in manifest["artifacts"]]
    primary_files.extend(f"metrics/{path}" for path in manifest["metrics"])
    primary_files.extend(f"parameters/{path}" for path in manifest["parameters"])
    summary_names = [PurePosixPath(path).name for path in primary_files]
    return sorted(primary_files + [f"summary/{name}" for name in summary_names])


def _all_files(run_directory: Path) -> list[str]:
    return sorted(path.relative_to(run_directory).as_posix() for path in run_directory.rglob("*") if path.is_file())


def _worksheet_values(path: Path) -> list[tuple[Any, ...]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(workbook.active.iter_rows(values_only=True))
    finally:
        workbook.close()


def _with_tracking_root(plan: Any, tracking_root: Path):
    """Keep direct CLI parity runs out of the user's default MLflow store."""
    if "data-mining" not in plan.public_command:
        return plan
    root = tracking_root.resolve()
    root.mkdir(parents=True, exist_ok=True)
    assert "--tracking-root" not in plan.public_command
    return replace(
        plan,
        public_command=(*plan.public_command, "--tracking-root", str(root)),
    )


def _stdio_environment(
    cli_executable: Path,
    root: Path,
    *,
    tracking_root: Path | None = None,
) -> dict[str, str]:
    """Give every stdio parity server a complete isolated user-state root."""
    state_root = root.resolve()
    return {
        "GEOCHEMISTRYPI_CLI_EXECUTABLE": str(cli_executable),
        "GEOCHEMISTRYPI_MCP_APP_ROOT": str(state_root / "app"),
        "GEOCHEMISTRYPI_MCP_RUNS_ROOT": str(state_root / "runs"),
        "GEOCHEMISTRYPI_MCP_TRACKING_ROOT": str((tracking_root or state_root / "tracking").resolve()),
        "GEOCHEMISTRYPI_MCP_SERVICE_STATE_ROOT": str(state_root / "service-state"),
        "MPLBACKEND": "Agg",
    }


@pytest.mark.anyio
@pytest.mark.mcp_cli_parity
async def test_stdio_mcp_attaches_a_real_run_to_an_existing_experiment(
    request: pytest.FixtureRequest,
) -> None:
    cli_executable = Path(os.environ["GEOCHEMISTRYPI_CLI_EXECUTABLE"]).resolve()
    cli_interpreter = resolve_cli_interpreter(cli_executable)
    temporary_root = tempfile.TemporaryDirectory(prefix="g9h-")
    request.addfinalizer(temporary_root.cleanup)
    parity_root = Path(temporary_root.name)
    tracking_root = (parity_root / "tracking").resolve()
    cli_environment = os.environ.copy()
    for inherited_name in ISOLATED_CLI_ENVIRONMENT_VARIABLES:
        cli_environment.pop(inherited_name, None)
    create = subprocess.run(
        (
            str(cli_interpreter),
            "-c",
            "import mlflow,sys; from pathlib import Path; "
            "root=Path(sys.argv[1]); root.mkdir(parents=True, exist_ok=True); "
            "mlflow.set_tracking_uri(root.as_uri()); "
            "print(mlflow.create_experiment('PR9H Existing Experiment'))",
            str(tracking_root),
        ),
        check=False,
        capture_output=True,
        text=True,
        encoding="utf-8",
        env=cli_environment,
        timeout=60,
    )
    assert create.returncode == 0, create.stdout + create.stderr
    experiment_id = create.stdout.strip()
    analysis = ClassificationRequest(
        training_dataset_path=DATASET_PATH,
        experiment_name="PR9H Existing Experiment",
        existing_experiment_id=experiment_id,
        run_name="Attached Through MCP",
        identifier_column="SampleID",
        feature_columns=(
            "SIO2(WT%)",
            "TIO2(WT%)",
            "AL2O3(WT%)",
            "FEOT(WT%)",
            "CAO(WT%)",
            "MGO(WT%)",
            "NA2O(WT%)",
        ),
        target_column="Label",
    )
    parameters = StdioServerParameters(
        command=sys.executable,
        args=["-m", "geochemistrypi_mcp"],
        env=_stdio_environment(
            cli_executable,
            parity_root,
            tracking_root=tracking_root,
        ),
    )
    async with Client(stdio_client(parameters)) as client:
        listed = await client.call_tool("list_experiments", {"maximum_experiments": 10})
        assert listed.is_error is False
        assert any(value["experiment_id"] == experiment_id for value in listed.structured_content["experiments"])
        before = await client.call_tool(
            "get_experiment",
            {"experiment_id": experiment_id, "maximum_runs": 10},
        )
        assert before.is_error is False
        assert before.structured_content["run_count"] == 0

        started = await client.call_tool(
            "start_analysis",
            analysis.model_dump(mode="json"),
        )
        assert started.is_error is False, started.content[0].text
        run_id = started.structured_content["run_id"]
        deadline = time.monotonic() + 360
        while True:
            status = await client.call_tool("get_run_status", {"run_id": run_id})
            assert status.is_error is False
            if status.structured_content["state"] in {
                "succeeded",
                "failed",
                "cancelled",
            }:
                break
            assert time.monotonic() < deadline
            await asyncio.sleep(0.25)
        assert status.structured_content["state"] == "succeeded", status.structured_content

        after = await client.call_tool(
            "get_experiment",
            {"experiment_id": experiment_id, "maximum_runs": 10},
        )
        assert after.is_error is False
        assert after.structured_content["run_count"] == 1
        assert after.structured_content["runs"][0]["run_name"] == analysis.run_name


@pytest.mark.anyio
@pytest.mark.mcp_cli_parity
async def test_stdio_mcp_time_series_matches_noninteractive_public_cli(
    tmp_path: Path,
) -> None:
    cli_executable = Path(os.environ["GEOCHEMISTRYPI_CLI_EXECUTABLE"]).resolve()
    source = tmp_path / "time-series.csv"
    source.write_text(
        "R_AGE,R_MAX_AGE,SBAP,LATITUDE,LONGITUDE\n" "10,12,0.9,-20,100\n" "20,25,0.1,5,110\n" "35,40,0.8,30,120\n" "50,55,0.2,45,130\n",
        encoding="utf-8",
    )
    request = TimeSeriesRequest(
        training_dataset_path=source,
        experiment_name="PR9F Time Series Parity",
        run_name="Seeded Bootstrap",
        bin_width=10,
        iterations=5,
        seed=7,
        fit_curve=False,
    )
    plan = TimeSeriesPlanCompiler().compile(request, cli_executable=cli_executable)
    plan = _with_tracking_root(plan, tmp_path / "direct-time-series-tracking")
    assert plan.steps == ()
    direct_workspace = tmp_path / "direct-time-series"
    direct_workspace.mkdir()
    direct_environment = os.environ.copy()
    for inherited_name in ISOLATED_CLI_ENVIRONMENT_VARIABLES:
        direct_environment.pop(inherited_name, None)
    direct_environment.update(
        {
            "COLUMNS": "200",
            "LINES": "60",
            "MPLBACKEND": "Agg",
            "PYTHONHASHSEED": "0",
            "PYTHONIOENCODING": "utf-8",
            "TERM": "dumb",
        }
    )
    direct = subprocess.run(
        plan.public_command,
        cwd=direct_workspace,
        env=direct_environment,
        capture_output=True,
        encoding="utf-8",
        errors="replace",
        timeout=180,
    )
    assert direct.returncode == 0, direct.stdout + direct.stderr

    parameters = StdioServerParameters(
        command=sys.executable,
        args=["-m", "geochemistrypi_mcp"],
        env=_stdio_environment(cli_executable, tmp_path / "time-series-state"),
    )
    async with Client(stdio_client(parameters)) as client:
        started = await client.call_tool("start_analysis", request.model_dump(mode="json"))
        assert started.is_error is False, started.content[0].text
        run_id = started.structured_content["run_id"]
        deadline = time.monotonic() + 180
        while True:
            status = await client.call_tool("get_run_status", {"run_id": run_id})
            assert status.is_error is False
            state = status.structured_content["state"]
            if state in {"succeeded", "partial_failure", "failed", "cancelled"}:
                break
            assert time.monotonic() < deadline
            await asyncio.sleep(0.2)
        assert state == "succeeded", status.structured_content
        result_call = await client.call_tool("get_run_result", {"run_id": run_id})
        assert result_call.is_error is False
        result = result_call.structured_content

    direct_run = direct_workspace / "geopi_output" / request.experiment_name / request.run_name
    wrapped_run = Path(result["output_directory"])
    assert result["task"] == "time_series"
    assert result["model"] == "subaerial_proportion_bootstrap"
    assert result["tuning"] == "not_applicable"
    assert result["input_sha256"] == _sha256(source)
    assert result["preprocessing_summary"] == {
        "input_row_count": 4,
        "analysis_row_count": 4,
        "dropped_row_count": 0,
    }
    assert _all_files(direct_run) == _all_files(wrapped_run)
    assert (direct_run / "artifacts" / "data" / "Subaerial Proportion.csv").read_bytes() == (wrapped_run / "artifacts" / "data" / "Subaerial Proportion.csv").read_bytes()
    assert _load_json(direct_run / "metrics" / "Time Series Metrics.json") == _load_json(wrapped_run / "metrics" / "Time Series Metrics.json")
    assert result["reported_metrics"]["Time Series Metrics.json"]["populated_bins"] > 0
    assert result["artifact_count"] == len(_all_files(wrapped_run))


@pytest.mark.anyio
@pytest.mark.mcp_cli_parity
async def test_stdio_mcp_embedding_label_overlay_matches_direct_public_cli(
    tmp_path: Path,
) -> None:
    cli_executable = Path(os.environ["GEOCHEMISTRYPI_CLI_EXECUTABLE"]).resolve()
    coordinates = tmp_path / "coordinates.csv"
    coordinates.write_text(
        "SampleID,PC1,PC2\nC,3.0,30.0\nA,1.0,10.0\nB,2.0,20.0\n",
        encoding="utf-8",
    )
    labels = tmp_path / "labels.csv"
    labels.write_text(
        "RecordID,Anomaly\nB,-1\nC,1\nA,-1\n",
        encoding="utf-8",
    )
    request = DecompositionRequest(
        training_dataset_path=coordinates,
        application_dataset_path=labels,
        mode="embedding_label_overlay",
        experiment_name="Artifact Composition Parity",
        run_name="Embedding Overlay",
        identifier_column="SampleID",
        feature_columns=("PC1", "PC2"),
        scaling="none",
        label_identifier_column="RecordID",
        label_column="Anomaly",
        positive_label_values=("-1",),
    )
    plan = AnalysisPlanCompiler().compile(request, cli_executable=cli_executable)
    plan = _with_tracking_root(plan, tmp_path / "direct-overlay-tracking")
    assert plan.steps == ()
    direct_workspace = tmp_path / "direct-overlay"
    direct_workspace.mkdir()
    direct_environment = os.environ.copy()
    for inherited_name in ISOLATED_CLI_ENVIRONMENT_VARIABLES:
        direct_environment.pop(inherited_name, None)
    direct_environment.update({"MPLBACKEND": "Agg", "PYTHONIOENCODING": "utf-8"})
    direct = subprocess.run(
        plan.public_command,
        cwd=direct_workspace,
        env=direct_environment,
        capture_output=True,
        encoding="utf-8",
        errors="replace",
        timeout=180,
    )
    assert direct.returncode == 0, direct.stdout + direct.stderr

    parameters = StdioServerParameters(
        command=sys.executable,
        args=["-m", "geochemistrypi_mcp"],
        env=_stdio_environment(cli_executable, tmp_path / "overlay-state"),
    )
    async with Client(stdio_client(parameters)) as client:
        started = await client.call_tool(
            "start_analysis",
            request.model_dump(mode="json"),
        )
        assert started.is_error is False, started.content[0].text
        run_id = started.structured_content["run_id"]
        deadline = time.monotonic() + 180
        while True:
            status = await client.call_tool("get_run_status", {"run_id": run_id})
            assert status.is_error is False
            state = status.structured_content["state"]
            if state in {"succeeded", "partial_failure", "failed", "cancelled"}:
                break
            assert time.monotonic() < deadline
            await asyncio.sleep(0.2)
        assert state == "succeeded", status.structured_content
        result_call = await client.call_tool("get_run_result", {"run_id": run_id})
        assert result_call.is_error is False
        result = result_call.structured_content

    direct_run = direct_workspace / "geopi_output" / request.experiment_name / request.run_name
    wrapped_run = Path(result["output_directory"])
    assert result["task"] == "decomposition"
    assert result["model"] == "embedding_label_overlay"
    assert result["tuning"] == "not_applicable"
    assert result["input_sha256"] == _sha256(coordinates)
    assert _all_files(direct_run) == _all_files(wrapped_run)
    assert (direct_run / "artifacts" / "data" / "Embedding Label Overlay.csv").read_bytes() == (wrapped_run / "artifacts" / "data" / "Embedding Label Overlay.csv").read_bytes()
    assert _load_json(direct_run / "metrics" / "Embedding Label Overlay Counts.json") == _load_json(wrapped_run / "metrics" / "Embedding Label Overlay Counts.json")
    assert result["reported_metrics"]["Embedding Label Overlay Counts.json"]["anomaly_count"] == 2


@pytest.mark.anyio
@pytest.mark.mcp_cli_parity
async def test_stdio_mcp_anomaly_all_models_matches_real_cli_aggregate(
    request: pytest.FixtureRequest,
) -> None:
    cli_executable = Path(os.environ["GEOCHEMISTRYPI_CLI_EXECUTABLE"]).resolve()
    temporary_root = tempfile.TemporaryDirectory(prefix="g9-")
    request.addfinalizer(temporary_root.cleanup)
    parity_root = Path(temporary_root.name)
    source = parity_root / "anomaly-all-models.csv"
    rows = ["SampleID,F1,F2,F3"]
    for index in range(50):
        rows.append(f"A{index:02d},{index + 1},{(index % 7) + 2},{(index % 11) + 3}")
    source.write_text("\n".join(rows) + "\n", encoding="utf-8")
    analysis_request = AnomalyDetectionRequest(
        training_dataset_path=source,
        experiment_name="E",
        run_name="R",
        identifier_column="SampleID",
        feature_columns=("F1", "F2", "F3"),
        model_selection={"mode": "all", "tuning": "manual"},
    )
    plan = AnalysisPlanCompiler().compile(analysis_request, cli_executable=cli_executable)
    plan = _with_tracking_root(plan, parity_root / "direct-tracking")
    assert next(step for step in plan.steps if step.id == "all_models").response == "3"
    direct_workspace = parity_root / "direct-anomaly-all"
    direct_workspace.mkdir()
    direct_environment = os.environ.copy()
    for inherited_name in ISOLATED_CLI_ENVIRONMENT_VARIABLES:
        direct_environment.pop(inherited_name, None)
    direct_environment.update(
        {
            "COLUMNS": "200",
            "LINES": "60",
            "MPLBACKEND": "Agg",
            "PYTHONHASHSEED": "0",
            "PYTHONIOENCODING": "utf-8",
            "TERM": "dumb",
        }
    )
    direct = subprocess.run(
        plan.public_command,
        cwd=direct_workspace,
        env=direct_environment,
        input="\n".join(step.response for step in plan.steps) + "\n",
        capture_output=True,
        encoding="utf-8",
        errors="replace",
        timeout=420,
    )
    assert direct.returncode == 0, direct.stdout[-12000:] + direct.stderr[-12000:]

    parameters = StdioServerParameters(
        command=sys.executable,
        args=["-m", "geochemistrypi_mcp"],
        env=_stdio_environment(cli_executable, parity_root / "mcp-state"),
    )
    mcp_payload = analysis_request.model_dump(mode="json")
    mcp_payload.pop("model", None)
    mcp_payload.pop("tuning", None)
    async with Client(stdio_client(parameters)) as client:
        started = await client.call_tool(
            "start_analysis",
            mcp_payload,
        )
        assert started.is_error is False, started.content[0].text
        run_id = started.structured_content["run_id"]
        deadline = time.monotonic() + 420
        while True:
            status = await client.call_tool("get_run_status", {"run_id": run_id})
            assert status.is_error is False
            state = status.structured_content["state"]
            if state in {"succeeded", "partial_failure", "failed", "cancelled"}:
                break
            assert time.monotonic() < deadline
            await asyncio.sleep(0.25)
        assert state == "succeeded", status.structured_content
        result_call = await client.call_tool("get_run_result", {"run_id": run_id})
        assert result_call.is_error is False
        result = result_call.structured_content

    direct_run = direct_workspace / "geopi_output" / analysis_request.experiment_name / analysis_request.run_name
    wrapped_run = Path(result["output_directory"])
    direct_manifest = _load_json(direct_run / "summary" / "Aggregate Model Results.json")
    wrapped_manifest = _load_json(wrapped_run / "summary" / "Aggregate Model Results.json")
    assert result["model"] == "all_models"
    assert result["aggregate_state"] == "complete"
    assert [child["model"] for child in result["children"]] == [
        "Isolation Forest",
        "Local Outlier Factor",
    ]
    assert all(child["state"] == "succeeded" for child in result["children"])
    assert direct_manifest["state"] == wrapped_manifest["state"] == "complete"
    assert direct_manifest["succeeded_count"] == wrapped_manifest["succeeded_count"] == 2
    assert _all_files(direct_run) == _all_files(wrapped_run)
    assert result["artifact_count"] == len(_all_files(wrapped_run))


@pytest.mark.anyio
@pytest.mark.mcp_cli_parity
async def test_stdio_mcp_lists_and_inspects_every_installed_builtin_dataset(
    tmp_path: Path,
) -> None:
    cli_executable = Path(os.environ["GEOCHEMISTRYPI_CLI_EXECUTABLE"]).resolve()
    parameters = StdioServerParameters(
        command=sys.executable,
        args=["-m", "geochemistrypi_mcp"],
        env=_stdio_environment(cli_executable, tmp_path / "dataset-state"),
    )

    async with Client(stdio_client(parameters)) as client:
        listed = await client.call_tool("list_datasets", {"source": "builtin"})
        assert listed.is_error is False
        datasets = listed.structured_content["datasets"]
        assert len(datasets) == 8
        inspections = []
        for dataset in datasets:
            inspected = await client.call_tool(
                "inspect_dataset",
                {
                    "dataset": {
                        "source": "builtin",
                        "dataset_id": dataset["dataset_id"],
                        "expected_sha256": dataset["sha256"],
                    },
                    "sample_rows": 0,
                },
            )
            assert inspected.is_error is False, inspected.content[0].text
            inspections.append(inspected.structured_content)

    assert {item["sha256"] for item in inspections} == {item["sha256"] for item in datasets}
    time_series = next(item for item in inspections if item["source_path"].endswith("Data_Time_Series.xlsx"))
    assert any("FEOT.1" in warning for warning in time_series["header_warnings"])


@pytest.mark.anyio
@pytest.mark.mcp_cli_parity
async def test_stdio_mcp_classification_application_feature_engineering_matches_direct_cli(
    tmp_path: Path,
) -> None:
    cli_executable = Path(os.environ["GEOCHEMISTRYPI_CLI_EXECUTABLE"]).resolve()
    fixture_hash_before = _sha256(DATASET_PATH)
    feature_columns = (
        "SIO2(WT%)",
        "TIO2(WT%)",
        "AL2O3(WT%)",
        "FEOT(WT%)",
        "CAO(WT%)",
        "MGO(WT%)",
        "NA2O(WT%)",
    )
    application_path = tmp_path / "classification-application.csv"
    with DATASET_PATH.open(encoding="utf-8", newline="") as source:
        rows = list(csv.DictReader(source))[:8]
    with application_path.open("w", encoding="utf-8", newline="") as destination:
        writer = csv.DictWriter(destination, fieldnames=("SampleID", *feature_columns))
        writer.writeheader()
        writer.writerows({column: row[column] for column in writer.fieldnames} for row in rows)

    request = ClassificationRequest(
        training_dataset_path=DATASET_PATH,
        application_dataset_path=application_path,
        experiment_name="PR3 Application Parity",
        run_name="Feature Engineering",
        identifier_column="SampleID",
        feature_columns=feature_columns,
        target_column="Label",
        engineered_features=(
            {
                "name": "SIO2_TIO2_RATIO",
                "formula": "{SIO2(WT%)} / ({TIO2(WT%)} + 1)",
            },
        ),
    )
    plan = ClassificationPlanCompiler().compile(
        request,
        cli_executable=cli_executable,
    )
    plan = _with_tracking_root(plan, tmp_path / "direct-tracking")
    direct_workspace = tmp_path / "direct"
    direct_workspace.mkdir()
    direct_environment = os.environ.copy()
    for inherited_name in ISOLATED_CLI_ENVIRONMENT_VARIABLES:
        direct_environment.pop(inherited_name, None)
    direct_environment.update(
        {
            "COLUMNS": "200",
            "LINES": "60",
            "MPLBACKEND": "Agg",
            "PYTHONHASHSEED": "0",
            "PYTHONIOENCODING": "utf-8",
            "TERM": "dumb",
        }
    )
    direct = subprocess.run(
        plan.public_command,
        cwd=direct_workspace,
        env=direct_environment,
        input="\n".join(step.response for step in plan.steps) + "\n",
        capture_output=True,
        encoding="utf-8",
        errors="replace",
        timeout=360,
    )
    assert direct.returncode == 0, direct.stdout[-12000:] + direct.stderr[-12000:]

    with tempfile.TemporaryDirectory(prefix="gpi-pr8b-inference-") as short_root:
        server_parameters = StdioServerParameters(
            command=sys.executable,
            args=["-m", "geochemistrypi_mcp"],
            env=_stdio_environment(cli_executable, Path(short_root)),
        )
        async with Client(stdio_client(server_parameters)) as client:
            started = await client.call_tool(
                "start_analysis",
                request.model_dump(mode="json"),
            )
            assert started.is_error is False
            run_id = started.structured_content["run_id"]
            deadline = time.monotonic() + 420
            while True:
                status = await client.call_tool(
                    "get_run_status",
                    {"run_id": run_id},
                )
                assert status.is_error is False
                state = status.structured_content["state"]
                if state in {"succeeded", "failed", "cancelled"}:
                    break
                assert time.monotonic() < deadline
                await asyncio.sleep(0.25)
            assert state == "succeeded", status.structured_content
            result_call = await client.call_tool(
                "get_run_result",
                {"run_id": run_id},
            )
            assert result_call.is_error is False
            result = result_call.structured_content
            healthy = await client.call_tool("get_capabilities", {})
            assert healthy.is_error is False

        direct_run = direct_workspace / "geopi_output" / request.experiment_name / request.run_name
        wrapped_run = Path(result["output_directory"])
        direct_data = direct_run / "artifacts" / "data"
        wrapped_data = wrapped_run / "artifacts" / "data"

        assert result["task"] == "classification"
        assert result["model"] == "logistic_regression"
        assert result["application_input_sha256"] == _sha256(application_path)
        assert result["application_input_hash_verified"] is True
        assert _all_files(direct_run) == _all_files(wrapped_run)
        for artifact_name in (
            "Application Data Feature-Engineering.xlsx",
            "Application Data Feature-Engineering Selected.xlsx",
            "Application Data Predicted.xlsx",
        ):
            assert _worksheet_values(direct_data / artifact_name) == _worksheet_values(wrapped_data / artifact_name)
        assert len(_worksheet_values(wrapped_data / "Application Data Predicted.xlsx")) == len(rows) + 1
        assert result["artifact_count"] == len(_all_files(wrapped_run))
    assert _sha256(DATASET_PATH) == fixture_hash_before


@pytest.mark.anyio
@pytest.mark.mcp_cli_parity
async def test_stdio_mcp_matches_direct_public_cli_and_preserves_protocol() -> None:
    cli_executable = Path(os.environ["GEOCHEMISTRYPI_CLI_EXECUTABLE"]).resolve()
    interaction = _load_json(INTERACTION_PATH)
    manifest = _load_json(MANIFEST_PATH)
    fixture_hash_before = _sha256(DATASET_PATH)
    request = ClassificationRequest(
        training_dataset_path=DATASET_PATH,
        experiment_name=interaction["experiment_name"],
        run_name=interaction["run_name"],
        identifier_column="SampleID",
        feature_columns=(
            "SIO2(WT%)",
            "TIO2(WT%)",
            "AL2O3(WT%)",
            "FEOT(WT%)",
            "CAO(WT%)",
            "MGO(WT%)",
            "NA2O(WT%)",
        ),
        target_column="Label",
    )
    plan = ClassificationPlanCompiler().compile(request, cli_executable=cli_executable)
    assert [step.response for step in plan.steps] == [step["response"] for step in interaction["steps"]]

    with tempfile.TemporaryDirectory(prefix="gpi-pr2-") as temporary_root:
        parity_root = Path(temporary_root)
        plan = _with_tracking_root(plan, parity_root / "direct-tracking")
        direct_workspace = parity_root / "direct"
        direct_workspace.mkdir()
        direct_environment = os.environ.copy()
        for inherited_name in ISOLATED_CLI_ENVIRONMENT_VARIABLES:
            direct_environment.pop(inherited_name, None)
        direct_environment.update(
            {
                "COLUMNS": "200",
                "LINES": "60",
                "MPLBACKEND": "Agg",
                "PYTHONHASHSEED": "0",
                "PYTHONIOENCODING": "utf-8",
                "TERM": "dumb",
            }
        )
        direct = subprocess.run(
            plan.public_command,
            cwd=direct_workspace,
            env=direct_environment,
            input="\n".join(step["response"] for step in interaction["steps"]) + "\n",
            capture_output=True,
            encoding="utf-8",
            errors="replace",
            timeout=300,
        )
        assert direct.returncode == 0, direct.stdout[-12000:] + direct.stderr[-12000:]

        server_parameters = StdioServerParameters(
            command=sys.executable,
            args=["-m", "geochemistrypi_mcp"],
            env=_stdio_environment(cli_executable, parity_root / "mcp-state"),
        )
        async with Client(stdio_client(server_parameters)) as client:
            capabilities = await client.call_tool("get_capabilities", {})
            assert capabilities.is_error is False
            started_at = time.monotonic()
            started = await client.call_tool("start_analysis", request.model_dump(mode="json"))
            assert time.monotonic() - started_at < 5
            assert started.is_error is False
            run_id = started.structured_content["run_id"]
            deadline = time.monotonic() + 360
            while True:
                status = await client.call_tool("get_run_status", {"run_id": run_id})
                assert status.is_error is False
                state = status.structured_content["state"]
                if state in {"succeeded", "failed", "cancelled"}:
                    break
                assert time.monotonic() < deadline
                await asyncio.sleep(0.25)
            assert state == "succeeded", status.structured_content
            result_call = await client.call_tool("get_run_result", {"run_id": run_id})
            assert result_call.is_error is False
            result = result_call.structured_content
            protocol_still_healthy = await client.call_tool("get_capabilities", {})
            assert protocol_still_healthy.is_error is False

        direct_run = direct_workspace / "geopi_output" / request.experiment_name / request.run_name
        wrapped_run = Path(result["output_directory"])
        expected_files = _expected_output_files(manifest)
        assert _sha256(DATASET_PATH) == fixture_hash_before == result["input_sha256"]
        assert result["input_hash_verified"] is True
        assert _all_files(direct_run) == expected_files
        assert _all_files(wrapped_run) == expected_files
        for directory in ("artifacts", "metrics", "parameters", "summary"):
            assert (wrapped_run / directory).is_dir()
        assert _load_json(direct_run / "metrics" / "Model Score - Logistic Regression.txt") == _load_json(wrapped_run / "metrics" / "Model Score - Logistic Regression.txt")
        assert _load_json(direct_run / "metrics" / "Cross Validation - Logistic Regression.txt") == _load_json(wrapped_run / "metrics" / "Cross Validation - Logistic Regression.txt")
        assert _load_json(direct_run / "parameters" / "Hyper Parameters - Logistic Regression.txt") == _load_json(wrapped_run / "parameters" / "Hyper Parameters - Logistic Regression.txt")
        assert _worksheet_values(direct_run / "artifacts" / "data" / "Y Test.xlsx") == _worksheet_values(wrapped_run / "artifacts" / "data" / "Y Test.xlsx")
        assert _worksheet_values(direct_run / "artifacts" / "data" / "Y Test Predict Decoded.xlsx") == _worksheet_values(wrapped_run / "artifacts" / "data" / "Y Test Predict Decoded.xlsx")
        assert result["artifact_count"] == len(expected_files)
        assert Path(result["interaction_trace"]).is_file()
        assert Path(result["cli_stdout_log"]).is_file()
        assert Path(result["cli_stderr_log"]).is_file()
        assert _load_json(Path(result["interaction_trace"]))["metadata"] == {
            "geochemistrypi_mcp_version": SERVER_VERSION,
            "geochemistrypi_cli_version": CLI_VERSION,
        }


@pytest.mark.anyio
@pytest.mark.mcp_cli_parity
async def test_stdio_mcp_regression_matches_direct_cli_with_application_data() -> None:
    cli_executable = Path(os.environ["GEOCHEMISTRYPI_CLI_EXECUTABLE"]).resolve()
    fixture_hash_before = _sha256(REGRESSION_DATASET_PATH)

    with tempfile.TemporaryDirectory(prefix="gpi-pr5-regression-parity-") as temporary_root:
        parity_root = Path(temporary_root)
        training_path = parity_root / "regression-multi-target.csv"
        application_path = parity_root / "regression-application.csv"
        with REGRESSION_DATASET_PATH.open(encoding="utf-8", newline="") as source:
            training_rows = list(csv.DictReader(source))
        with training_path.open("w", encoding="utf-8", newline="") as destination:
            fieldnames = ("SampleID", "Target", "TargetB", "SIO2", "TIO2")
            writer = csv.DictWriter(destination, fieldnames=fieldnames)
            writer.writeheader()
            for row in training_rows:
                writer.writerow(
                    {
                        **{column: row[column] for column in ("SampleID", "Target", "SIO2", "TIO2")},
                        "TargetB": float(row["Target"]) * 0.5 + float(row["TIO2"]),
                    }
                )
        with application_path.open("w", encoding="utf-8", newline="") as destination:
            writer = csv.DictWriter(destination, fieldnames=("SampleID", "SIO2", "TIO2"))
            writer.writeheader()
            writer.writerows({column: row[column] for column in writer.fieldnames} for row in training_rows[:5])

        request = RegressionRequest(
            task="regression",
            training_dataset_path=training_path,
            application_dataset_path=application_path,
            experiment_name="PR5 Regression Parity",
            run_name="Linear Regression",
            identifier_column="SampleID",
            feature_columns=("SIO2", "TIO2"),
            target_columns=("Target", "TargetB"),
            model={"type": "linear_regression"},
        )
        plan = RegressionPlanCompiler().compile(request, cli_executable=cli_executable)
        plan = _with_tracking_root(plan, parity_root / "direct-tracking")

        direct_workspace = parity_root / "direct"
        direct_workspace.mkdir()
        direct_environment = os.environ.copy()
        for inherited_name in ISOLATED_CLI_ENVIRONMENT_VARIABLES:
            direct_environment.pop(inherited_name, None)
        direct_environment.update(
            {
                "COLUMNS": "200",
                "LINES": "60",
                "MPLBACKEND": "Agg",
                "PYTHONHASHSEED": "0",
                "PYTHONIOENCODING": "utf-8",
                "TERM": "dumb",
            }
        )
        direct = subprocess.run(
            plan.public_command,
            cwd=direct_workspace,
            env=direct_environment,
            input="\n".join(step.response for step in plan.steps) + "\n",
            capture_output=True,
            encoding="utf-8",
            errors="replace",
            timeout=300,
        )
        assert direct.returncode == 0, direct.stdout[-12000:] + direct.stderr[-12000:]

        server_parameters = StdioServerParameters(
            command=sys.executable,
            args=["-m", "geochemistrypi_mcp"],
            env=_stdio_environment(cli_executable, parity_root / "mcp-state"),
        )
        async with Client(stdio_client(server_parameters)) as client:
            started = await client.call_tool("start_analysis", request.model_dump(mode="json"))
            assert started.is_error is False, started.content[0].text
            run_id = started.structured_content["run_id"]
            deadline = time.monotonic() + 360
            while True:
                status = await client.call_tool("get_run_status", {"run_id": run_id})
                assert status.is_error is False
                state = status.structured_content["state"]
                if state in {"succeeded", "failed", "cancelled"}:
                    break
                assert time.monotonic() < deadline
                await asyncio.sleep(0.25)
            assert state == "succeeded", status.structured_content
            result_call = await client.call_tool("get_run_result", {"run_id": run_id})
            assert result_call.is_error is False
            result = result_call.structured_content
            healthy = await client.call_tool("get_capabilities", {})
            assert healthy.is_error is False

        direct_run = direct_workspace / "geopi_output" / request.experiment_name / request.run_name
        wrapped_run = Path(result["output_directory"])
        assert _sha256(REGRESSION_DATASET_PATH) == fixture_hash_before
        assert _sha256(training_path) == result["input_sha256"]
        assert result["task"] == "regression"
        assert result["model"] == "linear_regression"
        assert result["application_input_hash_verified"] is True
        assert _all_files(direct_run) == _all_files(wrapped_run)
        direct_metrics = _load_json(direct_run / "metrics" / "Model Score - Linear Regression.txt")
        wrapped_metrics = _load_json(wrapped_run / "metrics" / "Model Score - Linear Regression.txt")
        assert direct_metrics == wrapped_metrics
        assert set(wrapped_metrics["Per Target"]) == {"Target", "TargetB"}
        assert _load_json(direct_run / "metrics" / "Cross Validation - Linear Regression.txt") == _load_json(wrapped_run / "metrics" / "Cross Validation - Linear Regression.txt")
        assert _load_json(direct_run / "parameters" / "Hyper Parameters - Linear Regression.txt") == _load_json(wrapped_run / "parameters" / "Hyper Parameters - Linear Regression.txt")
        assert _worksheet_values(direct_run / "artifacts" / "data" / "Y Test Predict.xlsx") == _worksheet_values(wrapped_run / "artifacts" / "data" / "Y Test Predict.xlsx")
        direct_application = _worksheet_values(direct_run / "artifacts" / "data" / "Application Data Predicted.xlsx")
        wrapped_application = _worksheet_values(wrapped_run / "artifacts" / "data" / "Application Data Predicted.xlsx")
        assert direct_application == wrapped_application
        assert {"Predicted_Target", "Predicted_TargetB"} <= set(wrapped_application[0])
        for plot_name in (
            "Predicted vs. Actual Diagram - Linear Regression.png",
            "Residuals Diagram - Linear Regression.png",
            "Permutation Importance Diagram - Linear Regression.png",
        ):
            assert (wrapped_run / "artifacts" / "image" / "model_output" / plot_name).is_file()
        assert result["artifact_count"] == len(_all_files(wrapped_run))


@pytest.mark.anyio
@pytest.mark.mcp_cli_parity
async def test_stdio_mcp_clustering_matches_direct_public_cli() -> None:
    cli_executable = Path(os.environ["GEOCHEMISTRYPI_CLI_EXECUTABLE"]).resolve()
    fixture_hash_before = _sha256(CLUSTERING_DATASET_PATH)
    request = ClusteringRequest(
        task="clustering",
        training_dataset_path=CLUSTERING_DATASET_PATH,
        experiment_name="PR6 Clustering Parity",
        run_name="KMeans",
        identifier_column="SampleID",
        feature_columns=("FeatureA", "FeatureB", "FeatureC"),
        model={"type": "kmeans"},
    )
    plan = ClusteringPlanCompiler().compile(
        request,
        cli_executable=cli_executable,
    )

    with tempfile.TemporaryDirectory(prefix="gpi-pr6-clustering-parity-") as temporary_root:
        parity_root = Path(temporary_root)
        plan = _with_tracking_root(plan, parity_root / "direct-tracking")
        direct_workspace = parity_root / "direct"
        direct_workspace.mkdir()
        direct_environment = os.environ.copy()
        for inherited_name in ISOLATED_CLI_ENVIRONMENT_VARIABLES:
            direct_environment.pop(inherited_name, None)
        direct_environment.update(
            {
                "COLUMNS": "200",
                "LINES": "60",
                "MPLBACKEND": "Agg",
                "PYTHONHASHSEED": "0",
                "PYTHONIOENCODING": "utf-8",
                "TERM": "dumb",
            }
        )
        direct = subprocess.run(
            plan.public_command,
            cwd=direct_workspace,
            env=direct_environment,
            input="\n".join(step.response for step in plan.steps) + "\n",
            capture_output=True,
            encoding="utf-8",
            errors="replace",
            timeout=360,
        )
        assert direct.returncode == 0, direct.stdout[-12000:] + direct.stderr[-12000:]

        server_parameters = StdioServerParameters(
            command=sys.executable,
            args=["-m", "geochemistrypi_mcp"],
            env=_stdio_environment(cli_executable, parity_root / "mcp-state"),
        )
        async with Client(stdio_client(server_parameters)) as client:
            started = await client.call_tool(
                "start_analysis",
                request.model_dump(mode="json"),
            )
            assert started.is_error is False
            run_id = started.structured_content["run_id"]
            deadline = time.monotonic() + 420
            while True:
                status = await client.call_tool("get_run_status", {"run_id": run_id})
                assert status.is_error is False
                state = status.structured_content["state"]
                if state in {"succeeded", "failed", "cancelled"}:
                    break
                assert time.monotonic() < deadline
                await asyncio.sleep(0.25)
            assert state == "succeeded", status.structured_content
            result_call = await client.call_tool("get_run_result", {"run_id": run_id})
            assert result_call.is_error is False
            result = result_call.structured_content
            healthy = await client.call_tool("get_capabilities", {})
            assert healthy.is_error is False

        direct_run = direct_workspace / "geopi_output" / request.experiment_name / request.run_name
        wrapped_run = Path(result["output_directory"])
        assert _sha256(CLUSTERING_DATASET_PATH) == fixture_hash_before == result["input_sha256"]
        assert result["task"] == "clustering"
        assert result["model"] == "kmeans"
        assert result["tuning"] == "not_applicable"
        assert result["application_input_sha256"] is None
        assert _all_files(direct_run) == _all_files(wrapped_run)
        assert _load_json(direct_run / "metrics" / "Model Score - KMeans.txt") == _load_json(wrapped_run / "metrics" / "Model Score - KMeans.txt")
        direct_silhouette_scores = _load_json(direct_run / "metrics" / "KMeans - Silhouette Scores.txt")
        wrapped_silhouette_scores = _load_json(wrapped_run / "metrics" / "KMeans - Silhouette Scores.txt")
        assert direct_silhouette_scores.keys() == wrapped_silhouette_scores.keys()
        assert direct_silhouette_scores == pytest.approx(
            wrapped_silhouette_scores,
            rel=CLUSTERING_METRIC_RELATIVE_TOLERANCE,
            abs=CLUSTERING_METRIC_ABSOLUTE_TOLERANCE,
        )
        assert _worksheet_values(direct_run / "artifacts" / "data" / "Cluster Labels - KMeans.xlsx") == _worksheet_values(wrapped_run / "artifacts" / "data" / "Cluster Labels - KMeans.xlsx")
        for artifact_name in (
            "Cluster Two-Dimensional Diagram - KMeans.png",
            "Cluster Three-Dimensional Diagram - KMeans.png",
            "Silhouette Diagram - KMeans.png",
            "Silhouette value Diagram - KMeans.png",
        ):
            assert (wrapped_run / "artifacts" / "image" / "model_output" / artifact_name).is_file()
        assert result["artifact_count"] == len(_all_files(wrapped_run))


@pytest.mark.anyio
@pytest.mark.mcp_cli_parity
async def test_stdio_mcp_decomposition_matches_direct_public_cli() -> None:
    cli_executable = Path(os.environ["GEOCHEMISTRYPI_CLI_EXECUTABLE"]).resolve()
    fixture_hash_before = _sha256(DECOMPOSITION_DATASET_PATH)
    request = DecompositionRequest(
        task="decomposition",
        training_dataset_path=DECOMPOSITION_DATASET_PATH,
        experiment_name="PR7 Decomposition Parity",
        run_name="PCA",
        identifier_column="SampleID",
        feature_columns=("FeatureA", "FeatureB", "FeatureC"),
        model={"type": "pca", "number_of_components": 2, "svd_solver": "auto"},
    )
    plan = DecompositionPlanCompiler().compile(
        request,
        cli_executable=cli_executable,
    )

    with tempfile.TemporaryDirectory(prefix="gpi-pr7-decomposition-parity-") as temporary_root:
        parity_root = Path(temporary_root)
        plan = _with_tracking_root(plan, parity_root / "direct-tracking")
        direct_workspace = parity_root / "direct"
        direct_workspace.mkdir()
        direct_environment = os.environ.copy()
        for inherited_name in ISOLATED_CLI_ENVIRONMENT_VARIABLES:
            direct_environment.pop(inherited_name, None)
        direct_environment.update(
            {
                "COLUMNS": "200",
                "LINES": "60",
                "MPLBACKEND": "Agg",
                "PYTHONHASHSEED": "0",
                "PYTHONIOENCODING": "utf-8",
                "TERM": "dumb",
            }
        )
        direct = subprocess.run(
            plan.public_command,
            cwd=direct_workspace,
            env=direct_environment,
            input="\n".join(step.response for step in plan.steps) + "\n",
            capture_output=True,
            encoding="utf-8",
            errors="replace",
            timeout=360,
        )
        assert direct.returncode == 0, direct.stdout[-12000:] + direct.stderr[-12000:]

        server_parameters = StdioServerParameters(
            command=sys.executable,
            args=["-m", "geochemistrypi_mcp"],
            env=_stdio_environment(cli_executable, parity_root / "mcp-state"),
        )
        async with Client(stdio_client(server_parameters)) as client:
            started = await client.call_tool(
                "start_analysis",
                request.model_dump(mode="json"),
            )
            assert started.is_error is False
            run_id = started.structured_content["run_id"]
            deadline = time.monotonic() + 420
            while True:
                status = await client.call_tool("get_run_status", {"run_id": run_id})
                assert status.is_error is False
                state = status.structured_content["state"]
                if state in {"succeeded", "failed", "cancelled"}:
                    break
                assert time.monotonic() < deadline
                await asyncio.sleep(0.25)
            assert state == "succeeded", status.structured_content
            result_call = await client.call_tool("get_run_result", {"run_id": run_id})
            assert result_call.is_error is False
            result = result_call.structured_content
            healthy = await client.call_tool("get_capabilities", {})
            assert healthy.is_error is False

        direct_run = direct_workspace / "geopi_output" / request.experiment_name / request.run_name
        wrapped_run = Path(result["output_directory"])
        assert _sha256(DECOMPOSITION_DATASET_PATH) == fixture_hash_before == result["input_sha256"]
        assert result["task"] == "decomposition"
        assert result["model"] == "pca"
        assert result["tuning"] == "not_applicable"
        assert result["application_input_sha256"] is None
        assert result["reported_metrics"] == {}
        assert _all_files(direct_run) == _all_files(wrapped_run)
        assert _load_json(direct_run / "parameters" / "Hyper Parameters - PCA.txt") == _load_json(wrapped_run / "parameters" / "Hyper Parameters - PCA.txt")
        assert _worksheet_values(direct_run / "artifacts" / "data" / "X Reduced.xlsx") == _worksheet_values(wrapped_run / "artifacts" / "data" / "X Reduced.xlsx")
        assert _worksheet_values(direct_run / "artifacts" / "image" / "model_output" / "Compositional Bi-plot - PC Data.xlsx") == _worksheet_values(
            wrapped_run / "artifacts" / "image" / "model_output" / "Compositional Bi-plot - PC Data.xlsx"
        )
        for artifact_name in (
            "Decomposition Two-Dimensional Diagram - PCA.png",
            "Decomposition Heatmap - PCA.png",
            "Dimensionality Reduction Contour Plot - PCA.png",
            "Compositional Bi-plot - PCA.png",
        ):
            assert (wrapped_run / "artifacts" / "image" / "model_output" / artifact_name).is_file()
        assert result["artifact_count"] == len(_all_files(wrapped_run))


@pytest.mark.anyio
@pytest.mark.mcp_cli_parity
async def test_stdio_mcp_anomaly_detection_matches_direct_public_cli() -> None:
    cli_executable = Path(os.environ["GEOCHEMISTRYPI_CLI_EXECUTABLE"]).resolve()
    fixture_hash_before = _sha256(ANOMALY_DETECTION_DATASET_PATH)
    request = AnomalyDetectionRequest(
        task="anomaly_detection",
        training_dataset_path=ANOMALY_DETECTION_DATASET_PATH,
        experiment_name="PR8A Anomaly Detection Parity",
        run_name="Isolation Forest",
        identifier_column="SampleID",
        feature_columns=("FeatureA", "FeatureB", "FeatureC"),
        model={
            "type": "isolation_forest",
            "number_of_estimators": 100,
            "contamination": 0.2,
            "maximum_features": 3,
            "bootstrap": False,
        },
    )
    plan = AnomalyDetectionPlanCompiler().compile(
        request,
        cli_executable=cli_executable,
    )

    # Keep the managed workspace below legacy Windows path limits. The run
    # already carries descriptive experiment and artifact names.
    with tempfile.TemporaryDirectory(prefix="g8a-") as temporary_root:
        parity_root = Path(temporary_root)
        plan = _with_tracking_root(plan, parity_root / "direct-tracking")
        direct_workspace = parity_root / "direct"
        direct_workspace.mkdir()
        direct_environment = os.environ.copy()
        for inherited_name in ISOLATED_CLI_ENVIRONMENT_VARIABLES:
            direct_environment.pop(inherited_name, None)
        direct_environment.update(
            {
                "COLUMNS": "200",
                "LINES": "60",
                "MPLBACKEND": "Agg",
                "PYTHONHASHSEED": "0",
                "PYTHONIOENCODING": "utf-8",
                "TERM": "dumb",
            }
        )
        direct = subprocess.run(
            plan.public_command,
            cwd=direct_workspace,
            env=direct_environment,
            input="\n".join(step.response for step in plan.steps) + "\n",
            capture_output=True,
            encoding="utf-8",
            errors="replace",
            timeout=360,
        )
        assert direct.returncode == 0, direct.stdout[-12000:] + direct.stderr[-12000:]

        server_parameters = StdioServerParameters(
            command=sys.executable,
            args=["-m", "geochemistrypi_mcp"],
            env=_stdio_environment(cli_executable, parity_root / "mcp-state"),
        )
        async with Client(stdio_client(server_parameters)) as client:
            started = await client.call_tool(
                "start_analysis",
                request.model_dump(mode="json"),
            )
            assert started.is_error is False
            run_id = started.structured_content["run_id"]
            deadline = time.monotonic() + 420
            while True:
                status = await client.call_tool(
                    "get_run_status",
                    {"run_id": run_id},
                )
                assert status.is_error is False
                state = status.structured_content["state"]
                if state in {"succeeded", "failed", "cancelled"}:
                    break
                assert time.monotonic() < deadline
                await asyncio.sleep(0.25)
            assert state == "succeeded", status.structured_content
            result_call = await client.call_tool(
                "get_run_result",
                {"run_id": run_id},
            )
            assert result_call.is_error is False
            result = result_call.structured_content
            healthy = await client.call_tool("get_capabilities", {})
            assert healthy.is_error is False

        direct_run = direct_workspace / "geopi_output" / request.experiment_name / request.run_name
        wrapped_run = Path(result["output_directory"])
        assert _sha256(ANOMALY_DETECTION_DATASET_PATH) == fixture_hash_before == result["input_sha256"]
        assert result["task"] == "anomaly_detection"
        assert result["model"] == "isolation_forest"
        assert result["tuning"] == "not_applicable"
        assert result["application_input_sha256"] is None
        assert result["reported_metrics"] == {}
        assert _all_files(direct_run) == _all_files(wrapped_run)
        assert _load_json(direct_run / "parameters" / "Hyper Parameters - Isolation Forest.txt") == _load_json(wrapped_run / "parameters" / "Hyper Parameters - Isolation Forest.txt")
        assert _worksheet_values(direct_run / "artifacts" / "data" / "X Abnormal Detection.xlsx") == _worksheet_values(wrapped_run / "artifacts" / "data" / "X Abnormal Detection.xlsx")
        assert _worksheet_values(direct_run / "artifacts" / "data" / "X Abnormal.xlsx") == _worksheet_values(wrapped_run / "artifacts" / "data" / "X Abnormal.xlsx")
        for artifact_name in (
            "Anomaly Detection Density Estimation - Isolation Forest.png",
            "Anomaly Detection Two-Dimensional Diagram - Isolation Forest.png",
            "Anomaly Detection Three-Dimensional Diagram - Isolation Forest.png",
        ):
            assert (wrapped_run / "artifacts" / "image" / "model_output" / artifact_name).is_file()
        assert result["artifact_count"] == len(_all_files(wrapped_run))
