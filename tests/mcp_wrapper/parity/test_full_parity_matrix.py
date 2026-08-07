"""PR9I executable inventory and opt-in real full-model parity gate."""

import hashlib
import json
import math
import os
import time
from dataclasses import replace
from pathlib import Path
from typing import Any

import pytest
from geochemistrypi_mcp import AnalysisPlanCompiler, AnomalyDetectionRequest, ClassificationRequest, ClusteringRequest, DecompositionRequest, RegressionRequest
from geochemistrypi_mcp.config.settings import McpSettings
from geochemistrypi_mcp.runtime.cli_driver import CliInteractionDriver
from geochemistrypi_mcp.runtime.runs import RunManager
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[3]
FIXTURES = ROOT / "tests" / "cli_contract" / "fixtures"
MATRIX_PATH = Path(__file__).parent / "fixtures" / "full_parity_matrix_v1.json"
ABS_TOLERANCE = 1e-9
REL_TOLERANCE = 1e-7


def _matrix() -> dict[str, Any]:
    return json.loads(MATRIX_PATH.read_text(encoding="utf-8"))


def _cases() -> list[tuple[str, str, str, str]]:
    matrix = _matrix()
    cases = []
    for task, models in matrix["manual_single_models"].items():
        cases.extend((f"manual.{task}.{model}", task, model, "manual") for model in models)
    for task, models in matrix["automl_models"].items():
        cases.extend((f"automl.{task}.{model}", task, model, "automl") for model in models)
    cases.extend((f"aggregate.{task}.all", task, "all", "aggregate") for task in matrix["aggregate_tasks"])
    return cases


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _inventory(root: Path) -> list[str]:
    return sorted(path.relative_to(root).as_posix() for path in root.rglob("*") if path.is_file())


def _assert_value_equal(left: Any, right: Any) -> None:
    if isinstance(left, (int, float)) and isinstance(right, (int, float)):
        assert math.isclose(float(left), float(right), abs_tol=ABS_TOLERANCE, rel_tol=REL_TOLERANCE)
    elif isinstance(left, list) and isinstance(right, list):
        assert len(left) == len(right)
        for left_item, right_item in zip(left, right):
            _assert_value_equal(left_item, right_item)
    elif isinstance(left, dict) and isinstance(right, dict):
        assert left.keys() == right.keys()
        for key in left:
            _assert_value_equal(left[key], right[key])
    else:
        assert left == right


def _worksheet(path: Path) -> list[list[Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return [list(row) for row in workbook.active.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _assert_outputs_equal(direct: Path, wrapped: Path) -> None:
    inventory = _inventory(direct)
    assert inventory == _inventory(wrapped)
    for relative in inventory:
        direct_path = direct / relative
        wrapped_path = wrapped / relative
        suffix = direct_path.suffix.lower()
        if suffix == ".xlsx":
            _assert_value_equal(_worksheet(direct_path), _worksheet(wrapped_path))
        elif suffix in {".json", ".txt"}:
            try:
                _assert_value_equal(
                    json.loads(direct_path.read_text(encoding="utf-8")),
                    json.loads(wrapped_path.read_text(encoding="utf-8")),
                )
            except json.JSONDecodeError:
                assert direct_path.read_text(encoding="utf-8") == wrapped_path.read_text(encoding="utf-8")
        elif suffix == ".csv":
            assert direct_path.read_text(encoding="utf-8-sig") == wrapped_path.read_text(encoding="utf-8-sig")
        elif suffix in {".png", ".pdf", ".svg"}:
            assert direct_path.stat().st_size > 100
            assert wrapped_path.stat().st_size > 100
            if suffix == ".png":
                assert direct_path.read_bytes()[:8] == wrapped_path.read_bytes()[:8] == b"\x89PNG\r\n\x1a\n"
            if suffix == ".pdf":
                assert direct_path.read_bytes()[:4] == wrapped_path.read_bytes()[:4] == b"%PDF"
        else:
            assert direct_path.stat().st_size > 0
            assert wrapped_path.stat().st_size > 0


def _request(case_id: str, task: str, model: str, mode: str):
    common = {
        # Full model names already exercise the CLI's longest generated output
        # paths. Keep test-owned display names short so pytest's temporary path
        # hierarchy does not consume the user's Windows path budget.
        "experiment_name": "P" + hashlib.sha1(case_id.encode()).hexdigest()[:8],
        "run_name": "R",
    }
    if task == "classification":
        values = dict(
            training_dataset_path=FIXTURES / "classification_baseline.csv",
            identifier_column="SampleID",
            feature_columns=("SIO2(WT%)", "TIO2(WT%)", "AL2O3(WT%)", "FEOT(WT%)"),
            target_column="Label",
            **common,
        )
        return ClassificationRequest(
            **values,
            **({"model_selection": {"mode": "all", "tuning": "manual"}} if mode == "aggregate" else {"tuning": mode, "model": {"type": model}}),
        )
    if task == "regression":
        values = dict(
            training_dataset_path=FIXTURES / "regression_baseline.csv",
            identifier_column="SampleID",
            feature_columns=("SIO2", "TIO2"),
            target_column="Target",
            **common,
        )
        return RegressionRequest(
            **values,
            **({"model_selection": {"mode": "all", "tuning": "manual"}} if mode == "aggregate" else {"tuning": mode, "model": {"type": model}}),
        )
    request_types = {
        "clustering": (ClusteringRequest, "clustering_baseline.csv"),
        "decomposition": (DecompositionRequest, "decomposition_baseline.csv"),
        "anomaly_detection": (AnomalyDetectionRequest, "anomaly_detection_baseline.csv"),
    }
    request_type, file_name = request_types[task]
    values = dict(
        training_dataset_path=FIXTURES / file_name,
        identifier_column="SampleID",
        feature_columns=("FeatureA", "FeatureB", "FeatureC"),
        **common,
    )
    return request_type(
        **values,
        **({"model_selection": {"mode": "all"}} if mode == "aggregate" else {"model": {"type": model}}),
    )


def test_full_matrix_inventory_is_complete_and_has_scientific_comparison_rules() -> None:
    matrix = _matrix()
    assert matrix["schema_version"] == 1
    assert len([case for case in _cases() if case[3] == "manual"]) == 36
    assert len([case for case in _cases() if case[3] == "automl" and case[1] == "classification"]) == 11
    assert len([case for case in _cases() if case[3] == "automl" and case[1] == "regression"]) == 13
    assert len([case for case in _cases() if case[3] == "aggregate"]) == 5
    contract = matrix["comparison_contract"]
    assert contract["input_sha256_before_and_after"] is True
    assert contract["complete_recursive_file_inventory"] is True
    assert set(contract["tabular_identity_fields"]) == {"identifiers", "row_order", "feature_order", "targets", "predictions"}
    assert contract["absolute_float_tolerance"] == ABS_TOLERANCE
    assert contract["relative_float_tolerance"] == REL_TOLERANCE
    assert contract["binary_hash_images_across_platforms"] is False
    assert len(matrix["branch_scenarios"]) == 11


@pytest.mark.mcp_cli_full_parity
@pytest.mark.parametrize("case_id,task,model,mode", _cases(), ids=lambda value: value if isinstance(value, str) else None)
def test_full_real_model_parity(case_id: str, task: str, model: str, mode: str, tmp_path_factory: pytest.TempPathFactory) -> None:
    if os.environ.get("GEOCHEMISTRYPI_FULL_PARITY") != "1":
        pytest.skip("Set GEOCHEMISTRYPI_FULL_PARITY=1 only in the scheduled or release-candidate full matrix.")
    shard = os.environ.get("GEOCHEMISTRYPI_PARITY_SHARD")
    expected_shard = "aggregates" if mode == "aggregate" else f"{task}-{mode}" if task in {"classification", "regression"} else "unsupervised-manual"
    if shard and shard != expected_shard:
        pytest.skip(f"Scenario belongs to shard {expected_shard}.")
    cli_executable = Path(os.environ["GEOCHEMISTRYPI_CLI_EXECUTABLE"]).resolve()
    parity_root = tmp_path_factory.mktemp("p")
    scientific_request = _request(case_id, task, model, mode)
    source = scientific_request.training_dataset_path
    input_hash = _sha256(source)
    plan = AnalysisPlanCompiler().compile(scientific_request, cli_executable=cli_executable)
    direct_tracking_root = (parity_root / "direct-tracking").resolve()
    direct_tracking_root.mkdir(parents=True, exist_ok=True)
    assert "--tracking-root" not in plan.public_command
    plan = replace(
        plan,
        public_command=(*plan.public_command, "--tracking-root", str(direct_tracking_root)),
    )
    direct_workspace = parity_root / "direct"
    direct_result = CliInteractionDriver(process_timeout_seconds=1800, automation_mode=True).run(plan, workspace=direct_workspace)
    assert direct_result.returncode == 0

    settings = McpSettings(
        runs_root=parity_root / "runs",
        cli_executable=cli_executable,
        tracking_root=parity_root / "tracking",
        maximum_process_seconds=1800,
    )
    manager = RunManager(settings, cli_resolver=lambda: (cli_executable, "0.8.0"))
    try:
        started = manager.start(scientific_request)
        deadline = time.monotonic() + 1800
        while True:
            status = manager.get_status(started.run_id)
            if status.state in {"succeeded", "partial_failure", "failed", "cancelled"}:
                break
            assert time.monotonic() < deadline
            time.sleep(0.2)
        assert status.state == "succeeded", status
        result = manager.get_result(started.run_id)
    finally:
        manager.close()

    direct_output = direct_workspace / "geopi_output" / scientific_request.experiment_name / scientific_request.run_name
    assert result.input_sha256 == input_hash == _sha256(source)
    assert result.input_hash_verified is True
    _assert_outputs_equal(direct_output, Path(result.output_directory))
