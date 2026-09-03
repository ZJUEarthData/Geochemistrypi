"""Deterministic, non-scientific dataset views for CLI-compatible inputs."""

import csv
import hashlib
import json
import math
import os
import tempfile
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Iterator, Sequence

from openpyxl import load_workbook

from ..api.schemas import DatasetFilterRule, DatasetPreparationContract
from .headers import HeaderValidationError, normalize_dataset_header
from .inspector import DatasetInspectionError, DatasetSnapshot, sha256_file, snapshot_dataset
from .source_rows import iter_cli_csv_rows, iter_cli_excel_rows


@dataclass(frozen=True)
class PreparedDataset:
    """Source identity, CLI input identity, and their immutable lineage record."""

    source_snapshot: DatasetSnapshot
    snapshot: DatasetSnapshot
    record: dict[str, Any]


class DatasetPreparationError(ValueError):
    """Raised when a requested dataset view is absent, ambiguous, or unverifiable."""


def _canonical_json_bytes(value: Any) -> bytes:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")


def _canonical_sha256(value: Any) -> str:
    return hashlib.sha256(_canonical_json_bytes(value)).hexdigest()


def _atomic_write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(prefix=f".{path.name}.", suffix=".tmp", dir=path.parent)
    temporary = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "w", encoding="utf-8", newline="\n") as stream:
            json.dump(value, stream, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
            stream.write("\n")
            stream.flush()
            os.fsync(stream.fileno())
        os.replace(temporary, path)
    finally:
        if temporary.exists():
            temporary.unlink()


def _cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return value


def _row_values(row: Sequence[Any], width: int) -> tuple[Any, ...]:
    values = tuple(row)
    if len(values) > width and any(value not in (None, "") for value in values[width:]):
        raise DatasetPreparationError("A data row contains values beyond the selected header width.")
    return (*values[:width], *("" for _ in range(max(0, width - len(values)))))


def _flatten_header_rows(rows: Sequence[Sequence[Any]], contract: DatasetPreparationContract) -> tuple[Any, ...]:
    width = max((len(row) for row in rows), default=0)
    normalized_rows: list[tuple[str, ...]] = []
    for row in rows:
        values: list[str] = []
        previous = ""
        for position in range(width):
            raw = row[position] if position < len(row) else None
            value = "" if raw is None else str(raw).strip()
            if not value and contract.empty_header_policy == "forward_fill":
                value = previous
            if value:
                previous = value
            values.append(value)
        normalized_rows.append(tuple(values))
    flattened: list[str] = []
    for position in range(width):
        tokens: list[str] = []
        for row in normalized_rows:
            token = row[position]
            if token and (not tokens or tokens[-1] != token):
                tokens.append(token)
        if not tokens and contract.empty_header_policy == "error":
            raise DatasetPreparationError(f"Compound header column {position + 1} is empty.")
        flattened.append(contract.header_join_separator.join(tokens))
    duplicates = {name for name in flattened if name and flattened.count(name) > 1}
    if duplicates and contract.duplicate_header_policy == "reject":
        raise DatasetPreparationError(f"Compound header contains duplicate names: {sorted(duplicates)}")
    if duplicates:
        counts: dict[str, int] = {}
        for position, name in enumerate(flattened):
            counts[name] = counts.get(name, 0) + 1
            if name in duplicates:
                flattened[position] = f"{name}{contract.header_join_separator}{counts[name]}"
    return tuple(flattened)


def _csv_table(
    path: Path,
    contract: DatasetPreparationContract,
) -> tuple[tuple[Any, ...], Iterator[tuple[Any, ...]], str | None, int]:
    stream = path.open(encoding="utf-8-sig", newline="")
    records = iter_cli_csv_rows(csv.reader(stream))
    indices = contract.header_row_indices or (contract.header_row_index,)
    selected_headers: list[tuple[Any, ...]] = []
    try:
        for index in range(indices[-1] + 1):
            row = next(records)
            if index in indices:
                selected_headers.append(tuple(row))
    except (StopIteration, csv.Error, UnicodeError) as exc:
        stream.close()
        raise DatasetPreparationError(f"CSV header rows {indices} do not identify readable records.") from exc
    header = _flatten_header_rows(selected_headers, contract) if contract.header_row_indices else selected_headers[0]

    def rows() -> Iterator[tuple[Any, ...]]:
        try:
            yield from records
        finally:
            stream.close()

    return header, rows(), None, indices[-1] + 2


def _excel_table(
    path: Path,
    worksheet: str | None,
    contract: DatasetPreparationContract,
) -> tuple[tuple[Any, ...], Iterator[tuple[Any, ...]], str, int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if worksheet is None and len(workbook.sheetnames) != 1:
        sheet_names = ", ".join(repr(name) for name in workbook.sheetnames)
        workbook.close()
        raise DatasetPreparationError(f"Excel worksheet must be selected explicitly; available worksheets: {sheet_names}.")
    try:
        selected = workbook[worksheet] if worksheet is not None else workbook[workbook.sheetnames[0]]
    except KeyError as exc:
        workbook.close()
        raise DatasetPreparationError(f"Excel worksheet does not exist: {worksheet!r}.") from exc
    rows = selected.iter_rows(values_only=True)
    indices = contract.header_row_indices or (contract.header_row_index,)
    selected_headers: list[tuple[Any, ...]] = []
    try:
        for index in range(indices[-1] + 1):
            row = tuple(next(rows))
            if index in indices:
                selected_headers.append(row)
    except StopIteration as exc:
        workbook.close()
        raise DatasetPreparationError(f"Excel header rows {indices} do not identify worksheet rows.") from exc
    header = _flatten_header_rows(selected_headers, contract) if contract.header_row_indices else selected_headers[0]

    def remaining() -> Iterator[tuple[Any, ...]]:
        try:
            yield from iter_cli_excel_rows(rows)
        finally:
            workbook.close()

    return header, remaining(), selected.title, indices[-1] + 2


def _open_table(
    snapshot: DatasetSnapshot,
    contract: DatasetPreparationContract,
) -> tuple[tuple[Any, ...], Iterator[tuple[Any, ...]], str | None, int]:
    if snapshot.format == "csv":
        if contract.worksheet is not None:
            raise DatasetPreparationError("worksheet selection is valid only for .xlsx datasets")
        return _csv_table(snapshot.resolved_path, contract)
    try:
        return _excel_table(snapshot.resolved_path, contract.worksheet, contract)
    except (OSError, ValueError) as exc:
        if isinstance(exc, DatasetPreparationError):
            raise
        raise DatasetPreparationError(f"Excel dataset cannot be prepared: {snapshot.resolved_path}") from exc


def _selected_header(
    raw_header: Sequence[Any],
    contract: DatasetPreparationContract,
    maximum_columns: int,
    allow_pandas_duplicate_mangling: bool,
) -> tuple[tuple[str, ...], tuple[int, ...], tuple[str, ...]]:
    strip_whitespace = contract.header_whitespace_policy == "strip"
    strip_bom = contract.header_bom_policy == "strip"
    selected_projection = bool(contract.selected_columns)
    try:
        header = normalize_dataset_header(
            raw_header,
            maximum_columns,
            allow_pandas_duplicate_mangling=(allow_pandas_duplicate_mangling or contract.duplicate_header_policy == "suffix" or selected_projection),
            strip_whitespace=strip_whitespace,
            strip_bom=strip_bom,
        )
    except HeaderValidationError as exc:
        raise DatasetPreparationError(str(exc)) from exc
    if selected_projection and contract.duplicate_header_policy == "reject":
        canonical_names = []
        for index, value in enumerate(raw_header):
            name = f"Unnamed: {index}" if value is None or value == "" else str(value)
            if strip_bom:
                name = name.lstrip("\ufeff")
            canonical_names.append(name.strip() if strip_whitespace else name)
        referenced_columns = {
            *contract.selected_columns,
            *contract.excluded_columns,
            *contract.row_identity.columns,
            *(rule.column for rule in contract.filters),
        }
        ambiguous = sorted(column for column in referenced_columns if canonical_names.count(column) > 1)
        if ambiguous:
            raise DatasetPreparationError("Selected or otherwise referenced dataset columns are ambiguous after header normalization: " f"{ambiguous}")
    requested_columns = contract.selected_columns
    if contract.excluded_columns:
        missing_exclusions = sorted(set(contract.excluded_columns) - set(header))
        if missing_exclusions:
            raise DatasetPreparationError(f"Excluded dataset columns are absent from the selected table: {missing_exclusions}")
        excluded = set(contract.excluded_columns)
        requested_columns = tuple(column for column in header if column not in excluded)
    if not requested_columns:
        if contract.excluded_columns:
            raise DatasetPreparationError("excluded_columns removed every dataset column")
        return header, tuple(range(len(header))), header
    missing = sorted(set(requested_columns) - set(header))
    if missing:
        raise DatasetPreparationError(f"Prepared dataset columns are absent from the selected table: {missing}")
    positions = tuple(header.index(column) for column in requested_columns)
    return requested_columns, positions, header


def _filter_positions(
    normalized_header: tuple[str, ...],
    contract: DatasetPreparationContract,
) -> tuple[tuple[int, DatasetFilterRule], ...]:
    missing = sorted({rule.column for rule in contract.filters} - set(normalized_header))
    if missing:
        raise DatasetPreparationError(f"Dataset filter columns are absent from the selected table: {missing}")
    return tuple((normalized_header.index(rule.column), rule) for rule in contract.filters)


def _is_null(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return not value.strip()
    return isinstance(value, float) and math.isnan(value)


def _comparable(value: Any, operand: str | int | float | bool) -> str | float | bool:
    if isinstance(operand, bool):
        if isinstance(value, bool):
            return value
        normalized = str(value).strip().lower()
        if normalized in {"true", "1"}:
            return True
        if normalized in {"false", "0"}:
            return False
        raise DatasetPreparationError(f"Filter value {value!r} is not boolean.")
    if isinstance(operand, (int, float)):
        try:
            numeric = float(value)
        except (TypeError, ValueError) as exc:
            raise DatasetPreparationError(f"Filter value {value!r} is not numeric.") from exc
        if not math.isfinite(numeric):
            raise DatasetPreparationError(f"Filter value {value!r} is not finite.")
        return numeric
    return str(value)


def _passes_filters(values: Sequence[Any], filters: tuple[tuple[int, DatasetFilterRule], ...]) -> bool:
    for position, rule in filters:
        value = values[position]
        if rule.operator == "not_null":
            if _is_null(value):
                return False
            continue
        if _is_null(value):
            return False
        if rule.operator == "between":
            assert rule.minimum is not None and rule.maximum is not None
            comparable = _comparable(value, rule.minimum)
            minimum = float(rule.minimum)
            maximum = float(rule.maximum)
            if rule.inclusive:
                if not minimum <= comparable <= maximum:
                    return False
            elif not minimum < comparable < maximum:
                return False
            continue
        if rule.operator == "in":
            if not any(_comparable(value, operand) == _comparable(operand, operand) for operand in rule.values):
                return False
            continue
        assert rule.value is not None
        comparable = _comparable(value, rule.value)
        expected = _comparable(rule.value, rule.value)
        if rule.operator == "equal" and comparable != expected:
            return False
        if rule.operator == "not_equal" and comparable == expected:
            return False
        if rule.operator == "greater_than" and not comparable > expected:
            return False
        if rule.operator == "greater_than_or_equal" and not comparable >= expected:
            return False
        if rule.operator == "less_than" and not comparable < expected:
            return False
        if rule.operator == "less_than_or_equal" and not comparable <= expected:
            return False
    return True


def _source_mapping_record(contract: DatasetPreparationContract, maximum_bytes: int) -> dict[str, Any] | None:
    mapping_path = contract.row_identity.source_mapping_path
    expected_sha256 = contract.row_identity.source_mapping_sha256
    if mapping_path is None or expected_sha256 is None:
        return None
    resolved = Path(mapping_path).expanduser().resolve()
    try:
        size_bytes = resolved.stat().st_size
    except OSError as exc:
        raise DatasetPreparationError(f"Source-row mapping is unavailable: {resolved}") from exc
    if not resolved.is_file() or size_bytes > maximum_bytes:
        raise DatasetPreparationError("Source-row mapping is not a bounded regular file.")
    actual_sha256 = sha256_file(resolved)
    if actual_sha256 != expected_sha256:
        raise DatasetPreparationError("Source-row mapping does not match source_mapping_sha256.")
    return {"path": str(resolved), "size_bytes": size_bytes, "sha256": actual_sha256}


def _identity_value(
    strategy: str,
    identity_positions: tuple[int, ...],
    selected_values: Sequence[Any],
    source_row_number: int,
    worksheet: str | None,
) -> str:
    if strategy == "source_row":
        scope = f"worksheet:{worksheet}" if worksheet is not None else "csv-record"
        return f"{scope}:row:{source_row_number}"
    values = [_cell_value(selected_values[position]) for position in identity_positions]
    if any(value == "" for value in values):
        raise DatasetPreparationError(f"Column-based row identity contains a missing value at source row {source_row_number}.")
    return _canonical_json_bytes(values).decode("utf-8")


def _materialize_worksheet_union(
    source: DatasetSnapshot,
    contract: DatasetPreparationContract,
    destination: Path,
    maximum_columns: int,
    allow_pandas_duplicate_mangling: bool,
) -> dict[str, Any]:
    """Union named worksheets by aligned column names with sheet/row lineage."""
    generated_columns = (
        str(contract.source_sheet_column),
        str(contract.source_row_column),
    )
    selected_columns = tuple(contract.selected_columns)
    source_columns = tuple(column for column in selected_columns if column not in generated_columns)
    if not source_columns:
        raise DatasetPreparationError("A worksheet union must select at least one source column.")
    identity_positions = tuple(selected_columns.index(column) for column in contract.row_identity.columns)
    identity_digest = hashlib.sha256()
    retained_source_rows_digest = hashlib.sha256()
    identities: set[str] = set()
    input_row_count = 0
    row_count = 0
    per_sheet: dict[str, dict[str, int]] = {}
    destination.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{destination.name}.",
        suffix=".tmp",
        dir=destination.parent,
    )
    temporary = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "w", encoding="utf-8", newline="") as stream:
            writer = csv.writer(stream, lineterminator="\n")
            writer.writerow(selected_columns)
            for worksheet in contract.worksheets:
                raw_header, rows, selected_worksheet, first_data_row_number = _excel_table(
                    source.resolved_path,
                    worksheet,
                    contract,
                )
                source_contract = contract.model_copy(
                    update={
                        "selected_columns": source_columns,
                        "row_identity": contract.row_identity.model_copy(update={"columns": ()}),
                    }
                )
                source_header, selected_positions, normalized_header = _selected_header(
                    raw_header,
                    source_contract,
                    maximum_columns,
                    allow_pandas_duplicate_mangling,
                )
                filters = _filter_positions(normalized_header, contract)
                sheet_input_count = 0
                sheet_output_count = 0
                try:
                    for source_row_number, row in enumerate(
                        rows,
                        start=first_data_row_number,
                    ):
                        input_row_count += 1
                        sheet_input_count += 1
                        values = _row_values(row, len(raw_header))
                        if not _passes_filters(values, filters):
                            continue
                        projected = {name: values[position] for name, position in zip(source_header, selected_positions)}
                        projected[generated_columns[0]] = selected_worksheet
                        projected[generated_columns[1]] = source_row_number
                        selected_values = tuple(projected[column] for column in selected_columns)
                        identity = _identity_value(
                            contract.row_identity.strategy,
                            identity_positions,
                            selected_values,
                            source_row_number,
                            selected_worksheet,
                        )
                        if identity in identities:
                            raise DatasetPreparationError("Prepared worksheet-union row identity is duplicated at " f"{selected_worksheet!r} row {source_row_number}.")
                        identities.add(identity)
                        identity_digest.update(identity.encode("utf-8"))
                        identity_digest.update(b"\n")
                        retained_source_rows_digest.update(f"{selected_worksheet}:{source_row_number}".encode("utf-8"))
                        retained_source_rows_digest.update(b"\n")
                        writer.writerow(tuple(_cell_value(value) for value in selected_values))
                        row_count += 1
                        sheet_output_count += 1
                finally:
                    close = getattr(rows, "close", None)
                    if close is not None:
                        close()
                per_sheet[selected_worksheet] = {
                    "input_row_count": sheet_input_count,
                    "source_row_count": sheet_output_count,
                    "filtered_row_count": sheet_input_count - sheet_output_count,
                }
            stream.flush()
            os.fsync(stream.fileno())
        if row_count < 1:
            raise DatasetPreparationError("The worksheet union contains no rows after filtering.")
        ordered_identity_sha256 = identity_digest.hexdigest()
        expected = contract.row_identity.expected_ordered_sha256
        if expected is not None and ordered_identity_sha256 != expected:
            raise DatasetPreparationError("Prepared worksheet-union identity does not match expected_ordered_sha256.")
        os.replace(temporary, destination)
    finally:
        if temporary.exists():
            temporary.unlink()
    return {
        "worksheet": None,
        "worksheets": list(contract.worksheets),
        "union_mode": contract.union_mode,
        "source_sheet_column": contract.source_sheet_column,
        "source_row_column": contract.source_row_column,
        "header_row_index": contract.header_row_index,
        "header_row_indices": list(contract.header_row_indices),
        "header_join_separator": contract.header_join_separator,
        "empty_header_policy": contract.empty_header_policy,
        "duplicate_header_policy": contract.duplicate_header_policy,
        "selected_columns": list(selected_columns),
        "input_row_count": input_row_count,
        "source_row_count": row_count,
        "filtered_row_count": input_row_count - row_count,
        "per_sheet": per_sheet,
        "filters": [rule.model_dump(mode="json", exclude_none=True, exclude_defaults=True) for rule in contract.filters],
        "filter_result_sha256": retained_source_rows_digest.hexdigest(),
        "row_identity": {
            "strategy": contract.row_identity.strategy,
            "columns": list(contract.row_identity.columns),
            "ordered_sha256": ordered_identity_sha256,
        },
    }


def _materialize(
    source: DatasetSnapshot,
    contract: DatasetPreparationContract,
    destination: Path,
    maximum_columns: int,
    allow_pandas_duplicate_mangling: bool,
    preserve_source_columns_for_cli: bool = False,
) -> dict[str, Any]:
    if contract.worksheets:
        return _materialize_worksheet_union(
            source,
            contract,
            destination,
            maximum_columns,
            allow_pandas_duplicate_mangling,
        )
    materialization_contract = contract
    if preserve_source_columns_for_cli:
        materialization_contract = contract.model_copy(update={"selected_columns": (), "excluded_columns": ()})
    raw_header, rows, worksheet, first_data_row_number = _open_table(source, materialization_contract)
    output_header = tuple(materialization_contract.selected_columns)
    source_row_column = materialization_contract.source_row_column
    source_contract = materialization_contract
    if source_row_column is not None:
        source_columns = tuple(column for column in output_header if column != source_row_column)
        if not source_columns:
            raise DatasetPreparationError("A generated source-row column must accompany at least one source column.")
        source_contract = contract.model_copy(
            update={
                "selected_columns": source_columns,
                "row_identity": contract.row_identity.model_copy(update={"columns": tuple(column for column in contract.row_identity.columns if column != source_row_column)}),
            }
        )
    selected_source_header, selected_positions, normalized_header = _selected_header(
        raw_header,
        source_contract,
        maximum_columns,
        allow_pandas_duplicate_mangling,
    )
    selected_header = output_header if source_row_column is not None else selected_source_header
    filters = _filter_positions(normalized_header, contract)
    identity_positions = tuple(selected_header.index(column) for column in contract.row_identity.columns)
    identity_digest = hashlib.sha256()
    identities: set[str] = set()
    row_count = 0
    input_row_count = 0
    retained_source_rows_digest = hashlib.sha256()
    destination.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(prefix=f".{destination.name}.", suffix=".tmp", dir=destination.parent)
    temporary = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "w", encoding="utf-8", newline="") as stream:
            writer = csv.writer(stream, lineterminator="\n")
            writer.writerow(selected_header)
            for source_row_number, row in enumerate(rows, start=first_data_row_number):
                input_row_count += 1
                values = _row_values(row, len(raw_header))
                if not _passes_filters(values, filters):
                    continue
                if source_row_column is None:
                    selected_values = tuple(values[position] for position in selected_positions)
                else:
                    projected = {
                        name: values[position]
                        for name, position in zip(
                            selected_source_header,
                            selected_positions,
                        )
                    }
                    projected[source_row_column] = source_row_number
                    selected_values = tuple(projected[column] for column in selected_header)
                identity = _identity_value(
                    contract.row_identity.strategy,
                    identity_positions,
                    selected_values,
                    source_row_number,
                    worksheet,
                )
                if identity in identities:
                    raise DatasetPreparationError(f"Prepared dataset row identity is duplicated at source row {source_row_number}.")
                identities.add(identity)
                identity_digest.update(identity.encode("utf-8"))
                identity_digest.update(b"\n")
                retained_source_rows_digest.update(str(source_row_number).encode("ascii"))
                retained_source_rows_digest.update(b"\n")
                writer.writerow(tuple(_cell_value(value) for value in selected_values))
                row_count += 1
            stream.flush()
            os.fsync(stream.fileno())
        if row_count < 1:
            raise DatasetPreparationError("The selected table contains no data rows after its header.")
        ordered_identity_sha256 = identity_digest.hexdigest()
        expected = contract.row_identity.expected_ordered_sha256
        if expected is not None and ordered_identity_sha256 != expected:
            raise DatasetPreparationError("Prepared source-row identity does not match expected_ordered_sha256.")
        os.replace(temporary, destination)
    finally:
        if temporary.exists():
            temporary.unlink()
    return {
        "worksheet": worksheet,
        "source_row_column": source_row_column,
        "header_row_index": contract.header_row_index,
        "header_row_indices": list(contract.header_row_indices),
        "header_join_separator": contract.header_join_separator,
        "empty_header_policy": contract.empty_header_policy,
        "duplicate_header_policy": contract.duplicate_header_policy,
        "selected_columns": list(selected_header),
        "cli_staging_preserves_source_columns": preserve_source_columns_for_cli,
        "input_row_count": input_row_count,
        "source_row_count": row_count,
        "filtered_row_count": input_row_count - row_count,
        "filters": [rule.model_dump(mode="json", exclude_none=True, exclude_defaults=True) for rule in contract.filters],
        "filter_result_sha256": retained_source_rows_digest.hexdigest(),
        "row_identity": {
            "strategy": contract.row_identity.strategy,
            "columns": list(contract.row_identity.columns),
            "ordered_sha256": ordered_identity_sha256,
        },
    }


def _load_cached(
    manifest_path: Path,
    destination: Path,
    expected_contract_hash: str,
    maximum_bytes: int,
) -> tuple[DatasetSnapshot, dict[str, Any]] | None:
    try:
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        if manifest.get("schema_version") != 1 or manifest.get("contract_hash") != expected_contract_hash:
            return None
        if not destination.is_file() or sha256_file(destination) != manifest["prepared_input"]["sha256"]:
            return None
        snapshot = snapshot_dataset(destination, maximum_bytes)
    except (OSError, KeyError, TypeError, ValueError, DatasetInspectionError, json.JSONDecodeError):
        return None
    return snapshot, manifest


def prepare_dataset_view(
    source_snapshot: DatasetSnapshot,
    contract: DatasetPreparationContract,
    state_root: Path,
    maximum_bytes: int,
    maximum_columns: int,
    *,
    allow_pandas_duplicate_mangling: bool = False,
    preserve_source_columns_for_cli: bool = False,
) -> PreparedDataset:
    """Create or reuse a deterministic CSV view without changing the source file."""
    contract_value = contract.model_dump(mode="json")
    source_mapping = _source_mapping_record(contract, maximum_bytes)
    if source_snapshot.format == "xlsx" and contract.worksheet is None and not contract.worksheets:
        workbook = load_workbook(source_snapshot.resolved_path, read_only=True, data_only=True)
        try:
            if len(workbook.sheetnames) != 1:
                sheet_names = ", ".join(repr(name) for name in workbook.sheetnames)
                raise DatasetPreparationError(f"Excel worksheet must be selected explicitly; available worksheets: {sheet_names}.")
        finally:
            workbook.close()
    materialize = bool(
        contract.worksheet is not None
        or contract.worksheets
        or contract.header_row_index
        or contract.header_row_indices
        or contract.selected_columns
        or contract.excluded_columns
        or contract.filters
        or contract.row_identity.strategy != "source_row"
    )
    if not materialize:
        expected_identity = contract.row_identity.expected_ordered_sha256
        actual_identity = source_snapshot.row_lineage.ordered_identity_sha256
        if expected_identity is not None and expected_identity != actual_identity:
            raise DatasetPreparationError("Source-row identity does not match expected_ordered_sha256.")
        record = {
            "schema_version": 1,
            "contract_hash": _canonical_sha256(
                {
                    "schema_version": 1,
                    "source_sha256": source_snapshot.sha256,
                    "source_format": source_snapshot.format,
                    "contract": contract_value,
                }
            ),
            "contract": contract_value,
            "source_file": {
                "path": str(source_snapshot.resolved_path),
                "format": source_snapshot.format,
                "size_bytes": source_snapshot.size_bytes,
                "sha256": source_snapshot.sha256,
            },
            "prepared_input": {
                "path": str(source_snapshot.resolved_path),
                "format": source_snapshot.format,
                "size_bytes": source_snapshot.size_bytes,
                "sha256": source_snapshot.sha256,
                "row_identity": source_snapshot.row_lineage.as_record(),
            },
            "table": {
                "worksheet": None,
                "header_row_index": 0,
                "selected_columns": [],
                "source_row_count": source_snapshot.row_lineage.source_row_count,
                "row_identity": {
                    "strategy": "source_row",
                    "columns": [],
                    "ordered_sha256": actual_identity,
                },
            },
            "declared_operations": list(contract.operations),
            "executed_view_operations": [],
            "source_mapping": source_mapping,
        }
        record["provenance"] = {
            "original_source": record["source_file"],
            "prepared_source": record["prepared_input"],
            "preparation_hash": record["contract_hash"],
        }
        return PreparedDataset(source_snapshot, source_snapshot, record)
    identity = {
        "schema_version": 1,
        "source_sha256": source_snapshot.sha256,
        "source_format": source_snapshot.format,
        "contract": contract_value,
        "preserve_source_columns_for_cli": preserve_source_columns_for_cli,
    }
    contract_hash = _canonical_sha256(identity)
    cache_root = Path(state_root).resolve() / "prepared-datasets"
    destination = cache_root / f"{contract_hash}.csv"
    manifest_path = cache_root / f"{contract_hash}.json"
    cached = _load_cached(manifest_path, destination, contract_hash, maximum_bytes)
    if cached is not None:
        prepared_snapshot, record = cached
        return PreparedDataset(source_snapshot, prepared_snapshot, record)

    table_record = _materialize(
        source_snapshot,
        contract,
        destination,
        maximum_columns,
        allow_pandas_duplicate_mangling,
        preserve_source_columns_for_cli,
    )
    prepared_snapshot = snapshot_dataset(destination, maximum_bytes)
    record = {
        "schema_version": 1,
        "contract_hash": contract_hash,
        "contract": contract_value,
        "source_file": {
            "path": str(source_snapshot.resolved_path),
            "format": source_snapshot.format,
            "size_bytes": source_snapshot.size_bytes,
            "sha256": source_snapshot.sha256,
        },
        "prepared_input": {
            "path": str(prepared_snapshot.resolved_path),
            "format": prepared_snapshot.format,
            "size_bytes": prepared_snapshot.size_bytes,
            "sha256": prepared_snapshot.sha256,
            "row_identity": prepared_snapshot.row_lineage.as_record(),
        },
        "table": table_record,
        "declared_operations": list(contract.operations),
        "executed_view_operations": [
            operation
            for enabled, operation in (
                (contract.worksheet is not None, "select_worksheet"),
                (bool(contract.worksheets), "union_worksheets_by_rows"),
                (contract.header_row_index != 0, "select_header_row"),
                (bool(contract.header_row_indices), "compose_header_rows"),
                (
                    bool(contract.selected_columns) and not preserve_source_columns_for_cli,
                    "select_columns",
                ),
                (
                    bool(contract.excluded_columns) and not preserve_source_columns_for_cli,
                    "exclude_columns",
                ),
                (bool(contract.filters), "filter_rows"),
                (contract.source_row_column is not None, "generate_source_row_column"),
                (contract.row_identity.strategy != "source_row", "verify_row_identity"),
            )
            if enabled
        ],
        "source_mapping": source_mapping,
    }
    record["provenance"] = {
        "original_source": record["source_file"],
        "prepared_source": record["prepared_input"],
        "preparation_hash": record["contract_hash"],
    }
    _atomic_write_json(manifest_path, record)
    return PreparedDataset(source_snapshot, prepared_snapshot, record)
