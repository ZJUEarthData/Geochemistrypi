"""Bounded, read-only inspection for explicitly named local datasets."""

import csv
import hashlib
import math
import os
import stat
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, BinaryIO, Iterable, Sequence

from openpyxl import load_workbook

from ..api.schemas import DatasetColumnSummary, DatasetInspectionRequest, DatasetInspectionResponse
from ..config.settings import McpSettings
from .headers import HeaderValidationError, header_normalization_warnings, normalize_dataset_header
from .row_identity import SourceRowIdentityError, SourceRowLineage, build_source_row_lineage
from .source_rows import iter_cli_csv_rows, iter_cli_excel_rows

_SUPPORTED_SUFFIXES = {".csv": "csv", ".xlsx": "xlsx"}
_TYPE_SAMPLE_ROWS = 50
_MAX_CELL_TEXT = 120


class DatasetInspectionError(ValueError):
    """Raised when a dataset cannot be inspected safely."""


@dataclass(frozen=True)
class DatasetSnapshot:
    """Identity recorded before a CLI run."""

    source_path: Path
    resolved_path: Path
    size_bytes: int
    sha256: str
    format: str
    row_lineage: SourceRowLineage


def _hash_stream(stream: BinaryIO) -> str:
    digest = hashlib.sha256()
    for chunk in iter(lambda: stream.read(1024 * 1024), b""):
        digest.update(chunk)
    return digest.hexdigest()


def _count_source_rows(path: Path, data_format: str) -> int:
    """Count physical data rows using the same CSV/XLSX boundaries as planning."""
    try:
        if data_format == "csv":
            with path.open("r", encoding="utf-8-sig", newline="") as stream:
                rows = csv.reader(stream)
                next(rows)
                return sum(1 for _ in iter_cli_csv_rows(rows))
        with path.open("rb") as stream:
            workbook = load_workbook(stream, read_only=True, data_only=True)
            try:
                rows = workbook.active.iter_rows(values_only=True)
                next(rows)
                return sum(1 for _ in iter_cli_excel_rows(rows))
            finally:
                workbook.close()
    except (OSError, StopIteration, UnicodeError, csv.Error, ValueError) as exc:
        raise DatasetInspectionError(f"Unable to establish source-row lineage: {exc}") from exc


def snapshot_dataset(path: Path, maximum_bytes: int) -> DatasetSnapshot:
    """Validate and hash one absolute regular file without modifying it."""
    source = Path(path).expanduser()
    if not source.is_absolute():
        raise DatasetInspectionError("dataset_path must be absolute.")
    try:
        resolved = source.resolve(strict=True)
        metadata = resolved.stat()
    except (OSError, RuntimeError) as exc:
        raise DatasetInspectionError(f"Dataset is unavailable: {source}") from exc
    if not stat.S_ISREG(metadata.st_mode):
        raise DatasetInspectionError("dataset_path must identify a regular file, not a directory or device.")
    data_format = _SUPPORTED_SUFFIXES.get(resolved.suffix.lower())
    if data_format is None:
        raise DatasetInspectionError("Supported dataset formats are .csv and .xlsx.")
    if metadata.st_size > maximum_bytes:
        raise DatasetInspectionError(f"Dataset size {metadata.st_size} bytes exceeds the configured {maximum_bytes}-byte limit.")
    try:
        with resolved.open("rb") as stream:
            opened_metadata = os.fstat(stream.fileno())
            if not stat.S_ISREG(opened_metadata.st_mode):
                raise DatasetInspectionError("The opened dataset is not a regular file.")
            if opened_metadata.st_size > maximum_bytes:
                raise DatasetInspectionError(f"Dataset size {opened_metadata.st_size} bytes exceeds the configured {maximum_bytes}-byte limit.")
            digest = _hash_stream(stream)
            if os.fstat(stream.fileno()).st_size != opened_metadata.st_size:
                raise DatasetInspectionError("Dataset size changed while it was being hashed; retry after writes have stopped.")
    except OSError as exc:
        raise DatasetInspectionError(f"Dataset cannot be read: {resolved}") from exc
    row_count = _count_source_rows(resolved, data_format)
    if sha256_file(resolved) != digest:
        raise DatasetInspectionError("Dataset changed while source-row lineage was being established; retry after writes have stopped.")
    try:
        row_lineage = build_source_row_lineage(digest, row_count)
    except SourceRowIdentityError as exc:
        raise DatasetInspectionError(str(exc)) from exc
    return DatasetSnapshot(source, resolved, opened_metadata.st_size, digest, data_format, row_lineage)


def sha256_file(path: Path) -> str:
    """Hash an existing run input for post-run integrity verification."""
    with Path(path).open("rb") as stream:
        return _hash_stream(stream)


def _safe_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return None if not math.isfinite(value) else value
    text = str(value)
    return text if len(text) <= _MAX_CELL_TEXT else f"{text[: _MAX_CELL_TEXT - 1]}…"


def _value_type(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, int):
        return "integer"
    if isinstance(value, float):
        return "number"
    text = str(value).strip()
    if not text:
        return None
    if text.lower() in {"true", "false"}:
        return "boolean"
    try:
        int(text)
    except ValueError:
        try:
            number = float(text)
        except ValueError:
            return "string"
        return "number" if math.isfinite(number) else "string"
    return "integer"


def _column_summaries(columns: Sequence[str], type_rows: Iterable[Sequence[Any]]) -> tuple[DatasetColumnSummary, ...]:
    observed: list[set[str]] = [set() for _ in columns]
    non_null = [0 for _ in columns]
    for row in type_rows:
        for index, value in enumerate(row[: len(columns)]):
            value_type = _value_type(value)
            if value_type is not None:
                observed[index].add(value_type)
                non_null[index] += 1
    summaries = []
    for index, name in enumerate(columns):
        types = observed[index]
        if not types:
            inferred = "empty"
        elif types == {"integer", "number"}:
            inferred = "number"
        elif len(types) == 1:
            inferred = next(iter(types))
        else:
            inferred = "mixed"
        summaries.append(DatasetColumnSummary(name=name, inferred_type=inferred, sampled_non_null=non_null[index]))
    return tuple(summaries)


def _validate_header(
    raw_header: Sequence[Any],
    maximum_columns: int,
    allow_pandas_duplicate_mangling: bool,
) -> tuple[tuple[str, ...], tuple[str, ...]]:
    try:
        columns = normalize_dataset_header(
            raw_header,
            maximum_columns,
            allow_pandas_duplicate_mangling=allow_pandas_duplicate_mangling,
        )
        return columns, header_normalization_warnings(raw_header, columns)
    except HeaderValidationError as exc:
        raise DatasetInspectionError(str(exc)) from exc


def _inspect_csv(
    snapshot: DatasetSnapshot,
    request: DatasetInspectionRequest,
    settings: McpSettings,
    allow_pandas_duplicate_mangling: bool,
) -> tuple[tuple[str, ...], tuple[str, ...], int, bool, list[list[Any]], list[list[Any]],]:
    try:
        with snapshot.resolved_path.open("r", encoding="utf-8-sig", newline="") as stream:
            reader = csv.reader(stream)
            columns, header_warnings = _validate_header(
                next(reader),
                settings.maximum_columns,
                allow_pandas_duplicate_mangling,
            )
            returned_rows: list[list[Any]] = []
            type_rows: list[list[Any]] = []
            row_count = 0
            for row in iter_cli_csv_rows(reader):
                row_count += 1
                normalized = list(row[: len(columns)]) + [None] * max(0, len(columns) - len(row))
                if len(type_rows) < _TYPE_SAMPLE_ROWS:
                    type_rows.append(normalized)
                if len(returned_rows) < request.sample_rows:
                    returned_rows.append(normalized)
    except (OSError, StopIteration, UnicodeError, csv.Error) as exc:
        raise DatasetInspectionError(f"Unable to inspect CSV dataset: {exc}") from exc
    return columns, header_warnings, row_count, True, returned_rows, type_rows


def _inspect_xlsx(
    snapshot: DatasetSnapshot,
    request: DatasetInspectionRequest,
    settings: McpSettings,
    allow_pandas_duplicate_mangling: bool,
) -> tuple[tuple[str, ...], tuple[str, ...], int, bool, list[list[Any]], list[list[Any]],]:
    try:
        with snapshot.resolved_path.open("rb") as stream:
            workbook = load_workbook(stream, read_only=True, data_only=True)
            try:
                worksheet = workbook.active
                rows = worksheet.iter_rows(values_only=True)
                columns, header_warnings = _validate_header(
                    next(rows),
                    settings.maximum_columns,
                    allow_pandas_duplicate_mangling,
                )
                returned_rows: list[list[Any]] = []
                type_rows: list[list[Any]] = []
                for row in iter_cli_excel_rows(rows):
                    normalized = list(row[: len(columns)]) + [None] * max(0, len(columns) - len(row))
                    if len(type_rows) < _TYPE_SAMPLE_ROWS:
                        type_rows.append(normalized)
                    if len(returned_rows) < request.sample_rows:
                        returned_rows.append(normalized)
                    if len(type_rows) >= _TYPE_SAMPLE_ROWS and len(returned_rows) >= request.sample_rows:
                        break
                row_count = snapshot.row_lineage.source_row_count
            finally:
                workbook.close()
    except (OSError, StopIteration, ValueError) as exc:
        raise DatasetInspectionError(f"Unable to inspect Excel dataset: {exc}") from exc
    return columns, header_warnings, row_count, True, returned_rows, type_rows


def inspect_dataset(
    request: DatasetInspectionRequest,
    settings: McpSettings,
    *,
    allow_pandas_duplicate_mangling: bool = False,
) -> DatasetInspectionResponse:
    """Return a bounded dataset view; never return the complete table."""
    snapshot = snapshot_dataset(request.dataset_path, settings.maximum_dataset_bytes)
    if snapshot.format == "csv":
        columns, header_warnings, row_count, exact, returned_rows, type_rows = _inspect_csv(
            snapshot,
            request,
            settings,
            allow_pandas_duplicate_mangling,
        )
    else:
        columns, header_warnings, row_count, exact, returned_rows, type_rows = _inspect_xlsx(
            snapshot,
            request,
            settings,
            allow_pandas_duplicate_mangling,
        )
    if sha256_file(snapshot.resolved_path) != snapshot.sha256:
        raise DatasetInspectionError("Dataset changed during inspection; retry after writes have stopped.")
    sample = tuple({column: _safe_value(row[index]) for index, column in enumerate(columns)} for row in returned_rows)
    return DatasetInspectionResponse(
        source_path=str(snapshot.source_path),
        resolved_path=str(snapshot.resolved_path),
        format=snapshot.format,
        size_bytes=snapshot.size_bytes,
        sha256=snapshot.sha256,
        row_count=row_count,
        row_count_exact=exact,
        column_count=len(columns),
        columns=_column_summaries(columns, type_rows),
        header_warnings=header_warnings,
        sample_rows=sample,
        sample_truncated=row_count > len(sample),
    )
