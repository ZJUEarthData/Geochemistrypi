"""Fail-closed pairing of original CLI rows to internal MCP source identities."""

import csv
import hashlib
import math
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from .headers import HeaderValidationError, normalize_dataset_header
from .row_identity import SourceRowIdentityError, SourceRowLineage
from .source_rows import iter_cli_csv_rows, iter_cli_excel_rows


def _canonical_number(number: Decimal) -> str:
    if not number.is_finite():
        return f"non-finite:{number}"
    return f"number:{number.normalize()}"


def _canonical_value(
    value: Any,
    *,
    parse_numeric_text: bool = False,
) -> str:
    if value is None or value == "" or isinstance(value, float) and math.isnan(value):
        return "missing"
    if isinstance(value, bool):
        return f"boolean:{str(value).lower()}"
    if isinstance(value, (int, float, Decimal)) and not isinstance(value, bool):
        number = Decimal(str(value))
        return _canonical_number(number)
    if isinstance(value, (datetime, date)):
        return f"date:{value.isoformat()}"
    if parse_numeric_text and isinstance(value, str):
        try:
            number = Decimal(value)
        except (InvalidOperation, ValueError):
            pass
        else:
            if number.is_finite():
                return _canonical_number(number)
    return f"string:{value}"


def _numeric_value(value: Any, *, parse_numeric_text: bool) -> Decimal | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    if parse_numeric_text and isinstance(value, str):
        try:
            return Decimal(value)
        except (InvalidOperation, ValueError):
            return None
    return None


def _values_equivalent(
    source_value: Any,
    output_value: Any,
    *,
    parse_source_numeric_text: bool,
    parse_output_numeric_text: bool,
    output_is_xlsx: bool,
) -> bool:
    source_number = _numeric_value(source_value, parse_numeric_text=parse_source_numeric_text)
    output_number = _numeric_value(output_value, parse_numeric_text=parse_output_numeric_text)
    if source_number is not None or output_number is not None:
        if source_number is None or output_number is None:
            return False
        if source_number.is_finite() and output_number.is_finite() and output_is_xlsx:
            magnitude = max(abs(source_number), abs(output_number))
            tolerance = magnitude * Decimal("1e-14")
            return abs(source_number - output_number) <= tolerance
        return source_number == output_number
    return _canonical_value(source_value) == _canonical_value(output_value)


def _column_position(header: Iterable[str], identifier_column: str, path: Path) -> int:
    header = tuple(header)
    positions = [index for index, column in enumerate(header) if column == identifier_column]
    if len(positions) != 1:
        raise SourceRowIdentityError(f"Cannot pair rows through scientific identifier {identifier_column!r} in {path.name}; " f"expected exactly one matching column, found {len(positions)}.")
    return positions[0]


def _read_table(path: Path) -> tuple[tuple[Any, ...], tuple[tuple[Any, ...], ...]]:
    try:
        if path.suffix.lower() == ".csv":
            with path.open("r", encoding="utf-8-sig", newline="") as stream:
                rows = csv.reader(stream)
                header = tuple(next(rows))
                values = []
                for row_number, row in enumerate(iter_cli_csv_rows(rows), start=2):
                    if len(row) > len(header):
                        raise SourceRowIdentityError(f"Cannot pair CSV row {row_number}; it has more values than the header.")
                    values.append(tuple(row) + (None,) * (len(header) - len(row)))
                return header, tuple(values)
        with path.open("rb") as stream:
            workbook = load_workbook(stream, read_only=True, data_only=True)
            try:
                rows = workbook.active.iter_rows(values_only=True)
                header = tuple(next(rows))
                return header, tuple(iter_cli_excel_rows(rows))
            finally:
                workbook.close()
    except SourceRowIdentityError:
        raise
    except (OSError, StopIteration, UnicodeError, csv.Error, ValueError) as exc:
        raise SourceRowIdentityError(f"Unable to read row-pairing data from {path}: {exc}") from exc


def verify_original_row_pairing(
    source_path: Path,
    output_directory: Path,
    identifier_column: str,
    lineage: SourceRowLineage,
) -> dict[str, Any]:
    """Verify source order/count and bind each original CLI row to an internal ID."""
    data_directory = output_directory / "artifacts" / "data"
    candidates = tuple(path for path in (data_directory / "Data Original.xlsx", data_directory / "Data Original.csv") if path.is_file())
    if len(candidates) != 1:
        raise SourceRowIdentityError("Original CLI row pairing requires exactly one Data Original.xlsx or Data Original.csv artifact; " f"found {len(candidates)}.")
    source_header, source_rows = _read_table(source_path)
    output_header, output_rows = _read_table(candidates[0])
    try:
        expected_header = normalize_dataset_header(
            source_header,
            len(source_header),
            allow_pandas_duplicate_mangling=True,
        )
        observed_header = normalize_dataset_header(output_header, len(output_header))
    except HeaderValidationError as exc:
        raise SourceRowIdentityError(f"Cannot pair CLI Data Original headers safely: {exc}") from exc
    if observed_header != expected_header:
        raise SourceRowIdentityError("CLI Data Original columns do not preserve the normalized source schema and order; row pairing is unsafe.")
    _column_position(expected_header, identifier_column, source_path)
    if len(source_rows) != lineage.source_row_count:
        raise SourceRowIdentityError(f"Internal source-row count {lineage.source_row_count} does not match the {len(source_rows)} source rows read for pairing.")
    if len(output_rows) != lineage.source_row_count:
        raise SourceRowIdentityError(f"CLI Data Original row count {len(output_rows)} does not match source row count {lineage.source_row_count}.")
    parse_source_numeric_text = source_path.suffix.lower() == ".csv"
    parse_output_numeric_text = candidates[0].suffix.lower() == ".csv"
    output_is_xlsx = candidates[0].suffix.lower() == ".xlsx"
    pairing_digest = hashlib.sha256()
    for row_number, (internal_identity, source_row, output_row) in enumerate(
        zip(lineage.identities, source_rows, output_rows),
        start=1,
    ):
        if any(
            not _values_equivalent(
                source_value,
                output_value,
                parse_source_numeric_text=parse_source_numeric_text,
                parse_output_numeric_text=parse_output_numeric_text,
                output_is_xlsx=output_is_xlsx,
            )
            for source_value, output_value in zip(source_row, output_row)
        ):
            raise SourceRowIdentityError(f"CLI Data Original changed or reordered source row {row_number}; row pairing is unsafe.")
        canonical_source = tuple(_canonical_value(value, parse_numeric_text=parse_source_numeric_text) for value in source_row)
        pairing_digest.update(internal_identity.encode("ascii"))
        for value in canonical_source:
            encoded = value.encode("utf-8")
            pairing_digest.update(len(encoded).to_bytes(8, "big"))
            pairing_digest.update(encoded)
        pairing_digest.update(b"\n")
    return {
        "verified": True,
        "artifact_relative_path": candidates[0].relative_to(output_directory).as_posix(),
        "scientific_identifier_column": identifier_column,
        "scientific_identifier_values_preserved": True,
        "source_rows_and_order_preserved": True,
        "numeric_comparison_policy": "xlsx_relative_1e-14" if output_is_xlsx else "exact",
        "source_row_count": lineage.source_row_count,
        "ordered_pairing_sha256": pairing_digest.hexdigest(),
    }
