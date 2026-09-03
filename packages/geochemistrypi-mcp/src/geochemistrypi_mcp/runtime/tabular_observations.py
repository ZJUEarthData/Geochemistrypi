"""Bounded, read-only observations over requirement-bound native output tables."""

from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path, PurePosixPath
from typing import Any, Iterable, Iterator

from openpyxl import load_workbook

from ..api.schemas import ArtifactReference, RequiredTabularObservation, RequiredTabularObservationSummary
from ..config.constants import ARTIFACT_INDEX_SCHEMA_VERSION
from ..data.inspector import sha256_file
from .result_views import partition_artifact_views

_SUPPORTED_SUFFIXES = {".csv", ".xlsx", ".json", ".txt"}
_MAX_REQUIRED_TABULAR_ARTIFACTS = 256
_MAX_TABULAR_FILE_BYTES = 64 * 1024 * 1024
_MAX_JSON_TABLE_BYTES = 16 * 1024 * 1024
_MAX_XLSX_MEMBERS = 4096
_MAX_XLSX_UNCOMPRESSED_BYTES = 256 * 1024 * 1024
_MAX_XLSX_COMPRESSION_RATIO = 2000
_MAX_OBSERVATIONS = 32
_MAX_COLUMNS_RETURNED = 64
_MAX_COMPLETE_TABLE_CELLS = 512
_MAX_TOTAL_RETURNED_CELLS = 512
_MAX_CELL_UTF8_BYTES = 256
_MAX_REQUIREMENT_IDS_RETURNED = 4
_MAX_OBSERVATION_JSON_BYTES = 16 * 1024

_SCIENTIFIC_TYPE_PRIORITY = {
    "feature_importance_table": 0,
    "time_series_bin_table": 1,
    "time_series_uncertainty": 1,
    "component_loadings": 2,
    "explained_variance": 2,
    "confusion_matrix_table": 3,
    "normalized_confusion_matrix_table": 3,
    "score_table": 4,
    "anomaly_scores": 4,
    "event_association_table": 5,
    "reference_anomaly_joined_table": 6,
    "embedding_label_joined_table": 6,
    "prediction_table": 20,
    "external_evaluation_table": 20,
    "evaluation_labels": 21,
    "residual_table": 22,
    "cluster_assignments": 30,
    "embedding_coordinates": 30,
    "anomaly_assignments": 30,
    "anomaly_labels": 30,
    "anomaly_subset": 31,
    "split_membership": 40,
    "machine_readable_table": 50,
}


class TabularObservationError(RuntimeError):
    """Raised when an indexed native table cannot be inspected safely or exactly."""


class TabularObservationUnavailable(RuntimeError):
    """A convenience view was safely omitted without changing scientific state."""

    def __init__(self, reason: str, message: str):
        super().__init__(message)
        self.reason = reason


@dataclass(frozen=True)
class _NativeTable:
    reference: ArtifactReference
    format: str
    sheet: str | None
    row_count: int
    column_count: int
    full_columns: tuple[str, ...]
    returned_columns: tuple[str, ...]
    columns_truncated: bool
    complete_rows: tuple[tuple[Any, ...], ...] | None
    intrinsic_omission_reason: str | None

    @property
    def complete_cell_count(self) -> int:
        if self.complete_rows is None:
            return 0
        return sum(len(row) for row in self.complete_rows)


def _canonical_json(value: Any) -> bytes:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    ).encode("utf-8")


def _canonical_sha256(value: Any) -> str:
    return hashlib.sha256(_canonical_json(value)).hexdigest()


def _json_size(value: Any) -> int:
    return len(
        json.dumps(
            value,
            ensure_ascii=True,
            separators=(",", ":"),
            allow_nan=False,
        ).encode("utf-8")
    )


def _normalize_cell(value: Any) -> Any:
    if value is None or isinstance(value, (bool, int, str)):
        return value
    if isinstance(value, float):
        if math.isnan(value):
            return "NaN"
        if math.isinf(value):
            return "Infinity" if value > 0 else "-Infinity"
        return value
    if isinstance(value, (date, datetime, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return str(value)


def _column_text(value: Any) -> str:
    normalized = _normalize_cell(value)
    return "" if normalized is None else str(normalized)


def _truncate_utf8(value: str, maximum_bytes: int) -> tuple[str, bool]:
    encoded = value.encode("utf-8")
    if len(encoded) <= maximum_bytes:
        return value, False
    prefix = encoded[:maximum_bytes]
    while prefix:
        try:
            return prefix.decode("utf-8"), True
        except UnicodeDecodeError:
            prefix = prefix[:-1]
    return "", True


def _is_nonempty_row(row: tuple[Any, ...]) -> bool:
    return any(value is not None and value != "" for value in row)


def _table_from_rows(
    reference: ArtifactReference,
    table_format: str,
    sheet: str | None,
    rows: Iterable[Iterable[Any]],
) -> _NativeTable | None:
    iterator: Iterator[tuple[Any, ...]] = (tuple(_normalize_cell(value) for value in row) for row in rows)
    header: tuple[Any, ...] | None = None
    for row in iterator:
        if _is_nonempty_row(row):
            header = row
            break
    if header is None:
        return None

    row_count = 0
    column_count = len(header)
    retained_rows: list[tuple[Any, ...]] | None = []
    cell_length_exceeded = False
    for row in iterator:
        if not _is_nonempty_row(row):
            continue
        row_count += 1
        column_count = max(column_count, len(row))
        if retained_rows is None:
            continue
        if any(isinstance(value, str) and len(value.encode("utf-8")) > _MAX_CELL_UTF8_BYTES for value in row):
            cell_length_exceeded = True
            retained_rows = None
            continue
        retained_rows.append(row)
        if row_count * column_count > _MAX_COMPLETE_TABLE_CELLS:
            retained_rows = None

    if column_count < 1:
        return None
    full_columns = tuple(_column_text(header[index]) if index < len(header) else "" for index in range(column_count))
    returned_column_values = []
    column_text_truncated = False
    for value in full_columns[:_MAX_COLUMNS_RETURNED]:
        bounded, truncated = _truncate_utf8(value, _MAX_CELL_UTF8_BYTES)
        returned_column_values.append(bounded)
        column_text_truncated = column_text_truncated or truncated
    returned_columns = tuple(returned_column_values)
    columns_truncated = len(returned_columns) < column_count or column_text_truncated

    omission_reason = None
    complete_rows = None
    if column_count > _MAX_COLUMNS_RETURNED:
        omission_reason = "column_limit"
    elif cell_length_exceeded:
        omission_reason = "cell_length_limit"
    elif retained_rows is None:
        omission_reason = "large_table"
    else:
        complete_rows = tuple(tuple((*row, *((None,) * (column_count - len(row))))) for row in retained_rows)
    return _NativeTable(
        reference=reference,
        format=table_format,
        sheet=sheet,
        row_count=row_count,
        column_count=column_count,
        full_columns=full_columns,
        returned_columns=returned_columns,
        columns_truncated=columns_truncated,
        complete_rows=complete_rows,
        intrinsic_omission_reason=omission_reason,
    )


def _csv_tables(reference: ArtifactReference, payload: bytes) -> tuple[_NativeTable, ...]:
    try:
        text = payload.decode("utf-8-sig", errors="strict")
    except UnicodeDecodeError as exc:
        raise TabularObservationUnavailable("parse_unavailable", f"Indexed CSV output {reference.relative_path!r} is not valid UTF-8.") from exc
    try:
        table = _table_from_rows(
            reference,
            "csv",
            None,
            csv.reader(io.StringIO(text, newline="")),
        )
    except csv.Error as exc:
        raise TabularObservationUnavailable("parse_unavailable", f"Indexed CSV output {reference.relative_path!r} is malformed.") from exc
    return () if table is None else (table,)


def _validate_xlsx_container(reference: ArtifactReference, payload: bytes) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(payload)) as archive:
            members = archive.infolist()
            if len(members) > _MAX_XLSX_MEMBERS:
                raise TabularObservationUnavailable("xlsx_safety_limit", f"Indexed XLSX output {reference.relative_path!r} exceeds the ZIP member limit.")
            total_uncompressed = 0
            for member in members:
                normalized = PurePosixPath(member.filename.replace("\\", "/"))
                if normalized.is_absolute() or ".." in normalized.parts or member.flag_bits & 0x1:
                    raise TabularObservationUnavailable("xlsx_safety_limit", f"Indexed XLSX output {reference.relative_path!r} has an unsafe ZIP member.")
                total_uncompressed += member.file_size
                if total_uncompressed > _MAX_XLSX_UNCOMPRESSED_BYTES:
                    raise TabularObservationUnavailable("xlsx_safety_limit", f"Indexed XLSX output {reference.relative_path!r} exceeds the uncompressed safety limit.")
                if member.compress_size and member.file_size / member.compress_size > _MAX_XLSX_COMPRESSION_RATIO:
                    raise TabularObservationUnavailable("xlsx_safety_limit", f"Indexed XLSX output {reference.relative_path!r} exceeds the compression-ratio safety limit.")
    except zipfile.BadZipFile as exc:
        raise TabularObservationUnavailable("parse_unavailable", f"Indexed XLSX output {reference.relative_path!r} is not a valid XLSX container.") from exc


def _xlsx_tables(reference: ArtifactReference, payload: bytes) -> tuple[_NativeTable, ...]:
    _validate_xlsx_container(reference, payload)
    try:
        workbook = load_workbook(
            io.BytesIO(payload),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:  # openpyxl exposes multiple parse exception types
        raise TabularObservationUnavailable("parse_unavailable", f"Indexed XLSX output {reference.relative_path!r} cannot be read safely.") from exc
    try:
        tables = []
        for worksheet in workbook.worksheets:
            table = _table_from_rows(
                reference,
                "xlsx",
                worksheet.title,
                worksheet.iter_rows(values_only=True),
            )
            if table is not None:
                tables.append(table)
        return tuple(tables)
    finally:
        workbook.close()


def _json_table_rows(value: Any) -> Iterable[Iterable[Any]] | None:
    if isinstance(value, dict):
        columns = value.get("columns")
        rows = value.get("rows")
        if isinstance(columns, list) and isinstance(rows, list) and all(isinstance(row, list) for row in rows):
            return [columns, *rows]
        if value and all(isinstance(column, str) and isinstance(items, list) for column, items in value.items()):
            lengths = {len(items) for items in value.values()}
            if len(lengths) == 1:
                ordered_columns = list(value)
                return [
                    ordered_columns,
                    *([value[column][index] for column in ordered_columns] for index in range(next(iter(lengths)))),
                ]
        return None
    if not isinstance(value, list) or not value:
        return None
    if all(isinstance(item, dict) for item in value):
        columns: list[str] = []
        for item in value:
            for key in item:
                key_text = str(key)
                if key_text not in columns:
                    columns.append(key_text)
        return [columns, *([item.get(column) for column in columns] for item in value)]
    if all(isinstance(item, list) for item in value):
        return value
    if all(item is None or isinstance(item, (bool, int, float, str)) for item in value):
        return [["value"], *([item] for item in value)]
    return None


def _json_tables(
    reference: ArtifactReference,
    payload: bytes,
    table_format: str,
) -> tuple[_NativeTable, ...]:
    if len(payload) > _MAX_JSON_TABLE_BYTES:
        raise TabularObservationUnavailable("json_size_limit", f"Indexed {table_format.upper()} output {reference.relative_path!r} exceeds the JSON table safety limit.")
    try:
        value = json.loads(payload.decode("utf-8-sig", errors="strict"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        # Plain-text parameter reports and other non-JSON TXT outputs are not
        # tables. They remain fully represented by the immutable artifact index.
        if table_format == "txt":
            return ()
        raise TabularObservationUnavailable(
            "parse_unavailable",
            f"Indexed JSON output {reference.relative_path!r} is malformed.",
        ) from exc
    rows = _json_table_rows(value)
    if rows is None:
        return ()
    table = _table_from_rows(reference, table_format, None, rows)
    return () if table is None else (table,)


def _reference_priority(reference: ArtifactReference) -> tuple[int, str, str]:
    return (
        _SCIENTIFIC_TYPE_PRIORITY.get(reference.scientific_type or "", 100),
        reference.relative_path.casefold(),
        reference.artifact_id,
    )


def _mirror_key(reference: ArtifactReference) -> tuple[tuple[str, ...], str, int, str] | None:
    parts = PurePosixPath(reference.relative_path.replace("\\", "/")).parts
    try:
        category_position = parts.index(reference.category)
    except ValueError:
        return None
    if category_position not in {0, 1}:
        return None
    tail = parts[category_position + 1 :]
    if not tail or (reference.category == "summary" and len(tail) != 1):
        return None
    return (tuple(parts[:category_position]), parts[-1], reference.size_bytes, reference.sha256)


def _without_duplicate_summary_tables(
    references: tuple[ArtifactReference, ...],
) -> tuple[ArtifactReference, ...]:
    sources: dict[tuple[tuple[str, ...], str, int, str], list[ArtifactReference]] = {}
    for reference in references:
        if reference.category == "summary":
            continue
        key = _mirror_key(reference)
        if key is not None:
            sources.setdefault(key, []).append(reference)
    retained = []
    for reference in references:
        key = _mirror_key(reference)
        if reference.category == "summary" and key is not None and len(sources.get(key, ())) == 1:
            continue
        retained.append(reference)
    return tuple(retained)


def _safe_indexed_path(output_directory: Path, reference: ArtifactReference) -> Path:
    normalized = reference.relative_path.replace("\\", "/")
    relative = PurePosixPath(normalized)
    if relative.is_absolute() or not relative.parts or ".." in relative.parts:
        raise TabularObservationError("The artifact index contains an unsafe tabular relative path.")
    output_root = output_directory.resolve()
    candidate = output_root.joinpath(*relative.parts)
    try:
        resolved = candidate.resolve(strict=True)
    except OSError as exc:
        raise TabularObservationError(f"Indexed tabular output {reference.relative_path!r} is unavailable.") from exc
    try:
        resolved.relative_to(output_root)
    except ValueError as exc:
        raise TabularObservationError("An indexed tabular output escapes the managed output root.") from exc
    current = candidate
    while current != output_root:
        if current.is_symlink():
            raise TabularObservationError("Indexed tabular outputs and their parent paths must not be symlinks.")
        current = current.parent
    if not candidate.is_file() or candidate.is_symlink():
        raise TabularObservationError("An indexed tabular output is not a regular file.")
    try:
        indexed_local = Path(reference.local_path).resolve(strict=True)
    except OSError as exc:
        raise TabularObservationError("The indexed tabular local path is unavailable.") from exc
    if indexed_local != resolved:
        raise TabularObservationError("The indexed tabular local path does not match its managed relative path.")
    return candidate


def _read_verified_payload(output_directory: Path, reference: ArtifactReference) -> bytes:
    path = _safe_indexed_path(output_directory, reference)
    before = path.stat()
    if before.st_size != reference.size_bytes:
        raise TabularObservationError(f"Indexed tabular output {reference.relative_path!r} changed size before observation.")
    before_sha256 = sha256_file(path)
    if before_sha256 != reference.sha256:
        raise TabularObservationError(f"Indexed tabular output {reference.relative_path!r} failed its pre-read SHA-256 check.")
    if before.st_size > _MAX_TABULAR_FILE_BYTES:
        raise TabularObservationUnavailable(
            "file_size_limit",
            f"Indexed tabular output {reference.relative_path!r} exceeds the read-only observation safety limit.",
        )
    payload = path.read_bytes()
    if hashlib.sha256(payload).hexdigest() != reference.sha256:
        raise TabularObservationError(f"Indexed tabular output {reference.relative_path!r} changed during observation.")
    after = path.stat()
    if after.st_size != reference.size_bytes or sha256_file(path) != reference.sha256:
        raise TabularObservationError(f"Indexed tabular output {reference.relative_path!r} failed its post-read SHA-256 check.")
    return payload


def _tables_for_reference(
    output_directory: Path,
    reference: ArtifactReference,
) -> tuple[_NativeTable, ...]:
    payload = _read_verified_payload(output_directory, reference)
    suffix = PurePosixPath(reference.relative_path).suffix.lower()
    if suffix == ".csv":
        return _csv_tables(reference, payload)
    if suffix == ".xlsx":
        return _xlsx_tables(reference, payload)
    if suffix in {".json", ".txt"}:
        return _json_tables(reference, payload, suffix[1:])
    return ()


def _observation_identity(table: _NativeTable) -> dict[str, Any]:
    reference = table.reference
    requirement_ids = tuple(
        dict.fromkeys(
            (
                *((reference.requirement_id,) if reference.requirement_id is not None else ()),
                *reference.requirement_ids,
            )
        )
    )
    return {
        "artifact_id": reference.artifact_id,
        "relative_path": reference.relative_path,
        "requirement_ids_sha256": _canonical_sha256(list(requirement_ids)),
        "sha256": reference.sha256,
        "size_bytes": reference.size_bytes,
        "format": table.format,
        "sheet": table.sheet,
        "row_count": table.row_count,
        "column_count": table.column_count,
        "columns_sha256": _canonical_sha256(list(table.full_columns)),
    }


def _make_observation(
    table: _NativeTable,
    *,
    rows: tuple[tuple[Any, ...], ...] | None,
    omission_reason: str | None,
) -> RequiredTabularObservation:
    reference = table.reference
    requirement_ids = tuple(
        dict.fromkeys(
            (
                *((reference.requirement_id,) if reference.requirement_id is not None else ()),
                *reference.requirement_ids,
            )
        )
    )
    compact_ids = requirement_ids[:_MAX_REQUIREMENT_IDS_RETURNED]
    delivered_rows = rows or ()
    return RequiredTabularObservation(
        artifact_id=reference.artifact_id,
        relative_path=reference.relative_path,
        requirement_ids=compact_ids,
        requirement_ids_total_count=len(requirement_ids),
        requirement_ids_truncated=len(compact_ids) < len(requirement_ids),
        requirement_ids_sha256=_canonical_sha256(list(requirement_ids)),
        sha256=reference.sha256,
        size_bytes=reference.size_bytes,
        format=table.format,
        sheet=table.sheet,
        row_count=table.row_count,
        column_count=table.column_count,
        columns=table.returned_columns,
        columns_truncated=table.columns_truncated,
        columns_sha256=_canonical_sha256(list(table.full_columns)),
        rows_included=rows is not None,
        rows=delivered_rows,
        returned_cell_count=sum(len(row) for row in delivered_rows),
        rows_omission_reason=omission_reason,
    )


def build_required_tabular_observations(
    output_directory: Path,
    artifact_index_path: Path,
    expected_artifact_index_sha256: str,
) -> RequiredTabularObservationSummary:
    """Build one deterministic, globally bounded view from the immutable index.

    The function never accepts caller-provided artifact paths. It resolves only
    indexed paths below ``output_directory`` and validates the artifact-index
    and file SHA-256 values before and after every read.
    """

    artifact_index_path = Path(artifact_index_path)
    if artifact_index_path.is_symlink() or not artifact_index_path.is_file():
        raise TabularObservationError("The immutable artifact index is unavailable or unsafe.")
    if sha256_file(artifact_index_path) != expected_artifact_index_sha256:
        raise TabularObservationError("The artifact index failed its pre-read SHA-256 check.")
    try:
        index = json.loads(artifact_index_path.read_text(encoding="utf-8"))
    except (OSError, UnicodeError, json.JSONDecodeError) as exc:
        raise TabularObservationError("The immutable artifact index is malformed.") from exc
    if index.get("schema_version") != ARTIFACT_INDEX_SCHEMA_VERSION or not isinstance(index.get("artifacts"), list):
        raise TabularObservationError("The immutable artifact index has an unsupported identity.")
    references = tuple(ArtifactReference.model_validate(item) for item in index["artifacts"])
    canonical = partition_artifact_views(references).canonical_entries
    eligible = tuple(
        reference for reference in canonical if (reference.requirement_id is not None or reference.requirement_ids) and PurePosixPath(reference.relative_path).suffix.lower() in _SUPPORTED_SUFFIXES
    )
    eligible = tuple(sorted(_without_duplicate_summary_tables(eligible), key=_reference_priority))

    tables: list[_NativeTable] = []
    omissions: list[dict[str, str]] = []
    for reference in eligible[:_MAX_REQUIRED_TABULAR_ARTIFACTS]:
        try:
            observed_tables = _tables_for_reference(Path(output_directory), reference)
        except TabularObservationUnavailable as exc:
            omissions.append(
                {
                    "artifact_id": reference.artifact_id,
                    "relative_path": reference.relative_path,
                    "sha256": reference.sha256,
                    "reason": exc.reason,
                }
            )
            continue
        if not observed_tables:
            omissions.append(
                {
                    "artifact_id": reference.artifact_id,
                    "relative_path": reference.relative_path,
                    "sha256": reference.sha256,
                    "reason": "not_tabular",
                }
            )
            continue
        tables.extend(observed_tables)
    for reference in eligible[_MAX_REQUIRED_TABULAR_ARTIFACTS:]:
        omissions.append(
            {
                "artifact_id": reference.artifact_id,
                "relative_path": reference.relative_path,
                "sha256": reference.sha256,
                "reason": "artifact_limit",
            }
        )
    tables.sort(
        key=lambda table: (
            *_reference_priority(table.reference),
            (table.sheet or "").casefold(),
        )
    )
    identities = [_observation_identity(table) for table in tables]
    observations_sha256 = _canonical_sha256(identities)

    observations: list[RequiredTabularObservation] = []
    used_cells = 0
    for table in tables:
        if len(observations) >= _MAX_OBSERVATIONS:
            break
        rows = table.complete_rows
        omission_reason = table.intrinsic_omission_reason
        if rows is not None and used_cells + table.complete_cell_count > _MAX_TOTAL_RETURNED_CELLS:
            rows = None
            omission_reason = "total_cell_budget"
        candidate = _make_observation(
            table,
            rows=rows,
            omission_reason=omission_reason,
        )
        trial = [
            *[item.model_dump(mode="json") for item in observations],
            candidate.model_dump(mode="json"),
        ]
        if _json_size(trial) > _MAX_OBSERVATION_JSON_BYTES and rows is not None:
            candidate = _make_observation(
                table,
                rows=None,
                omission_reason="response_byte_budget",
            )
            trial = [
                *[item.model_dump(mode="json") for item in observations],
                candidate.model_dump(mode="json"),
            ]
        if _json_size(trial) > _MAX_OBSERVATION_JSON_BYTES:
            break
        observations.append(candidate)
        used_cells += candidate.returned_cell_count

    if sha256_file(artifact_index_path) != expected_artifact_index_sha256:
        raise TabularObservationError("The artifact index changed during tabular observation.")
    serialized = [item.model_dump(mode="json") for item in observations]
    omission_reason_counts: dict[str, int] = {}
    for omission in omissions:
        reason = omission["reason"]
        omission_reason_counts[reason] = omission_reason_counts.get(reason, 0) + 1
    return RequiredTabularObservationSummary(
        artifact_index_sha256=expected_artifact_index_sha256,
        observations=tuple(observations),
        total_count=len(tables),
        returned_count=len(observations),
        truncated=len(observations) < len(tables),
        observations_sha256=observations_sha256,
        returned_cell_count=sum(item.returned_cell_count for item in observations),
        returned_utf8_bytes=_json_size(serialized),
        omitted_artifact_count=len(omissions),
        omission_reason_counts=omission_reason_counts,
        omissions_sha256=_canonical_sha256(omissions),
    )
