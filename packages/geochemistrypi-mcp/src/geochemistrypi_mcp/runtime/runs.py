"""Durable, non-blocking lifecycle control for local CLI subprocess runs."""

import hashlib
import hmac
import json
import math
import os
import re
import secrets
import tempfile
import threading
import time
import uuid
from concurrent.futures import Future, ThreadPoolExecutor
from concurrent.futures import TimeoutError as FutureTimeoutError
from dataclasses import asdict, dataclass, replace
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable

from pydantic import ValidationError

from ..api.schemas import (
    AnalysisRequest,
    AnalysisValidationResponse,
    AnomalyDetectionRequest,
    ArtifactReference,
    CancelRunResponse,
    ClassificationRequest,
    ClusteringRequest,
    DatasetInspectionRequest,
    DatasetPreparationContract,
    DecompositionRequest,
    PendingRunResultResponse,
    RegressionRequest,
    RunResultResponse,
    RunStatusResponse,
    StartAnalysisResponse,
    TimeSeriesRequest,
)
from ..api.terminal_receipts import TerminalEvidenceIdentity, TerminalRunReceipt, normalize_terminal_error, sanitize_terminal_error, terminal_error_projection
from ..config.constants import ARTIFACT_INDEX_SCHEMA_VERSION, SERVER_VERSION
from ..config.settings import McpSettings
from ..contracts.anomaly_detection import MODEL_DISPLAY_NAMES as ANOMALY_DETECTION_MODEL_DISPLAY_NAMES
from ..contracts.anomaly_detection import MODEL_ORDER as ANOMALY_DETECTION_MODEL_ORDER
from ..contracts.classification import MODEL_DISPLAY_NAMES as CLASSIFICATION_MODEL_DISPLAY_NAMES
from ..contracts.classification import MODEL_ORDER as CLASSIFICATION_MODEL_ORDER
from ..contracts.clustering import MODEL_DISPLAY_NAMES as CLUSTERING_MODEL_DISPLAY_NAMES
from ..contracts.clustering import MODEL_ORDER as CLUSTERING_MODEL_ORDER
from ..contracts.decomposition import MODEL_DISPLAY_NAMES as DECOMPOSITION_MODEL_DISPLAY_NAMES
from ..contracts.decomposition import MODEL_ORDER as DECOMPOSITION_MODEL_ORDER
from ..contracts.regression import MODEL_DISPLAY_NAMES as REGRESSION_MODEL_DISPLAY_NAMES
from ..contracts.regression import MODEL_ORDER as REGRESSION_MODEL_ORDER
from ..data.catalog import DatasetCatalog, ResolvedDataset
from ..data.headers import source_allows_pandas_duplicate_mangling
from ..data.inspector import DatasetSnapshot
from ..data.inspector import inspect_dataset as inspect_local_dataset
from ..data.inspector import sha256_file, snapshot_dataset
from ..data.preparation import DatasetPreparationError, PreparedDataset, prepare_dataset_view
from ..data.row_pairing import verify_original_row_pairing
from ..planning.interaction_plan import AnalysisPlanCompiler, DatasetCompilationContext, InteractionPlan, PlanCompilationError
from ..planning.scientific_contract import assess_scientific_compatibility, canonical_scientific_contract, canonical_sha256, planned_artifact_requirements, resolved_environment_profile
from ..tracking.experiments import ExperimentManager
from .artifacts import discover_artifacts, read_time_series_preprocessing_summary
from .cli_capabilities import CliCapabilityProbe, probe_cli_capabilities
from .cli_driver import CliDriverError, CliInteractionDriver, CliRunCancelledError, validate_workspace_path
from .environment import EnvironmentSnapshot, inspect_cli_environment
from .result_views import partition_artifact_views
from .tabular_observations import build_required_tabular_observations

_RUN_ID = re.compile(r"^run-[0-9a-f]{16}$")
_VALIDATION_ID = re.compile(r"^val-[0-9a-f]{32}$")
_TERMINAL_STATES = {"succeeded", "partial_failure", "failed", "cancelled"}
_ATOMIC_REPLACE_RETRY_DELAYS_SECONDS = (0.01, 0.02, 0.04)
_VALIDATION_TTL_SECONDS = 1800
_VALIDATION_RECEIPT_FIELDS_V1 = {
    "schema_version",
    "validation_id",
    "request_hash",
    "canonical_contract",
    "canonical_contract_hash",
    "compiled_plan_hash",
    "dataset_hash",
    "adapter_identity",
    "artifact_requirements",
    "execution_readiness",
    "request",
    "task",
    "created_at_epoch",
    "expires_at_epoch",
    "training",
    "application",
    "cli",
    "interaction_plan",
    "environment",
    "integrity_hmac_sha256",
}
_VALIDATION_RECEIPT_FIELDS = _VALIDATION_RECEIPT_FIELDS_V1 | {"event"}
_VALIDATION_DETAIL_FIELDS = {
    "schema_version",
    "validation_id",
    "request_hash",
    "response",
    "integrity_hmac_sha256",
}


def _selected_models(request: Any) -> tuple[str, ...]:
    if request.task == "time_series":
        return (
            {
                "subaerial_proportion": "subaerial_proportion_bootstrap",
                "continuous": "spatiotemporal_weighted_continuous_bootstrap",
                "element_mean": "element_mean",
                "reference_anomaly_series": "reference_label_event_overlay",
            }[request.mode],
        )
    if request.task == "decomposition" and request.mode == "embedding_label_overlay":
        return ("embedding_label_overlay",)
    orders = {
        "classification": CLASSIFICATION_MODEL_ORDER,
        "regression": REGRESSION_MODEL_ORDER,
        "clustering": CLUSTERING_MODEL_ORDER,
        "decomposition": DECOMPOSITION_MODEL_ORDER,
        "anomaly_detection": ANOMALY_DETECTION_MODEL_ORDER,
    }
    return orders[request.task] if request.model_selection.mode == "all" else (request.model.type,)


def _selected_tuning(request: Any) -> str:
    if request.task == "time_series" or (request.task == "decomposition" and request.mode == "embedding_label_overlay"):
        return "not_applicable"
    if request.model_selection.mode == "all":
        return request.model_selection.tuning
    return getattr(request, "tuning", "not_applicable")


def _resolved_model_parameters(request: Any) -> dict[str, Any]:
    if request.task == "time_series":
        if request.mode == "subaerial_proportion":
            return {
                "mode": request.mode,
                "bin_width": request.bin_width,
                "iterations": request.iterations,
                "seed": request.seed,
                "age_unit": request.age_unit,
                "fit_curve": request.fit_curve,
                "sheet": request.sheet,
            }
        if request.mode == "continuous":
            return {
                "mode": request.mode,
                "bin_width": request.bin_width,
                "iterations": request.iterations,
                "seed": request.seed,
                "aggregation": request.aggregation,
                "uncertainty": request.uncertainty,
                "minimum_samples_per_bin": request.minimum_samples_per_bin,
                "filter_minimum": request.filter_minimum,
                "filter_maximum": request.filter_maximum,
                "relative_value_two_sigma": request.relative_value_two_sigma,
                "age_unit": request.age_unit,
                "fit_curve": request.fit_curve,
                "compact_y_axis": request.compact_y_axis,
                "sheet": request.sheet,
            }
        if request.mode == "element_mean":
            return {
                "mode": request.mode,
                "bin_width": request.bin_width,
                "aggregation": request.aggregation,
                "uncertainty": request.uncertainty,
                "minimum_samples_per_bin": request.minimum_samples_per_bin,
                "filter_minimum": request.filter_minimum,
                "filter_maximum": request.filter_maximum,
                "age_unit": request.age_unit,
                "sheet": request.sheet,
            }
        return {
            "mode": request.mode,
            "sheet": request.sheet,
            "reference_label_provenance": request.reference_label_provenance,
            "comparison_label_provenance": request.comparison_label_provenance,
            **(
                {
                    "event_sheet": request.event_sheet,
                    "association_window_days": request.association_window_days,
                    "association_direction": request.association_direction,
                }
                if request.event_dataset_path is not None
                else {}
            ),
        }
    if request.task == "decomposition" and request.mode == "embedding_label_overlay":
        return {
            "mode": request.mode,
            "join_policy": "exact_identifier_set_one_to_one",
        }
    if request.model_selection.mode == "all" or _selected_tuning(request) == "automl":
        return {}
    model = getattr(request, "model", None)
    if model is None:
        return {}
    return model.model_dump(mode="python", exclude={"type"})


def _typed_semantic_label(value: Any) -> dict[str, Any] | None:
    if value is None:
        return None
    label_type = "boolean" if isinstance(value, bool) else "integer" if isinstance(value, int) else "number" if isinstance(value, float) else "string"
    return {"type": label_type, "value": value}


def _workflow_specific_execution_contract(request: Any) -> dict[str, Any] | None:
    """Keep active role bindings without replaying inactive workflow defaults."""

    if request.task == "decomposition" and request.mode == "embedding_label_overlay":
        return {
            "contract_type": "decomposition_embedding_label_overlay",
            "coordinate_sheet": request.coordinate_sheet,
            "label_sheet": request.label_sheet,
            "coordinate_identifier_column": request.identifier_column,
            "label_identifier_column": request.label_identifier_column,
            "label_column": request.label_column,
            "positive_label_values": tuple(request.positive_label_values),
            "join_policy": "exact_identifier_set_one_to_one",
        }
    if request.task != "time_series":
        return None

    common = {
        "identifier_column": request.identifier_column,
        "selected_columns": tuple(request.resolved_selected_columns),
        "sheet": request.sheet,
    }
    if request.mode == "subaerial_proportion":
        return {
            "contract_type": "time_series_subaerial_proportion",
            **common,
            "age_column": request.age_column,
            "maximum_age_column": request.maximum_age_column,
            "probability_column": request.probability_column,
            "latitude_column": request.latitude_column,
            "longitude_column": request.longitude_column,
        }
    if request.mode == "continuous":
        return {
            "contract_type": "time_series_continuous",
            **common,
            "age_column": request.age_column,
            "minimum_age_column": request.minimum_age_column,
            "maximum_age_column": request.maximum_age_column,
            "value_column": request.value_column,
            "latitude_column": request.latitude_column,
            "longitude_column": request.longitude_column,
            "filter_column": request.filter_column,
        }
    if request.mode == "element_mean":
        return {
            "contract_type": "time_series_element_mean",
            **common,
            "age_column": request.age_column,
            "element_columns": tuple(request.element_columns),
            "filter_column": request.filter_column,
        }
    return {
        "contract_type": "time_series_reference_anomaly_series",
        **common,
        "time_column": request.time_column,
        "signal_columns": tuple(request.signal_columns),
        "reference_label_column": request.reference_label_column,
        "reference_positive_values": tuple(request.reference_positive_values),
        "comparison_label_column": request.comparison_label_column,
        "comparison_positive_values": tuple(request.comparison_positive_values),
        "event_sheet": request.event_sheet if request.event_dataset_path is not None else None,
        "event_time_column": request.event_time_column,
        "event_identifier_column": request.event_identifier_column,
        "event_filter_column": request.event_filter_column,
        "event_filter_values": tuple(request.event_filter_values),
    }


def _validation_execution_decisions(
    request: Any,
    plan: InteractionPlan,
    *,
    application_enabled: bool,
    secondary_is_evaluation: bool,
) -> dict[str, Any]:
    """Expose exact active scientific decisions without replaying provenance bulk."""
    execution_contract: dict[str, Any] = {}
    if plan.scientific_execution_contract_json:
        parsed = json.loads(plan.scientific_execution_contract_json)
        if isinstance(parsed, dict):
            execution_contract = parsed
    evaluation = getattr(request, "evaluation", None)
    requested_mode = getattr(evaluation, "mode", "not_applicable")
    requested_split = getattr(evaluation, "split_strategy", None)
    supervised = request.task in {"classification", "regression"}
    overlay = request.task == "decomposition" and request.mode == "embedding_label_overlay"
    if supervised and execution_contract:
        effective_mode = execution_contract.get("evaluation_mode") or "not_reported"
    elif overlay:
        effective_mode = "not_applicable"
    else:
        effective_mode = requested_mode
    internal_holdout = supervised and effective_mode == "internal_holdout"
    external_labeled = supervised and effective_mode == "external_labeled"
    effective_split = execution_contract.get("split_strategy") if internal_holdout else None
    effective_folds = execution_contract.get("cross_validation_folds") if supervised and not external_labeled and execution_contract else None
    requested_confusion_normalization = getattr(
        evaluation,
        "confusion_matrix_normalization",
        None,
    )
    effective_confusion_normalization = None
    if request.task == "classification" and execution_contract:
        effective_confusion_normalization = execution_contract.get("confusion_matrix_normalization") or "none"
    requested_metric_average = getattr(request, "metric_average", None)
    effective_metric_average = execution_contract.get("classification_metric_average") if request.task == "classification" and execution_contract else None
    requested_positive_label = _typed_semantic_label(getattr(request, "positive_label", None))
    effective_positive_label = execution_contract.get("classification_positive_label") if request.task == "classification" and execution_contract else None
    if overlay:
        application_role = "artifact_overlay"
        secondary_identifier_column = request.label_identifier_column
        application_target_columns = (request.label_column,)
    elif application_enabled and secondary_is_evaluation:
        application_role = "external_evaluation"
        secondary_identifier_column = getattr(
            evaluation,
            "external_identifier_column",
            None,
        )
        application_target_columns = tuple(getattr(request, "resolved_target_columns", ()))
    elif application_enabled:
        application_role = "inference"
        secondary_identifier_column = getattr(request, "identifier_column", None)
        application_target_columns = ()
    else:
        application_role = "none"
        secondary_identifier_column = None
        application_target_columns = ()

    missing_values = getattr(request, "missing_values", None)
    feature_selection = getattr(request, "feature_selection", None)
    label_customization = getattr(request, "label_customization", None)
    world_map = getattr(request, "world_map", None)
    return {
        "evaluation": {
            "requested_mode": requested_mode,
            "effective_mode": effective_mode,
            "requested_test_ratio": getattr(request, "test_ratio", None),
            "effective_test_ratio": (getattr(request, "test_ratio", None) if internal_holdout else None),
            "requested_split_strategy": requested_split,
            "effective_split_strategy": effective_split,
            "requested_cross_validation_folds": getattr(evaluation, "folds", None),
            "effective_cross_validation_folds": effective_folds,
            "requested_metrics": tuple(getattr(evaluation, "metrics", ())),
            "metric_artifact_bindings": dict(getattr(evaluation, "metric_artifact_bindings", {})),
            "required_artifact_ids": tuple(getattr(evaluation, "required_artifact_ids", ())),
            "class_order": tuple(getattr(evaluation, "class_order", ())),
            "requested_confusion_matrix_normalization": requested_confusion_normalization,
            "effective_confusion_matrix_normalization": effective_confusion_normalization,
            "requested_metric_average": requested_metric_average,
            "effective_metric_average": effective_metric_average,
            "requested_positive_label": requested_positive_label,
            "effective_positive_label": effective_positive_label,
        },
        "preprocessing": {
            "missing_values": (missing_values.model_dump(mode="json") if missing_values is not None else None),
            "scaling": getattr(request, "scaling", None),
            "feature_selection": (feature_selection.model_dump(mode="json") if feature_selection is not None else None),
            "engineered_features": tuple(item.model_dump(mode="json") for item in getattr(request, "engineered_features", ())),
            "label_customization": (label_customization.model_dump(mode="json") if label_customization is not None else None),
            "world_map": (world_map.model_dump(mode="json") if world_map is not None else None),
            "target_transformations": {
                column: transformation.model_dump(mode="json")
                for column, transformation in getattr(
                    request,
                    "target_transformations",
                    {},
                ).items()
            },
            "sample_balancing": getattr(request, "sample_balancing", None),
            "metadata_columns": tuple(getattr(request, "metadata_columns", ())),
            "feature_engineering": getattr(request, "feature_engineering", None),
        },
        "application": {
            "enabled": application_enabled,
            "role": application_role,
            "training_identifier_column": getattr(
                request,
                "identifier_column",
                None,
            ),
            "secondary_identifier_column": secondary_identifier_column,
            "target_columns": application_target_columns,
            "label_used_as_feature": False,
        },
        "bindings": {
            "model": plan.model_parameter_binding or "not_bound",
            "preprocessing": plan.preprocessing_parameter_binding or "not_bound",
            "scientific_execution_contract_bound": bool(plan.scientific_execution_contract_json),
            "workflow_specific_contract": _workflow_specific_execution_contract(request),
        },
    }


class RunNotFoundError(ValueError):
    """Raised when a run ID is not owned by this wrapper."""


class RunStateError(ValueError):
    """Raised when an operation is invalid for the run's current state."""


class InputIntegrityError(RuntimeError):
    """Raised when the user's source dataset changes during a run."""


def _terminal_failure_result_type(exc: BaseException) -> str:
    """Project asynchronous exceptions onto the public recovery taxonomy."""

    if isinstance(exc, InputIntegrityError):
        return "input_integrity_changed"
    if isinstance(exc, CliDriverError):
        return "cli_execution_failed"
    if isinstance(exc, RunStateError):
        return "run_state_invalid"
    return "internal_error"


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _new_run_id() -> str:
    return f"run-{uuid.uuid4().hex[:16]}"


def _canonical_json_bytes(value: Any) -> bytes:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")


def _json_sha256(value: Any) -> str:
    return hashlib.sha256(_canonical_json_bytes(value)).hexdigest()


def _validation_request_value(request: Any) -> dict[str, Any]:
    """Serialize a strict request without reintroducing mutually exclusive defaults."""

    value = request.model_dump(mode="json")
    for field in getattr(request, "mode_inapplicable_fields", ()):
        # The public schema rejects explicitly supplied fields owned by another
        # mode. Pydantic still materializes their class defaults, so omit those
        # inactive defaults from the signed request to preserve strict
        # schema/runtime round trips.
        value.pop(field, None)
    selection = value.get("model_selection")
    if isinstance(selection, dict) and selection.get("mode") == "all":
        # Pydantic materializes the backward-compatible single-model defaults
        # even though the all-models discriminator replaces both fields.  A
        # signed receipt must round-trip through the same strict public schema.
        value.pop("model", None)
        value.pop("tuning", None)

    def normalize(item: Any) -> None:
        if isinstance(item, dict):
            if item.get("header_row_indices"):
                item.pop("header_row_index", None)
            for child in item.values():
                normalize(child)
        elif isinstance(item, list):
            for child in item:
                normalize(child)

    normalize(value)
    return value


def _plan_identity(plan: InteractionPlan) -> dict[str, Any]:
    value = asdict(plan)
    return {
        "name": plan.name,
        "schema_version": plan.schema_version,
        "environment_profile_id": plan.environment_profile_id,
        "environment_profile_identity_sha256": plan.environment_profile_identity_sha256,
        "sha256": _json_sha256(value),
    }


def _materialize_tracking_root(
    plan: InteractionPlan,
    tracking_root: Path | None,
) -> InteractionPlan:
    """Bind installer-owned runtime paths before a plan receives an identity."""

    if "data-mining" not in plan.public_command:
        return plan
    if tracking_root is None:
        raise RunStateError("The installer-owned MLflow tracking root is not configured.")
    if "--tracking-root" in plan.public_command:
        return plan
    return replace(
        plan,
        public_command=(
            *plan.public_command,
            "--tracking-root",
            str(tracking_root),
        ),
    )


def _replace_with_bounded_permission_retry(source: Path, destination: Path) -> None:
    """Retry only transient permission failures while publishing metadata."""
    for attempt in range(len(_ATOMIC_REPLACE_RETRY_DELAYS_SECONDS) + 1):
        try:
            os.replace(source, destination)
            return
        except PermissionError as exc:
            if attempt == len(_ATOMIC_REPLACE_RETRY_DELAYS_SECONDS):
                raise PermissionError(f"Atomic metadata replacement failed after {attempt + 1} attempts for {destination}: {exc}") from exc
            time.sleep(_ATOMIC_REPLACE_RETRY_DELAYS_SECONDS[attempt])


def _atomic_write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    serialized = json.dumps(value, indent=2, ensure_ascii=False) + "\n"
    with tempfile.NamedTemporaryFile(
        "w",
        encoding="utf-8",
        dir=path.parent,
        prefix=f".{path.name}.",
        suffix=".tmp",
        delete=False,
    ) as stream:
        temporary_path = Path(stream.name)
        stream.write(serialized)
        stream.flush()
        os.fsync(stream.fileno())
    try:
        _replace_with_bounded_permission_retry(temporary_path, path)
    finally:
        temporary_path.unlink(missing_ok=True)


def _atomic_write_json_once(path: Path, value: Any) -> str:
    """Atomically publish complete JSON without ever replacing an existing path."""

    path.parent.mkdir(parents=True, exist_ok=True)
    serialized = (json.dumps(value, indent=2, ensure_ascii=False) + "\n").encode("utf-8")
    with tempfile.NamedTemporaryFile(
        "wb",
        dir=path.parent,
        prefix=f".{path.name}.",
        suffix=".tmp",
        delete=False,
    ) as stream:
        temporary_path = Path(stream.name)
        stream.write(serialized)
        stream.flush()
        os.fsync(stream.fileno())
    try:
        for attempt in range(len(_ATOMIC_REPLACE_RETRY_DELAYS_SECONDS) + 1):
            try:
                os.link(temporary_path, path)
                return hashlib.sha256(serialized).hexdigest()
            except FileExistsError:
                raise
            except PermissionError as exc:
                if attempt == len(_ATOMIC_REPLACE_RETRY_DELAYS_SECONDS):
                    raise PermissionError(f"Immutable metadata publication failed after {attempt + 1} attempts for {path}: {exc}") from exc
                time.sleep(_ATOMIC_REPLACE_RETRY_DELAYS_SECONDS[attempt])
    finally:
        temporary_path.unlink(missing_ok=True)


def _read_json(path: Path) -> dict[str, Any]:
    try:
        with path.open(encoding="utf-8") as stream:
            value = json.load(stream)
    except (OSError, UnicodeError, json.JSONDecodeError) as exc:
        raise RunStateError(f"Run metadata is unavailable or corrupt: {path}") from exc
    if not isinstance(value, dict):
        raise RunStateError(f"Run metadata must be a JSON object: {path}")
    return value


def _normalize_stored_scientific_identity(
    value: Any,
    *,
    legacy_schema: bool,
    include_execution_bound: bool,
    record_label: str,
) -> dict[str, Any]:
    """Migrate only a recognized legacy stored record before strict validation."""

    if not isinstance(value, dict):
        raise RunStateError(f"{record_label} must be a JSON object.")
    required = {"scientific_contract_id"}
    if include_execution_bound:
        required.add("scientific_execution_contract_bound")
    present = required & set(value)
    if present == required:
        return value
    if present:
        raise RunStateError(f"{record_label} contains an incomplete scientific identity.")
    if not legacy_schema:
        raise RunStateError(f"{record_label} is missing its required scientific identity.")
    normalized = dict(value)
    normalized["scientific_contract_id"] = "scientific-contract-v1/legacy"
    if include_execution_bound:
        normalized["scientific_execution_contract_bound"] = False
    return normalized


def _read_terminal_receipt(path: Path) -> tuple[TerminalRunReceipt, str]:
    try:
        raw = path.read_bytes()
        value = json.loads(raw.decode("utf-8"))
        schema_version = value.get("schema_version") if isinstance(value, dict) else None
        normalized = _normalize_stored_scientific_identity(
            value,
            legacy_schema=schema_version == 1,
            include_execution_bound=True,
            record_label="The stored terminal receipt",
        )
        receipt = TerminalRunReceipt.model_validate(normalized)
    except (OSError, UnicodeError, json.JSONDecodeError, ValueError) as exc:
        raise RunStateError(f"Terminal result receipt is unavailable or corrupt: {path}") from exc
    if receipt.result_record_path is not None or receipt.result_record_sha256 is not None:
        raise RunStateError("The stored terminal receipt must not contain a recursive result-record identity.")
    return receipt, hashlib.sha256(raw).hexdigest()


def _terminal_evidence_identity(path: Path, kind: str) -> TerminalEvidenceIdentity | None:
    """Hash one exact wrapper-owned evidence file without reading its contents into a response."""

    if not path.exists():
        return None
    if path.is_symlink() or not path.is_file():
        raise RunStateError(f"Allowlisted terminal evidence is not a regular file: {path}")
    digest = hashlib.sha256()
    size_bytes = 0
    try:
        with path.open("rb") as stream:
            before = os.fstat(stream.fileno())
            for chunk in iter(lambda: stream.read(1024 * 1024), b""):
                digest.update(chunk)
                size_bytes += len(chunk)
            after = os.fstat(stream.fileno())
    except OSError as exc:
        raise RunStateError(f"Allowlisted terminal evidence cannot be read: {path}") from exc
    if (before.st_size, before.st_mtime_ns) != (after.st_size, after.st_mtime_ns) or size_bytes != after.st_size:
        raise RunStateError(f"Allowlisted terminal evidence changed while it was hashed: {path}")
    return TerminalEvidenceIdentity(
        kind=kind,
        path=str(path),
        size_bytes=size_bytes,
        sha256=digest.hexdigest(),
    )


def _read_cli_trace_evidence(path: Path) -> tuple[int | None, str | None, str | None, float | None]:
    """Read exit and child timing evidence once, without managed-run fallback."""

    if path.is_symlink() or not path.is_file():
        return None, None, None, None
    try:
        value = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, UnicodeError, json.JSONDecodeError):
        return None, None, None, None
    if not isinstance(value, dict) or value.get("schema_version") != 1:
        return None, None, None, None
    raw_returncode = value.get("returncode")
    returncode = raw_returncode if isinstance(raw_returncode, int) and not isinstance(raw_returncode, bool) else None

    explicit_fields = (
        "cli_started_at",
        "cli_finished_at",
        "cli_execution_duration_seconds",
    )
    if any(field in value for field in explicit_fields):
        started_at = value.get("cli_started_at")
        finished_at = value.get("cli_finished_at")
        duration = value.get("cli_execution_duration_seconds")
        if started_at is None and finished_at is None and duration is None:
            if returncode is not None:
                raise RunStateError("The interaction trace records a CLI exit without a child-process interval.")
            return None, None, None, None
        if not isinstance(started_at, str) or not isinstance(finished_at, str):
            raise RunStateError("The interaction trace contains an incomplete CLI child-process interval.")
        if isinstance(duration, bool) or not isinstance(duration, (int, float)) or not math.isfinite(float(duration)) or duration < 0:
            raise RunStateError("The interaction trace contains an invalid CLI child-process duration.")
        duration_value = float(duration)
    else:
        # Compatibility with sealed schema-v1 traces written before explicit
        # child timing was added.  A numeric exit code is the evidence that a
        # child existed; the legacy wrapper timestamps remain trace-native.
        if returncode is None:
            return None, None, None, None
        started_at = value.get("started_at")
        finished_at = value.get("finished_at")
        if not isinstance(started_at, str) or not isinstance(finished_at, str):
            raise RunStateError("The legacy interaction trace has no complete CLI child-process interval.")
        duration_value = -1.0

    try:
        started = datetime.fromisoformat(started_at.replace("Z", "+00:00"))
        finished = datetime.fromisoformat(finished_at.replace("Z", "+00:00"))
    except ValueError as exc:
        raise RunStateError("The interaction trace CLI timestamps are not ISO 8601.") from exc
    if started.tzinfo is None or started.utcoffset() is None or finished.tzinfo is None or finished.utcoffset() is None:
        raise RunStateError("The interaction trace CLI timestamps must include a timezone.")
    if finished < started:
        raise RunStateError("The interaction trace CLI finish precedes its start.")
    if duration_value < 0:
        duration_value = (finished - started).total_seconds()
    return returncode, started_at, finished_at, duration_value


def _read_terminal_cli_trace_evidence(
    path: Path,
) -> tuple[int | None, str | None, str | None, float | None]:
    """Best-effort trace evidence for a failure receipt that must seal.

    Success publication keeps using the strict reader above.  Once execution
    has already failed, malformed or incomplete child timing is itself part of
    the failure evidence and must not prevent the durable run from reaching a
    terminal state.  A trustworthy numeric return code may still be retained;
    an incomplete timing tuple is always discarded as a whole.
    """

    try:
        return _read_cli_trace_evidence(path)
    except RunStateError:
        pass
    if path.is_symlink() or not path.is_file():
        return None, None, None, None
    try:
        value = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, UnicodeError, json.JSONDecodeError):
        return None, None, None, None
    if not isinstance(value, dict) or value.get("schema_version") != 1:
        return None, None, None, None
    raw_returncode = value.get("returncode")
    returncode = raw_returncode if isinstance(raw_returncode, int) and not isinstance(raw_returncode, bool) else None
    return returncode, None, None, None


def _safe_error(exc: BaseException) -> str:
    return normalize_terminal_error(exc) or type(exc).__name__


def _aggregate_children(output_directory: Path, task: str) -> tuple[str, tuple[dict[str, Any], ...]]:
    manifest_path = output_directory / "summary" / "Aggregate Model Results.json"
    value = _read_json(manifest_path)
    expected_fields = {
        "schema_version",
        "task",
        "selection_mode",
        "tuning",
        "state",
        "expected_model_count",
        "succeeded_count",
        "failed_count",
        "children",
    }
    if set(value) != expected_fields:
        raise RunStateError("Aggregate model manifest has unknown or missing fields.")
    if value["schema_version"] != 1 or value["task"] != task or value["selection_mode"] != "all" or value["state"] not in {"complete", "partial_failure"} or not isinstance(value["children"], list):
        raise RunStateError("Aggregate model manifest identity is invalid.")
    contracts = {
        "classification": (
            CLASSIFICATION_MODEL_ORDER,
            CLASSIFICATION_MODEL_DISPLAY_NAMES,
        ),
        "regression": (REGRESSION_MODEL_ORDER, REGRESSION_MODEL_DISPLAY_NAMES),
        "clustering": (CLUSTERING_MODEL_ORDER, CLUSTERING_MODEL_DISPLAY_NAMES),
        "decomposition": (
            DECOMPOSITION_MODEL_ORDER,
            DECOMPOSITION_MODEL_DISPLAY_NAMES,
        ),
        "anomaly_detection": (
            ANOMALY_DETECTION_MODEL_ORDER,
            ANOMALY_DETECTION_MODEL_DISPLAY_NAMES,
        ),
    }
    model_order, display_names = contracts[task]
    expected_models = [display_names[model] for model in model_order]
    children = []
    seen = set()
    for child in value["children"]:
        if not isinstance(child, dict) or set(child) != {
            "model",
            "state",
            "output_relative_path",
            "artifact_count",
            "error",
        }:
            raise RunStateError("Aggregate child metadata is invalid.")
        model = child.get("model")
        relative = child.get("output_relative_path")
        if not isinstance(model, str) or not model or model in seen:
            raise RunStateError("Aggregate child model names must be unique.")
        if not isinstance(relative, str):
            raise RunStateError("Aggregate child output path must be a string.")
        relative_path = Path(relative)
        if relative_path.is_absolute() or len(relative_path.parts) != 1:
            raise RunStateError("Aggregate child output must be one direct child directory.")
        child_directory = (output_directory / relative_path).resolve()
        try:
            child_directory.relative_to(output_directory.resolve())
        except ValueError as exc:
            raise RunStateError("Aggregate child output escapes the run directory.") from exc
        if not child_directory.is_dir():
            raise RunStateError("Aggregate child output directory is missing.")
        actual_count = sum(1 for path in child_directory.rglob("*") if path.is_file())
        if child.get("artifact_count") != actual_count:
            raise RunStateError(f"Aggregate child artifact count changed for {model!r}.")
        seen.add(model)
        children.append(child)
    succeeded = sum(child.get("state") == "succeeded" for child in children)
    failed = sum(child.get("state") == "failed" for child in children)
    aggregate_is_consistent = all(
        (
            [child["model"] for child in children] == expected_models,
            value["expected_model_count"] == len(expected_models),
            value["succeeded_count"] == succeeded,
            value["failed_count"] == failed,
            succeeded + failed == len(children),
            value["state"] == ("complete" if failed == 0 else "partial_failure"),
        )
    )
    if not aggregate_is_consistent:
        raise RunStateError("Aggregate model counts or state are inconsistent.")
    return value["state"], tuple(children)


@dataclass(frozen=True)
class RunPaths:
    root: Path
    wrapper: Path
    workspace: Path
    request: Path
    status: Path
    result: Path
    terminal_result: Path
    artifact_index: Path
    provenance_manifest: Path

    @classmethod
    def create(cls, runs_root: Path, run_id: str) -> "RunPaths":
        root = runs_root / run_id
        wrapper = root / "wrapper"
        return cls(
            root=root,
            wrapper=wrapper,
            workspace=root / "workspace",
            request=wrapper / "request.json",
            status=wrapper / "status.json",
            result=wrapper / "result.json",
            terminal_result=wrapper / "terminal-result.json",
            artifact_index=wrapper / "artifact-index.json",
            provenance_manifest=wrapper / "scientific-run-manifest.json",
        )


@dataclass
class _RunControl:
    cancellation: threading.Event
    future: Future[None] | None = None


class RunManager:
    """Own local run transitions while the real CLI owns scientific execution."""

    def __init__(
        self,
        settings: McpSettings,
        plan_compiler: AnalysisPlanCompiler | None = None,
        driver_factory: Callable[[], CliInteractionDriver] | None = None,
        cli_resolver: Callable[[], tuple[Path, str]] | None = None,
        dataset_catalog: DatasetCatalog | None = None,
        experiment_manager: ExperimentManager | None = None,
        environment_resolver: Callable[[Path], EnvironmentSnapshot] | None = None,
        capability_resolver: Callable[[Path, tuple[str, ...]], CliCapabilityProbe] | None = None,
    ):
        self.settings = settings
        self.plan_compiler = plan_compiler or AnalysisPlanCompiler()
        self.driver_factory = driver_factory or (
            lambda: CliInteractionDriver(
                process_timeout_seconds=settings.maximum_process_seconds,
                automation_mode=True,
            )
        )
        self.cli_resolver = cli_resolver or settings.require_supported_cli
        self.dataset_catalog = dataset_catalog or DatasetCatalog(settings)
        self.experiment_manager = experiment_manager or ExperimentManager(settings)
        self.environment_resolver = environment_resolver or inspect_cli_environment
        self.capability_resolver = capability_resolver
        self._executor = ThreadPoolExecutor(max_workers=settings.concurrency, thread_name_prefix="geochemistrypi-run")
        self._lock = threading.RLock()
        self._active: dict[str, _RunControl] = {}
        self._initialized = False
        self._closed = False

    def _ensure_initialized(self) -> None:
        with self._lock:
            if self._initialized:
                return
            self.settings.runs_root.mkdir(parents=True, exist_ok=True)
            for run_directory in self.settings.runs_root.glob("run-*"):
                status_path = run_directory / "wrapper" / "status.json"
                if not status_path.is_file():
                    continue
                try:
                    status = _read_json(status_path)
                except RunStateError:
                    continue
                if status.get("state") in {"queued", "running"}:
                    self._finish(
                        RunPaths.create(self.settings.runs_root, run_directory.name),
                        "failed",
                        "The previous MCP server stopped before this run finished.",
                        "Run state recovered after an unclean server shutdown; no stale PID was terminated.",
                    )
            self._initialized = True

    def _paths(self, run_id: str) -> RunPaths:
        if not _RUN_ID.fullmatch(run_id):
            raise RunNotFoundError("Invalid GeochemistryPi run ID.")
        paths = RunPaths.create(self.settings.runs_root, run_id)
        if not paths.root.is_dir():
            raise RunNotFoundError(f"GeochemistryPi run does not exist: {run_id}")
        return paths

    def _write_status(self, paths: RunPaths, status: dict[str, Any]) -> None:
        _atomic_write_json(paths.status, status)

    def _status_response(self, status: dict[str, Any]) -> RunStatusResponse:
        public_status = {field: status[field] for field in RunStatusResponse.model_fields if field in status}
        return RunStatusResponse.model_validate(public_status)

    def _prepare_dataset(
        self,
        request: Any,
        role: str,
        resolved: ResolvedDataset,
        source_snapshot: DatasetSnapshot,
    ) -> PreparedDataset:
        reference = getattr(request, f"{role}_dataset", None)
        if role == "application" and reference is None:
            reference, _, _ = self._secondary_dataset(request)
        contract = reference.preparation if reference is not None else DatasetPreparationContract()
        requested_sheet = getattr(request, "sheet", "0")
        if request.task == "time_series" and role == "training" and source_snapshot.resolved_path.suffix.lower() == ".xlsx" and (requested_sheet != "0" or contract.worksheet is None):
            try:
                from openpyxl import load_workbook

                with source_snapshot.resolved_path.open("rb") as stream:
                    workbook = load_workbook(stream, read_only=True, data_only=True)
                    try:
                        if str(requested_sheet).isdigit():
                            sheet_index = int(requested_sheet)
                            if sheet_index >= len(workbook.sheetnames):
                                raise RunStateError(f"Excel sheet index {sheet_index} is out of range.")
                            resolved_sheet = workbook.sheetnames[sheet_index]
                        else:
                            resolved_sheet = str(requested_sheet)
                            if resolved_sheet not in workbook.sheetnames:
                                raise RunStateError(f"Excel sheet {resolved_sheet!r} does not exist.")
                    finally:
                        workbook.close()
            except OSError as exc:
                raise RunStateError(f"Unable to resolve Time Series worksheet {requested_sheet!r}.") from exc
            if contract.worksheet is not None and contract.worksheet != resolved_sheet:
                raise RunStateError("Time Series sheet conflicts with training_dataset.preparation.worksheet.")
            contract = contract.model_copy(update={"worksheet": resolved_sheet})
        state_root = self.settings.service_state_root
        if state_root is None:
            raise RunStateError("The MCP service-state root is not configured.")
        try:
            return prepare_dataset_view(
                source_snapshot,
                contract,
                state_root,
                self.settings.maximum_dataset_bytes,
                self.settings.maximum_columns,
                allow_pandas_duplicate_mangling=source_allows_pandas_duplicate_mangling(resolved.source),
                preserve_source_columns_for_cli=bool(contract.selected_columns or contract.excluded_columns)
                and not contract.worksheets
                and contract.source_sheet_column is None
                and contract.source_row_column is None,
            )
        except DatasetPreparationError as exc:
            raise PlanCompilationError(str(exc)) from exc

    @staticmethod
    def _secondary_dataset(request: Any) -> tuple[Any | None, Path | None, bool]:
        """Resolve ordinary application or nested external-evaluation input."""

        application_reference = getattr(request, "application_dataset", None)
        application_path = getattr(request, "application_dataset_path", None)
        if application_reference is not None or application_path is not None:
            return application_reference, application_path, False
        evaluation = getattr(request, "evaluation", None)
        if evaluation is not None and evaluation.mode == "external_labeled":
            return (
                evaluation.evaluation_dataset,
                evaluation.evaluation_dataset_path,
                True,
            )
        return None, None, False

    def _resolve_request_datasets(
        self,
        request: Any,
    ) -> tuple[ResolvedDataset, ResolvedDataset | None, bool]:
        """Resolve all named inputs together so one source needs one exact query."""

        training_reference = getattr(request, "training_dataset", None)
        application_reference, application_path, secondary_is_evaluation = self._secondary_dataset(request)
        specifications: list[tuple[Any, str | None, str | None]] = []
        training_index: int | None = None
        application_index: int | None = None
        if training_reference is not None:
            training_index = len(specifications)
            specifications.append((training_reference, request.task, "training"))
        if application_reference is not None:
            application_index = len(specifications)
            specifications.append((application_reference, request.task, "application"))
        named = self.dataset_catalog.resolve_many(tuple(specifications)) if specifications else ()
        resolved_training = (
            named[training_index]
            if training_index is not None
            else ResolvedDataset(
                path=request.training_dataset_path,
                expected_sha256=None,
                dataset_id=None,
                source="path",
                file_name=Path(request.training_dataset_path).name,
            )
        )
        resolved_application = (
            named[application_index]
            if application_index is not None
            else (
                ResolvedDataset(
                    path=application_path,
                    expected_sha256=None,
                    dataset_id=None,
                    source="path",
                    file_name=Path(application_path).name,
                )
                if application_path is not None
                else None
            )
        )
        return resolved_training, resolved_application, secondary_is_evaluation

    @staticmethod
    def _bind_source_snapshot(
        resolved: ResolvedDataset,
        snapshot: DatasetSnapshot,
        role: str,
        *,
        since_validation: bool = False,
    ) -> ResolvedDataset:
        expected_hashes = tuple(value for value in (resolved.expected_sha256, resolved.observed_sha256) if value is not None)
        if any(snapshot.sha256 != value for value in expected_hashes):
            raise InputIntegrityError(f"The {role} dataset changed " + ("since validation." if since_validation else "between source resolution and validation."))
        if resolved.observed_size_bytes is not None and snapshot.size_bytes != resolved.observed_size_bytes:
            raise InputIntegrityError(f"The {role} dataset size changed " + ("since validation." if since_validation else "between source resolution and validation."))
        return replace(
            resolved,
            path=snapshot.resolved_path,
            observed_size_bytes=snapshot.size_bytes,
            observed_sha256=snapshot.sha256,
        )

    @staticmethod
    def _resolution_identity(
        resolved: ResolvedDataset,
        *,
        requested_task: str,
        requested_role: str,
    ) -> dict[str, Any]:
        if resolved.observed_size_bytes is None or resolved.observed_sha256 is None:
            raise RunStateError("A resolved dataset has no observed source identity.")
        return {
            "source": resolved.source,
            "dataset_id": resolved.dataset_id,
            "file_name": resolved.file_name,
            "canonical_path": str(resolved.path),
            "root_path": str(resolved.root_path) if resolved.root_path is not None else None,
            "catalog_task": resolved.catalog_task,
            "catalog_role": resolved.catalog_role,
            "requested_task": requested_task,
            "requested_role": requested_role,
            "expected_sha256": resolved.expected_sha256,
            "observed_size_bytes": resolved.observed_size_bytes,
            "observed_sha256": resolved.observed_sha256,
        }

    def _resolved_from_validation_receipt(
        self,
        request: Any,
        prepared_identity: dict[str, Any],
        *,
        requested_role: str,
    ) -> ResolvedDataset:
        """Rebuild one selected input from the signed receipt without catalog discovery."""

        resolution = prepared_identity.get("resolution")
        expected_fields = {
            "source",
            "dataset_id",
            "file_name",
            "canonical_path",
            "root_path",
            "catalog_task",
            "catalog_role",
            "requested_task",
            "requested_role",
            "expected_sha256",
            "observed_size_bytes",
            "observed_sha256",
        }
        if not isinstance(resolution, dict) or set(resolution) != expected_fields:
            raise RunStateError("The validation receipt does not bind an exact dataset resolution; validate the request again.")
        if resolution["requested_task"] != request.task or resolution["requested_role"] != requested_role:
            raise RunStateError("The validation receipt dataset role or task is inconsistent with the request.")

        reference = getattr(request, f"{requested_role}_dataset", None)
        requested_path = getattr(request, f"{requested_role}_dataset_path", None)
        if requested_role == "application" and reference is None and requested_path is None:
            reference, requested_path, _ = self._secondary_dataset(request)
        expected_source = getattr(reference, "source", "path") if reference is not None else "path"
        expected_dataset_id = getattr(reference, "dataset_id", None)
        expected_file_name = (
            getattr(reference, "file_name", None) if expected_source == "desktop" else Path(getattr(reference, "path", requested_path)).name if expected_source == "path" else resolution["file_name"]
        )
        expected_sha256 = getattr(reference, "expected_sha256", None)
        if (
            resolution["source"] != expected_source
            or resolution["dataset_id"] != expected_dataset_id
            or resolution["file_name"] != expected_file_name
            or resolution["expected_sha256"] != expected_sha256
        ):
            raise RunStateError("The validation receipt dataset selector is inconsistent with the request.")
        if expected_source == "builtin" and (resolution["catalog_task"] != request.task or resolution["catalog_role"] != requested_role):
            raise RunStateError("The validation receipt built-in dataset task or role is invalid.")
        if expected_source == "desktop" and (resolution["catalog_task"] is not None or resolution["catalog_role"] != "unspecified"):
            raise RunStateError("The validation receipt Desktop catalog identity is invalid.")
        if expected_source == "path" and (resolution["catalog_task"] is not None or resolution["catalog_role"] is not None):
            raise RunStateError("The validation receipt explicit-path catalog identity is invalid.")

        try:
            canonical_path = Path(resolution["canonical_path"])
            if not canonical_path.is_absolute():
                raise ValueError("non-absolute path")
            canonical_path = canonical_path.resolve(strict=True)
        except (OSError, RuntimeError, TypeError, ValueError) as exc:
            raise InputIntegrityError(f"The {requested_role} dataset selected by validation is no longer available.") from exc
        root_value = resolution["root_path"]
        try:
            root_path = Path(root_value).resolve(strict=True) if isinstance(root_value, str) else None
        except (OSError, RuntimeError, TypeError, ValueError) as exc:
            raise InputIntegrityError(f"The validated {requested_role} dataset root is no longer available.") from exc
        if expected_source in {"builtin", "desktop"}:
            if root_path is None or not isinstance(resolution["file_name"], str):
                raise RunStateError("The validation receipt catalog root identity is incomplete.")
            selected_path = root_path / resolution["file_name"]
            if selected_path.is_symlink():
                raise InputIntegrityError(f"The {requested_role} dataset selector became a symbolic link after validation.")
            try:
                selected_resolved = selected_path.resolve(strict=True)
                selected_resolved.relative_to(root_path)
            except (OSError, RuntimeError, ValueError) as exc:
                raise InputIntegrityError(f"The {requested_role} dataset escaped its validated catalog root.") from exc
            if selected_path.parent.resolve() != root_path or selected_resolved != canonical_path:
                raise InputIntegrityError(f"The {requested_role} dataset path changed after validation.")
        elif root_path is not None:
            raise RunStateError("An explicit-path validation receipt must not bind a catalog root.")
        elif requested_path is not None:
            try:
                explicit_path = Path(requested_path).resolve(strict=True)
            except (OSError, RuntimeError, TypeError, ValueError) as exc:
                raise InputIntegrityError(f"The explicit {requested_role} dataset is no longer available.") from exc
            if explicit_path != canonical_path:
                raise RunStateError("The validation receipt explicit dataset path is inconsistent with the request.")

        observed_size = resolution["observed_size_bytes"]
        observed_sha = resolution["observed_sha256"]
        if not isinstance(observed_size, int) or observed_size < 0:
            raise RunStateError("The validation receipt dataset size identity is invalid.")
        if not isinstance(observed_sha, str) or not re.fullmatch(r"[0-9a-f]{64}", observed_sha):
            raise RunStateError("The validation receipt dataset hash identity is invalid.")
        return ResolvedDataset(
            path=canonical_path,
            expected_sha256=expected_sha256,
            dataset_id=expected_dataset_id,
            source=expected_source,
            file_name=resolution["file_name"],
            catalog_task=resolution["catalog_task"],
            catalog_role=resolution["catalog_role"],
            root_path=root_path,
            observed_size_bytes=observed_size,
            observed_sha256=observed_sha,
        )

    @staticmethod
    def _execution_dataset_updates(
        request: Any,
        training_path: Path,
        secondary_path: Path | None,
        secondary_is_evaluation: bool,
    ) -> dict[str, Any]:
        updates: dict[str, Any] = {
            "training_dataset_path": training_path,
            "training_dataset": None,
        }
        if secondary_is_evaluation:
            updates["evaluation"] = request.evaluation.model_copy(
                update={
                    "evaluation_dataset_path": secondary_path,
                    "evaluation_dataset": None,
                }
            )
        elif hasattr(request, "application_dataset_path"):
            updates.update(
                {
                    "application_dataset_path": secondary_path,
                    "application_dataset": None,
                }
            )
        return updates

    @staticmethod
    def _prepared_identity(
        prepared: PreparedDataset,
        resolved: ResolvedDataset,
        *,
        requested_task: str,
        requested_role: str,
    ) -> dict[str, Any]:
        snapshot = prepared.snapshot
        return {
            "resolved_path": str(snapshot.resolved_path),
            "size_bytes": snapshot.size_bytes,
            "sha256": snapshot.sha256,
            "source": resolved.source,
            "dataset_id": resolved.dataset_id,
            "resolution": RunManager._resolution_identity(
                resolved,
                requested_task=requested_task,
                requested_role=requested_role,
            ),
            "source_file": prepared.record["source_file"],
            "preparation": prepared.record,
        }

    def _event_dataset_snapshot(self, request: Any) -> DatasetSnapshot | None:
        """Hash an optional Time Series event input without treating it as inference data."""

        event_path = getattr(request, "event_dataset_path", None)
        if event_path is None:
            return None
        return snapshot_dataset(Path(event_path), self.settings.maximum_dataset_bytes)

    @staticmethod
    def _event_identity(snapshot: DatasetSnapshot | None) -> dict[str, Any] | None:
        if snapshot is None:
            return None
        return {
            "resolved_path": str(snapshot.resolved_path),
            "size_bytes": snapshot.size_bytes,
            "sha256": snapshot.sha256,
            "format": snapshot.format,
        }

    def _validation_secret(self) -> bytes:
        state_root = self.settings.service_state_root
        if state_root is None:  # McpSettings normalizes this in __post_init__.
            raise RunStateError("The MCP validation state root is not configured.")
        secret_path = state_root / "validation-receipt.key"
        secret_path.parent.mkdir(parents=True, exist_ok=True)
        try:
            descriptor = os.open(
                secret_path,
                os.O_WRONLY | os.O_CREAT | os.O_EXCL | getattr(os, "O_BINARY", 0),
                0o600,
            )
        except FileExistsError:
            descriptor = None
        if descriptor is not None:
            try:
                secret = secrets.token_bytes(32)
                remaining = memoryview(secret)
                while remaining:
                    written = os.write(descriptor, remaining)
                    if written <= 0:
                        raise OSError("Validation receipt integrity key write made no progress.")
                    remaining = remaining[written:]
                os.fsync(descriptor)
            finally:
                os.close(descriptor)
        try:
            secret = secret_path.read_bytes()
        except OSError as exc:
            raise RunStateError("The validation receipt integrity key is unavailable.") from exc
        if len(secret) != 32:
            raise RunStateError("The validation receipt integrity key is corrupt.")
        return secret

    def _receipt_integrity(self, receipt_without_integrity: dict[str, Any]) -> str:
        return hmac.new(
            self._validation_secret(),
            _canonical_json_bytes(receipt_without_integrity),
            hashlib.sha256,
        ).hexdigest()

    def _write_validation_receipt(
        self,
        request: Any,
        prepared_training: PreparedDataset,
        resolved_training: ResolvedDataset,
        prepared_application: PreparedDataset | None,
        resolved_application: ResolvedDataset | None,
        event_snapshot: DatasetSnapshot | None,
        cli_executable: Path,
        cli_version: str,
        plan: InteractionPlan,
        environment_snapshot: EnvironmentSnapshot,
    ) -> dict[str, Any]:
        request_value = _validation_request_value(request)
        request_hash = _json_sha256(request_value)
        training = self._prepared_identity(
            prepared_training,
            resolved_training,
            requested_task=request.task,
            requested_role="training",
        )
        application = (
            self._prepared_identity(
                prepared_application,
                resolved_application,
                requested_task=request.task,
                requested_role="application",
            )
            if prepared_application is not None and resolved_application is not None
            else None
        )
        event = self._event_identity(event_snapshot)
        cli = {
            "executable": str(cli_executable.resolve()),
            "executable_sha256": sha256_file(cli_executable),
            "version": cli_version,
        }
        plan_identity = _plan_identity(plan)
        artifact_requirements = planned_artifact_requirements(request, plan)
        canonical_contract = canonical_scientific_contract(request, plan)
        canonical_contract_hash = canonical_sha256(canonical_contract)
        assessment = assess_scientific_compatibility(
            request,
            plan,
            artifact_requirements,
            environment_snapshot,
        )
        environment = {
            "identity_sha256": environment_snapshot.identity_sha256,
            "record": environment_snapshot.record,
        }
        stable_identity = {
            "schema_version": 2,
            "request_hash": request_hash,
            "request": request_value,
            "task": request.task,
            "training": training,
            "application": application,
            "event": event,
            "cli": cli,
            "interaction_plan": plan_identity,
            "canonical_contract_hash": canonical_contract_hash,
            "environment": environment,
        }
        validation_id = f"val-{_json_sha256(stable_identity)[:32]}"
        created_at_epoch = int(time.time())
        receipt_without_integrity = {
            "schema_version": 2,
            "validation_id": validation_id,
            "request_hash": request_hash,
            "canonical_contract": canonical_contract,
            "canonical_contract_hash": canonical_contract_hash,
            "compiled_plan_hash": plan_identity["sha256"],
            "dataset_hash": training["sha256"],
            "adapter_identity": {
                "id": plan.adapter_id,
                "version": plan.adapter_version,
                "status": assessment.adapter_status,
            },
            "artifact_requirements": [requirement.model_dump(mode="json") for requirement in artifact_requirements],
            "execution_readiness": {
                "execution_ready": assessment.execution_ready,
                "comparison_ready": assessment.comparison_ready,
                "claim_ready": assessment.claim_ready,
                "scientific_status": assessment.scientific_status,
                "adapter_status": assessment.adapter_status,
                "artifact_status": assessment.artifact_status,
                "environment_status": assessment.environment_status,
                "blocking_issues": list(assessment.blocking_issues),
            },
            "request": request_value,
            "task": request.task,
            "created_at_epoch": created_at_epoch,
            "expires_at_epoch": created_at_epoch + _VALIDATION_TTL_SECONDS,
            "training": training,
            "application": application,
            "event": event,
            "cli": cli,
            "interaction_plan": plan_identity,
            "environment": environment,
        }
        receipt = {
            **receipt_without_integrity,
            "integrity_hmac_sha256": self._receipt_integrity(receipt_without_integrity),
        }
        state_root = self.settings.service_state_root
        if state_root is None:
            raise RunStateError("The MCP validation state root is not configured.")
        _atomic_write_json(state_root / "validations" / f"{validation_id}.json", receipt)
        return receipt

    def _write_validation_detail(self, response: AnalysisValidationResponse) -> None:
        """Persist the exact full validation response behind its public reference."""
        state_root = self.settings.service_state_root
        if state_root is None:
            raise RunStateError("The MCP validation state root is not configured.")
        detail_without_integrity = {
            "schema_version": 2,
            "validation_id": response.validation_id,
            "request_hash": response.request_hash,
            "response": response.model_dump(mode="json"),
        }
        detail = {
            **detail_without_integrity,
            "integrity_hmac_sha256": self._receipt_integrity(detail_without_integrity),
        }
        _atomic_write_json(
            state_root / "validations" / f"{response.validation_id}.detail.json",
            detail,
        )

    def _load_validation_receipt(
        self,
        validation_id: str,
        request_hash: str,
        *,
        expected_task: str | None,
        require_unexpired: bool = True,
    ) -> tuple[Any, dict[str, Any]]:
        if not _VALIDATION_ID.fullmatch(validation_id):
            raise RunStateError("Invalid validation ID.")
        if not re.fullmatch(r"[0-9a-f]{64}", request_hash):
            raise RunStateError("Invalid validation request hash.")
        state_root = self.settings.service_state_root
        if state_root is None:
            raise RunStateError("The MCP validation state root is not configured.")
        receipt_path = state_root / "validations" / f"{validation_id}.json"
        try:
            receipt = _read_json(receipt_path)
        except RunStateError as exc:
            raise RunStateError(f"Validation receipt does not exist or is unreadable: {validation_id}") from exc
        receipt_fields = set(receipt)
        legacy_receipt = receipt.get("schema_version") == 1 and receipt_fields == _VALIDATION_RECEIPT_FIELDS_V1
        current_receipt = receipt.get("schema_version") == 2 and receipt_fields == _VALIDATION_RECEIPT_FIELDS
        if not legacy_receipt and not current_receipt:
            raise RunStateError("Validation receipt integrity check failed: unknown or missing fields.")
        recorded_integrity = receipt.pop("integrity_hmac_sha256")
        if not isinstance(recorded_integrity, str) or not hmac.compare_digest(
            recorded_integrity,
            self._receipt_integrity(receipt),
        ):
            raise RunStateError("Validation receipt integrity check failed.")
        receipt["integrity_hmac_sha256"] = recorded_integrity
        if receipt["validation_id"] != validation_id:
            raise RunStateError("Validation receipt identity is invalid.")
        if not hmac.compare_digest(str(receipt["request_hash"]), request_hash):
            raise RunStateError("Validation request hash does not match the receipt.")
        if not isinstance(receipt["expires_at_epoch"], int):
            raise RunStateError("Validation receipt expiry is invalid.")
        if require_unexpired and time.time() > receipt["expires_at_epoch"]:
            raise RunStateError("Validation receipt expired; call validate_analysis again.")
        if expected_task is not None and receipt["task"] != expected_task:
            raise RunStateError(f"Validation receipt task must be {expected_task!r} in this scoped session.")
        if _json_sha256(receipt["request"]) != request_hash:
            raise RunStateError("Validation request integrity check failed.")
        if canonical_sha256(receipt["canonical_contract"]) != receipt["canonical_contract_hash"]:
            raise RunStateError("Validation canonical scientific contract integrity check failed.")
        try:
            request = AnalysisRequest.model_validate(receipt["request"]).root
        except Exception as exc:
            raise RunStateError("Validated analysis request is no longer readable by the strict protocol.") from exc
        if request.task != receipt["task"]:
            raise RunStateError("Validation receipt task identity is invalid.")
        if legacy_receipt and getattr(request, "event_dataset_path", None) is not None:
            raise RunStateError("The legacy validation receipt does not bind the Time Series event dataset; " "call validate_analysis again.")
        if legacy_receipt:
            receipt["event"] = None
        return request, receipt

    def get_validation_detail(
        self,
        validation_id: str,
        request_hash: str,
        *,
        expected_task: str | None = None,
    ) -> AnalysisValidationResponse:
        """Read a stored full validation response without validating again."""
        _, receipt = self._load_validation_receipt(
            validation_id,
            request_hash,
            expected_task=expected_task,
            require_unexpired=False,
        )
        state_root = self.settings.service_state_root
        if state_root is None:
            raise RunStateError("The MCP validation state root is not configured.")
        detail_path = state_root / "validations" / f"{validation_id}.detail.json"
        try:
            detail = _read_json(detail_path)
        except RunStateError as exc:
            raise RunStateError("Complete validation detail is unavailable for this legacy validation " "reference; call validate_analysis once to create a current reference.") from exc
        detail_schema_version = detail.get("schema_version")
        if set(detail) != _VALIDATION_DETAIL_FIELDS or detail_schema_version not in {1, 2}:
            raise RunStateError("Complete validation detail integrity check failed: unknown or missing fields.")
        recorded_integrity = detail.pop("integrity_hmac_sha256")
        if not isinstance(recorded_integrity, str) or not hmac.compare_digest(
            recorded_integrity,
            self._receipt_integrity(detail),
        ):
            raise RunStateError("Complete validation detail integrity check failed.")
        detail["integrity_hmac_sha256"] = recorded_integrity
        if detail["validation_id"] != validation_id or not hmac.compare_digest(
            str(detail["request_hash"]),
            request_hash,
        ):
            raise RunStateError("Complete validation detail identity is invalid.")
        try:
            response_value = _normalize_stored_scientific_identity(
                detail["response"],
                legacy_schema=detail_schema_version == 1,
                include_execution_bound=False,
                record_label="The complete validation detail",
            )
            response = AnalysisValidationResponse.model_validate(response_value)
        except Exception as exc:
            raise RunStateError("Complete validation detail is no longer readable by the strict protocol.") from exc
        expected_identity = (
            receipt["validation_id"],
            receipt["request_hash"],
            receipt["canonical_contract_hash"],
            receipt["compiled_plan_hash"],
            receipt["task"],
        )
        observed_identity = (
            response.validation_id,
            response.request_hash,
            response.canonical_contract_hash,
            response.compiled_plan_hash,
            response.task,
        )
        if observed_identity != expected_identity:
            raise RunStateError("Complete validation detail does not match its validation receipt.")
        return response

    def _assert_validation_still_matches(
        self,
        receipt: dict[str, Any],
        prepared_training: PreparedDataset,
        resolved_training: ResolvedDataset,
        prepared_application: PreparedDataset | None,
        resolved_application: ResolvedDataset | None,
        event_snapshot: DatasetSnapshot | None,
        cli_executable: Path,
        cli_version: str,
        plan: InteractionPlan,
        environment_snapshot: EnvironmentSnapshot,
    ) -> None:
        requested_task = receipt["task"]
        current_training = self._prepared_identity(
            prepared_training,
            resolved_training,
            requested_task=requested_task,
            requested_role="training",
        )
        if receipt["training"] != current_training:
            raise RunStateError("The training dataset changed since validation; validate the request again.")
        current_application = (
            self._prepared_identity(
                prepared_application,
                resolved_application,
                requested_task=requested_task,
                requested_role="application",
            )
            if prepared_application is not None and resolved_application is not None
            else None
        )
        if receipt["application"] != current_application:
            raise RunStateError("The application dataset changed since validation; validate the request again.")
        if receipt["event"] != self._event_identity(event_snapshot):
            raise RunStateError("The event dataset changed since validation; validate the request again.")
        current_cli = {
            "executable": str(cli_executable.resolve()),
            "executable_sha256": sha256_file(cli_executable),
            "version": cli_version,
        }
        if receipt["cli"] != current_cli:
            raise RunStateError("The configured GeochemistryPi CLI changed since validation; validate the request again.")
        if receipt["interaction_plan"] != _plan_identity(plan):
            raise RunStateError("The compiled interaction plan changed since validation; validate the request again.")
        current_environment = {
            "identity_sha256": environment_snapshot.identity_sha256,
            "record": environment_snapshot.record,
        }
        if receipt["environment"] != current_environment:
            raise RunStateError("The isolated CLI environment changed since validation; validate the request again.")

    def validate(
        self,
        request: ClassificationRequest | RegressionRequest | ClusteringRequest | DecompositionRequest | AnomalyDetectionRequest | TimeSeriesRequest,
    ) -> AnalysisValidationResponse:
        """Resolve and compile an analysis without creating a run or CLI process."""
        cli_executable, cli_version = self.cli_resolver()
        environment_snapshot = self.environment_resolver(cli_executable)
        existing_experiment_id = getattr(request, "existing_experiment_id", None)
        if existing_experiment_id:
            self.experiment_manager.require_matching_name(existing_experiment_id, request.experiment_name)
        resolved_training, resolved_application, secondary_is_evaluation = self._resolve_request_datasets(request)
        source_snapshot = snapshot_dataset(resolved_training.path, self.settings.maximum_dataset_bytes)
        resolved_training = self._bind_source_snapshot(
            resolved_training,
            source_snapshot,
            "training",
        )
        prepared_training = self._prepare_dataset(request, "training", resolved_training, source_snapshot)
        snapshot = prepared_training.snapshot
        application_reference, application_path, secondary_is_evaluation = self._secondary_dataset(request)
        application_source_snapshot = snapshot_dataset(resolved_application.path, self.settings.maximum_dataset_bytes) if resolved_application is not None else None
        if resolved_application is not None and application_source_snapshot is not None:
            resolved_application = self._bind_source_snapshot(
                resolved_application,
                application_source_snapshot,
                "application",
            )
        prepared_application = (
            self._prepare_dataset(request, "application", resolved_application, application_source_snapshot) if resolved_application is not None and application_source_snapshot is not None else None
        )
        application_snapshot = prepared_application.snapshot if prepared_application is not None else None
        event_snapshot = self._event_dataset_snapshot(request)
        execution_request = request.model_copy(
            update=self._execution_dataset_updates(
                request,
                snapshot.resolved_path,
                application_snapshot.resolved_path if application_snapshot else None,
                secondary_is_evaluation,
            )
        )
        dataset_context = DatasetCompilationContext(
            training_source=(resolved_training.source if snapshot.resolved_path == source_snapshot.resolved_path else "path"),
            application_source=(
                resolved_application.source
                if resolved_application is not None
                and application_source_snapshot is not None
                and application_snapshot is not None
                and application_snapshot.resolved_path == application_source_snapshot.resolved_path
                else "path"
                if resolved_application is not None
                else None
            ),
        )
        plan = self.plan_compiler.compile(
            execution_request,
            cli_executable=cli_executable,
            dataset_context=dataset_context,
        )
        plan = AnalysisPlanCompiler.bind_scientific_adapter(plan, execution_request)
        capability_probe = (self.capability_resolver or probe_cli_capabilities)(
            cli_executable,
            tuple(plan.required_cli_capabilities),
        )
        if capability_probe.missing:
            missing = ", ".join(capability_probe.missing)
            plan = replace(
                plan,
                adapter_status="unavailable",
                execution_ready=False,
                blocking_issues=tuple(
                    dict.fromkeys(
                        (
                            *plan.blocking_issues,
                            f"The configured CLI is missing required public capabilities: {missing}.",
                        )
                    )
                ),
            )
        plan = _materialize_tracking_root(plan, self.settings.tracking_root)
        artifact_requirements = planned_artifact_requirements(request, plan)
        assessment = assess_scientific_compatibility(request, plan, artifact_requirements, environment_snapshot)
        if plan.execution_ready:
            validation_paths = RunPaths.create(self.settings.runs_root, "run-0000000000000000")
            validate_workspace_path(plan, validation_paths.workspace)
        inspection = inspect_local_dataset(
            DatasetInspectionRequest(dataset_path=snapshot.resolved_path, sample_rows=0),
            self.settings,
            allow_pandas_duplicate_mangling=dataset_context.allows_pandas_duplicate_mangling("training"),
        )
        inspection_columns = inspection.column_names or tuple(column.name for column in inspection.columns)
        models = _selected_models(request)
        tuning = _selected_tuning(request)
        warnings = []
        if len(models) > 1:
            warnings.append(f"This aggregate workload will execute {len(models)} child models in CLI order.")
        if tuning == "automl":
            warnings.append("AutoML can take substantially longer than a manual model run.")
        if application_reference is None and application_path is None and request.task in {"classification", "regression"}:
            warnings.append("No application dataset was selected, so model inference will be skipped.")
        if getattr(request, "world_map", None) is not None and request.world_map.enabled:
            warnings.append("World-map rendering is enabled and may add platform-dependent image artifacts.")
        if existing_experiment_id:
            warnings.append("The new run will be attached to the verified existing MLflow experiment ID.")
        if not assessment.execution_ready:
            warnings.append("The request is scientifically representable but the exact validated contract is not execution-ready.")
        receipt = self._write_validation_receipt(
            request,
            prepared_training,
            resolved_training,
            prepared_application,
            resolved_application,
            event_snapshot,
            cli_executable,
            cli_version,
            plan,
            environment_snapshot,
        )
        response = AnalysisValidationResponse(
            validation_id=receipt["validation_id"],
            request_hash=receipt["request_hash"],
            canonical_contract_hash=receipt["canonical_contract_hash"],
            compiled_plan_hash=receipt["compiled_plan_hash"],
            validation_expires_at=datetime.fromtimestamp(
                receipt["expires_at_epoch"],
                timezone.utc,
            ).isoformat(),
            execution_ready=assessment.execution_ready,
            comparison_ready=assessment.comparison_ready,
            claim_ready=assessment.claim_ready,
            scientific_status=assessment.scientific_status,
            adapter_status=assessment.adapter_status,
            artifact_status=assessment.artifact_status,
            environment_status=assessment.environment_status,
            workflow_family=plan.workflow_family,
            workflow_mode=plan.workflow_mode,
            method=plan.method,
            scientific_contract_id=plan.scientific_contract_id,
            adapter_id=plan.adapter_id,
            adapter_version=plan.adapter_version,
            adapter_identity=(f"{plan.adapter_id}@{plan.adapter_version}" if plan.adapter_id and plan.adapter_version else plan.adapter_id),
            artifact_requirements=artifact_requirements,
            blocking_issues=assessment.blocking_issues,
            task=request.task,
            models=models,
            estimated_model_count=len(models),
            tuning=tuning,
            training_source=resolved_training.source,
            training_dataset_path=str(snapshot.resolved_path),
            training_sha256=snapshot.sha256,
            training_size_bytes=snapshot.size_bytes,
            source_dataset_path=str(source_snapshot.resolved_path),
            source_dataset_sha256=source_snapshot.sha256,
            dataset_preparation=prepared_training.record,
            dataset_preparation_sha256=prepared_training.record["contract_hash"],
            environment_identity_sha256=environment_snapshot.identity_sha256,
            environment_profile={
                "requested": resolved_environment_profile(request),
                "observed": environment_snapshot.record,
            },
            environment_profile_identity_sha256=plan.environment_profile_identity_sha256,
            requested_seeds=dict(plan.requested_seeds),
            effective_seeds=dict(plan.effective_seeds),
            execution_decisions=_validation_execution_decisions(
                request,
                plan,
                application_enabled=application_snapshot is not None,
                secondary_is_evaluation=secondary_is_evaluation,
            ),
            parameter_binding={
                "requested_model": {name: json.loads(value) for name, value in plan.requested_model_parameters},
                "effective_model": {name: json.loads(value) for name, value in plan.effective_model_parameters},
                "model_binding": plan.model_parameter_binding,
                "requested_preprocessing": {name: json.loads(value) for name, value in plan.requested_preprocessing_parameters},
                "effective_preprocessing": {name: json.loads(value) for name, value in plan.effective_preprocessing_parameters},
                "preprocessing_binding": plan.preprocessing_parameter_binding,
            },
            adapter_artifact_mappings=tuple(
                {
                    "mapping_id": mapping.mapping_id,
                    "scientific_type": mapping.scientific_type,
                    "output_role": mapping.output_role,
                    "relative_path": mapping.relative_path,
                    "availability": mapping.availability,
                    "reason": mapping.reason,
                }
                for mapping in plan.artifact_mappings
            ),
            source_row_count=snapshot.row_lineage.source_row_count,
            row_identity_scheme=snapshot.row_lineage.scheme,
            row_identity_sha256=snapshot.row_lineage.ordered_identity_sha256,
            columns=inspection_columns,
            identifier_column=getattr(request, "identifier_column", None),
            feature_columns=tuple(getattr(request, "feature_columns", ())),
            selected_columns=tuple(getattr(request, "resolved_selected_columns", ())),
            target_column=getattr(request, "target_column", None),
            target_columns=(
                tuple(column for column in inspection_columns if column in set(request.resolved_target_columns))
                if isinstance(request, RegressionRequest)
                else ((request.target_column,) if isinstance(request, ClassificationRequest) else ())
            ),
            resolved_model_parameters=_resolved_model_parameters(request),
            application_source=resolved_application.source if resolved_application else None,
            application_dataset_path=str(application_snapshot.resolved_path) if application_snapshot else None,
            application_sha256=application_snapshot.sha256 if application_snapshot else None,
            application_source_sha256=(application_source_snapshot.sha256 if application_source_snapshot else None),
            application_preparation=(prepared_application.record if prepared_application else None),
            application_source_row_count=(application_snapshot.row_lineage.source_row_count if application_snapshot else None),
            application_row_identity_sha256=(application_snapshot.row_lineage.ordered_identity_sha256 if application_snapshot else None),
            event_dataset_path=(str(event_snapshot.resolved_path) if event_snapshot else None),
            event_source_sha256=(event_snapshot.sha256 if event_snapshot else None),
            event_size_bytes=(event_snapshot.size_bytes if event_snapshot else None),
            experiment_mode=("not_applicable" if request.task == "time_series" else "existing" if existing_experiment_id else "new"),
            experiment_name=request.experiment_name,
            existing_experiment_id=existing_experiment_id,
            interaction_plan=plan.name,
            warnings=tuple(warnings),
        )
        self._write_validation_detail(response)
        return response

    def start_validated(
        self,
        validation_id: str,
        request_hash: str,
        *,
        expected_task: str | None = None,
    ) -> StartAnalysisResponse:
        """Start only the immutable request and inputs covered by a validation receipt."""
        request, receipt = self._load_validation_receipt(
            validation_id,
            request_hash,
            expected_task=expected_task,
        )
        return self._start(
            request,
            validation={**receipt, "mode": "validation_reference"},
        )

    def start(
        self,
        request: ClassificationRequest | RegressionRequest | ClusteringRequest | DecompositionRequest | AnomalyDetectionRequest | TimeSeriesRequest,
    ) -> StartAnalysisResponse:
        """Retain the strict legacy full-request start path for compatibility."""
        return self._start(request, validation=None)

    def _start(
        self,
        request: ClassificationRequest | RegressionRequest | ClusteringRequest | DecompositionRequest | AnomalyDetectionRequest | TimeSeriesRequest,
        *,
        validation: dict[str, Any] | None,
    ) -> StartAnalysisResponse:
        """Validate synchronously, then queue the long-running CLI work."""
        self._ensure_initialized()
        with self._lock:
            if self._closed:
                raise RunStateError("The GeochemistryPi run manager is shutting down.")
        cli_executable, cli_version = self.cli_resolver()
        environment_snapshot = self.environment_resolver(cli_executable)
        existing_experiment_id = getattr(request, "existing_experiment_id", None)
        if existing_experiment_id:
            self.experiment_manager.require_matching_name(existing_experiment_id, request.experiment_name)
        application_reference, application_dataset_path, secondary_is_evaluation = self._secondary_dataset(request)
        if validation is not None:
            resolved_training = self._resolved_from_validation_receipt(
                request,
                validation["training"],
                requested_role="training",
            )
            recorded_application = validation["application"]
            has_application = application_reference is not None or application_dataset_path is not None
            if has_application != (recorded_application is not None):
                raise RunStateError("The validation receipt application-dataset identity is inconsistent with the request.")
            resolved_application = (
                self._resolved_from_validation_receipt(
                    request,
                    recorded_application,
                    requested_role="application",
                )
                if recorded_application is not None
                else None
            )
        else:
            resolved_training, resolved_application, secondary_is_evaluation = self._resolve_request_datasets(request)
        source_snapshot = snapshot_dataset(resolved_training.path, self.settings.maximum_dataset_bytes)
        resolved_training = self._bind_source_snapshot(
            resolved_training,
            source_snapshot,
            "training",
            since_validation=validation is not None,
        )
        prepared_training = self._prepare_dataset(request, "training", resolved_training, source_snapshot)
        snapshot = prepared_training.snapshot
        application_source_snapshot = (
            snapshot_dataset(
                resolved_application.path,
                self.settings.maximum_dataset_bytes,
            )
            if resolved_application is not None
            else None
        )
        if resolved_application is not None and application_source_snapshot is not None:
            resolved_application = self._bind_source_snapshot(
                resolved_application,
                application_source_snapshot,
                "application",
                since_validation=validation is not None,
            )
        prepared_application = (
            self._prepare_dataset(request, "application", resolved_application, application_source_snapshot) if resolved_application is not None and application_source_snapshot is not None else None
        )
        application_snapshot = prepared_application.snapshot if prepared_application is not None else None
        event_snapshot = self._event_dataset_snapshot(request)
        execution_request = request.model_copy(
            update=self._execution_dataset_updates(
                request,
                snapshot.resolved_path,
                application_snapshot.resolved_path if application_snapshot else None,
                secondary_is_evaluation,
            )
        )
        dataset_context = DatasetCompilationContext(
            training_source=(resolved_training.source if snapshot.resolved_path == source_snapshot.resolved_path else "path"),
            application_source=(
                resolved_application.source
                if resolved_application is not None
                and application_source_snapshot is not None
                and application_snapshot is not None
                and application_snapshot.resolved_path == application_source_snapshot.resolved_path
                else "path"
                if resolved_application is not None
                else None
            ),
        )
        plan = self.plan_compiler.compile(
            execution_request,
            cli_executable=cli_executable,
            dataset_context=dataset_context,
        )
        plan = AnalysisPlanCompiler.bind_scientific_adapter(plan, execution_request)
        capability_probe = (self.capability_resolver or probe_cli_capabilities)(
            cli_executable,
            tuple(plan.required_cli_capabilities),
        )
        if capability_probe.missing:
            missing = ", ".join(capability_probe.missing)
            plan = replace(
                plan,
                adapter_status="unavailable",
                execution_ready=False,
                blocking_issues=tuple(
                    dict.fromkeys(
                        (
                            *plan.blocking_issues,
                            f"The configured CLI is missing required public capabilities: {missing}.",
                        )
                    )
                ),
            )
        plan = _materialize_tracking_root(plan, self.settings.tracking_root)
        if validation is not None:
            self._assert_validation_still_matches(
                validation,
                prepared_training,
                resolved_training,
                prepared_application,
                resolved_application,
                event_snapshot,
                cli_executable,
                cli_version,
                plan,
                environment_snapshot,
            )
        artifact_requirements = planned_artifact_requirements(request, plan)
        assessment = assess_scientific_compatibility(request, plan, artifact_requirements, environment_snapshot)
        if not assessment.execution_ready:
            explanation = "; ".join(assessment.blocking_issues) or "The exact scientific contract has no executable CLI adapter."
            raise RunStateError(f"Validated scientific contract is not execution-ready: {explanation}")
        if validation is None:
            validation = {
                **self._write_validation_receipt(
                    request,
                    prepared_training,
                    resolved_training,
                    prepared_application,
                    resolved_application,
                    event_snapshot,
                    cli_executable,
                    cli_version,
                    plan,
                    environment_snapshot,
                ),
                "mode": "legacy_full_request",
            }
        if "data-mining" in plan.public_command:
            assert self.settings.tracking_root is not None
            self.settings.tracking_root.mkdir(parents=True, exist_ok=True)
        run_id = _new_run_id()
        paths = RunPaths.create(self.settings.runs_root, run_id)
        validate_workspace_path(plan, paths.workspace)
        created_at = _utc_now()
        request_value = _validation_request_value(request)
        request_hash = _json_sha256(request_value)
        if validation is not None and not hmac.compare_digest(
            request_hash,
            str(validation["request_hash"]),
        ):
            raise RunStateError("Validated request serialization changed before run creation.")
        request_record = {
            "schema_version": 2,
            "run_id": run_id,
            "request_hash": request_hash,
            "canonical_contract_hash": canonical_sha256(canonical_scientific_contract(request, plan)),
            "request": request_value,
            "validation": (
                {
                    "mode": validation.get("mode", "validation_reference"),
                    "validation_id": validation["validation_id"],
                    "request_hash": validation["request_hash"],
                }
                if validation is not None
                else None
            ),
            "input": {
                "source_path": str(snapshot.source_path),
                "resolved_path": str(snapshot.resolved_path),
                "size_bytes": snapshot.size_bytes,
                "sha256": snapshot.sha256,
                "format": snapshot.format,
                "source": resolved_training.source,
                "dataset_id": resolved_training.dataset_id,
                "row_identity": snapshot.row_lineage.as_record(),
                "source_file": prepared_training.record["source_file"],
                "preparation": prepared_training.record,
            },
            "application_input": (
                {
                    "source_path": str(application_snapshot.source_path),
                    "resolved_path": str(application_snapshot.resolved_path),
                    "size_bytes": application_snapshot.size_bytes,
                    "sha256": application_snapshot.sha256,
                    "format": application_snapshot.format,
                    "source": resolved_application.source,
                    "dataset_id": resolved_application.dataset_id,
                    "row_identity": application_snapshot.row_lineage.as_record(),
                    "source_file": prepared_application.record["source_file"],
                    "preparation": prepared_application.record,
                }
                if application_snapshot is not None
                else None
            ),
            "event_input": self._event_identity(event_snapshot),
            "interaction_plan": {
                "name": plan.name,
                "schema_version": plan.schema_version,
                "sha256": _plan_identity(plan)["sha256"],
                "adapter_id": plan.adapter_id,
                "adapter_version": plan.adapter_version,
                "workflow_family": plan.workflow_family,
                "workflow_mode": plan.workflow_mode,
                "method": plan.method,
                "scientific_contract_id": plan.scientific_contract_id,
                "scientific_execution_contract_bound": plan.scientific_execution_contract_json is not None,
                "environment_profile_id": plan.environment_profile_id,
                "environment_profile_identity_sha256": plan.environment_profile_identity_sha256,
                "requested_seeds": dict(plan.requested_seeds),
                "effective_seeds": dict(plan.effective_seeds),
                "seed_binding": plan.seed_binding,
                "requested_model_parameters": {name: json.loads(value) for name, value in plan.requested_model_parameters},
                "effective_model_parameters": {name: json.loads(value) for name, value in plan.effective_model_parameters},
                "requested_preprocessing_parameters": {name: json.loads(value) for name, value in plan.requested_preprocessing_parameters},
                "effective_preprocessing_parameters": {name: json.loads(value) for name, value in plan.effective_preprocessing_parameters},
                "artifact_mappings": [
                    {
                        "mapping_id": mapping.mapping_id,
                        "scientific_type": mapping.scientific_type,
                        "output_role": mapping.output_role,
                        "relative_path": mapping.relative_path,
                        "availability": mapping.availability,
                        "reason": mapping.reason,
                    }
                    for mapping in plan.artifact_mappings
                ],
            },
            "versions": {
                "geochemistrypi_mcp": SERVER_VERSION,
                "geochemistrypi_cli": cli_version,
                "environment_identity_sha256": environment_snapshot.identity_sha256,
                "environment": environment_snapshot.record,
            },
        }
        status = {
            "schema_version": 1,
            "run_id": run_id,
            "state": "queued",
            "stage": "queued",
            "created_at": created_at,
            "started_at": None,
            "finished_at": None,
            "cli_pid": None,
            "recorded_cli_pid": None,
            "recorded_process_create_time": None,
            "progress_message": "Waiting for the local GeochemistryPi CLI execution slot.",
            "error": None,
            "error_truncated": False,
            "error_sha256": None,
            "error_total_utf8_bytes": None,
            "result_type": None,
            "retryable": None,
        }
        control = _RunControl(cancellation=threading.Event())
        with self._lock:
            if self._closed:
                raise RunStateError("The GeochemistryPi run manager is shutting down.")
            if len(self._active) >= self.settings.maximum_pending_runs:
                raise RunStateError(f"The local run queue is full ({self.settings.maximum_pending_runs} active or queued runs). " "Wait for a run to finish or cancel one before starting another.")
            paths.wrapper.mkdir(parents=True)
            paths.workspace.mkdir()
            _atomic_write_json(paths.request, request_record)
            self._write_status(paths, status)
            self._active[run_id] = control
            control.future = self._executor.submit(
                self._execute,
                run_id,
                execution_request,
                snapshot,
                application_snapshot,
                event_snapshot,
                plan,
                cli_version,
                control,
            )
        return StartAnalysisResponse(
            run_id=run_id,
            state="queued",
            models=_selected_models(request),
            estimated_model_count=len(_selected_models(request)),
            status_hint=(
                f"Call get_run_result once with run_id {run_id} and wait_seconds=300; "
                "the call returns immediately if the run finishes sooner. Use get_run_status "
                "only when progress detail is explicitly needed."
            ),
            request_hash=(validation["request_hash"] if validation is not None else None),
            started_from_validation=(validation is not None and validation.get("mode") == "validation_reference"),
        )

    def _mark_running(self, paths: RunPaths, control: _RunControl, pid: int, create_time: float) -> None:
        with self._lock:
            if control.cancellation.is_set():
                raise CliRunCancelledError(
                    "The run was cancelled while the CLI process was starting.",
                    paths.workspace,
                )
            status = _read_json(paths.status)
            status.update(
                {
                    "state": "running",
                    "stage": "running_cli",
                    "started_at": status.get("started_at") or _utc_now(),
                    "cli_pid": pid,
                    "recorded_cli_pid": pid,
                    "recorded_process_create_time": create_time,
                    "progress_message": "The existing GeochemistryPi CLI is running in the managed workspace.",
                }
            )
            self._write_status(paths, status)

    def _build_terminal_receipt(
        self,
        paths: RunPaths,
        status: dict[str, Any],
        *,
        state: str,
        message: str,
        error: str | None,
        finished_at: str,
        result_type: str | None,
        retryable: bool | None,
    ) -> TerminalRunReceipt:
        if status.get("run_id") != paths.root.name:
            raise RunStateError("Durable run status does not match its managed run directory.")
        trace_path = paths.wrapper / "interaction-trace.json"
        cli_exit_code, cli_started_at, cli_finished_at, cli_execution_duration_seconds = _read_terminal_cli_trace_evidence(trace_path)
        analysis_process_started = bool(status.get("started_at") is not None or isinstance(status.get("recorded_cli_pid"), int) or cli_exit_code is not None)
        bounded_message = sanitize_terminal_error(message) or "The managed run reached a terminal state."
        bounded_error, error_truncated, error_sha256, error_total_utf8_bytes = terminal_error_projection(error)
        # A failed/cancelled receipt must still be publishable when the very
        # failure being recorded is missing or corrupt request metadata.  Do
        # not infer a scientific contract from the method name in that case:
        # publish an explicit unavailable identity and make no v4 binding
        # claim.  Successful results remain strict and require their complete
        # immutable request/plan identity.
        try:
            request_record = _read_json(paths.request)
            plan_identity = request_record["interaction_plan"]
            scientific_contract_id = plan_identity["scientific_contract_id"]
            scientific_execution_contract_bound = plan_identity["scientific_execution_contract_bound"]
            if not isinstance(scientific_contract_id, str) or not scientific_contract_id:
                raise KeyError("scientific_contract_id")
            if not isinstance(scientific_execution_contract_bound, bool):
                raise KeyError("scientific_execution_contract_bound")
        except (KeyError, TypeError, RunStateError):
            scientific_contract_id = "scientific-contract-unavailable/corrupt-or-missing-request-record"
            scientific_execution_contract_bound = False
        if state == "failed" and bounded_error is None:
            bounded_error, error_truncated, error_sha256, error_total_utf8_bytes = terminal_error_projection("The managed run failed without an additional diagnostic.")
        return TerminalRunReceipt(
            run_id=str(status.get("run_id")),
            state=state,
            stage=state,
            created_at=status.get("created_at"),
            started_at=status.get("started_at"),
            finished_at=finished_at,
            progress_message=bounded_message,
            error=bounded_error,
            error_truncated=error_truncated,
            error_sha256=error_sha256,
            error_total_utf8_bytes=error_total_utf8_bytes,
            result_type=result_type,
            retryable=retryable,
            scientific_contract_id=scientific_contract_id,
            scientific_execution_contract_bound=scientific_execution_contract_bound,
            analysis_process_started=analysis_process_started,
            cli_exit_code=cli_exit_code,
            cli_started_at=cli_started_at,
            cli_finished_at=cli_finished_at,
            cli_execution_duration_seconds=cli_execution_duration_seconds,
            interaction_trace=_terminal_evidence_identity(trace_path, "interaction_trace"),
            cli_stdout_log=_terminal_evidence_identity(paths.wrapper / "stdout.log", "cli_stdout_log"),
            cli_stderr_log=_terminal_evidence_identity(paths.wrapper / "stderr.log", "cli_stderr_log"),
        )

    def _publish_terminal_receipt(
        self,
        paths: RunPaths,
        status: dict[str, Any],
        desired: TerminalRunReceipt,
    ) -> tuple[TerminalRunReceipt, str]:
        payload = desired.model_dump(
            mode="json",
            exclude={"result_record_path", "result_record_sha256"},
        )
        try:
            digest = _atomic_write_json_once(paths.terminal_result, payload)
            return desired, digest
        except FileExistsError:
            existing, digest = _read_terminal_receipt(paths.terminal_result)
            if existing.run_id != desired.run_id:
                raise RunStateError("An immutable terminal result receipt belongs to a different run.")
            current_state = status.get("state")
            if current_state not in {"queued", "running"} and existing.state != desired.state:
                raise RunStateError("An immutable terminal result receipt cannot be replaced by a different terminal state.")
            return existing, digest

    def _terminal_result(self, paths: RunPaths, status: dict[str, Any]) -> TerminalRunReceipt:
        receipt, digest = _read_terminal_receipt(paths.terminal_result)
        if receipt.run_id != paths.root.name:
            raise RunStateError("The terminal result receipt does not match its managed run directory.")
        expected_path = status.get("result_record_path")
        expected_sha256 = status.get("result_record_sha256")
        try:
            path_matches = isinstance(expected_path, str) and Path(expected_path).resolve() == paths.terminal_result.resolve()
        except OSError:
            path_matches = False
        if not path_matches or expected_sha256 != digest:
            raise RunStateError("The terminal result receipt identity is inconsistent with durable run status.")
        status_identity = (
            status.get("run_id"),
            status.get("state"),
            status.get("stage"),
            status.get("created_at"),
            status.get("started_at"),
            status.get("finished_at"),
            status.get("progress_message"),
            status.get("error"),
            status.get("error_truncated", False),
            status.get("error_sha256"),
            status.get("error_total_utf8_bytes"),
            status.get("result_type"),
            status.get("retryable"),
        )
        receipt_identity = (
            receipt.run_id,
            receipt.state,
            receipt.stage,
            receipt.created_at,
            receipt.started_at,
            receipt.finished_at,
            receipt.progress_message,
            receipt.error,
            receipt.error_truncated,
            receipt.error_sha256,
            receipt.error_total_utf8_bytes,
            receipt.result_type,
            receipt.retryable,
        )
        if status_identity != receipt_identity:
            raise RunStateError("The terminal result receipt content is inconsistent with durable run status.")
        allowlisted_evidence = {
            "interaction_trace": (paths.wrapper / "interaction-trace.json", "interaction_trace"),
            "cli_stdout_log": (paths.wrapper / "stdout.log", "cli_stdout_log"),
            "cli_stderr_log": (paths.wrapper / "stderr.log", "cli_stderr_log"),
        }
        for field_name, (evidence_path, evidence_kind) in allowlisted_evidence.items():
            recorded = getattr(receipt, field_name)
            if recorded is None:
                continue
            current = _terminal_evidence_identity(evidence_path, evidence_kind)
            if current != recorded:
                raise RunStateError(f"The terminal {field_name} identity is no longer valid.")
        return receipt.model_copy(
            update={
                "result_record_path": str(paths.terminal_result),
                "result_record_sha256": digest,
            }
        )

    def _finish(
        self,
        paths: RunPaths,
        state: str,
        message: str,
        error: str | None = None,
        *,
        result_record_sha256: str | None = None,
        result_type: str | None = None,
        retryable: bool | None = None,
    ) -> None:
        with self._lock:
            status = _read_json(paths.status)
            finished_at = _utc_now()
            terminal_receipt = None
            terminal_receipt_sha256 = None
            success_result_sha256 = None
            error_truncated = False
            error_sha256 = None
            error_total_utf8_bytes = None
            terminal_result_type = None
            terminal_retryable = None
            if state in {"succeeded", "partial_failure"}:
                if result_record_sha256 is None or not paths.result.is_file():
                    raise RunStateError("A successful terminal state requires an immutable result record.")
                observed_result_sha256 = sha256_file(paths.result)
                if not hmac.compare_digest(observed_result_sha256, result_record_sha256):
                    raise RunStateError("The successful result record changed before terminal status publication.")
                success_result_sha256 = observed_result_sha256
            if state in {"failed", "cancelled"}:
                if state == "failed" and result_type is None:
                    result_type = "internal_error"
                if state == "failed" and retryable is None:
                    retryable = False
                desired = self._build_terminal_receipt(
                    paths,
                    status,
                    state=state,
                    message=message,
                    error=error,
                    finished_at=finished_at,
                    result_type=result_type,
                    retryable=retryable,
                )
                terminal_receipt, terminal_receipt_sha256 = self._publish_terminal_receipt(
                    paths,
                    status,
                    desired,
                )
                state = terminal_receipt.state
                message = terminal_receipt.progress_message
                error = terminal_receipt.error
                error_truncated = terminal_receipt.error_truncated
                error_sha256 = terminal_receipt.error_sha256
                error_total_utf8_bytes = terminal_receipt.error_total_utf8_bytes
                terminal_result_type = terminal_receipt.result_type
                terminal_retryable = terminal_receipt.retryable
                finished_at = terminal_receipt.finished_at
            status.update(
                {
                    "state": state,
                    "stage": ("completed" if state in {"succeeded", "partial_failure"} else "cancelled" if state == "cancelled" else "failed"),
                    "finished_at": finished_at,
                    "cli_pid": None,
                    "progress_message": message,
                    "error": error,
                    "error_truncated": error_truncated,
                    "error_sha256": error_sha256,
                    "error_total_utf8_bytes": error_total_utf8_bytes,
                    "result_type": terminal_result_type,
                    "retryable": terminal_retryable,
                }
            )
            if terminal_receipt is not None and terminal_receipt_sha256 is not None:
                status.update(
                    {
                        "result_record_path": str(paths.terminal_result),
                        "result_record_sha256": terminal_receipt_sha256,
                    }
                )
            elif success_result_sha256 is not None:
                status.update(
                    {
                        "result_record_path": str(paths.result),
                        "result_record_sha256": success_result_sha256,
                    }
                )
            self._write_status(paths, status)

    def _execute(
        self,
        run_id: str,
        request: ClassificationRequest | RegressionRequest | ClusteringRequest | DecompositionRequest | AnomalyDetectionRequest | TimeSeriesRequest,
        snapshot: DatasetSnapshot,
        application_snapshot: DatasetSnapshot | None,
        event_snapshot: DatasetSnapshot | None,
        plan: InteractionPlan,
        cli_version: str,
        control: _RunControl,
    ) -> None:
        paths = RunPaths.create(self.settings.runs_root, run_id)
        try:
            if control.cancellation.is_set():
                raise CliRunCancelledError("The queued run was cancelled before execution.", paths.workspace)
            if sha256_file(snapshot.resolved_path) != snapshot.sha256:
                raise InputIntegrityError("The input dataset changed after validation and before CLI execution.")
            if application_snapshot is not None and sha256_file(application_snapshot.resolved_path) != application_snapshot.sha256:
                raise InputIntegrityError("The application dataset changed after validation and before CLI execution.")
            if event_snapshot is not None and sha256_file(event_snapshot.resolved_path) != event_snapshot.sha256:
                raise InputIntegrityError("The event dataset changed after validation and before CLI execution.")
            driver = self.driver_factory()
            cli_result = driver.run(
                plan,
                workspace=paths.workspace,
                capture_directory=paths.wrapper,
                cancellation_event=control.cancellation,
                process_started=lambda pid, create_time: self._mark_running(paths, control, pid, create_time),
                trace_metadata={
                    "geochemistrypi_mcp_version": SERVER_VERSION,
                    "geochemistrypi_cli_version": cli_version,
                },
            )
            if control.cancellation.is_set():
                raise CliRunCancelledError(
                    "The run was cancelled before its result was published.",
                    paths.workspace,
                )
            with self._lock:
                status = _read_json(paths.status)
                status.update(
                    {
                        "stage": "indexing_outputs",
                        "progress_message": "The CLI finished; original outputs are being indexed without recalculation.",
                    }
                )
                self._write_status(paths, status)
            final_hash = sha256_file(snapshot.resolved_path)
            if final_hash != snapshot.sha256:
                raise InputIntegrityError("The input dataset changed during CLI execution; the result was not published as valid.")
            application_hash_verified = None
            if application_snapshot is not None:
                application_hash_verified = sha256_file(application_snapshot.resolved_path) == application_snapshot.sha256
                if not application_hash_verified:
                    raise InputIntegrityError("The application dataset changed during CLI execution; the result was not published as valid.")
            event_hash_verified = None
            if event_snapshot is not None:
                event_hash_verified = sha256_file(event_snapshot.resolved_path) == event_snapshot.sha256
                if not event_hash_verified:
                    raise InputIntegrityError("The event dataset changed during CLI execution; the result was not published as valid.")
            output_directory = paths.workspace / "geopi_output" / request.experiment_name / request.run_name
            source_row_pairing = (
                verify_original_row_pairing(
                    snapshot.resolved_path,
                    output_directory,
                    request.identifier_column,
                    snapshot.row_lineage,
                )
                if plan.requires_source_row_pairing
                else None
            )
            artifact_requirements = planned_artifact_requirements(request, plan)
            expected_attestation_source_sha256 = None
            expected_attestation_source_contract = None
            if plan.scientific_execution_contract_json is not None:
                scientific_execution_path = paths.wrapper / "scientific-execution.json"
                if scientific_execution_path.is_file():
                    try:
                        expected_attestation_source_contract = json.loads(plan.scientific_execution_contract_json)
                        observed_attestation_source_contract = json.loads(scientific_execution_path.read_text(encoding="utf-8"))
                    except (OSError, UnicodeError, json.JSONDecodeError) as exc:
                        raise InputIntegrityError("The wrapper-owned scientific execution sidecar is not valid JSON.") from exc
                    if _canonical_json_bytes(observed_attestation_source_contract) != _canonical_json_bytes(expected_attestation_source_contract):
                        raise InputIntegrityError("The wrapper-owned scientific execution sidecar changed during CLI execution.")
                    expected_attestation_source_sha256 = sha256_file(scientific_execution_path)
            discovered = discover_artifacts(
                output_directory,
                self.settings.maximum_artifact_references,
                artifact_requirements,
                workflow_family=plan.workflow_family,
                artifact_mappings=plan.artifact_mappings,
                expected_attestation_source_sha256=expected_attestation_source_sha256,
                expected_attestation_source_contract=expected_attestation_source_contract,
            )
            preprocessing_summary = (
                read_time_series_preprocessing_summary(
                    output_directory,
                    source_row_count=snapshot.row_lineage.source_row_count,
                    indexed_relative_paths=tuple(entry["relative_path"] for entry in discovered.all_index_entries),
                )
                if request.task == "time_series" and request.mode in {"subaerial_proportion", "continuous"}
                else None
            )
            is_aggregate = request.task != "time_series" and request.model_selection.mode == "all"
            aggregate_state, children = _aggregate_children(output_directory, request.task) if is_aggregate else (None, ())
            request_record = _read_json(paths.request)
            validation_identity = request_record["validation"] or {
                "mode": "legacy_full_request",
                "validation_id": None,
                "request_hash": request_record["request_hash"],
            }
            missing_artifact_requirement_ids = discovered.missing_requirement_ids
            result_state = "partial_failure" if aggregate_state == "partial_failure" or missing_artifact_requirement_ids else "succeeded"
            artifact_index = {
                "schema_version": ARTIFACT_INDEX_SCHEMA_VERSION,
                "run_id": run_id,
                "request_identity": {
                    "request_hash": request_record["request_hash"],
                    "canonical_contract_hash": request_record["canonical_contract_hash"],
                },
                "validation_identity": validation_identity,
                "plan_identity": request_record["interaction_plan"],
                "output_directory": str(output_directory),
                "source_row_lineage": snapshot.row_lineage.as_record(),
                "source_row_pairing": source_row_pairing,
                "artifact_contract": {
                    "required": [item.model_dump(mode="json") for item in artifact_requirements],
                    "missing_required_ids": list(missing_artifact_requirement_ids),
                    "matches": {key: list(value) for key, value in discovered.requirement_matches.items()},
                    "failures": discovered.requirement_failures,
                },
                "artifacts": discovered.all_index_entries,
            }
            _atomic_write_json(paths.artifact_index, artifact_index)
            artifact_index_sha256 = sha256_file(paths.artifact_index)
            required_tabular_observations = build_required_tabular_observations(
                output_directory,
                paths.artifact_index,
                artifact_index_sha256,
            )
            provenance_manifest = {
                "schema_version": 2,
                "request_identity": {
                    "request_hash": request_record["request_hash"],
                    "canonical_contract_hash": request_record["canonical_contract_hash"],
                },
                "validation_identity": validation_identity,
                "run_identity": {"run_id": run_id},
                "plan_identity": request_record["interaction_plan"],
                "runtime": request_record["versions"],
                "input": request_record["input"],
                "application_input": request_record["application_input"],
                "event_input": request_record["event_input"],
                "artifact_index": {
                    "path": str(paths.artifact_index),
                    "sha256": artifact_index_sha256,
                },
                "required_tabular_observations": {
                    "artifact_index_sha256": required_tabular_observations.artifact_index_sha256,
                    "total_count": required_tabular_observations.total_count,
                    "returned_count": required_tabular_observations.returned_count,
                    "truncated": required_tabular_observations.truncated,
                    "observations_sha256": required_tabular_observations.observations_sha256,
                    "returned_cell_count": required_tabular_observations.returned_cell_count,
                    "returned_utf8_bytes": required_tabular_observations.returned_utf8_bytes,
                    "omitted_artifact_count": required_tabular_observations.omitted_artifact_count,
                    "omission_reason_counts": required_tabular_observations.omission_reason_counts,
                    "omissions_sha256": required_tabular_observations.omissions_sha256,
                },
                "artifacts": list(discovered.all_index_entries),
                "artifact_contract": artifact_index["artifact_contract"],
                "metadata": {
                    "task": request.task,
                    "workflow_family": plan.workflow_family,
                    "workflow_mode": plan.workflow_mode,
                    "method": plan.method,
                    "adapter_id": plan.adapter_id,
                    "adapter_version": plan.adapter_version,
                },
            }
            _atomic_write_json(paths.provenance_manifest, provenance_manifest)
            provenance_manifest_sha256 = sha256_file(paths.provenance_manifest)
            trace_exit_code, cli_started_at, cli_finished_at, cli_execution_duration_seconds = _read_cli_trace_evidence(cli_result.trace_path)
            if trace_exit_code != cli_result.returncode or cli_started_at is None:
                raise RunStateError("The completed CLI interaction trace has no matching child-process interval.")
            result = RunResultResponse(
                run_id=run_id,
                request_hash=request_record["request_hash"],
                validation_id=(request_record["validation"] or {}).get("validation_id"),
                canonical_contract_hash=request_record["canonical_contract_hash"],
                compiled_plan_hash=request_record["interaction_plan"]["sha256"],
                scientific_contract_id=plan.scientific_contract_id,
                scientific_execution_contract_bound=plan.scientific_execution_contract_json is not None,
                provenance_manifest_path=str(paths.provenance_manifest),
                provenance_manifest_sha256=provenance_manifest_sha256,
                contract_status=("incomplete" if missing_artifact_requirement_ids else "complete"),
                missing_artifact_requirement_ids=missing_artifact_requirement_ids,
                state=result_state,
                task=request.task,
                model=("all_models" if is_aggregate else _selected_models(request)[0]),
                tuning=(request.model_selection.tuning if is_aggregate else _selected_tuning(request)),
                output_directory=str(output_directory),
                interaction_trace=str(cli_result.trace_path),
                cli_stdout_log=str(cli_result.stdout_path),
                cli_stderr_log=str(cli_result.stderr_path),
                cli_exit_code=cli_result.returncode,
                cli_started_at=cli_started_at,
                cli_finished_at=cli_finished_at,
                cli_execution_duration_seconds=cli_execution_duration_seconds,
                cli_version=cli_version,
                input_sha256=snapshot.sha256,
                input_hash_verified=True,
                source_input_sha256=request_record["input"]["source_file"]["sha256"],
                dataset_preparation=request_record["input"]["preparation"],
                environment_identity_sha256=request_record["versions"]["environment_identity_sha256"],
                effective_seeds=request_record["interaction_plan"]["effective_seeds"],
                source_row_count=snapshot.row_lineage.source_row_count,
                row_identity_scheme=snapshot.row_lineage.scheme,
                row_identity_sha256=snapshot.row_lineage.ordered_identity_sha256,
                source_row_pairing_verified=(source_row_pairing["verified"] if source_row_pairing else None),
                source_row_pairing_sha256=(source_row_pairing["ordered_pairing_sha256"] if source_row_pairing else None),
                preprocessing_summary=preprocessing_summary,
                application_input_sha256=(application_snapshot.sha256 if application_snapshot is not None else None),
                application_input_hash_verified=application_hash_verified,
                event_input_sha256=(event_snapshot.sha256 if event_snapshot is not None else None),
                event_input_hash_verified=event_hash_verified,
                reported_metrics=discovered.reported_metrics,
                required_tabular_observations=required_tabular_observations,
                artifact_count=len(discovered.all_index_entries),
                artifact_offset=0,
                returned_artifact_count=len(discovered.response_references),
                next_artifact_offset=(len(discovered.response_references) if discovered.truncated else None),
                artifacts=discovered.response_references,
                artifacts_truncated=discovered.truncated,
                aggregate_state=aggregate_state,
                aggregate_summary=(
                    {
                        "expected_model_count": len(children),
                        "succeeded_count": sum(child["state"] == "succeeded" for child in children),
                        "failed_count": sum(child["state"] == "failed" for child in children),
                    }
                    if is_aggregate
                    else None
                ),
                children=children,
                limitations=(
                    "Metrics are read from original CLI files and are not recalculated by MCP.",
                    "The MCP wrapper does not recreate models, metrics, predictions, plots, or transformed datasets.",
                    {
                        "classification": "The public sample-balancing helper remains unwired in the CLI workflow.",
                        "regression": (
                            "Multiple-target regression exposes per-target holdout metrics, while cross-validation metrics remain uniformly "
                            "averaged across targets; univariate feature selection is rejected for this branch."
                        ),
                        "clustering": "Application inference, targets, unsupervised AutoML, and internal-only OPTICS remain outside the public workflow.",
                        "decomposition": "Application inference, targets, unsupervised AutoML, and unresolved missing values remain outside the public workflow.",
                        "anomaly_detection": "Application inference, targets, unsupervised AutoML, feature selection, and unresolved missing values remain outside the public workflow.",
                        "time_series": "The Time Series workflow uses the validated CLI numerical implementation and does not perform model inference.",
                    }[request.task],
                ),
            )
            result_record_sha256 = _atomic_write_json_once(
                paths.result,
                result.model_dump(mode="json"),
            )
            self._finish(
                paths,
                result_state,
                "The existing GeochemistryPi CLI completed and its original outputs were indexed."
                if result_state == "succeeded"
                else ("The GeochemistryPi CLI process completed, but the scientific result is a partial failure because " "one or more required artifacts are missing or an aggregate child failed."),
                result_record_sha256=result_record_sha256,
            )
        except CliRunCancelledError:
            self._finish(
                paths,
                "cancelled",
                "The recorded GeochemistryPi CLI process tree was cancelled.",
            )
        except Exception as exc:
            self._finish(
                paths,
                "failed",
                "The GeochemistryPi CLI run failed; inspect the bounded error and wrapper logs.",
                _safe_error(exc),
                result_type=_terminal_failure_result_type(exc),
                retryable=False,
            )
        finally:
            with self._lock:
                self._active.pop(run_id, None)

    def _wait_for_run(self, run_id: str, wait_seconds: float) -> None:
        if wait_seconds < 0 or wait_seconds > 300:
            raise RunStateError("wait_seconds must be between 0 and 300.")
        if wait_seconds == 0:
            return
        with self._lock:
            control = self._active.get(run_id)
            future = control.future if control is not None else None
        if future is None:
            return
        try:
            future.result(timeout=wait_seconds)
        except FutureTimeoutError:
            return

    def get_status(self, run_id: str, *, wait_seconds: float = 0) -> RunStatusResponse:
        self._ensure_initialized()
        paths = self._paths(run_id)
        self._wait_for_run(run_id, wait_seconds)
        return self._status_response(_read_json(paths.status))

    def get_result(
        self,
        run_id: str,
        *,
        wait_seconds: float = 0,
        artifact_offset: int = 0,
        artifact_limit: int | None = None,
        artifact_view: str = "canonical",
    ) -> RunResultResponse | PendingRunResultResponse | TerminalRunReceipt:
        self._ensure_initialized()
        paths = self._paths(run_id)
        self._wait_for_run(run_id, wait_seconds)
        status = _read_json(paths.status)
        state = status.get("state")
        if state in {"failed", "cancelled"}:
            return self._terminal_result(paths, status)
        if state in {"queued", "running"}:
            return PendingRunResultResponse.from_status(
                self._status_response(status),
                wait_seconds=wait_seconds,
            )
        if state not in {"succeeded", "partial_failure"}:
            raise RunStateError(f"Run {run_id} has unsupported durable state {state!r}.")
        if artifact_offset < 0:
            raise RunStateError("artifact_offset must be non-negative.")
        if artifact_limit is not None and (artifact_limit < 1 or artifact_limit > 200):
            raise RunStateError("artifact_limit must be between 1 and 200.")
        if artifact_view not in {"canonical", "all"}:
            raise RunStateError("artifact_view must be 'canonical' or 'all'.")
        if paths.result.is_symlink() or not paths.result.is_file():
            raise RunStateError("The successful result record is unavailable or is not a regular file.")
        result_record_sha256 = sha256_file(paths.result)
        expected_result_path = status.get("result_record_path")
        expected_result_sha256 = status.get("result_record_sha256")
        if expected_result_path is not None or expected_result_sha256 is not None:
            try:
                result_path_matches = isinstance(expected_result_path, str) and Path(expected_result_path).resolve() == paths.result.resolve()
            except OSError:
                result_path_matches = False
            if not result_path_matches or expected_result_sha256 != result_record_sha256:
                raise RunStateError("The successful result record identity is inconsistent with durable run status.")
        try:
            request_record = _read_json(paths.request)
            request_record_schema_version = request_record.get("schema_version")
            if request_record_schema_version not in {1, 2}:
                raise RunStateError("The stored run request has an unsupported schema version.")
            result_value = _normalize_stored_scientific_identity(
                _read_json(paths.result),
                legacy_schema=request_record_schema_version == 1,
                include_execution_bound=True,
                record_label="The durable run result",
            )
            response = RunResultResponse.model_validate(result_value)
        except ValidationError as exc:
            raise RunStateError("The durable run result does not match the public result contract.") from exc
        artifact_index = _read_json(paths.artifact_index)
        if artifact_index.get("schema_version") != ARTIFACT_INDEX_SCHEMA_VERSION or artifact_index.get("run_id") != run_id or not isinstance(artifact_index.get("artifacts"), list):
            raise RunStateError("The complete artifact index identity is invalid.")
        entries = artifact_index["artifacts"]
        if len(entries) != response.artifact_count:
            raise RunStateError("The complete artifact index count is inconsistent with the run result.")
        try:
            validated_entries = tuple(ArtifactReference.model_validate(entry) for entry in entries)
        except ValidationError as exc:
            raise RunStateError("The durable artifact index does not match the public artifact contract.") from exc
        views = partition_artifact_views(validated_entries)
        selected_entries = views.canonical_entries if artifact_view == "canonical" else views.all_entries
        if artifact_offset > len(selected_entries):
            raise RunStateError(f"artifact_offset is beyond the {artifact_view} artifact view.")
        artifact_index_sha256 = sha256_file(paths.artifact_index)
        observation_summary = response.required_tabular_observations
        if observation_summary.artifact_index_sha256 is not None and observation_summary.artifact_index_sha256 != artifact_index_sha256:
            raise RunStateError("The required tabular observations are not bound to the current artifact index.")
        if response.provenance_manifest_path is not None and response.provenance_manifest_sha256 is not None:
            if Path(response.provenance_manifest_path).resolve() != paths.provenance_manifest.resolve():
                raise RunStateError("The provenance manifest path is inconsistent with the managed run.")
            if sha256_file(paths.provenance_manifest) != response.provenance_manifest_sha256:
                raise RunStateError("The provenance manifest integrity check failed.")
            provenance_manifest = _read_json(paths.provenance_manifest)
            if provenance_manifest.get("artifact_index", {}).get("sha256") != artifact_index_sha256:
                raise RunStateError("The artifact index integrity check failed against the provenance manifest.")
            if observation_summary.artifact_index_sha256 is not None:
                expected_observation_identity = {
                    "artifact_index_sha256": observation_summary.artifact_index_sha256,
                    "total_count": observation_summary.total_count,
                    "returned_count": observation_summary.returned_count,
                    "truncated": observation_summary.truncated,
                    "observations_sha256": observation_summary.observations_sha256,
                    "returned_cell_count": observation_summary.returned_cell_count,
                    "returned_utf8_bytes": observation_summary.returned_utf8_bytes,
                    "omitted_artifact_count": observation_summary.omitted_artifact_count,
                    "omission_reason_counts": observation_summary.omission_reason_counts,
                    "omissions_sha256": observation_summary.omissions_sha256,
                }
                if provenance_manifest.get("required_tabular_observations") != expected_observation_identity:
                    raise RunStateError("The required tabular observation identity failed against the provenance manifest.")
        page_limit = artifact_limit or self.settings.maximum_artifact_references
        page_end = min(len(selected_entries), artifact_offset + page_limit)
        page = selected_entries[artifact_offset:page_end]
        next_offset = page_end if page_end < len(selected_entries) else None
        return response.model_copy(
            update={
                "result_record_path": str(paths.result),
                "result_record_sha256": result_record_sha256,
                "artifact_index_path": str(paths.artifact_index),
                "artifact_index_sha256": artifact_index_sha256,
                "canonical_artifact_count": len(views.canonical_entries),
                "summary_mirror_count": views.summary_mirror_count,
                "artifact_view": artifact_view,
                "artifact_view_count": len(selected_entries),
                "artifact_offset": artifact_offset,
                "returned_artifact_count": len(page),
                "next_artifact_offset": next_offset,
                "artifacts": page,
                "artifacts_truncated": next_offset is not None,
            }
        )

    def cancel(self, run_id: str) -> CancelRunResponse:
        self._ensure_initialized()
        paths = self._paths(run_id)
        with self._lock:
            status = _read_json(paths.status)
            state = status.get("state")
            if state in _TERMINAL_STATES:
                raise RunStateError(f"Run {run_id} is already {state} and cannot be cancelled.")
            control = self._active.get(run_id)
            if control is None:
                raise RunStateError(f"Run {run_id} has no live process handle and will not be terminated by PID alone.")
            control.cancellation.set()
            if state == "queued" and control.future is not None and control.future.cancel():
                self._active.pop(run_id, None)
                self._finish(
                    paths,
                    "cancelled",
                    "The queued run was cancelled before the CLI process started.",
                )
                return CancelRunResponse(
                    run_id=run_id,
                    state="cancelled",
                    message="No CLI process was started.",
                )
            status["progress_message"] = "Cancellation requested for this run's recorded CLI process tree."
            self._write_status(paths, status)
            return CancelRunResponse(
                run_id=run_id,
                state="cancellation_requested",
                message="Cancellation is in progress; poll get_run_status until the state becomes cancelled.",
            )

    def close(self) -> None:
        """Cancel live owned runs and wait for their drivers to release subprocesses."""
        with self._lock:
            if self._closed:
                return
            self._closed = True
            active = list(self._active.items())
            for _, control in active:
                control.cancellation.set()
        self._executor.shutdown(wait=True, cancel_futures=True)
        with self._lock:
            for run_id, _ in active:
                paths = RunPaths.create(self.settings.runs_root, run_id)
                if not paths.status.is_file():
                    continue
                status = _read_json(paths.status)
                if status.get("state") not in _TERMINAL_STATES:
                    self._finish(
                        paths,
                        "cancelled",
                        "The run was cancelled because the MCP server stopped.",
                    )
            self._active.clear()
