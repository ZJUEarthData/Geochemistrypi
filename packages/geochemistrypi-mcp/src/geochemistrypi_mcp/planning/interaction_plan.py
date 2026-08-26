"""Compile semantic learning requests into versioned CLI interactions."""

import ast
import bisect
import csv
import hashlib
import json
import math
import os
import re
import shutil
import sysconfig
from collections import Counter
from dataclasses import dataclass
from dataclasses import replace as dataclass_replace
from pathlib import Path
from typing import Any, Iterable, List, Literal, Optional, Sequence, Tuple

from pydantic import TypeAdapter

from ..api.schemas import (
    AnomalyDetectionModelSettings,
    AnomalyDetectionRequest,
    ClassificationModelSettings,
    ClassificationRequest,
    ClusteringModelSettings,
    ClusteringRequest,
    DecompositionModelSettings,
    DecompositionRequest,
    RegressionModelSettings,
    RegressionRequest,
    TimeSeriesRequest,
)
from ..config.constants import INTERACTION_PLAN_VERSION
from ..contracts.anomaly_detection import MODEL_DISPLAY_NAMES as ANOMALY_DETECTION_MODEL_DISPLAY_NAMES
from ..contracts.anomaly_detection import MODEL_NUMBERS as ANOMALY_DETECTION_MODEL_NUMBERS
from ..contracts.anomaly_detection import MODEL_ORDER as ANOMALY_DETECTION_MODEL_ORDER
from ..contracts.classification import MODEL_DISPLAY_NAMES, MODEL_NUMBERS, MODEL_ORDER, MODELS_SUPPORTING_MISSING_VALUES
from ..contracts.clustering import MODEL_DISPLAY_NAMES as CLUSTERING_MODEL_DISPLAY_NAMES
from ..contracts.clustering import MODEL_NUMBERS as CLUSTERING_MODEL_NUMBERS
from ..contracts.clustering import MODEL_ORDER as CLUSTERING_MODEL_ORDER
from ..contracts.decomposition import MODEL_DISPLAY_NAMES as DECOMPOSITION_MODEL_DISPLAY_NAMES
from ..contracts.decomposition import MODEL_NUMBERS as DECOMPOSITION_MODEL_NUMBERS
from ..contracts.decomposition import MODEL_ORDER as DECOMPOSITION_MODEL_ORDER
from ..contracts.regression import MODEL_DISPLAY_NAMES as REGRESSION_MODEL_DISPLAY_NAMES
from ..contracts.regression import MODEL_NUMBERS as REGRESSION_MODEL_NUMBERS
from ..contracts.regression import MODEL_ORDER as REGRESSION_MODEL_ORDER
from ..contracts.regression import MODELS_SUPPORTING_MISSING_VALUES as REGRESSION_MODELS_SUPPORTING_MISSING_VALUES
from ..contracts.regression import MODELS_WITH_FEATURE_IMPORTANCE as REGRESSION_MODELS_WITH_FEATURE_IMPORTANCE
from ..contracts.regression import MODELS_WITH_INTERACTIVE_PLOT_SELECTION as REGRESSION_MODELS_WITH_INTERACTIVE_PLOT_SELECTION
from ..contracts.regression import MODELS_WITHOUT_AUTOML as REGRESSION_MODELS_WITHOUT_AUTOML
from ..data.headers import HeaderValidationError, normalize_dataset_header, source_allows_pandas_duplicate_mangling
from ..data.source_rows import iter_cli_csv_rows, iter_cli_excel_rows
from .artifact_mapping import AdapterArtifactMapping, build_adapter_artifact_mappings

_CLI_FIXED_RANDOM_SEED = 42


class PlanCompilationError(ValueError):
    """Raised when a semantic request cannot be represented by this driver."""


DatasetSource = Literal["path", "builtin", "desktop"]
DatasetRole = Literal["training", "application"]


def _canonical_sha256(value: Any) -> str:
    serialized = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(serialized.encode("utf-8")).hexdigest()


def _parameter_entries(value: dict[str, Any]) -> Tuple[Tuple[str, str], ...]:
    return tuple((name, json.dumps(parameter, ensure_ascii=False, sort_keys=True, separators=(",", ":"))) for name, parameter in sorted(value.items()))


def _request_model_parameters(request: Any) -> dict[str, Any]:
    if request.task == "time_series":
        if request.mode == "reference_anomaly_series":
            return {
                "mode": request.mode,
                "reference_label_provenance": request.reference_label_provenance,
                "comparison_label_provenance": (request.comparison_label_provenance if request.comparison_label_column is not None else None),
                "association_window_days": request.association_window_days,
                "association_direction": request.association_direction,
            }
        return {
            "mode": request.mode,
            "bin_width": request.bin_width,
            **(
                {
                    "iterations": request.iterations,
                    "seed": request.seed,
                    "age_unit": request.age_unit,
                    "fit_curve": request.fit_curve,
                }
                if request.mode == "subaerial_proportion"
                else {
                    "aggregation": request.aggregation,
                    "uncertainty": request.uncertainty,
                    "minimum_samples_per_bin": request.minimum_samples_per_bin,
                    "filter_minimum": request.filter_minimum,
                    "filter_maximum": request.filter_maximum,
                }
            ),
        }
    if request.task == "decomposition" and request.mode == "embedding_label_overlay":
        return {
            "mode": request.mode,
            "join_policy": "exact_identifier_set_one_to_one",
            "positive_label_values": list(request.positive_label_values),
        }
    all_models = request.model_selection.mode == "all"
    tuning = request.model_selection.tuning if all_models else getattr(request, "tuning", "manual")
    if all_models or tuning == "automl":
        return {}
    return request.model.model_dump(mode="json")


def _resolved_model_parameters(request: Any) -> dict[str, Any]:
    if request.task == "time_series":
        return _request_model_parameters(request)
    all_models = request.model_selection.mode == "all"
    tuning = request.model_selection.tuning if all_models else getattr(request, "tuning", "manual")
    if all_models or tuning == "automl":
        return {}
    return request.model.model_dump(mode="json")


def _native_scientific_model_parameters(request: Any) -> dict[str, Any] | None:
    model_type = request.model.type
    supported = {
        ("classification", "xgboost"),
        ("regression", "xgboost"),
        ("regression", "extra_trees"),
        ("anomaly_detection", "local_outlier_factor"),
    }
    if (request.task, model_type) not in supported:
        return None
    parameters = _resolved_model_parameters(request)
    mappings = {
        "xgboost": {
            "number_of_estimators": "n_estimators",
            "maximum_depth": "max_depth",
            "column_subsample": "colsample_bytree",
            "column_subsample_by_level": "colsample_bylevel",
            "column_subsample_by_node": "colsample_bynode",
            "l1_regularization": "reg_alpha",
            "l2_regularization": "reg_lambda",
            "minimum_child_weight": "min_child_weight",
            "maximum_delta_step": "max_delta_step",
            "number_of_jobs": "n_jobs",
        },
        "extra_trees": {
            "number_of_estimators": "n_estimators",
            "maximum_depth": "max_depth",
            "minimum_samples_split": "min_samples_split",
            "minimum_samples_leaf": "min_samples_leaf",
            "maximum_features": "max_features",
            "maximum_samples": "max_samples",
            "out_of_bag_score": "oob_score",
        },
        "local_outlier_factor": {
            "number_of_neighbors": "n_neighbors",
            "number_of_jobs": "n_jobs",
            "power": "p",
        },
    }[model_type]
    ignored = {"type", "detection_mode"}
    return {mappings.get(name, name): value for name, value in parameters.items() if name not in ignored}


def _effective_scientific_model_parameters(
    request: Any,
    effective_seeds: Tuple[Tuple[str, int], ...],
) -> dict[str, Any]:
    """Expose public controls plus constructor values injected by the adapter."""
    parameters = _resolved_model_parameters(request)
    model_seed = dict(effective_seeds).get("model")
    if model_seed is not None:
        parameters["random_state"] = model_seed
    if request.model.type == "local_outlier_factor":
        parameters["novelty"] = request.model.detection_mode == "novelty_detection"
    return parameters


def _scientific_execution_contract_json(
    request: Any,
    workflow_family: str,
    workflow_mode: str,
) -> tuple[str, Tuple[Tuple[str, int], ...]] | None:
    if request.model_selection.mode == "all" or getattr(request, "tuning", "manual") != "manual":
        return None
    native_parameters = _native_scientific_model_parameters(request)
    if native_parameters is None:
        return None
    model_type = request.model.type
    supervised = request.task in {"classification", "regression"}
    external_labeled_training = request.task == "regression" and request.evaluation.mode == "external_labeled"
    split_seed = (request.reproducibility.split_seed if request.reproducibility.split_seed is not None else _CLI_FIXED_RANDOM_SEED) if supervised and not external_labeled_training else None
    seeded_model = model_type in {"xgboost", "extra_trees"}
    model_seed = (request.reproducibility.model_seed if request.reproducibility.model_seed is not None else _CLI_FIXED_RANDOM_SEED) if seeded_model else None
    folds = request.evaluation.folds if request.evaluation.folds is not None else 10
    evaluation_mode = request.model.detection_mode if model_type == "local_outlier_factor" else "external_labeled" if request.evaluation.mode == "external_labeled" else "internal_holdout"
    payload = {
        "schema_version": 2,
        "workflow_family": workflow_family,
        "workflow_mode": workflow_mode,
        "method": model_type,
        "split_seed": split_seed,
        "model_seed": model_seed,
        "cross_validation_folds": folds,
        "evaluation_mode": evaluation_mode,
        "confusion_matrix_normalization": (None if request.evaluation.confusion_matrix_normalization in {None, "none"} else request.evaluation.confusion_matrix_normalization),
        "external_evaluation_identifier_column": (request.evaluation.external_identifier_column if evaluation_mode == "external_labeled" else None),
        "external_evaluation_target_columns": (list(request.resolved_target_columns) if evaluation_mode == "external_labeled" else []),
        "target_transformations": {
            column: transformation.model_dump(mode="json")
            for column, transformation in getattr(
                request,
                "target_transformations",
                {},
            ).items()
        },
        "model_parameters": native_parameters,
    }
    effective_seeds = tuple((role, value) for role, value in (("split", split_seed), ("model", model_seed)) if value is not None)
    return json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")), effective_seeds


def _request_preprocessing_parameters(request: Any) -> dict[str, Any]:
    parameters = {
        "missing_values": request.missing_values.model_dump(mode="json"),
        "engineered_features": [item.model_dump(mode="json") for item in getattr(request, "engineered_features", ())],
        "selected_columns": list(getattr(request, "resolved_selected_columns", ())),
        "scaling": getattr(request, "scaling", "none"),
    }
    if hasattr(request, "feature_selection"):
        parameters["feature_selection"] = request.feature_selection.model_dump(mode="json")
    if hasattr(request, "label_customization"):
        parameters["label_customization"] = request.label_customization.model_dump(mode="json")
    if hasattr(request, "target_transformations"):
        parameters["target_transformations"] = {column: transformation.model_dump(mode="json") for column, transformation in request.target_transformations.items()}
    if request.task == "time_series":
        parameters["feature_engineering"] = request.feature_engineering
    return parameters


def _environment_profile_binding(request: Any) -> tuple[str, str | None, str | None]:
    profile = getattr(request, "environment_profile", None)
    if profile is not None:
        value = profile.model_dump(mode="json")
        return "request.environment_profile", profile.profile_id, _canonical_sha256(value)
    environment = request.reproducibility.environment.model_dump(mode="json")
    specified = any(value not in (None, {}, (), []) for value in environment.values())
    if not specified:
        return "request.reproducibility.environment", None, None
    profile_id = getattr(request.reproducibility, "dependency_profile", None) or "legacy-inline-environment"
    return "request.reproducibility.environment", profile_id, _canonical_sha256(environment)


def _scientific_binding_fields(
    request: Any,
    workflow_family: str,
    workflow_mode: str,
    method: str,
    paths: Tuple[str, ...],
    effective_seeds: Tuple[Tuple[str, int], ...],
    effective_model_parameters_override: dict[str, Any] | None = None,
) -> dict[str, Any]:
    environment_pointer, environment_profile_id, environment_identity = _environment_profile_binding(request)
    if request.task == "time_series":
        requested_seeds = (("model", request.seed),) if request.mode == "subaerial_proportion" else ()
    else:
        reproducibility = request.reproducibility
        requested_seeds = tuple(
            (role, value)
            for role, value in (
                ("split", reproducibility.split_seed),
                ("model", reproducibility.model_seed),
                ("tuning", reproducibility.tuning_seed),
            )
            if value is not None
        )
    model_parameters = _request_model_parameters(request)
    effective_model_parameters = dict(effective_model_parameters_override) if effective_model_parameters_override is not None else dict(model_parameters)
    preprocessing_parameters = _request_preprocessing_parameters(request)
    return {
        "environment_profile": environment_pointer,
        "environment_profile_id": environment_profile_id,
        "environment_profile_identity_sha256": environment_identity,
        "artifact_contract": paths,
        "artifact_mappings": build_adapter_artifact_mappings(
            workflow_family,
            workflow_mode,
            paths,
            method,
        ),
        "effective_seeds": effective_seeds,
        "requested_seeds": requested_seeds,
        "requested_model_parameters": _parameter_entries(model_parameters),
        "effective_model_parameters": _parameter_entries(effective_model_parameters),
        "requested_preprocessing_parameters": _parameter_entries(preprocessing_parameters),
        "effective_preprocessing_parameters": _parameter_entries(preprocessing_parameters),
        "model_parameter_binding": "interaction_plan" if model_parameters else "runtime_selected",
        "preprocessing_parameter_binding": "interaction_plan",
    }


@dataclass(frozen=True)
class DatasetCompilationContext:
    """Preserve resolved dataset provenance after references become local paths."""

    training_source: DatasetSource = "path"
    application_source: DatasetSource | None = None

    def source_for(self, role: DatasetRole) -> DatasetSource:
        if role == "training":
            return self.training_source
        return self.application_source or "path"

    def allows_pandas_duplicate_mangling(self, role: DatasetRole) -> bool:
        return source_allows_pandas_duplicate_mangling(self.source_for(role))


@dataclass(frozen=True)
class InteractionStep:
    """One response guarded by ordered output anchors ending at a CLI prompt."""

    id: str
    output_anchors: Tuple[str, ...]
    response: str
    timeout_seconds: Optional[float] = None

    def __post_init__(self) -> None:
        if not self.id:
            raise ValueError("Interaction step id must not be empty.")
        if not self.output_anchors or any(not anchor for anchor in self.output_anchors):
            raise ValueError(f"Interaction step {self.id!r} must define non-empty output anchors.")
        if "\n" in self.response or "\r" in self.response:
            raise ValueError(f"Interaction step {self.id!r} response must be a single line.")
        if self.timeout_seconds is not None and self.timeout_seconds <= 0:
            raise ValueError(f"Interaction step {self.id!r} timeout_seconds must be positive.")


@dataclass(frozen=True)
class InteractionPlan:
    """Complete public command and ordered prompt/response contract."""

    schema_version: int
    name: str
    public_command: Tuple[str, ...]
    steps: Tuple[InteractionStep, ...]
    expected_output_relative_paths: Tuple[str, ...] = ()
    requires_source_row_pairing: bool = False
    workflow_family: str = "legacy"
    workflow_mode: str = "legacy"
    method: str = "legacy"
    scientific_contract_id: str = "scientific-contract-v1/legacy"
    adapter_id: str | None = "geochemistrypi-cli.public"
    adapter_version: str | None = "1"
    environment_profile: str = "request.reproducibility.environment"
    environment_profile_id: str | None = None
    environment_profile_identity_sha256: str | None = None
    artifact_contract: Tuple[str, ...] = ()
    artifact_mappings: Tuple[AdapterArtifactMapping, ...] = ()
    adapter_status: Literal["available", "unavailable", "requirements_unmet"] = "available"
    execution_ready: bool = True
    blocking_issues: Tuple[str, ...] = ()
    effective_seeds: Tuple[Tuple[str, int], ...] = ()
    requested_seeds: Tuple[Tuple[str, int], ...] = ()
    seed_binding: Literal["cli_fixed", "request_parameter", "mixed", "not_applicable"] = "not_applicable"
    requested_model_parameters: Tuple[Tuple[str, str], ...] = ()
    effective_model_parameters: Tuple[Tuple[str, str], ...] = ()
    requested_preprocessing_parameters: Tuple[Tuple[str, str], ...] = ()
    effective_preprocessing_parameters: Tuple[Tuple[str, str], ...] = ()
    model_parameter_binding: Literal["interaction_plan", "runtime_selected", "not_applicable"] = "not_applicable"
    preprocessing_parameter_binding: Literal["interaction_plan", "not_applicable"] = "not_applicable"
    scientific_execution_contract_json: str | None = None
    required_cli_capabilities: Tuple[str, ...] = ()

    def __post_init__(self) -> None:
        if self.schema_version < 1:
            raise ValueError("Interaction plan schema_version must be positive.")
        if not self.name:
            raise ValueError("Interaction plan name is required.")
        if self.execution_ready and (not self.public_command or self.adapter_status != "available" or self.adapter_id is None):
            raise ValueError("An execution-ready interaction plan requires an available adapter and command.")
        if not self.execution_ready and not self.blocking_issues:
            raise ValueError("A blocked interaction plan must explain why it cannot execute.")
        step_ids = [step.id for step in self.steps]
        if len(step_ids) != len(set(step_ids)):
            raise ValueError("Interaction plan step ids must be unique.")
        mapping_ids = [mapping.mapping_id for mapping in self.artifact_mappings]
        if len(mapping_ids) != len(set(mapping_ids)):
            raise ValueError("Interaction plan artifact mapping ids must be unique.")
        if any(not requirement for requirement in self.required_cli_capabilities):
            raise ValueError("Required CLI capabilities must not contain blank values.")
        if len(self.required_cli_capabilities) != len(set(self.required_cli_capabilities)):
            raise ValueError("Required CLI capabilities must be unique.")
        if self.scientific_execution_contract_json is not None:
            try:
                scientific_execution = json.loads(self.scientific_execution_contract_json)
            except json.JSONDecodeError as exc:
                raise ValueError("Scientific execution contract must be valid canonical JSON.") from exc
            if not isinstance(scientific_execution, dict):
                raise ValueError("Scientific execution contract must be a JSON object.")


@dataclass(frozen=True)
class _DatasetProfile:
    row_count: int
    class_counts: Counter[str]
    missing_columns: frozenset[str]
    unresolved_missing_columns: frozenset[str]


@dataclass(frozen=True)
class _RegressionDatasetProfile:
    row_count: int
    missing_columns: frozenset[str]
    unresolved_missing_columns: frozenset[str]


@dataclass(frozen=True)
class _ClusteringDatasetProfile:
    row_count: int
    missing_columns: frozenset[str]
    unresolved_missing_columns: frozenset[str]


@dataclass(frozen=True)
class _DecompositionDatasetProfile:
    row_count: int
    missing_columns: frozenset[str]
    unresolved_missing_columns: frozenset[str]


@dataclass(frozen=True)
class _AnomalyDetectionDatasetProfile:
    row_count: int
    missing_columns: frozenset[str]
    unresolved_missing_columns: frozenset[str]


def _console_script_name(os_name: str = os.name) -> str:
    return "geochemistrypi.exe" if os_name == "nt" else "geochemistrypi"


def resolve_public_cli_executable() -> Path:
    """Resolve the public console script from the current Python environment."""
    script_name = _console_script_name()
    current_environment_command = Path(sysconfig.get_path("scripts")) / script_name
    if current_environment_command.is_file():
        return current_environment_command.resolve()
    discovered = shutil.which("geochemistrypi")
    if discovered:
        return Path(discovered).resolve()
    raise PlanCompilationError("The public 'geochemistrypi' command is unavailable. Install the GeochemistryPi package in the current Python environment.")


def _read_dataset_columns(
    path: Path,
    *,
    source: DatasetSource = "path",
    sheet: str = "0",
) -> Tuple[str, ...]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        try:
            with path.open(encoding="utf-8-sig", newline="") as stream:
                row = next(csv.reader(stream))
        except (OSError, StopIteration, UnicodeError, csv.Error) as exc:
            raise PlanCompilationError(f"Unable to read the CSV header from {path}: {exc}") from exc
    elif suffix == ".xlsx":
        try:
            from openpyxl import load_workbook

            with path.open("rb") as stream:
                workbook = load_workbook(stream, read_only=True, data_only=True)
                try:
                    if str(sheet).isdigit():
                        index = int(sheet)
                        if index >= len(workbook.worksheets):
                            raise ValueError(f"Excel sheet index {index} is out of range.")
                        worksheet = workbook.worksheets[index]
                    else:
                        if sheet not in workbook.sheetnames:
                            raise ValueError(f"Excel sheet {sheet!r} does not exist.")
                        worksheet = workbook[sheet]
                    row = [value.lstrip("\ufeff") if isinstance(value, str) else value for value in next(worksheet.iter_rows(max_row=1, values_only=True))]
                finally:
                    workbook.close()
        except (OSError, StopIteration, ValueError) as exc:
            raise PlanCompilationError(f"Unable to read the Excel header from {path}: {exc}") from exc
    else:
        raise PlanCompilationError(f"PR3 supports .csv and .xlsx data; received {path.suffix or 'a file without an extension'}.")
    try:
        return normalize_dataset_header(
            row,
            256,
            allow_pandas_duplicate_mangling=source_allows_pandas_duplicate_mangling(source),
        )
    except HeaderValidationError as exc:
        raise PlanCompilationError(str(exc)) from exc


def _validate_world_map(path: Path, columns: Sequence[str], configuration: Any) -> None:
    """Validate semantic map roles and every rendered value before CLI startup."""
    if not configuration.enabled:
        return
    requested = (
        configuration.longitude_column,
        configuration.latitude_column,
        *configuration.value_columns,
    )
    missing = sorted({column for column in requested if column not in columns})
    if missing:
        raise PlanCompilationError(f"World-map columns are absent from the training dataset: {missing}")
    positions = [columns.index(column) for column in requested]
    row_count = 0
    for row_count, raw_values in enumerate(_iter_selected_rows(path, positions), start=1):
        parsed = []
        for column, raw in zip(requested, raw_values):
            if raw is None or (isinstance(raw, str) and not raw.strip()):
                raise PlanCompilationError(f"World-map column {column!r} contains a missing value at data row {row_count + 1}.")
            try:
                number = float(raw)
            except (TypeError, ValueError) as exc:
                raise PlanCompilationError(f"World-map column {column!r} contains a non-numeric value at data row {row_count + 1}.") from exc
            if not math.isfinite(number):
                raise PlanCompilationError(f"World-map column {column!r} contains a non-finite value at data row {row_count + 1}.")
            parsed.append(number)
        longitude, latitude = parsed[:2]
        if longitude < -180 or longitude > 180:
            raise PlanCompilationError(f"World-map longitude must be between -180 and 180 degrees; data row {row_count + 1} contains {longitude}.")
        if latitude < -90 or latitude > 90:
            raise PlanCompilationError(f"World-map latitude must be between -90 and 90 degrees; data row {row_count + 1} contains {latitude}.")
    if row_count == 0:
        raise PlanCompilationError("World-map rendering requires at least one data row.")


def _world_map_option(configuration: Any) -> str:
    payload = {
        "schema_version": 1,
        "enabled": bool(configuration.enabled),
        "longitude_column": getattr(configuration, "longitude_column", None),
        "latitude_column": getattr(configuration, "latitude_column", None),
        "value_columns": list(getattr(configuration, "value_columns", ())),
    }
    return json.dumps(payload, ensure_ascii=False, separators=(",", ":"))


def _command_with_world_map(command: Tuple[str, ...], configuration: Any) -> Tuple[str, ...]:
    return (*command, "--world-map-config", _world_map_option(configuration))


def _command_with_analysis_options(command: Tuple[str, ...], configuration: Any, existing_experiment_id: Optional[str]) -> Tuple[str, ...]:
    resolved = _command_with_world_map(command, configuration)
    if existing_experiment_id:
        resolved = (*resolved, "--existing-experiment-id", existing_experiment_id)
    return resolved


def _experiment_steps(request: Any, enter_prompt: str) -> List[InteractionStep]:
    steps: List[InteractionStep] = []
    if not request.existing_experiment_id:
        steps.extend(
            (
                InteractionStep("use_previous_experiment", ("Use Previous Experiment", "[y/n]"), ""),
                InteractionStep("experiment_name", ("New Experiment",), request.experiment_name),
            )
        )
    steps.extend(
        (
            InteractionStep("run_name", ("Run Name",), request.run_name),
            InteractionStep("continue_after_run_setup", (enter_prompt,), ""),
        )
    )
    return steps


def _iter_selected_rows(
    path: Path,
    positions: Sequence[int],
    *,
    sheet: str = "0",
) -> Iterable[tuple[Any, ...]]:
    if path.suffix.lower() == ".csv":
        try:
            with path.open(encoding="utf-8-sig", newline="") as stream:
                reader = csv.reader(stream)
                next(reader)
                for row_number, row in enumerate(iter_cli_csv_rows(reader), start=2):
                    if len(row) <= max(positions):
                        raise PlanCompilationError(f"CSV row {row_number} has fewer values than the header.")
                    yield tuple(row[position] for position in positions)
        except (OSError, StopIteration, UnicodeError, csv.Error) as exc:
            if isinstance(exc, PlanCompilationError):
                raise
            raise PlanCompilationError(f"Unable to read CSV rows from {path}: {exc}") from exc
        return

    try:
        from openpyxl import load_workbook

        with path.open("rb") as stream:
            workbook = load_workbook(stream, read_only=True, data_only=True)
            try:
                if str(sheet).isdigit():
                    index = int(sheet)
                    if index >= len(workbook.worksheets):
                        raise ValueError(f"Excel sheet index {index} is out of range.")
                    worksheet = workbook.worksheets[index]
                else:
                    if sheet not in workbook.sheetnames:
                        raise ValueError(f"Excel sheet {sheet!r} does not exist.")
                    worksheet = workbook[sheet]
                rows = worksheet.iter_rows(values_only=True)
                next(rows)
                for row_number, row in enumerate(iter_cli_excel_rows(rows), start=2):
                    if len(row) <= max(positions):
                        raise PlanCompilationError(f"Excel row {row_number} has fewer values than the header.")
                    yield tuple(row[position] for position in positions)
            finally:
                workbook.close()
    except (OSError, StopIteration, ValueError) as exc:
        if isinstance(exc, PlanCompilationError):
            raise
        raise PlanCompilationError(f"Unable to read Excel rows from {path}: {exc}") from exc


def _selection_expression(one_based_indices: Sequence[int]) -> str:
    indices = sorted(set(one_based_indices))
    if not indices:
        raise PlanCompilationError("A CLI column selection cannot be empty.")
    groups: List[Tuple[int, int]] = []
    start = previous = indices[0]
    for index in indices[1:]:
        if index == previous + 1:
            previous = index
            continue
        groups.append((start, previous))
        start = previous = index
    groups.append((start, previous))
    return "; ".join(str(start) if start == end else f"[{start},{end}]" for start, end in groups)


def _float_response(value: float, default: Optional[float] = None) -> str:
    if default is not None and value == default:
        return ""
    return format(value, ".15g")


def _choice(value: Any, choices: Sequence[Any]) -> str:
    try:
        return str(choices.index(value) + 1)
    except ValueError as exc:
        raise PlanCompilationError(f"The CLI does not offer choice {value!r}; expected one of {tuple(choices)!r}.") from exc


def _is_missing(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip().lower() in {"", "na", "n/a", "nan", "null", "none"}
    return isinstance(value, float) and math.isnan(value)


def _numeric(value: Any, column: str, row_number: int) -> float:
    if isinstance(value, bool):
        return float(value)
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise PlanCompilationError(f"Feature column {column!r} contains a non-numeric value at data row {row_number}: {value!r}.") from exc
    if not math.isfinite(number):
        raise PlanCompilationError(f"Feature column {column!r} contains a non-finite value at data row {row_number}.")
    return number


def _quantile_counts(values: list[float], number_of_classes: int, labels: tuple[str, ...] | None) -> Counter[str]:
    ordered = sorted(values)
    if len(set(ordered)) < number_of_classes:
        raise PlanCompilationError("Quantile label customization requires at least as many unique target values as final classes.")
    cuts = []
    for index in range(1, number_of_classes):
        position = (len(ordered) - 1) * index / number_of_classes
        lower = math.floor(position)
        upper = math.ceil(position)
        fraction = position - lower
        cuts.append(ordered[lower] + (ordered[upper] - ordered[lower]) * fraction)
    if any(left >= right for left, right in zip(cuts, cuts[1:])):
        raise PlanCompilationError("Quantile label customization produces duplicate bin edges for this target distribution.")
    final_labels = labels or tuple(f"Class_{index}" for index in range(number_of_classes))
    return Counter(final_labels[bisect.bisect_left(cuts, value)] for value in values)


def _scan_training_dataset(path: Path, columns: tuple[str, ...], request: ClassificationRequest) -> _DatasetProfile:
    selected_names = tuple(column for column in columns if column == request.target_column or column in request.feature_columns)
    scan_names = selected_names
    positions = [columns.index(column) for column in scan_names]
    drop_columns = set(getattr(request.missing_values, "columns", ()))
    unknown_drop_columns = sorted(drop_columns - set(selected_names))
    if unknown_drop_columns:
        raise PlanCompilationError(f"Missing-value drop columns are not selected training columns: {unknown_drop_columns}")

    missing_columns: set[str] = set()
    unresolved_missing: set[str] = set()
    labels: list[str] = []
    numeric_targets: list[float] = []
    kept_rows = 0
    label_strategy = request.label_customization.strategy
    mapping = getattr(request.label_customization, "mapping", {})
    cut_points = getattr(request.label_customization, "cut_points", ())
    interval_labels = getattr(request.label_customization, "labels", None)

    for row_number, row in enumerate(_iter_selected_rows(path, positions), start=1):
        values = dict(zip(scan_names, row))
        row_missing = {column for column in selected_names if _is_missing(values[column])}
        missing_columns.update(row_missing)
        should_drop = request.missing_values.method == "drop_rows" and bool(row_missing & (drop_columns or set(selected_names)))
        if should_drop:
            continue
        target = values[request.target_column]
        if _is_missing(target):
            raise PlanCompilationError(f"Target column {request.target_column!r} contains a missing label at data row {row_number}.")
        unresolved_missing.update(row_missing)
        for feature in request.feature_columns:
            if not _is_missing(values[feature]):
                _numeric(values[feature], feature, row_number)
        kept_rows += 1
        if label_strategy == "encode_original":
            labels.append(str(target))
        elif label_strategy == "map":
            key = str(target)
            if key not in mapping:
                raise PlanCompilationError(f"Label mapping does not cover observed target value {key!r}.")
            labels.append(mapping[key])
        else:
            try:
                target_number = float(target)
            except (TypeError, ValueError) as exc:
                raise PlanCompilationError(f"{label_strategy} label customization requires a numeric target column.") from exc
            if not math.isfinite(target_number):
                raise PlanCompilationError(f"{label_strategy} label customization requires finite target values.")
            numeric_targets.append(target_number)
            if label_strategy == "interval":
                final_labels = interval_labels or tuple(f"Class_{index}" for index in range(len(cut_points) + 1))
                labels.append(final_labels[bisect.bisect_left(cut_points, target_number)])

    if kept_rows == 0:
        raise PlanCompilationError("Missing-value handling removes every training row.")
    method = request.missing_values.method
    if not missing_columns and method != "error":
        raise PlanCompilationError(f"missing_values.method={method!r} would be silently skipped because the selected data has no missing values; use 'error'.")
    if missing_columns and method == "error":
        raise PlanCompilationError(f"Selected feature data contains missing values in {sorted(missing_columns)}; choose keep, drop_rows, or impute explicitly.")
    if method == "impute":
        unresolved_missing.clear()
        if request.missing_values.strategy in {"mean", "median"}:
            # The CLI applies its imputer to the complete selected table, including Y.
            for target in numeric_targets or []:
                if not math.isfinite(target):
                    raise PlanCompilationError("Mean/median imputation requires a numeric target column in the current CLI workflow.")
            if label_strategy in {"encode_original", "map"}:
                try:
                    for row in _iter_selected_rows(path, [columns.index(request.target_column)]):
                        float(row[0])
                except (TypeError, ValueError) as exc:
                    raise PlanCompilationError("Mean/median imputation requires a numeric target column because the CLI imputes the selected table before splitting X and Y.") from exc

    if label_strategy == "quantile":
        labels = list(
            _quantile_counts(
                numeric_targets,
                request.label_customization.number_of_classes,
                request.label_customization.labels,
            ).elements()
        )
    class_counts = Counter(labels)
    if len(class_counts) < 2:
        raise PlanCompilationError("Classification requires at least two final classes after label customization.")
    too_small = {label: count for label, count in class_counts.items() if count < 2}
    if too_small:
        raise PlanCompilationError(f"Each final class must have at least 2 samples for stratified splitting; too-small classes: {too_small}")
    test_size = math.ceil(kept_rows * request.test_ratio)
    train_size = kept_rows - test_size
    if min(test_size, train_size) < len(class_counts):
        raise PlanCompilationError(
            f"test_ratio={request.test_ratio} produces {train_size} training and {test_size} test rows, but each split needs at least {len(class_counts)} rows for stratification."
        )
    if unresolved_missing and request.model.type not in MODELS_SUPPORTING_MISSING_VALUES:
        raise PlanCompilationError(f"Unprocessed missing values remain in {sorted(unresolved_missing)}; the public CLI only offers XGBoost for this classification branch.")
    if request.model.type == "k_nearest_neighbors" and request.model.number_of_neighbors > train_size:
        raise PlanCompilationError(f"number_of_neighbors={request.model.number_of_neighbors} exceeds the {train_size} training rows produced by the requested split.")
    return _DatasetProfile(
        kept_rows,
        class_counts,
        frozenset(missing_columns),
        frozenset(unresolved_missing),
    )


def _scan_regression_training_dataset(
    path: Path,
    columns: tuple[str, ...],
    request: RegressionRequest,
) -> _RegressionDatasetProfile:
    target_columns = request.resolved_target_columns
    target_set = set(target_columns)
    selected_names = tuple(column for column in columns if column in target_set or column in request.feature_columns)
    scan_names = selected_names
    positions = [columns.index(column) for column in scan_names]
    drop_columns = set(getattr(request.missing_values, "columns", ()))
    unknown_drop_columns = sorted(drop_columns - set(selected_names))
    if unknown_drop_columns:
        raise PlanCompilationError(f"Missing-value drop columns are not selected training columns: {unknown_drop_columns}")

    missing_columns: set[str] = set()
    unresolved_missing: set[str] = set()
    kept_rows = 0
    has_negative_target = False
    for row_number, row in enumerate(_iter_selected_rows(path, positions), start=1):
        values = dict(zip(scan_names, row))
        row_missing = {column for column in selected_names if _is_missing(values[column])}
        missing_columns.update(row_missing)
        should_drop = request.missing_values.method == "drop_rows" and bool(row_missing & (drop_columns or set(selected_names)))
        if should_drop:
            continue
        for target_column in target_columns:
            target = values[target_column]
            if _is_missing(target):
                raise PlanCompilationError(f"Regression target column {target_column!r} is missing at data row {row_number}.")
            try:
                target_number = float(target)
            except (TypeError, ValueError) as exc:
                raise PlanCompilationError(f"Regression target column {target_column!r} contains a non-numeric value at data row {row_number}: {target!r}.") from exc
            if not math.isfinite(target_number):
                raise PlanCompilationError(f"Regression target column {target_column!r} contains a non-finite value at data row {row_number}.")
            has_negative_target = has_negative_target or target_number < 0

        unresolved_missing.update(row_missing)
        for feature in request.feature_columns:
            if feature not in row_missing:
                _numeric(values[feature], feature, row_number)
        kept_rows += 1

    if kept_rows == 0:
        raise PlanCompilationError("Missing-value handling removes every training row.")
    method = request.missing_values.method
    if not missing_columns and method != "error":
        raise PlanCompilationError(f"missing_values.method={method!r} would be silently skipped because the selected data has no missing values; use 'error'.")
    if missing_columns and method == "error":
        raise PlanCompilationError(f"Selected feature data contains missing values in {sorted(missing_columns)}; choose keep, drop_rows, or impute explicitly.")
    if method == "impute":
        unresolved_missing.clear()

    if request.evaluation.mode == "external_labeled":
        train_size = kept_rows
        validation_folds = 10
    else:
        test_size = math.ceil(kept_rows * request.test_ratio)
        train_size = kept_rows - test_size
        if min(test_size, train_size) < 1:
            raise PlanCompilationError(f"test_ratio={request.test_ratio} produces {train_size} training and {test_size} test rows; regression requires both splits to be non-empty.")
        validation_folds = request.evaluation.folds if request.evaluation.mode == "cross_validation" else 10
    if train_size < validation_folds:
        scope = "the complete external-evaluation training cohort" if request.evaluation.mode == "external_labeled" else f"the split produced by test_ratio={request.test_ratio}"
        validation_description = (
            "the existing regression workflow performs fixed 10-fold cross-validation" if validation_folds == 10 else f"the requested workflow performs {validation_folds}-fold cross-validation"
        )
        raise PlanCompilationError(f"{scope} contains only {train_size} training rows; " f"{validation_description} and requires at least {validation_folds}.")
    if unresolved_missing and request.model.type not in REGRESSION_MODELS_SUPPORTING_MISSING_VALUES:
        raise PlanCompilationError(f"Unprocessed missing values remain in {sorted(unresolved_missing)}; the public CLI only offers XGBoost for this regression branch.")
    if request.model.type == "k_nearest_neighbors" and request.model.number_of_neighbors > train_size:
        raise PlanCompilationError(f"number_of_neighbors={request.model.number_of_neighbors} exceeds the {train_size} training rows produced by the requested split.")
    if request.model.type == "decision_tree" and request.model.criterion == "poisson" and has_negative_target:
        raise PlanCompilationError("Decision Tree criterion='poisson' requires non-negative regression target values.")
    return _RegressionDatasetProfile(kept_rows, frozenset(missing_columns), frozenset(unresolved_missing))


def _scan_clustering_training_dataset(
    path: Path,
    columns: tuple[str, ...],
    request: ClusteringRequest,
) -> _ClusteringDatasetProfile:
    scan_names = request.feature_columns
    positions = [columns.index(column) for column in scan_names]
    drop_columns = set(getattr(request.missing_values, "columns", ()))
    unknown_drop_columns = sorted(drop_columns - set(request.feature_columns))
    if unknown_drop_columns:
        raise PlanCompilationError(f"Missing-value drop columns are not selected clustering features: {unknown_drop_columns}")

    missing_columns: set[str] = set()
    unresolved_missing: set[str] = set()
    kept_rows = 0
    for row_number, row in enumerate(_iter_selected_rows(path, positions), start=1):
        values = dict(zip(scan_names, row))
        row_missing = {feature for feature in request.feature_columns if _is_missing(values[feature])}
        missing_columns.update(row_missing)
        should_drop = request.missing_values.method == "drop_rows" and bool(row_missing & (drop_columns or set(request.feature_columns)))
        if should_drop:
            continue
        unresolved_missing.update(row_missing)
        for feature in request.feature_columns:
            if feature not in row_missing:
                _numeric(values[feature], feature, row_number)
        kept_rows += 1

    if kept_rows == 0:
        raise PlanCompilationError("Missing-value handling removes every clustering row.")
    method = request.missing_values.method
    if not missing_columns and method != "error":
        raise PlanCompilationError(f"missing_values.method={method!r} would be silently skipped because the selected data has no missing values; use 'error'.")
    if missing_columns and method == "error":
        raise PlanCompilationError(f"Selected clustering features contain missing values in {sorted(missing_columns)}; choose drop_rows or impute explicitly.")
    if method == "impute":
        unresolved_missing.clear()
    if unresolved_missing:
        raise PlanCompilationError(f"Unprocessed missing values remain in {sorted(unresolved_missing)}; the public CLI exposes no clustering models for this branch.")
    if kept_rows < 3:
        raise PlanCompilationError(f"Clustering requires at least 3 retained rows; missing-value handling leaves {kept_rows}.")
    if request.model.type in {"kmeans", "agglomerative"} and kept_rows < 11:
        raise PlanCompilationError(f"{CLUSTERING_MODEL_DISPLAY_NAMES[request.model.type]} requires at least 11 retained rows because the public CLI evaluates silhouette scores for k=2 through k=10.")
    cluster_count = getattr(request.model, "number_of_clusters", None)
    if cluster_count is not None and cluster_count >= kept_rows:
        raise PlanCompilationError(f"number_of_clusters={cluster_count} must be less than the {kept_rows} retained rows so silhouette scoring remains defined.")
    if request.model.type == "dbscan" and request.model.minimum_samples > kept_rows:
        raise PlanCompilationError(f"minimum_samples={request.model.minimum_samples} exceeds the {kept_rows} retained rows.")
    if request.model.type == "mean_shift" and request.model.bin_seeding and request.model.minimum_bin_frequency > kept_rows:
        raise PlanCompilationError(f"minimum_bin_frequency={request.model.minimum_bin_frequency} exceeds the {kept_rows} retained rows.")
    return _ClusteringDatasetProfile(
        kept_rows,
        frozenset(missing_columns),
        frozenset(unresolved_missing),
    )


def _scan_decomposition_training_dataset(
    path: Path,
    columns: tuple[str, ...],
    request: DecompositionRequest,
) -> _DecompositionDatasetProfile:
    scan_names = request.feature_columns
    positions = [columns.index(column) for column in scan_names]
    drop_columns = set(getattr(request.missing_values, "columns", ()))
    unknown_drop_columns = sorted(drop_columns - set(request.feature_columns))
    if unknown_drop_columns:
        raise PlanCompilationError(f"Missing-value drop columns are not selected decomposition features: {unknown_drop_columns}")

    missing_columns: set[str] = set()
    unresolved_missing: set[str] = set()
    kept_rows = 0
    for row_number, row in enumerate(_iter_selected_rows(path, positions), start=1):
        values = dict(zip(scan_names, row))
        row_missing = {feature for feature in request.feature_columns if _is_missing(values[feature])}
        missing_columns.update(row_missing)
        should_drop = request.missing_values.method == "drop_rows" and bool(row_missing & (drop_columns or set(request.feature_columns)))
        if should_drop:
            continue
        unresolved_missing.update(row_missing)
        for feature in request.feature_columns:
            if feature not in row_missing:
                _numeric(values[feature], feature, row_number)
        kept_rows += 1

    if kept_rows < 2:
        raise PlanCompilationError(f"Decomposition requires at least 2 retained rows; missing-value handling leaves {kept_rows}.")
    method = request.missing_values.method
    if not missing_columns and method != "error":
        raise PlanCompilationError(f"missing_values.method={method!r} would be silently skipped because the selected data has no missing values; use 'error'.")
    if missing_columns and method == "error":
        raise PlanCompilationError(f"Selected decomposition features contain missing values in {sorted(missing_columns)}; choose drop_rows or impute explicitly.")
    if method == "impute":
        unresolved_missing.clear()
    if unresolved_missing:
        raise PlanCompilationError(f"Unprocessed missing values remain in {sorted(unresolved_missing)}; the public CLI exposes no decomposition models for this branch.")
    return _DecompositionDatasetProfile(
        kept_rows,
        frozenset(missing_columns),
        frozenset(unresolved_missing),
    )


def _scan_anomaly_detection_training_dataset(
    path: Path,
    columns: tuple[str, ...],
    request: AnomalyDetectionRequest,
) -> _AnomalyDetectionDatasetProfile:
    scan_names = request.feature_columns
    positions = [columns.index(column) for column in scan_names]
    drop_columns = set(getattr(request.missing_values, "columns", ()))
    unknown_drop_columns = sorted(drop_columns - set(request.feature_columns))
    if unknown_drop_columns:
        raise PlanCompilationError("Missing-value drop columns are not selected anomaly-detection features: " f"{unknown_drop_columns}")

    missing_columns: set[str] = set()
    unresolved_missing: set[str] = set()
    kept_rows = 0
    for row_number, row in enumerate(_iter_selected_rows(path, positions), start=1):
        values = dict(zip(scan_names, row))
        row_missing = {feature for feature in request.feature_columns if _is_missing(values[feature])}
        missing_columns.update(row_missing)
        should_drop = request.missing_values.method == "drop_rows" and bool(row_missing & (drop_columns or set(request.feature_columns)))
        if should_drop:
            continue
        unresolved_missing.update(row_missing)
        for feature in request.feature_columns:
            if feature not in row_missing:
                _numeric(values[feature], feature, row_number)
        kept_rows += 1

    if kept_rows < 2:
        raise PlanCompilationError("Anomaly detection requires at least 2 retained rows; " f"missing-value handling leaves {kept_rows}.")
    method = request.missing_values.method
    if not missing_columns and method != "error":
        raise PlanCompilationError(f"missing_values.method={method!r} would be silently skipped because " "the selected data has no missing values; use 'error'.")
    if missing_columns and method == "error":
        raise PlanCompilationError("Selected anomaly-detection features contain missing values in " f"{sorted(missing_columns)}; choose drop_rows or impute explicitly.")
    if method == "impute":
        unresolved_missing.clear()
    if unresolved_missing:
        raise PlanCompilationError(f"Unprocessed missing values remain in {sorted(unresolved_missing)}; " "the public CLI exposes no anomaly-detection models for this branch.")
    return _AnomalyDetectionDatasetProfile(
        kept_rows,
        frozenset(missing_columns),
        frozenset(unresolved_missing),
    )


def _validate_application_dataset(
    path: Path,
    columns: tuple[str, ...],
    request: ClassificationRequest | RegressionRequest,
) -> None:
    scan_names = request.feature_columns
    positions = [columns.index(column) for column in scan_names]
    row_count = 0
    usable_rows = 0
    for row_number, row in enumerate(_iter_selected_rows(path, positions), start=1):
        row_count += 1
        values = dict(zip(scan_names, row))
        missing_features = {feature for feature in request.feature_columns if _is_missing(values[feature])}
        for feature in request.feature_columns:
            if feature not in missing_features:
                _numeric(values[feature], feature, row_number)
        if missing_features and request.missing_values.method == "error":
            raise PlanCompilationError(f"Application feature data contains missing values in {sorted(missing_features)}, but the training request has no missing-value transform.")
        if request.missing_values.method == "drop_rows" and missing_features:
            continue
        usable_rows += 1
    if row_count == 0:
        raise PlanCompilationError("Application dataset must contain at least one data row.")
    if usable_rows == 0:
        raise PlanCompilationError("Missing-value handling removes every application row, so inference would produce no predictions.")


_PLACEHOLDER = re.compile(r"\{([^{}]+)\}")
_ALLOWED_FORMULA_NAMES = {"pow", "sin", "cos", "tan", "pi", "mean", "std", "var", "log"}
_FORMULA_FUNCTION_ARITY = {
    "pow": 2,
    "sin": 1,
    "cos": 1,
    "tan": 1,
    "mean": 1,
    "std": 1,
    "var": 1,
    "log": 1,
}
_ALLOWED_FORMULA_NODES = (
    ast.Expression,
    ast.BinOp,
    ast.UnaryOp,
    ast.Call,
    ast.Name,
    ast.Constant,
    ast.Add,
    ast.Sub,
    ast.Mult,
    ast.Div,
    ast.Pow,
    ast.UAdd,
    ast.USub,
    ast.Load,
)


def _compile_engineered_features(
    request: ClassificationRequest | RegressionRequest | ClusteringRequest | DecompositionRequest | AnomalyDetectionRequest,
    selected_names: tuple[str, ...],
) -> tuple[tuple[str, str], ...]:
    if not request.engineered_features:
        return ()
    if len(selected_names) + len(request.engineered_features) > 26:
        raise PlanCompilationError("The CLI letter-based feature builder can address at most 26 selected and engineered columns.")
    request_target_columns = set(getattr(request, "resolved_target_columns", ()))
    if not request_target_columns:
        target_column = getattr(request, "target_column", None)
        request_target_columns = {target_column} if target_column is not None else set()
    available = list(selected_names)
    compiled: list[tuple[str, str]] = []
    for feature in request.engineered_features:
        referenced: set[str] = set()

        def replace(match: re.Match[str]) -> str:
            column = match.group(1).strip()
            if column not in available:
                raise PlanCompilationError(f"Engineered feature {feature.name!r} references unavailable column {column!r}.")
            if column in request_target_columns:
                raise PlanCompilationError(f"Engineered feature {feature.name!r} must not use the target column; that would leak labels into model inputs.")
            referenced.add(column)
            return chr(ord("a") + available.index(column))

        expression = _PLACEHOLDER.sub(replace, feature.formula)
        if not referenced:
            raise PlanCompilationError(f"Engineered feature {feature.name!r} must reference at least one column with {{column name}} syntax.")
        try:
            tree = ast.parse(expression, mode="eval")
        except SyntaxError as exc:
            raise PlanCompilationError(f"Engineered feature {feature.name!r} has an invalid arithmetic formula.") from exc
        for node in ast.walk(tree):
            if not isinstance(node, _ALLOWED_FORMULA_NODES):
                raise PlanCompilationError(f"Engineered feature {feature.name!r} uses unsupported formula syntax: {type(node).__name__}.")
            if isinstance(node, ast.Name) and node.id not in _ALLOWED_FORMULA_NAMES and node.id not in {chr(ord("a") + i) for i in range(len(available))}:
                raise PlanCompilationError(f"Engineered feature {feature.name!r} contains unknown formula name {node.id!r}.")
            if isinstance(node, ast.Call) and (not isinstance(node.func, ast.Name) or node.func.id not in _ALLOWED_FORMULA_NAMES or node.keywords):
                raise PlanCompilationError(f"Engineered feature {feature.name!r} uses an unsupported function call.")
            if isinstance(node, ast.Call) and (node.func.id not in _FORMULA_FUNCTION_ARITY or len(node.args) != _FORMULA_FUNCTION_ARITY[node.func.id]):
                raise PlanCompilationError(f"Engineered feature {feature.name!r} uses the wrong number of function arguments.")
            if isinstance(node, ast.Constant) and (isinstance(node.value, bool) or not isinstance(node.value, (int, float))):
                raise PlanCompilationError(f"Engineered feature {feature.name!r} may only contain numeric constants.")
        compiled.append((feature.name, expression))
        available.append(feature.name)
    return tuple(compiled)


def _model_steps(request: ClassificationRequest) -> list[InteractionStep]:
    model = request.model
    model_name = model.type
    prefix = MODEL_DISPLAY_NAMES[model_name]
    prompt = f"{prefix} - Hyper-parameters Specification"
    steps: list[InteractionStep] = []

    def add(step_id: str, label: str, response: Any) -> None:
        anchors = (prompt, label, "(Model) ➜") if not steps else (label, "(Model) ➜")
        steps.append(InteractionStep(step_id, anchors, str(response)))

    def add_float(step_id: str, label: str, value: float, default: float) -> None:
        add(step_id, label, _float_response(value, default))

    if model_name == "logistic_regression":
        penalty_ids = {
            "l1": "l1_penalty",
            "l2": "l2_penalty",
            "elasticnet": "elasticnet_penalty",
        }
        add(
            penalty_ids[model.penalty],
            "Penalty:",
            _choice(model.penalty, ("l1", "l2", "elasticnet", "None")),
        )
        add_float(
            "default_regularization_strength",
            "C: This hyperparameter",
            model.regularization_strength,
            1.0,
        )
        if model.penalty == "l1":
            add(
                "solver",
                "Solver: This hyperparameter",
                _choice(model.solver, ("liblinear", "saga")),
            )
        elif model.penalty == "l2":
            solver_id = "lbfgs_solver" if model.solver == "lbfgs" else "solver"
            add(
                solver_id,
                "Solver: This hyperparameter",
                _choice(model.solver, ("newton-cg", "lbfgs", "sag", "saga")),
            )
        else:
            add_float("l1_ratio", "L1 Ratio:", model.l1_ratio, 0.5)
        add("maximum_iterations", "Max Iter:", model.maximum_iterations)
        class_id = "no_class_weight" if model.class_weight == "none" else "balanced_class_weight"
        add(
            class_id,
            "Class Weight:",
            _choice(
                "None" if model.class_weight == "none" else "balanced",
                ("None", "balanced"),
            ),
        )
    elif model_name == "support_vector_machine":
        add(
            "kernel",
            "Kernel:",
            _choice(model.kernel, ("linear", "poly", "rbf", "sigmoid")),
        )
        if model.kernel == "poly":
            add("degree", "Degree:", model.degree)
        if model.kernel in {"poly", "rbf", "sigmoid"}:
            add_float("gamma", "Gamma:", model.gamma, 0.1)
        add_float(
            "regularization_strength",
            "C: This hyperparameter",
            model.regularization_strength,
            1.0,
        )
        add("shrinking", "Shrinking:", "1" if model.shrinking else "2")
    elif model_name == "decision_tree":
        add(
            "criterion",
            "Criterion:",
            _choice(model.criterion, ("gini", "entropy", "log_loss")),
        )
        add("maximum_depth", "Max Depth:", model.maximum_depth)
        add("minimum_samples_split", "Min Samples Split:", model.minimum_samples_split)
        add("minimum_samples_leaf", "Min Samples Leaf:", model.minimum_samples_leaf)
        add("maximum_features", "Max Features:", model.maximum_features)
    elif model_name in {"random_forest", "extra_trees"}:
        add("number_of_estimators", "N Estimators:", model.number_of_estimators)
        add("maximum_depth", "Max Depth:", model.maximum_depth)
        add("minimum_samples_split", "Min Samples Split:", model.minimum_samples_split)
        add("minimum_samples_leaf", "Min Samples Leaf:", model.minimum_samples_leaf)
        add("maximum_features", "Max Features:", model.maximum_features)
        add("bootstrap", "Bootstrap:", "1" if model.bootstrap else "2")
        if model.bootstrap:
            add_float("maximum_samples", "Max Samples:", model.maximum_samples, 0.8)
        add("out_of_bag_score", "oob_score:", "1" if model.out_of_bag_score else "2")
    elif model_name == "xgboost":
        add("number_of_estimators", "N Estimators:", model.number_of_estimators)
        add_float("learning_rate", "Learning Rate:", model.learning_rate, 0.01)
        add("maximum_depth", "Max Depth:", model.maximum_depth)
        add_float("subsample", "Subsample:", model.subsample, 1.0)
        add_float("column_subsample", "Colsample Bytree:", model.column_subsample, 1.0)
        add_float("l1_regularization", "Alpha:", model.l1_regularization, 0.0)
        add_float("l2_regularization", "Lambda:", model.l2_regularization, 1.0)
    elif model_name == "multi_layer_perceptron":
        add(
            "hidden_layer_sizes",
            "Hidden Layer Sizes:",
            repr(tuple(model.hidden_layer_sizes)),
        )
        add(
            "activation",
            "Activation:",
            _choice(model.activation, ("identity", "logistic", "tanh", "relu")),
        )
        add("solver", "Solver:", _choice(model.solver, ("lbfgs", "sgd", "adam")))
        add_float("alpha", "Alpha:", model.alpha, 0.0001)
        add(
            "learning_rate",
            "Learning Rate:",
            _choice(model.learning_rate, ("constant", "invscaling", "adaptive")),
        )
        add("maximum_iterations", "Max Iterations:", model.maximum_iterations)
    elif model_name == "gradient_boosting":
        add("number_of_estimators", "N Estimators:", model.number_of_estimators)
        add_float("learning_rate", "Learning Rate:", model.learning_rate, 0.1)
        add("maximum_depth", "Max Depth:", model.maximum_depth)
        add("minimum_samples_split", "Min Samples Split:", model.minimum_samples_split)
        add("minimum_samples_leaf", "Min Samples Leaf:", model.minimum_samples_leaf)
        add("maximum_features", "Max Features:", model.maximum_features)
        add_float("subsample", "Subsample:", model.subsample, 1.0)
        add("loss", "Loss:", _choice(model.loss, ("log_loss", "exponential")))
    elif model_name == "k_nearest_neighbors":
        add("number_of_neighbors", "N Neighbors:", model.number_of_neighbors)
        add("weights", "Weights:", _choice(model.weights, ("uniform", "distance")))
        add(
            "algorithm",
            "Algorithm:",
            _choice(model.algorithm, ("auto", "ball_tree", "kd_tree", "brute")),
        )
        if model.algorithm in {"ball_tree", "kd_tree"}:
            add("leaf_size", "Leaf Size:", model.leaf_size)
        add(
            "metric",
            "Metric:",
            _choice(model.metric, ("euclidean", "manhattan", "minkowski")),
        )
        if model.metric == "minkowski":
            add("power", "P:", model.power)
    elif model_name == "stochastic_gradient_descent":
        add(
            "loss",
            "Loss Function:",
            _choice(model.loss, ("log_loss", "modified_huber")),
        )
        add(
            "penalty",
            "Penalty:",
            _choice(
                "None" if model.penalty == "none" else model.penalty,
                ("l2", "l1", "elasticnet", "None"),
            ),
        )
        if model.penalty == "elasticnet":
            add_float("l1_ratio", "L1 Ratio:", model.l1_ratio, 0.15)
        add_float("alpha", "Alpha:", model.alpha, 0.0001)
        add("fit_intercept", "Fit Intercept:", "1" if model.fit_intercept else "2")
        add(
            "maximum_iterations",
            "Maximum Number of Iterations:",
            model.maximum_iterations,
        )
        add_float("tolerance", "Tolerance:", model.tolerance, 0.001)
        add("shuffle", "Shuffle:", "1" if model.shuffle else "2")
        add(
            "learning_rate",
            "Learning Rate:",
            _choice(model.learning_rate, ("constant", "optimal", "invscaling", "adaptive")),
        )
        add_float(
            "initial_learning_rate",
            "Initial Learning Rate:",
            model.initial_learning_rate,
            0.0,
        )
        add_float("power", "Power T:", model.power, 0.5)
        add("early_stopping", "Early Stopping:", "1" if model.early_stopping else "2")
        add_float(
            "validation_fraction",
            "Validation Fraction:",
            model.validation_fraction,
            0.1,
        )
        add(
            "iterations_without_improvement",
            "Iterations With No Improvement:",
            model.iterations_without_improvement,
        )
    elif model_name == "adaboost":
        add("number_of_estimators", "N Estimators:", model.number_of_estimators)
        add_float("learning_rate", "Learning Rate:", model.learning_rate, 0.01)
        add("maximum_depth", "Max Depth:", model.maximum_depth)
    return steps


def _regression_model_steps(request: RegressionRequest) -> list[InteractionStep]:
    model = request.model
    model_name = model.type
    prefix = REGRESSION_MODEL_DISPLAY_NAMES[model_name]
    prompt = f"{prefix} - Hyper-parameters Specification"
    steps: list[InteractionStep] = []

    def add(step_id: str, label: str, response: Any) -> None:
        anchors = (prompt, label, "(Model)") if not steps else (label, "(Model)")
        steps.append(InteractionStep(step_id, anchors, str(response)))

    def add_float(step_id: str, label: str, value: float, default: float) -> None:
        add(step_id, label, _float_response(value, default))

    if model_name == "linear_regression":
        add("fit_intercept", "Fit Intercept:", "1" if model.fit_intercept else "2")
    elif model_name == "polynomial_regression":
        add("degree", "Degree:", model.degree)
        add(
            "interaction_only",
            "Interaction Only:",
            "1" if model.interaction_only else "2",
        )
        add("include_bias", "Include Bias:", "1" if model.include_bias else "2")
    elif model_name == "k_nearest_neighbors":
        add("number_of_neighbors", "N Neighbors:", model.number_of_neighbors)
        add("weights", "Weights:", _choice(model.weights, ("uniform", "distance")))
        add(
            "algorithm",
            "Algorithm:",
            _choice(model.algorithm, ("auto", "ball_tree", "kd_tree", "brute")),
        )
        if model.algorithm in {"ball_tree", "kd_tree"}:
            add("leaf_size", "Leaf Size:", model.leaf_size)
        add(
            "metric",
            "Metric:",
            _choice(model.metric, ("euclidean", "manhattan", "minkowski")),
        )
        if model.metric == "minkowski":
            add("power", "P:", model.power)
    elif model_name == "support_vector_machine":
        add(
            "kernel",
            "Kernel:",
            _choice(model.kernel, ("linear", "poly", "rbf", "sigmoid")),
        )
        if model.kernel == "poly":
            add("degree", "Degree:", model.degree)
        if model.kernel in {"poly", "rbf", "sigmoid"}:
            add_float("gamma", "Gamma:", model.gamma, 0.1)
        add_float(
            "regularization_strength",
            "C: This hyperparameter",
            model.regularization_strength,
            1.0,
        )
        add("shrinking", "Shrinking:", "1" if model.shrinking else "2")
    elif model_name == "decision_tree":
        add(
            "criterion",
            "Criterion:",
            _choice(
                model.criterion,
                ("squared_error", "friedman_mse", "absolute_error", "poisson"),
            ),
        )
        add("maximum_depth", "Max Depth:", model.maximum_depth)
        add("minimum_samples_split", "Min Samples Split:", model.minimum_samples_split)
        add("minimum_samples_leaf", "Min Samples Leaf:", model.minimum_samples_leaf)
        add("maximum_features", "Max Features:", model.maximum_features)
    elif model_name in {"random_forest", "extra_trees"}:
        add("number_of_estimators", "N Estimators:", model.number_of_estimators)
        add("maximum_depth", "Max Depth:", model.maximum_depth)
        add("minimum_samples_split", "Min Samples Split:", model.minimum_samples_split)
        add("minimum_samples_leaf", "Min Samples Leaf:", model.minimum_samples_leaf)
        add("maximum_features", "Max Features:", model.maximum_features)
        add("bootstrap", "Bootstrap:", "1" if model.bootstrap else "2")
        if model.bootstrap:
            add_float("maximum_samples", "Max Samples:", model.maximum_samples, 0.8)
        add("out_of_bag_score", "oob_score:", "1" if model.out_of_bag_score else "2")
    elif model_name == "gradient_boosting":
        add("number_of_estimators", "N Estimators:", model.number_of_estimators)
        add_float("learning_rate", "Learning Rate:", model.learning_rate, 0.1)
        add("maximum_depth", "Max Depth:", model.maximum_depth)
        add("minimum_samples_split", "Min Samples Split:", model.minimum_samples_split)
        add("minimum_samples_leaf", "Min Samples Leaf:", model.minimum_samples_leaf)
        add("maximum_features", "Max Features:", model.maximum_features)
        add_float("subsample", "Subsample:", model.subsample, 1.0)
        add(
            "loss",
            "Loss:",
            _choice(model.loss, ("squared_error", "absolute_error", "huber", "quantile")),
        )
    elif model_name == "xgboost":
        add("number_of_estimators", "N Estimators:", model.number_of_estimators)
        add_float("learning_rate", "Learning Rate:", model.learning_rate, 0.01)
        add("maximum_depth", "Max Depth:", model.maximum_depth)
        add_float("subsample", "Subsample:", model.subsample, 1.0)
        add_float("column_subsample", "Colsample Bytree:", model.column_subsample, 1.0)
        add_float("l1_regularization", "Alpha:", model.l1_regularization, 0.0)
        add_float("l2_regularization", "Lambda:", model.l2_regularization, 1.0)
    elif model_name == "multi_layer_perceptron":
        add(
            "hidden_layer_sizes",
            "Hidden Layer Sizes:",
            repr(tuple(model.hidden_layer_sizes)),
        )
        add(
            "activation",
            "Activation:",
            _choice(model.activation, ("identity", "logistic", "tanh", "relu")),
        )
        add("solver", "Solver:", _choice(model.solver, ("lbfgs", "sgd", "adam")))
        add_float("alpha", "Alpha:", model.alpha, 0.0001)
        add(
            "learning_rate",
            "Learning Rate:",
            _choice(model.learning_rate, ("constant", "invscaling", "adaptive")),
        )
        add("maximum_iterations", "Max Iterations:", model.maximum_iterations)
    elif model_name == "lasso_regression":
        add_float("alpha", "Alpha:", model.alpha, 0.01)
        add("fit_intercept", "Fit Intercept:", "1" if model.fit_intercept else "2")
        add("maximum_iterations", "Max Iter:", model.maximum_iterations)
        add_float("tolerance", "Tolerance:", model.tolerance, 0.0001)
        add("selection", "Selection:", _choice(model.selection, ("cyclic", "random")))
    elif model_name == "elastic_net":
        add_float("alpha", "Alpha:", model.alpha, 1.0)
        add_float("l1_ratio", "L1 Ratio:", model.l1_ratio, 0.5)
        add("fit_intercept", "Fit Intercept:", "1" if model.fit_intercept else "2")
        add("maximum_iterations", "Max Iter:", model.maximum_iterations)
        add_float("tolerance", "Tolerance:", model.tolerance, 0.0001)
        add("selection", "Selection:", _choice(model.selection, ("cyclic", "random")))
    elif model_name == "stochastic_gradient_descent":
        add(
            "loss",
            "Loss Function:",
            _choice(
                model.loss,
                (
                    "squared_error",
                    "huber",
                    "epsilon_insensitive",
                    "squared_epsilon_insensitive",
                ),
            ),
        )
        add(
            "penalty",
            "Penalty:",
            _choice(
                "None" if model.penalty == "none" else model.penalty,
                ("l2", "l1", "elasticnet", "None"),
            ),
        )
        if model.penalty == "elasticnet":
            add_float("l1_ratio", "L1 Ratio:", model.l1_ratio, 0.15)
        add_float("alpha", "Alpha:", model.alpha, 0.0001)
        add("fit_intercept", "Fit Intercept:", "1" if model.fit_intercept else "2")
        add(
            "maximum_iterations",
            "Maximum Number of Iterations:",
            model.maximum_iterations,
        )
        add_float("tolerance", "Tolerance:", model.tolerance, 0.001)
        add("shuffle", "Shuffle:", "1" if model.shuffle else "2")
        add(
            "learning_rate",
            "Learning Rate:",
            _choice(model.learning_rate, ("constant", "optimal", "invscaling", "adaptive")),
        )
        add_float(
            "initial_learning_rate",
            "Initial Learning Rate:",
            model.initial_learning_rate,
            0.01,
        )
        add_float("power", "Power T:", model.power, 0.25)
    elif model_name == "bayesian_ridge":
        add_float("tolerance", "Tolerance:", model.tolerance, 0.0001)
        add_float("alpha_1", "Alpha 1:", model.alpha_1, 0.000001)
        add_float("alpha_2", "Alpha 2:", model.alpha_2, 0.000001)
        add_float("lambda_1", "Lambda 1:", model.lambda_1, 0.000001)
        add_float("lambda_2", "Lambda 2:", model.lambda_2, 0.000001)
        add_float("alpha_initial", "Alpha Init:", model.alpha_initial, 1.0)
        add_float("lambda_initial", "Lambda Init:", model.lambda_initial, 1.0)
        add("compute_score", "Compute Score:", "1" if model.compute_score else "2")
        add("fit_intercept", "Fit Intercept:", "1" if model.fit_intercept else "2")
        add("copy_x", "Copy X:", "1" if model.copy_x else "2")
        add("verbose", "Verbose:", "1" if model.verbose else "2")
    elif model_name == "ridge_regression":
        add_float("alpha", "Alpha:", model.alpha, 0.01)
        add("fit_intercept", "Fit Intercept:", "1" if model.fit_intercept else "2")
        add("maximum_iterations", "Max Iter:", model.maximum_iterations)
        add_float("tolerance", "Tolerance:", model.tolerance, 0.0001)
    return steps


def _clustering_model_steps(request: ClusteringRequest) -> list[InteractionStep]:
    model = request.model
    model_name = model.type
    prefix = CLUSTERING_MODEL_DISPLAY_NAMES[model_name]
    prompt = f"{prefix} - Hyper-parameters Specification"
    steps: list[InteractionStep] = []

    def add(step_id: str, label: str, response: Any) -> None:
        anchors = (prompt, label, "(Model)") if not steps else (label, "(Model)")
        steps.append(InteractionStep(step_id, anchors, str(response)))

    def add_float(step_id: str, label: str, value: float, default: float) -> None:
        add(step_id, label, _float_response(value, default))

    if model_name == "kmeans":
        add("number_of_clusters", "N Clusters:", model.number_of_clusters)
        add(
            "initialization",
            "Init: Method for initialization",
            _choice(model.initialization, ("k-means++", "random")),
        )
        add("maximum_iterations", "Max Iter:", model.maximum_iterations)
        add_float("tolerance", "Tolerance:", model.tolerance, 0.0005)
        add(
            "algorithm",
            "Algorithm: The algorithm to use",
            _choice(model.algorithm, ("auto", "full", "elkan")),
        )
    elif model_name == "dbscan":
        add_float("epsilon", "Eps:", model.epsilon, 0.5)
        add("minimum_samples", "Min Samples:", model.minimum_samples)
        add(
            "algorithm",
            "Algorithm: The algorithm to be used",
            _choice(model.algorithm, ("auto", "ball_tree", "kd_tree", "brute")),
        )
        if model.algorithm == "kd_tree":
            metrics = (
                "euclidean",
                "l2",
                "minkowski",
                "p",
                "manhattan",
                "cityblock",
                "l1",
                "chebyshev",
                "infinity",
            )
        elif model.algorithm == "ball_tree":
            metrics = (
                "euclidean",
                "l2",
                "minkowski",
                "p",
                "manhattan",
                "cityblock",
                "l1",
                "chebyshev",
                "infinity",
                "seuclidean",
                "mahalanobis",
                "hamming",
                "canberra",
                "braycurtis",
                "jaccard",
                "dice",
                "rogerstanimoto",
                "russellrao",
                "sokalmichener",
                "sokalsneath",
                "haversine",
            )
        else:
            metrics = (
                "euclidean",
                "manhattan",
                "chebyshev",
                "minkowski",
                "cosine",
                "correlation",
            )
        add(
            "metric",
            "Metric: The metric to use",
            _choice(model.metric, metrics),
        )
        add("leaf_size", "Leaf Size:", model.leaf_size)
        if model.metric == "minkowski":
            add("power", "P:", model.power)
    elif model_name == "agglomerative":
        add("number_of_clusters", "N Clusters:", model.number_of_clusters)
        add(
            "linkage",
            "linkage: The linkage criterion",
            _choice(model.linkage, ("ward", "complete", "average", "single")),
        )
    elif model_name == "affinity_propagation":
        add_float("damping", "damping:", model.damping, 0.5)
        add("maximum_iterations", "Max Iter:", model.maximum_iterations)
        add(
            "convergence_iterations",
            "convergence_iter:",
            model.convergence_iterations,
        )
        add(
            "affinity",
            "affinity: Different affinity methods",
            _choice(model.affinity, ("euclidean", "precomputed")),
        )
    elif model_name == "mean_shift":
        add("bandwidth", "Bandwidth:", 0 if model.bandwidth is None else model.bandwidth)
        add(
            "cluster_all",
            "Cluster All:",
            _choice("True" if model.cluster_all else "False", ("True", "False")),
        )
        add(
            "bin_seeding",
            "Bin Seeding:",
            _choice("True" if model.bin_seeding else "False", ("True", "False")),
        )
        add("minimum_bin_frequency", "Min Bin Frequency:", model.minimum_bin_frequency)
        add("number_of_jobs", "Number of Jobs:", model.number_of_jobs)
        add("maximum_iterations", "Max Iterations:", model.maximum_iterations)
    return steps


def _decomposition_model_steps(request: DecompositionRequest) -> list[InteractionStep]:
    model = request.model
    prompt = f"{DECOMPOSITION_MODEL_DISPLAY_NAMES[model.type]} - Hyper-parameters Specification"
    steps: list[InteractionStep] = []

    def add(step_id: str, label: str, response: Any) -> None:
        anchors = (prompt, label, "(Model)") if not steps else (label, "(Model)")
        steps.append(InteractionStep(step_id, anchors, str(response)))

    if model.type == "pca":
        add("number_of_components", "N Components:", model.number_of_components)
        add(
            "svd_solver",
            "SVD Solver:",
            _choice(model.svd_solver, ("auto", "full", "arpack", "randomized")),
        )
    elif model.type == "tsne":
        add("number_of_components", "N Components:", model.number_of_components)
        add("perplexity", "Perplexity:", model.perplexity)
        add(
            "learning_rate",
            "Learning Rate:",
            _float_response(model.learning_rate, 200.0),
        )
        add("number_of_iterations", "Number of Iterations:", model.number_of_iterations)
        add(
            "early_exaggeration",
            "Early Exaggeration:",
            _float_response(model.early_exaggeration, 12.0),
        )
    else:
        add("number_of_components", "N Components:", model.number_of_components)
        add("metric", "Metric:", "1" if model.metric else "2")
        add("number_of_initializations", "N Init:", model.number_of_initializations)
        add("maximum_iterations", "Max Iter:", model.maximum_iterations)
    return steps


def _anomaly_detection_model_steps(
    request: AnomalyDetectionRequest,
) -> list[InteractionStep]:
    model = request.model
    prompt = f"{ANOMALY_DETECTION_MODEL_DISPLAY_NAMES[model.type]} " "- Hyper-parameters Specification"
    steps: list[InteractionStep] = []

    def add(step_id: str, label: str, response: Any) -> None:
        anchors = (prompt, label, "(Model)") if not steps else (label, "(Model)")
        steps.append(InteractionStep(step_id, anchors, str(response)))

    if model.type == "isolation_forest":
        add("number_of_estimators", "N Estimators:", model.number_of_estimators)
        add(
            "contamination",
            "Contamination:",
            _float_response(model.contamination, 0.3),
        )
        add("maximum_features", "Max Features:", model.maximum_features)
        add("bootstrap", "Bootstrap:", "1" if model.bootstrap else "2")
        if model.bootstrap:
            add("maximum_samples", "Max Samples:", model.maximum_samples)
    else:
        add("number_of_neighbors", "N neighbors:", model.number_of_neighbors)
        add("leaf_size", "Leaf size:", model.leaf_size)
        add("power", "P: The power parameter", _float_response(model.power, 2.0))
        add(
            "contamination",
            "Contamination:",
            _float_response(model.contamination, 0.3),
        )
        add("number_of_jobs", "N jobs:", model.number_of_jobs)
    return steps


class ClassificationPlanCompiler:
    """Compile supported classification branches without importing ML code."""

    def compile(
        self,
        request: ClassificationRequest,
        cli_executable: Optional[Path] = None,
        *,
        dataset_context: DatasetCompilationContext | None = None,
    ) -> InteractionPlan:
        dataset_context = dataset_context or DatasetCompilationContext()
        data_path = request.training_dataset_path.expanduser().resolve()
        if not data_path.is_file():
            raise PlanCompilationError(f"Training data file does not exist: {data_path}")
        columns = _read_dataset_columns(data_path, source=dataset_context.training_source)
        _validate_world_map(data_path, columns, request.world_map)
        requested_columns = (
            request.identifier_column,
            request.target_column,
            *request.feature_columns,
        )
        missing = sorted({column for column in requested_columns if column not in columns})
        if missing:
            raise PlanCompilationError(f"Requested columns are absent from the training dataset: {missing}")
        application_path = None
        external_evaluation_path = request.evaluation.evaluation_dataset_path if request.evaluation.mode == "external_labeled" else None
        if request.evaluation.mode == "external_labeled" and request.evaluation.evaluation_dataset is not None:
            raise PlanCompilationError("External evaluation dataset references must be resolved to a local path before CLI compilation.")
        if request.application_dataset_path is not None and external_evaluation_path is not None:
            raise PlanCompilationError("A distinct external evaluation dataset cannot also be used as ordinary application data.")
        selected_application_path = request.application_dataset_path or external_evaluation_path
        if selected_application_path is not None:
            application_path = selected_application_path.expanduser().resolve()
            if not application_path.is_file():
                raise PlanCompilationError(f"Application or external evaluation data file does not exist: {application_path}")
            application_columns = _read_dataset_columns(
                application_path,
                source=dataset_context.source_for("application"),
            )
            application_identifier = request.evaluation.external_identifier_column or request.identifier_column if external_evaluation_path is not None else request.identifier_column
            required_application = {application_identifier, *request.feature_columns}
            application_missing = sorted(required_application - set(application_columns))
            if application_missing:
                raise PlanCompilationError(f"Application dataset is missing required identifier or feature columns: {application_missing}")
            if external_evaluation_path is not None:
                evaluation_missing = sorted({request.target_column} - set(application_columns))
                if evaluation_missing:
                    raise PlanCompilationError("External evaluation dataset is missing target columns: " f"{evaluation_missing}")
            _validate_application_dataset(application_path, application_columns, request)

        selected_names = tuple(column for column in columns if column == request.target_column or column in request.feature_columns)
        engineered = _compile_engineered_features(request, selected_names)
        profile = _scan_training_dataset(data_path, columns, request)
        final_feature_names = (
            *request.feature_columns,
            *(name for name, _ in engineered),
        )
        final_feature_count = len(final_feature_names)
        selected_feature_count = getattr(request.feature_selection, "retain_count", final_feature_count)
        if request.feature_selection.method != "none" and selected_feature_count >= final_feature_count:
            raise PlanCompilationError(f"feature_selection.retain_count must be less than the {final_feature_count} input features because that is the CLI's enforced contract.")
        maximum_features = getattr(request.model, "maximum_features", None)
        if maximum_features is not None and maximum_features > selected_feature_count:
            raise PlanCompilationError(f"model.maximum_features={maximum_features} exceeds the {selected_feature_count} features available after preprocessing.")

        original_positions = {column: index + 1 for index, column in enumerate(columns)}
        selected_positions = {column: index + 1 for index, column in enumerate((*selected_names, *(name for name, _ in engineered)))}
        selected_expression = _selection_expression([original_positions[column] for column in selected_names])
        feature_expression = _selection_expression([selected_positions[column] for column in final_feature_names])
        target_expression = _selection_expression([selected_positions[request.target_column]])
        executable = Path(cli_executable).expanduser().resolve() if cli_executable else resolve_public_cli_executable()
        if not executable.is_file():
            raise PlanCompilationError(f"CLI executable does not exist: {executable}")

        enter_prompt = "(Press Enter key to move forward)"
        steps: list[InteractionStep] = [
            *_experiment_steps(request, enter_prompt),
            InteractionStep(
                "identifier_column",
                ("output data identifier column", "(Data) ➜ @Number:"),
                str(original_positions[request.identifier_column]),
            ),
            InteractionStep("continue_after_identifier", (enter_prompt,), ""),
            InteractionStep("continue_after_map_skip", ("World Map Projection", enter_prompt), ""),
            InteractionStep(
                "selected_data_columns",
                (
                    "Data Selection",
                    "Select the data range you want to process.",
                    "@input:",
                ),
                selected_expression,
            ),
            InteractionStep(
                "continue_after_data_selection",
                ("Index - Column Name", enter_prompt),
                "",
            ),
            InteractionStep(
                "continue_after_selected_data_preview",
                ("The Selected Data Set:", enter_prompt),
                "",
            ),
            InteractionStep(
                "continue_after_statistics",
                ("Data Selected.xlsx", enter_prompt),
                "",
                timeout_seconds=240,
            ),
        ]
        steps.extend(self._missing_steps(request, profile, selected_positions, enter_prompt))
        steps.extend(self._feature_engineering_steps(engineered, enter_prompt))
        steps.extend(
            [
                InteractionStep(
                    "classification_mode",
                    ("Mode Selection", "2 - Classification", "(Model) ➜ @Number:"),
                    "2",
                ),
                InteractionStep("continue_after_mode", (enter_prompt,), ""),
                InteractionStep(
                    "feature_columns",
                    (
                        "Selected sub data set to create X data set",
                        "Select the data range you want to process.",
                        "@input:",
                    ),
                    feature_expression,
                ),
                InteractionStep(
                    "continue_after_features",
                    ("X Without Scaling.xlsx", enter_prompt),
                    "",
                    timeout_seconds=180,
                ),
                InteractionStep(
                    "target_column",
                    (
                        "The selected Y data set",
                        "Select the data range you want to process.",
                        "@input:",
                    ),
                    target_expression,
                ),
            ]
        )
        steps.extend(self._label_steps(request, len(profile.class_counts), enter_prompt))
        steps.extend(self._preprocessing_steps(request, enter_prompt))
        model_number = 1 if profile.unresolved_missing_columns else MODEL_NUMBERS[request.model.type]
        steps.extend(
            [
                InteractionStep(
                    request.model.type,
                    ("Which model do you want to apply?", "(Model) ➜ @Number:"),
                    str(model_number),
                ),
                InteractionStep("continue_after_model", (enter_prompt,), ""),
                InteractionStep(
                    "enable_automl" if request.tuning == "automl" else "disable_automl",
                    ("automated machine learning", "(Model) ➜ @Number:"),
                    "1" if request.tuning == "automl" else "2",
                ),
                InteractionStep("continue_after_automl", (enter_prompt,), ""),
            ]
        )
        if application_path is None:
            steps.append(
                InteractionStep(
                    "continue_after_inference_skip",
                    ("You did not provide application data.", enter_prompt),
                    "",
                )
            )
        else:
            application_output = "Application Data Feature-Engineering Selected.xlsx" if request.engineered_features else "Application Data Selected.xlsx"
            steps.append(
                InteractionStep(
                    "continue_after_application_preparation",
                    (application_output, enter_prompt),
                    "",
                    timeout_seconds=180,
                )
            )
        if request.tuning == "manual":
            steps.extend(_model_steps(request))
            steps.append(InteractionStep("continue_after_hyperparameters", (enter_prompt,), ""))
        model_display = MODEL_DISPLAY_NAMES[request.model.type]
        steps.extend(
            [
                InteractionStep(
                    "continue_after_training",
                    (f"{model_display}.joblib", enter_prompt),
                    "",
                    timeout_seconds=600,
                ),
                InteractionStep(
                    "continue_after_transform_pipeline",
                    ("Transform Pipeline.joblib", enter_prompt),
                    "",
                    timeout_seconds=180,
                ),
            ]
        )
        if application_path is not None:
            steps.append(
                InteractionStep(
                    "continue_after_inference",
                    ("Application Data Predicted.xlsx", enter_prompt),
                    "",
                    timeout_seconds=300,
                )
            )

        command = (str(executable), "data-mining", "--data", str(data_path))
        if application_path is not None:
            command = (
                str(executable),
                "data-mining",
                "--training",
                str(data_path),
                "--application",
                str(application_path),
            )
        command = _command_with_analysis_options(command, request.world_map, request.existing_experiment_id)
        return InteractionPlan(
            schema_version=INTERACTION_PLAN_VERSION,
            name=f"classification-{request.model.type}-v1",
            public_command=command,
            steps=tuple(steps),
            expected_output_relative_paths=(
                str(Path("geopi_output") / request.experiment_name / request.run_name / "artifacts" / "model" / f"{model_display}.joblib"),
                str(Path("geopi_output") / request.experiment_name / request.run_name / "artifacts" / "model" / "Transform Pipeline.joblib"),
                str(Path("geopi_output") / request.experiment_name / request.run_name / "artifacts" / "image" / "model_output" / f"Precision-Recall vs. Threshold Diagram - {model_display}.png"),
                str(Path("geopi_output") / request.experiment_name / request.run_name / "metrics" / f"Model Score - {model_display}.txt"),
                str(Path("geopi_output") / request.experiment_name / request.run_name / "metrics" / f"Classification Report - {model_display}.txt"),
                str(Path("geopi_output") / request.experiment_name / request.run_name / "artifacts" / "image" / "model_output" / f"Confusion Matrix - {model_display}.png"),
                str(Path("geopi_output") / request.experiment_name / request.run_name / "artifacts" / "image" / "model_output" / f"Confusion Matrix - {model_display}.xlsx"),
                str(Path("geopi_output") / request.experiment_name / request.run_name / "artifacts" / "data" / "Y Test.xlsx"),
                str(Path("geopi_output") / request.experiment_name / request.run_name / "artifacts" / "data" / "Y Test Predict Decoded.xlsx"),
                str(Path("geopi_output") / request.experiment_name / request.run_name / "artifacts" / "image" / "model_output" / "ROC Curve - Probabilities.xlsx"),
                *(
                    (
                        str(Path("geopi_output") / request.experiment_name / request.run_name / "artifacts" / "image" / "model_output" / f"Feature Importance Diagram - {model_display}.png"),
                        str(Path("geopi_output") / request.experiment_name / request.run_name / "artifacts" / "image" / "model_output" / f"Feature Importance Diagram - {model_display}.xlsx"),
                    )
                    if request.model.type == "xgboost"
                    else ()
                ),
                str(Path("geopi_output") / request.experiment_name / request.run_name / "parameters" / f"Hyper Parameters - {model_display}.txt"),
                *((str(Path("geopi_output") / request.experiment_name / request.run_name / "artifacts" / "data" / "Application Data Predicted.xlsx"),) if application_path is not None else ()),
            ),
            requires_source_row_pairing=True,
        )

    @staticmethod
    def _missing_steps(
        request: ClassificationRequest | RegressionRequest | ClusteringRequest | DecompositionRequest | AnomalyDetectionRequest,
        profile: _DatasetProfile | _RegressionDatasetProfile | _ClusteringDatasetProfile | _DecompositionDatasetProfile | _AnomalyDetectionDatasetProfile,
        selected_positions: dict[str, int],
        enter_prompt: str,
    ) -> list[InteractionStep]:
        if not profile.missing_columns:
            return [
                InteractionStep(
                    "continue_after_missing_value_check",
                    ("complete without missing values", enter_prompt),
                    "",
                )
            ]
        steps = [
            InteractionStep(
                "continue_after_missing_value_check",
                ("Missing Value Check", enter_prompt),
                "",
            )
        ]
        method = request.missing_values.method
        steps.append(
            InteractionStep(
                ("process_missing_values" if method in {"drop_rows", "impute"} else "keep_missing_values"),
                ("Do you want to deal with the missing values?", "(Data) ➜ @Number:"),
                "1" if method in {"drop_rows", "impute"} else "2",
            )
        )
        if method == "keep":
            steps.append(
                InteractionStep(
                    "continue_after_keep_missing",
                    ("Data Selected Dropped-Imputed.xlsx", enter_prompt),
                    "",
                )
            )
            return steps
        steps.append(InteractionStep("continue_to_missing_strategy", (enter_prompt,), ""))
        if method == "drop_rows":
            steps.extend(
                [
                    InteractionStep(
                        "drop_missing_rows",
                        ("Which strategy do you want to apply?", "(Data) ➜ @Number:"),
                        "1",
                    ),
                    InteractionStep("continue_to_drop_strategy", (enter_prompt,), ""),
                    InteractionStep(
                        ("drop_all_missing_rows" if not request.missing_values.columns else "drop_missing_by_columns"),
                        ("Drop the rows with Missing Values", "(Data) ➜ @Number:"),
                        "1" if not request.missing_values.columns else "2",
                    ),
                ]
            )
            if request.missing_values.columns:
                expression = _selection_expression([selected_positions[column] for column in request.missing_values.columns])
                steps.append(
                    InteractionStep(
                        "drop_columns",
                        ("Select the data range you want to process.", "@input:"),
                        expression,
                    )
                )
            steps.append(
                InteractionStep(
                    "continue_after_drop_missing",
                    ("Data Selected Dropped-Imputed.xlsx", enter_prompt),
                    "",
                )
            )
            return steps
        steps.extend(
            [
                InteractionStep(
                    "impute_missing_values",
                    ("Which strategy do you want to apply?", "(Data) ➜ @Number:"),
                    "2",
                ),
                InteractionStep("continue_to_imputation_method", (enter_prompt,), ""),
                InteractionStep(
                    "imputation_method",
                    (
                        "Imputation Method Option",
                        "Which method do you want to apply?",
                        "(Data) ➜ @Number:",
                    ),
                    _choice(
                        request.missing_values.strategy,
                        ("mean", "median", "most_frequent", "constant"),
                    ),
                ),
            ]
        )
        if request.missing_values.strategy == "constant":
            steps.append(
                InteractionStep(
                    "imputation_fill_value",
                    ("Specified Value:", "(Model) ➜"),
                    _float_response(request.missing_values.fill_value, 0.0),
                )
            )
        steps.extend(
            [
                InteractionStep(
                    "continue_after_imputation",
                    ("Successfully fill the missing values", enter_prompt),
                    "",
                    timeout_seconds=240,
                ),
                InteractionStep(
                    "continue_after_imputation_statistics",
                    ("Data Selected Dropped-Imputed.xlsx", enter_prompt),
                    "",
                    timeout_seconds=240,
                ),
            ]
        )
        return steps

    @staticmethod
    def _feature_engineering_steps(engineered: tuple[tuple[str, str], ...], enter_prompt: str) -> list[InteractionStep]:
        if not engineered:
            return [
                InteractionStep(
                    "skip_feature_engineering",
                    ("Feature Engineering Option", "(Data) ➜ @Number:"),
                    "2",
                ),
                InteractionStep(
                    "continue_after_feature_engineering",
                    (
                        "Data Selected Dropped-Imputed Feature-Engineering.xlsx",
                        enter_prompt,
                    ),
                    "",
                ),
            ]
        steps: list[InteractionStep] = []
        for index, (name, expression) in enumerate(engineered, start=1):
            if index == 1:
                steps.append(
                    InteractionStep(
                        "enable_feature_engineering",
                        ("Feature Engineering Option", "(Data) ➜ @Number:"),
                        "1",
                    )
                )
            steps.extend(
                [
                    InteractionStep(
                        f"engineered_feature_{index}_name",
                        ("Name the constructed feature", "@input:"),
                        name,
                    ),
                    InteractionStep(
                        f"engineered_feature_{index}_formula",
                        ("Build up new feature", "@input:"),
                        expression,
                    ),
                    InteractionStep(f"engineered_feature_{index}_formula_ack", (enter_prompt,), ""),
                    InteractionStep(
                        f"engineered_feature_{index}_constructed",
                        ("Successfully construct a new feature", enter_prompt),
                        "",
                    ),
                    InteractionStep(
                        f"engineered_feature_{index}_statistics",
                        (
                            "Some basic statistic information of the designated data set:",
                            enter_prompt,
                        ),
                        "",
                    ),
                    InteractionStep(
                        f"engineered_feature_{index}_continue",
                        (
                            "Do you want to continue to build a new feature?",
                            "(Data) ➜ @Number:",
                        ),
                        "1" if index < len(engineered) else "2",
                    ),
                ]
            )
            if index < len(engineered):
                steps.append(
                    InteractionStep(
                        f"engineered_feature_{index}_continue_to_next",
                        (enter_prompt,),
                        "",
                    )
                )
        steps.append(
            InteractionStep(
                "continue_after_feature_engineering",
                (
                    "Data Selected Dropped-Imputed Feature-Engineering.xlsx",
                    enter_prompt,
                ),
                "",
            )
        )
        return steps

    @staticmethod
    def _label_steps(request: ClassificationRequest, class_count: int, enter_prompt: str) -> list[InteractionStep]:
        customization = request.label_customization
        strategy_number = {
            "encode_original": "1",
            "map": "2",
            "interval": "3",
            "quantile": "4",
        }[customization.strategy]
        step_id = "keep_and_encode_labels" if customization.strategy == "encode_original" else f"{customization.strategy}_labels"
        steps = [
            InteractionStep(
                step_id,
                (
                    "Classification Label Customization",
                    "Keep Original Labels and Encode",
                    "(Data) ➜ @Number:",
                ),
                strategy_number,
            )
        ]
        if customization.strategy == "map":
            mapping = "; ".join(f"{source}:{target}" for source, target in customization.mapping.items())
            steps.append(InteractionStep("label_mapping", ("Map every original label", "@Mapping:"), mapping))
        elif customization.strategy == "interval":
            steps.extend(
                [
                    InteractionStep(
                        "number_of_classes",
                        ("Number of Classes:", "(Data) ➜"),
                        str(len(customization.cut_points) + 1),
                    ),
                    InteractionStep(
                        "class_labels",
                        ("Input class labels separated by", "@Labels:"),
                        "; ".join(customization.labels or ()),
                    ),
                    InteractionStep(
                        "interval_cut_points",
                        ("internal cut points separated by", "@Cut Points:"),
                        "; ".join(format(value, ".15g") for value in customization.cut_points),
                    ),
                ]
            )
        elif customization.strategy == "quantile":
            steps.extend(
                [
                    InteractionStep(
                        "number_of_classes",
                        ("Number of Classes:", "(Data) ➜"),
                        str(customization.number_of_classes),
                    ),
                    InteractionStep(
                        "class_labels",
                        ("Input class labels separated by", "@Labels:"),
                        "; ".join(customization.labels or ()),
                    ),
                ]
            )
        if class_count > 2:
            steps.append(
                InteractionStep(
                    "metric_average",
                    (
                        "Please select calculation method for multiclass metrics",
                        "(Model) ➜ @Number:",
                    ),
                    _choice(request.metric_average, ("micro", "macro", "weighted")),
                )
            )
        steps.append(InteractionStep("continue_after_target", ("contiguous integer codes", enter_prompt), ""))
        return steps

    @staticmethod
    def _preprocessing_steps(request: ClassificationRequest, enter_prompt: str) -> list[InteractionStep]:
        steps: list[InteractionStep] = []
        if request.scaling == "none":
            steps.append(
                InteractionStep(
                    "skip_feature_scaling",
                    ("Feature Scaling on X Set", "(Data) ➜ @Number:"),
                    "2",
                )
            )
        else:
            steps.extend(
                [
                    InteractionStep(
                        "enable_feature_scaling",
                        ("Feature Scaling on X Set", "(Data) ➜ @Number:"),
                        "1",
                    ),
                    InteractionStep(
                        request.scaling,
                        ("Which strategy do you want to apply?", "(Data) ➜ @Number:"),
                        _choice(
                            request.scaling,
                            ("min_max", "standardization", "mean_normalization"),
                        ),
                    ),
                ]
            )
        scientific_execution_active = (
            getattr(request.model_selection, "mode", "single") != "all" and getattr(request, "tuning", "manual") == "manual" and _native_scientific_model_parameters(request) is not None
        )
        scaling_anchor = (
            "Scientific execution: scaler fitting is deferred until the model-fitting cohort is defined."
            if request.scaling != "none" and scientific_execution_active
            else "X With Scaling.xlsx"
            if request.scaling != "none"
            else "Feature Selection on X set"
        )
        steps.append(
            InteractionStep(
                "continue_after_scaling",
                (scaling_anchor, enter_prompt),
                "",
                timeout_seconds=180,
            )
        )
        if request.feature_selection.method == "none":
            steps.append(
                InteractionStep(
                    "skip_feature_selection",
                    ("Feature Selection on X set", "(Data) ➜ @Number:"),
                    "2",
                )
            )
        else:
            steps.extend(
                [
                    InteractionStep(
                        "enable_feature_selection",
                        ("Feature Selection on X set", "(Data) ➜ @Number:"),
                        "1",
                    ),
                    InteractionStep(
                        "feature_selection_method",
                        ("Which strategy do you want to apply?", "(Data) ➜ @Number:"),
                        _choice(
                            request.feature_selection.method,
                            ("generic_univariate", "select_k_best"),
                        ),
                    ),
                    InteractionStep(
                        "feature_selection_retain_count",
                        ("Please enter the number of features to retain", "@input:"),
                        str(request.feature_selection.retain_count),
                    ),
                ]
            )
        steps.append(
            InteractionStep(
                "continue_after_feature_selection",
                (enter_prompt,),
                "",
                timeout_seconds=180,
            )
        )
        if getattr(request, "task", None) == "regression" and request.evaluation.mode == "external_labeled":
            steps.append(
                InteractionStep(
                    "continue_after_external_training_scope",
                    (
                        "External labeled evaluation uses every prepared training row for model fitting.",
                        enter_prompt,
                    ),
                    "",
                    timeout_seconds=180,
                )
            )
        else:
            steps.extend(
                [
                    InteractionStep(
                        "default_test_ratio",
                        ("Data Split - Train Set and Test Set", "(Data) ➜ @Test Ratio:"),
                        _float_response(request.test_ratio, 0.2),
                    ),
                    InteractionStep(
                        "continue_after_split",
                        ("Y Test.xlsx", enter_prompt),
                        "",
                        timeout_seconds=180,
                    ),
                ]
            )
        return steps


class RegressionPlanCompiler:
    """Compile every supported single-model regression branch without importing ML code."""

    def compile(
        self,
        request: RegressionRequest,
        cli_executable: Optional[Path] = None,
        *,
        dataset_context: DatasetCompilationContext | None = None,
    ) -> InteractionPlan:
        dataset_context = dataset_context or DatasetCompilationContext()
        data_path = request.training_dataset_path.expanduser().resolve()
        if not data_path.is_file():
            raise PlanCompilationError(f"Training data file does not exist: {data_path}")
        columns = _read_dataset_columns(data_path, source=dataset_context.training_source)
        _validate_world_map(data_path, columns, request.world_map)
        target_columns = request.resolved_target_columns
        target_set = set(target_columns)
        requested_columns = (
            request.identifier_column,
            *target_columns,
            *request.feature_columns,
        )
        missing = sorted({column for column in requested_columns if column not in columns})
        if missing:
            raise PlanCompilationError(f"Requested columns are absent from the training dataset: {missing}")

        application_path = None
        external_evaluation_path = request.evaluation.evaluation_dataset_path if request.evaluation.mode == "external_labeled" else None
        if request.evaluation.mode == "external_labeled" and request.evaluation.evaluation_dataset is not None:
            raise PlanCompilationError("External evaluation dataset references must be resolved to a local path before CLI compilation.")
        if request.application_dataset_path is not None and external_evaluation_path is not None:
            raise PlanCompilationError("A distinct external evaluation dataset cannot also be used as ordinary application data.")
        selected_application_path = request.application_dataset_path or external_evaluation_path
        if selected_application_path is not None:
            application_path = selected_application_path.expanduser().resolve()
            if not application_path.is_file():
                raise PlanCompilationError(f"Application or external evaluation data file does not exist: {application_path}")
            application_columns = _read_dataset_columns(
                application_path,
                source=dataset_context.source_for("application"),
            )
            application_identifier = request.evaluation.external_identifier_column or request.identifier_column if external_evaluation_path is not None else request.identifier_column
            required_application = {application_identifier, *request.feature_columns}
            application_missing = sorted(required_application - set(application_columns))
            if application_missing:
                raise PlanCompilationError(f"Application dataset is missing required identifier or feature columns: {application_missing}")
            if external_evaluation_path is not None:
                evaluation_missing = sorted(target_set - set(application_columns))
                if evaluation_missing:
                    raise PlanCompilationError("External evaluation dataset is missing target columns: " f"{evaluation_missing}")
            _validate_application_dataset(application_path, application_columns, request)

        selected_names = tuple(column for column in columns if column in target_set or column in request.feature_columns)
        engineered = _compile_engineered_features(request, selected_names)
        profile = _scan_regression_training_dataset(data_path, columns, request)
        final_feature_names = (
            *request.feature_columns,
            *(name for name, _ in engineered),
        )
        final_feature_count = len(final_feature_names)
        selected_feature_count = getattr(request.feature_selection, "retain_count", final_feature_count)
        if len(target_columns) > 1 and request.feature_selection.method != "none":
            raise PlanCompilationError("Multiple-target regression cannot use the CLI's univariate feature-selection methods.")
        if request.feature_selection.method != "none" and selected_feature_count >= final_feature_count:
            raise PlanCompilationError(f"feature_selection.retain_count must be less than the {final_feature_count} input features because that is the CLI's enforced contract.")
        maximum_features = getattr(request.model, "maximum_features", None)
        if maximum_features is not None and maximum_features > selected_feature_count:
            raise PlanCompilationError(f"model.maximum_features={maximum_features} exceeds the {selected_feature_count} features available after preprocessing.")

        original_positions = {column: index + 1 for index, column in enumerate(columns)}
        selected_positions = {column: index + 1 for index, column in enumerate((*selected_names, *(name for name, _ in engineered)))}
        selected_expression = _selection_expression([original_positions[column] for column in selected_names])
        feature_expression = _selection_expression([selected_positions[column] for column in final_feature_names])
        target_expression = _selection_expression([selected_positions[column] for column in target_columns])
        executable = Path(cli_executable).expanduser().resolve() if cli_executable else resolve_public_cli_executable()
        if not executable.is_file():
            raise PlanCompilationError(f"CLI executable does not exist: {executable}")

        enter_prompt = "(Press Enter key to move forward)"
        steps: list[InteractionStep] = [
            *_experiment_steps(request, enter_prompt),
            InteractionStep(
                "identifier_column",
                ("output data identifier column", "(Data)"),
                str(original_positions[request.identifier_column]),
            ),
            InteractionStep("continue_after_identifier", (enter_prompt,), ""),
            InteractionStep("continue_after_map_skip", ("World Map Projection", enter_prompt), ""),
            InteractionStep(
                "selected_data_columns",
                (
                    "Data Selection",
                    "Select the data range you want to process.",
                    "@input:",
                ),
                selected_expression,
            ),
            InteractionStep(
                "continue_after_data_selection",
                ("Index - Column Name", enter_prompt),
                "",
            ),
            InteractionStep(
                "continue_after_selected_data_preview",
                ("The Selected Data Set:", enter_prompt),
                "",
            ),
            InteractionStep(
                "continue_after_statistics",
                ("Data Selected.xlsx", enter_prompt),
                "",
                timeout_seconds=240,
            ),
        ]
        steps.extend(ClassificationPlanCompiler._missing_steps(request, profile, selected_positions, enter_prompt))
        steps.extend(ClassificationPlanCompiler._feature_engineering_steps(engineered, enter_prompt))
        steps.extend(
            [
                InteractionStep(
                    "regression_mode",
                    ("Mode Selection", "1 - Regression", "(Model)"),
                    "1",
                ),
                InteractionStep("continue_after_mode", (enter_prompt,), ""),
                InteractionStep(
                    "feature_columns",
                    (
                        "Selected sub data set to create X data set",
                        "Select the data range you want to process.",
                        "@input:",
                    ),
                    feature_expression,
                ),
                InteractionStep(
                    "continue_after_features",
                    ("X Without Scaling.xlsx", enter_prompt),
                    "",
                    timeout_seconds=180,
                ),
                InteractionStep(
                    "target_columns" if len(target_columns) > 1 else "target_column",
                    (
                        "The selected Y data set",
                        "Select the data range you want to process.",
                        "@input:",
                    ),
                    target_expression,
                ),
                InteractionStep("continue_after_target", ("Y.xlsx", enter_prompt), ""),
            ]
        )
        steps.extend(ClassificationPlanCompiler._preprocessing_steps(request, enter_prompt))
        model_number = 1 if profile.unresolved_missing_columns else REGRESSION_MODEL_NUMBERS[request.model.type]
        steps.extend(
            [
                InteractionStep(
                    request.model.type,
                    ("Which model do you want to apply?", "(Model)"),
                    str(model_number),
                ),
                InteractionStep("continue_after_model", (enter_prompt,), ""),
            ]
        )
        if request.model.type not in REGRESSION_MODELS_WITHOUT_AUTOML:
            steps.extend(
                [
                    InteractionStep(
                        ("enable_automl" if request.tuning == "automl" else "disable_automl"),
                        ("automated machine learning", "(Model)"),
                        "1" if request.tuning == "automl" else "2",
                    ),
                    InteractionStep("continue_after_automl", (enter_prompt,), ""),
                ]
            )
        if application_path is None:
            steps.append(
                InteractionStep(
                    "continue_after_inference_skip",
                    ("You did not provide application data.", enter_prompt),
                    "",
                )
            )
        else:
            application_output = "Application Data Feature-Engineering Selected.xlsx" if request.engineered_features else "Application Data Selected.xlsx"
            steps.append(
                InteractionStep(
                    "continue_after_application_preparation",
                    (application_output, enter_prompt),
                    "",
                    timeout_seconds=180,
                )
            )
        if request.tuning == "manual":
            steps.extend(_regression_model_steps(request))
            steps.append(InteractionStep("continue_after_hyperparameters", (enter_prompt,), ""))
        model_display = REGRESSION_MODEL_DISPLAY_NAMES[request.model.type]
        if request.model.type in REGRESSION_MODELS_WITH_INTERACTIVE_PLOT_SELECTION and selected_feature_count >= 2:
            steps.append(
                InteractionStep(
                    "one_dimensional_plot_feature",
                    (
                        "1 Dimensions Data Selection",
                        "Choose dimension - 1 data:",
                        "(Plot)",
                    ),
                    "1",
                    timeout_seconds=600,
                )
            )
            if selected_feature_count > 2:
                steps.extend(
                    [
                        InteractionStep(
                            "two_dimensional_plot_feature_1",
                            (
                                "2 Dimensions Data Selection",
                                "Choose dimension - 1 data:",
                                "(Plot)",
                            ),
                            "1",
                        ),
                        InteractionStep(
                            "two_dimensional_plot_feature_2",
                            ("Choose dimension - 2 data:", "(Plot)"),
                            "2",
                        ),
                    ]
                )
        steps.extend(
            [
                InteractionStep(
                    "continue_after_training",
                    (f"{model_display}.joblib", enter_prompt),
                    "",
                    timeout_seconds=600,
                ),
                InteractionStep(
                    "continue_after_transform_pipeline",
                    ("Transform Pipeline.joblib", enter_prompt),
                    "",
                    timeout_seconds=180,
                ),
            ]
        )
        if application_path is not None:
            steps.append(
                InteractionStep(
                    "continue_after_inference",
                    ("Application Data Predicted.xlsx", enter_prompt),
                    "",
                    timeout_seconds=300,
                )
            )

        command = (str(executable), "data-mining", "--data", str(data_path))
        if application_path is not None:
            command = (
                str(executable),
                "data-mining",
                "--training",
                str(data_path),
                "--application",
                str(application_path),
            )
        command = _command_with_analysis_options(command, request.world_map, request.existing_experiment_id)
        output_root = Path("geopi_output") / request.experiment_name / request.run_name
        external_labeled_evaluation = request.evaluation.mode == "external_labeled"
        expected_outputs = [
            str(output_root / "artifacts" / "model" / f"{model_display}.joblib"),
            str(output_root / "artifacts" / "model" / "Transform Pipeline.joblib"),
        ]
        if external_labeled_evaluation:
            expected_outputs.extend(
                (
                    str(output_root / "parameters" / f"Hyper Parameters - {model_display}.txt"),
                    str(output_root / "artifacts" / "data" / "Application Data Predicted.xlsx"),
                )
            )
        else:
            expected_outputs.extend(
                (
                    str(output_root / "artifacts" / "image" / "model_output" / f"Predicted vs. Actual Diagram - {model_display}.png"),
                    str(output_root / "metrics" / f"Model Score - {model_display}.txt"),
                    str(output_root / "artifacts" / "data" / "Y Test Predict.xlsx"),
                    str(output_root / "artifacts" / "image" / "model_output" / f"Residuals Diagram - {model_display}.png"),
                    str(output_root / "artifacts" / "image" / "model_output" / f"Residuals Diagram - {model_display}.xlsx"),
                    str(output_root / "parameters" / f"Hyper Parameters - {model_display}.txt"),
                )
            )
            if application_path is not None:
                expected_outputs.append(str(output_root / "artifacts" / "data" / "Application Data Predicted.xlsx"))
        if request.model.type in REGRESSION_MODELS_WITH_FEATURE_IMPORTANCE:
            expected_outputs.extend(
                (
                    str(output_root / "artifacts" / "image" / "model_output" / f"Feature Importance Diagram - {model_display}.png"),
                    str(output_root / "artifacts" / "image" / "model_output" / f"Feature Importance Diagram - {model_display}.xlsx"),
                )
            )
        return InteractionPlan(
            schema_version=INTERACTION_PLAN_VERSION,
            name=(f"regression-{request.model.type}-multi-output-v1" if len(target_columns) > 1 else f"regression-{request.model.type}-v1"),
            public_command=command,
            steps=tuple(steps),
            expected_output_relative_paths=tuple(expected_outputs),
            requires_source_row_pairing=True,
        )


class ClusteringPlanCompiler:
    """Compile every public single-model clustering branch without importing ML code."""

    def compile(
        self,
        request: ClusteringRequest,
        cli_executable: Optional[Path] = None,
        *,
        dataset_context: DatasetCompilationContext | None = None,
    ) -> InteractionPlan:
        dataset_context = dataset_context or DatasetCompilationContext()
        data_path = request.training_dataset_path.expanduser().resolve()
        if not data_path.is_file():
            raise PlanCompilationError(f"Training data file does not exist: {data_path}")
        columns = _read_dataset_columns(data_path, source=dataset_context.training_source)
        _validate_world_map(data_path, columns, request.world_map)
        requested_columns = (request.identifier_column, *request.feature_columns)
        missing = sorted({column for column in requested_columns if column not in columns})
        if missing:
            raise PlanCompilationError(f"Requested columns are absent from the training dataset: {missing}")

        selected_names = tuple(column for column in columns if column in request.feature_columns)
        engineered = _compile_engineered_features(request, selected_names)
        profile = _scan_clustering_training_dataset(data_path, columns, request)
        final_feature_names = (
            *selected_names,
            *(name for name, _ in engineered),
        )
        final_feature_count = len(final_feature_names)
        if final_feature_count < 2:
            raise PlanCompilationError("The public clustering workflow requires at least 2 final features to produce its mandatory silhouette and cluster diagrams.")
        if request.model.type == "affinity_propagation" and request.model.affinity == "precomputed" and final_feature_count != profile.row_count:
            raise PlanCompilationError("AffinityPropagation affinity='precomputed' requires a square feature matrix: the final feature count must equal the retained row count.")

        original_positions = {column: index + 1 for index, column in enumerate(columns)}
        selected_positions = {column: index + 1 for index, column in enumerate((*selected_names, *(name for name, _ in engineered)))}
        selected_expression = _selection_expression([original_positions[column] for column in selected_names])
        executable = Path(cli_executable).expanduser().resolve() if cli_executable else resolve_public_cli_executable()
        if not executable.is_file():
            raise PlanCompilationError(f"CLI executable does not exist: {executable}")

        enter_prompt = "(Press Enter key to move forward)"
        steps: list[InteractionStep] = [
            *_experiment_steps(request, enter_prompt),
            InteractionStep(
                "identifier_column",
                ("output data identifier column", "(Data)"),
                str(original_positions[request.identifier_column]),
            ),
            InteractionStep("continue_after_identifier", (enter_prompt,), ""),
            InteractionStep(
                "continue_after_map_skip",
                ("World Map Projection", enter_prompt),
                "",
            ),
            InteractionStep(
                "selected_data_columns",
                (
                    "Data Selection",
                    "Select the data range you want to process.",
                    "@input:",
                ),
                selected_expression,
            ),
            InteractionStep(
                "continue_after_data_selection",
                ("Index - Column Name", enter_prompt),
                "",
            ),
            InteractionStep(
                "continue_after_selected_data_preview",
                ("The Selected Data Set:", enter_prompt),
                "",
            ),
            InteractionStep(
                "continue_after_statistics",
                ("Data Selected.xlsx", enter_prompt),
                "",
                timeout_seconds=240,
            ),
        ]
        steps.extend(
            ClassificationPlanCompiler._missing_steps(
                request,
                profile,
                selected_positions,
                enter_prompt,
            )
        )
        steps.extend(
            ClassificationPlanCompiler._feature_engineering_steps(
                engineered,
                enter_prompt,
            )
        )
        steps.extend(
            [
                InteractionStep(
                    "clustering_mode",
                    ("Mode Selection", "3 - Clustering", "(Model)"),
                    "3",
                ),
                InteractionStep("continue_after_mode", (enter_prompt,), ""),
            ]
        )
        if request.scaling == "none":
            steps.extend(
                [
                    InteractionStep(
                        "skip_feature_scaling",
                        ("Feature Scaling on X Set", "(Data) ➜ @Number:"),
                        "2",
                    ),
                    InteractionStep("continue_after_scaling", (enter_prompt,), ""),
                ]
            )
        else:
            steps.extend(
                [
                    InteractionStep(
                        "enable_feature_scaling",
                        ("Feature Scaling on X Set", "(Data) ➜ @Number:"),
                        "1",
                    ),
                    InteractionStep(
                        request.scaling,
                        (
                            "Which strategy do you want to apply?",
                            "(Data) ➜ @Number:",
                        ),
                        _choice(
                            request.scaling,
                            ("min_max", "standardization", "mean_normalization"),
                        ),
                    ),
                    InteractionStep(
                        "continue_after_scaling",
                        ("X With Scaling.xlsx", enter_prompt),
                        "",
                        timeout_seconds=180,
                    ),
                ]
            )

        model_display = CLUSTERING_MODEL_DISPLAY_NAMES[request.model.type]
        steps.extend(
            [
                InteractionStep(
                    request.model.type,
                    ("Which model do you want to apply?", "(Model)"),
                    str(CLUSTERING_MODEL_NUMBERS[request.model.type]),
                ),
                InteractionStep("continue_after_model", (enter_prompt,), ""),
            ]
        )
        steps.extend(_clustering_model_steps(request))
        steps.append(InteractionStep("continue_after_hyperparameters", (enter_prompt,), ""))
        if final_feature_count >= 3:
            steps.extend(
                [
                    InteractionStep(
                        "clustering_plot_2d_feature_1",
                        (
                            "2 Dimensions Data Selection",
                            "Choose dimension - 1 data:",
                            "(Plot)",
                        ),
                        "1",
                        timeout_seconds=600,
                    ),
                    InteractionStep(
                        "clustering_plot_2d_feature_2",
                        ("Choose dimension - 2 data:", "(Plot)"),
                        "2",
                    ),
                    InteractionStep(
                        "clustering_plot_3d_feature_1",
                        (
                            "3 Dimensions Data Selection",
                            "Choose dimension - 1 data:",
                            "(Plot)",
                        ),
                        "1",
                    ),
                    InteractionStep(
                        "clustering_plot_3d_feature_2",
                        ("Choose dimension - 2 data:", "(Plot)"),
                        "2",
                    ),
                    InteractionStep(
                        "clustering_plot_3d_feature_3",
                        ("Choose dimension - 3 data:", "(Plot)"),
                        "3",
                    ),
                ]
            )
        steps.append(
            InteractionStep(
                "continue_after_training",
                (f"{model_display}.joblib", enter_prompt),
                "",
                timeout_seconds=600,
            )
        )
        transform_pipeline_expected = request.scaling != "none" or request.missing_values.method == "impute"
        steps.append(
            InteractionStep(
                "continue_after_transform_pipeline",
                (
                    ("Transform Pipeline.joblib" if transform_pipeline_expected else "Transform Pipeline Configuration.txt"),
                    enter_prompt,
                ),
                "",
                timeout_seconds=180,
            )
        )

        base = Path("geopi_output") / request.experiment_name / request.run_name
        expected_outputs = [
            str(base / "artifacts" / "model" / f"{model_display}.joblib"),
            str(base / "artifacts" / "data" / f"Cluster Labels - {model_display}.xlsx"),
            str(base / "metrics" / f"Model Score - {model_display}.txt"),
            str(base / "artifacts" / "image" / "model_output" / f"Cluster Two-Dimensional Diagram - {model_display}.png"),
            str(base / "artifacts" / "Transform Pipeline Configuration.txt"),
        ]
        if final_feature_count >= 3:
            expected_outputs.append(str(base / "artifacts" / "image" / "model_output" / f"Cluster Three-Dimensional Diagram - {model_display}.png"))
        if transform_pipeline_expected:
            expected_outputs.append(str(base / "artifacts" / "model" / "Transform Pipeline.joblib"))
        return InteractionPlan(
            schema_version=INTERACTION_PLAN_VERSION,
            name=f"clustering-{request.model.type}-v1",
            public_command=_command_with_analysis_options(
                (
                    str(executable),
                    "data-mining",
                    "--data",
                    str(data_path),
                ),
                request.world_map,
                request.existing_experiment_id,
            ),
            steps=tuple(steps),
            expected_output_relative_paths=tuple(expected_outputs),
            requires_source_row_pairing=True,
        )


class DecompositionPlanCompiler:
    """Compile every public single-model dimensional-reduction branch."""

    def compile(
        self,
        request: DecompositionRequest,
        cli_executable: Optional[Path] = None,
        *,
        dataset_context: DatasetCompilationContext | None = None,
    ) -> InteractionPlan:
        dataset_context = dataset_context or DatasetCompilationContext()
        if request.mode == "embedding_label_overlay":
            coordinate_path = request.training_dataset_path.expanduser().resolve()
            label_path = request.application_dataset_path.expanduser().resolve()
            if not coordinate_path.is_file():
                raise PlanCompilationError(f"Coordinate data file does not exist: {coordinate_path}")
            if not label_path.is_file():
                raise PlanCompilationError(f"Label data file does not exist: {label_path}")
            coordinate_columns = _read_dataset_columns(
                coordinate_path,
                source=dataset_context.training_source,
                sheet=request.coordinate_sheet,
            )
            label_columns = _read_dataset_columns(
                label_path,
                source=dataset_context.application_source or "path",
                sheet=request.label_sheet,
            )
            missing_coordinate_columns = sorted({request.identifier_column, *request.feature_columns} - set(coordinate_columns))
            if missing_coordinate_columns:
                raise PlanCompilationError("Embedding coordinate columns are absent from the coordinate dataset: " f"{missing_coordinate_columns}")
            missing_label_columns = sorted({request.label_identifier_column, request.label_column} - set(label_columns))
            if missing_label_columns:
                raise PlanCompilationError("Embedding label columns are absent from the label dataset: " f"{missing_label_columns}")
            executable = Path(cli_executable).expanduser().resolve() if cli_executable else resolve_public_cli_executable()
            if not executable.is_file():
                raise PlanCompilationError(f"CLI executable does not exist: {executable}")
            command = [
                str(executable),
                "embedding-label-overlay",
                "--coordinates",
                str(coordinate_path),
                "--labels",
                str(label_path),
                "--coordinate-sheet",
                request.coordinate_sheet,
                "--label-sheet",
                request.label_sheet,
                "--coordinate-identifier-column",
                request.identifier_column,
                "--label-identifier-column",
                request.label_identifier_column,
                "--x-column",
                request.feature_columns[0],
                "--y-column",
                request.feature_columns[1],
                "--label-column",
                request.label_column,
                "--experiment-name",
                request.experiment_name,
                "--run-name",
                request.run_name,
            ]
            for value in request.positive_label_values:
                command.extend(("--positive-label-value", value))
            base = Path(request.experiment_name) / request.run_name
            return InteractionPlan(
                schema_version=INTERACTION_PLAN_VERSION,
                name="decomposition-embedding-label-overlay-v1",
                public_command=tuple(command),
                steps=(),
                expected_output_relative_paths=(
                    (base / "artifacts" / "data" / "Embedding Label Overlay.csv").as_posix(),
                    (base / "artifacts" / "image" / "model_output" / "Embedding Label Overlay.png").as_posix(),
                    (base / "artifacts" / "image" / "model_output" / "Embedding Label Overlay.pdf").as_posix(),
                    (base / "metrics" / "Embedding Label Overlay Counts.json").as_posix(),
                    (base / "parameters" / "Embedding Label Overlay Parameters.json").as_posix(),
                    (base / "summary" / "Embedding Label Overlay Artifact Index.json").as_posix(),
                    (base / "summary" / "Embedding Label Overlay Manifest.json").as_posix(),
                ),
                workflow_family="artifact_composition",
                workflow_mode="embedding_label_overlay",
                method="exact_identifier_join",
                adapter_id="geochemistrypi-cli.artifact-composition.embedding-label-overlay",
                adapter_version="1",
                required_cli_capabilities=("command:embedding-label-overlay",),
            )
        data_path = request.training_dataset_path.expanduser().resolve()
        if not data_path.is_file():
            raise PlanCompilationError(f"Training data file does not exist: {data_path}")
        columns = _read_dataset_columns(data_path, source=dataset_context.training_source)
        _validate_world_map(data_path, columns, request.world_map)
        requested_columns = (request.identifier_column, *request.feature_columns)
        missing = sorted({column for column in requested_columns if column not in columns})
        if missing:
            raise PlanCompilationError(f"Requested columns are absent from the training dataset: {missing}")

        selected_names = tuple(column for column in columns if column in request.feature_columns)
        engineered = _compile_engineered_features(request, selected_names)
        profile = _scan_decomposition_training_dataset(data_path, columns, request)
        final_feature_names = (
            *selected_names,
            *(name for name, _ in engineered),
        )
        final_feature_count = len(final_feature_names)
        if final_feature_count < 2:
            raise PlanCompilationError("The public decomposition workflow requires at least 2 final features for its mandatory diagrams.")
        if request.model.type == "pca":
            component_limit = min(profile.row_count, final_feature_count)
            if request.model.number_of_components > component_limit:
                raise PlanCompilationError(f"PCA number_of_components={request.model.number_of_components} exceeds min(retained rows, final features)={component_limit}.")
            if request.model.svd_solver == "arpack" and request.model.number_of_components >= component_limit:
                raise PlanCompilationError("PCA svd_solver='arpack' requires number_of_components to be strictly less than min(retained rows, final features).")
        if request.model.type == "tsne" and request.model.perplexity >= profile.row_count:
            raise PlanCompilationError(f"T-SNE perplexity={request.model.perplexity} must be less than the {profile.row_count} retained rows.")

        original_positions = {column: index + 1 for index, column in enumerate(columns)}
        selected_positions = {column: index + 1 for index, column in enumerate((*selected_names, *(name for name, _ in engineered)))}
        selected_expression = _selection_expression([original_positions[column] for column in selected_names])
        executable = Path(cli_executable).expanduser().resolve() if cli_executable else resolve_public_cli_executable()
        if not executable.is_file():
            raise PlanCompilationError(f"CLI executable does not exist: {executable}")

        enter_prompt = "(Press Enter key to move forward)"
        steps: list[InteractionStep] = [
            *_experiment_steps(request, enter_prompt),
            InteractionStep(
                "identifier_column",
                ("output data identifier column", "(Data)"),
                str(original_positions[request.identifier_column]),
            ),
            InteractionStep("continue_after_identifier", (enter_prompt,), ""),
            InteractionStep(
                "continue_after_map_skip",
                ("World Map Projection", enter_prompt),
                "",
            ),
            InteractionStep(
                "selected_data_columns",
                (
                    "Data Selection",
                    "Select the data range you want to process.",
                    "@input:",
                ),
                selected_expression,
            ),
            InteractionStep(
                "continue_after_data_selection",
                ("Index - Column Name", enter_prompt),
                "",
            ),
            InteractionStep(
                "continue_after_selected_data_preview",
                ("The Selected Data Set:", enter_prompt),
                "",
            ),
            InteractionStep(
                "continue_after_statistics",
                ("Data Selected.xlsx", enter_prompt),
                "",
                timeout_seconds=240,
            ),
        ]
        steps.extend(
            ClassificationPlanCompiler._missing_steps(
                request,
                profile,
                selected_positions,
                enter_prompt,
            )
        )
        steps.extend(
            ClassificationPlanCompiler._feature_engineering_steps(
                engineered,
                enter_prompt,
            )
        )
        steps.extend(
            [
                InteractionStep(
                    "decomposition_mode",
                    ("Mode Selection", "4 - Dimensional Reduction", "(Model)"),
                    "4",
                ),
                InteractionStep("continue_after_mode", (enter_prompt,), ""),
            ]
        )
        if request.scaling == "none":
            steps.extend(
                [
                    InteractionStep(
                        "skip_feature_scaling",
                        ("Feature Scaling on X Set", "(Data) ➜ @Number:"),
                        "2",
                    ),
                    InteractionStep("continue_after_scaling", (enter_prompt,), ""),
                ]
            )
        else:
            steps.extend(
                [
                    InteractionStep(
                        "enable_feature_scaling",
                        ("Feature Scaling on X Set", "(Data) ➜ @Number:"),
                        "1",
                    ),
                    InteractionStep(
                        request.scaling,
                        ("Which strategy do you want to apply?", "(Data) ➜ @Number:"),
                        _choice(
                            request.scaling,
                            ("min_max", "standardization", "mean_normalization"),
                        ),
                    ),
                    InteractionStep(
                        "continue_after_scaling",
                        ("X With Scaling.xlsx", enter_prompt),
                        "",
                        timeout_seconds=180,
                    ),
                ]
            )

        model_display = DECOMPOSITION_MODEL_DISPLAY_NAMES[request.model.type]
        steps.extend(
            [
                InteractionStep(
                    request.model.type,
                    ("Which model do you want to apply?", "(Model)"),
                    str(DECOMPOSITION_MODEL_NUMBERS[request.model.type]),
                ),
                InteractionStep("continue_after_model", (enter_prompt,), ""),
            ]
        )
        steps.extend(_decomposition_model_steps(request))
        steps.append(InteractionStep("continue_after_hyperparameters", (enter_prompt,), ""))
        if request.model.type == "pca" and request.model.number_of_components >= 3:
            steps.extend(
                [
                    InteractionStep(
                        "pca_biplot_component_1",
                        (
                            "2 Dimensions Data Selection",
                            "Choose dimension - 1 data:",
                            "(Plot)",
                        ),
                        "1",
                        timeout_seconds=600,
                    ),
                    InteractionStep(
                        "pca_biplot_component_2",
                        ("Choose dimension - 2 data:", "(Plot)"),
                        "2",
                    ),
                ]
            )
        if request.model.type == "pca" and request.model.number_of_components > 3:
            steps.extend(
                [
                    InteractionStep(
                        "pca_triplot_component_1",
                        (
                            "3 Dimensions Data Selection",
                            "Choose dimension - 1 data:",
                            "(Plot)",
                        ),
                        "1",
                    ),
                    InteractionStep(
                        "pca_triplot_component_2",
                        ("Choose dimension - 2 data:", "(Plot)"),
                        "2",
                    ),
                    InteractionStep(
                        "pca_triplot_component_3",
                        ("Choose dimension - 3 data:", "(Plot)"),
                        "3",
                    ),
                ]
            )
        steps.append(
            InteractionStep(
                "continue_after_training",
                (f"{model_display}.joblib", enter_prompt),
                "",
                timeout_seconds=600,
            )
        )
        transform_pipeline_expected = request.scaling != "none" or request.missing_values.method == "impute"
        steps.append(
            InteractionStep(
                "continue_after_transform_pipeline",
                (
                    ("Transform Pipeline.joblib" if transform_pipeline_expected else "Transform Pipeline Configuration.txt"),
                    enter_prompt,
                ),
                "",
                timeout_seconds=180,
            )
        )

        base = Path("geopi_output") / request.experiment_name / request.run_name
        expected_outputs = [
            str(base / "artifacts" / "model" / f"{model_display}.joblib"),
            str(base / "artifacts" / "data" / "X Reduced.xlsx"),
            str(base / "artifacts" / "image" / "model_output" / f"Decomposition Two-Dimensional Diagram - {model_display}.png"),
            str(base / "artifacts" / "image" / "model_output" / f"Decomposition Heatmap - {model_display}.png"),
            str(base / "artifacts" / "image" / "model_output" / f"Dimensionality Reduction Contour Plot - {model_display}.png"),
            str(base / "artifacts" / "Transform Pipeline Configuration.txt"),
        ]
        if request.model.type == "pca" and request.model.number_of_components >= 2:
            expected_outputs.extend(
                [
                    str(base / "artifacts" / "image" / "model_output" / "Compositional Bi-plot - PCA.png"),
                    str(base / "artifacts" / "image" / "model_output" / "Compositional Bi-plot - PC Data.xlsx"),
                ]
            )
        if request.model.type == "pca" and request.model.number_of_components >= 3:
            expected_outputs.append(str(base / "artifacts" / "image" / "model_output" / "Compositional Tri-plot - PCA.png"))
        if transform_pipeline_expected:
            expected_outputs.append(str(base / "artifacts" / "model" / "Transform Pipeline.joblib"))
        return InteractionPlan(
            schema_version=INTERACTION_PLAN_VERSION,
            name=f"decomposition-{request.model.type}-v1",
            public_command=_command_with_analysis_options(
                (str(executable), "data-mining", "--data", str(data_path)),
                request.world_map,
                request.existing_experiment_id,
            ),
            steps=tuple(steps),
            expected_output_relative_paths=tuple(expected_outputs),
            requires_source_row_pairing=True,
        )


class AnomalyDetectionPlanCompiler:
    """Compile every public single-model anomaly-detection branch."""

    def compile(
        self,
        request: AnomalyDetectionRequest,
        cli_executable: Optional[Path] = None,
        *,
        dataset_context: DatasetCompilationContext | None = None,
    ) -> InteractionPlan:
        dataset_context = dataset_context or DatasetCompilationContext()
        data_path = request.training_dataset_path.expanduser().resolve()
        if not data_path.is_file():
            raise PlanCompilationError(f"Training data file does not exist: {data_path}")
        columns = _read_dataset_columns(data_path, source=dataset_context.training_source)
        _validate_world_map(data_path, columns, request.world_map)
        requested_columns = (request.identifier_column, *request.feature_columns)
        missing = sorted({column for column in requested_columns if column not in columns})
        if missing:
            raise PlanCompilationError(f"Requested columns are absent from the training dataset: {missing}")

        selected_names = tuple(column for column in columns if column in request.feature_columns)
        engineered = _compile_engineered_features(request, selected_names)
        profile = _scan_anomaly_detection_training_dataset(
            data_path,
            columns,
            request,
        )
        final_feature_names = (
            *selected_names,
            *(name for name, _ in engineered),
        )
        final_feature_count = len(final_feature_names)
        if request.model.type == "isolation_forest":
            if request.model.maximum_features > final_feature_count:
                raise PlanCompilationError(f"maximum_features={request.model.maximum_features} exceeds " f"the {final_feature_count} final features.")
            if request.model.maximum_samples is not None and request.model.maximum_samples > profile.row_count:
                raise PlanCompilationError(f"maximum_samples={request.model.maximum_samples} exceeds " f"the {profile.row_count} retained rows.")
        elif request.model.number_of_neighbors >= profile.row_count:
            raise PlanCompilationError(f"number_of_neighbors={request.model.number_of_neighbors} must be " f"less than the {profile.row_count} retained rows.")

        original_positions = {column: index + 1 for index, column in enumerate(columns)}
        selected_positions = {column: index + 1 for index, column in enumerate((*selected_names, *(name for name, _ in engineered)))}
        selected_expression = _selection_expression([original_positions[column] for column in selected_names])
        executable = Path(cli_executable).expanduser().resolve() if cli_executable else resolve_public_cli_executable()
        if not executable.is_file():
            raise PlanCompilationError(f"CLI executable does not exist: {executable}")

        enter_prompt = "(Press Enter key to move forward)"
        steps: list[InteractionStep] = [
            *_experiment_steps(request, enter_prompt),
            InteractionStep(
                "identifier_column",
                ("output data identifier column", "(Data)"),
                str(original_positions[request.identifier_column]),
            ),
            InteractionStep("continue_after_identifier", (enter_prompt,), ""),
            InteractionStep(
                "continue_after_map_skip",
                ("World Map Projection", enter_prompt),
                "",
            ),
            InteractionStep(
                "selected_data_columns",
                (
                    "Data Selection",
                    "Select the data range you want to process.",
                    "@input:",
                ),
                selected_expression,
            ),
            InteractionStep(
                "continue_after_data_selection",
                ("Index - Column Name", enter_prompt),
                "",
            ),
            InteractionStep(
                "continue_after_selected_data_preview",
                ("The Selected Data Set:", enter_prompt),
                "",
            ),
            InteractionStep(
                "continue_after_statistics",
                ("Data Selected.xlsx", enter_prompt),
                "",
                timeout_seconds=240,
            ),
        ]
        steps.extend(
            ClassificationPlanCompiler._missing_steps(
                request,
                profile,
                selected_positions,
                enter_prompt,
            )
        )
        steps.extend(
            ClassificationPlanCompiler._feature_engineering_steps(
                engineered,
                enter_prompt,
            )
        )
        steps.extend(
            [
                InteractionStep(
                    "anomaly_detection_mode",
                    ("Mode Selection", "5 - Anomaly Detection", "(Model)"),
                    "5",
                ),
                InteractionStep("continue_after_mode", (enter_prompt,), ""),
            ]
        )
        if request.scaling == "none":
            steps.extend(
                [
                    InteractionStep(
                        "skip_feature_scaling",
                        ("Feature Scaling on X Set", "@Number:"),
                        "2",
                    ),
                    InteractionStep(
                        "continue_after_scaling",
                        (enter_prompt,),
                        "",
                    ),
                ]
            )
        else:
            steps.extend(
                [
                    InteractionStep(
                        "enable_feature_scaling",
                        ("Feature Scaling on X Set", "@Number:"),
                        "1",
                    ),
                    InteractionStep(
                        request.scaling,
                        ("Which strategy do you want to apply?", "@Number:"),
                        _choice(
                            request.scaling,
                            (
                                "min_max",
                                "standardization",
                                "mean_normalization",
                            ),
                        ),
                    ),
                    InteractionStep(
                        "continue_after_scaling",
                        ("X With Scaling.xlsx", enter_prompt),
                        "",
                        timeout_seconds=180,
                    ),
                ]
            )

        model_display = ANOMALY_DETECTION_MODEL_DISPLAY_NAMES[request.model.type]
        steps.extend(
            [
                InteractionStep(
                    request.model.type,
                    ("Which model do you want to apply?", "(Model)"),
                    str(ANOMALY_DETECTION_MODEL_NUMBERS[request.model.type]),
                ),
                InteractionStep("continue_after_model", (enter_prompt,), ""),
            ]
        )
        steps.extend(_anomaly_detection_model_steps(request))
        steps.append(
            InteractionStep(
                "continue_after_hyperparameters",
                (enter_prompt,),
                "",
            )
        )
        if final_feature_count >= 3:
            steps.extend(
                [
                    InteractionStep(
                        "anomaly_plot_2d_feature_1",
                        (
                            "2 Dimensions Data Selection",
                            "Choose dimension - 1 data:",
                            "(Plot)",
                        ),
                        "1",
                        timeout_seconds=600,
                    ),
                    InteractionStep(
                        "anomaly_plot_2d_feature_2",
                        ("Choose dimension - 2 data:", "(Plot)"),
                        "2",
                    ),
                    InteractionStep(
                        "anomaly_plot_3d_feature_1",
                        (
                            "3 Dimensions Data Selection",
                            "Choose dimension - 1 data:",
                            "(Plot)",
                        ),
                        "1",
                    ),
                    InteractionStep(
                        "anomaly_plot_3d_feature_2",
                        ("Choose dimension - 2 data:", "(Plot)"),
                        "2",
                    ),
                    InteractionStep(
                        "anomaly_plot_3d_feature_3",
                        ("Choose dimension - 3 data:", "(Plot)"),
                        "3",
                    ),
                ]
            )
        steps.append(
            InteractionStep(
                "continue_after_training",
                (f"{model_display}.joblib", enter_prompt),
                "",
                timeout_seconds=600,
            )
        )
        transform_pipeline_expected = request.scaling != "none" or request.missing_values.method == "impute"
        steps.append(
            InteractionStep(
                "continue_after_transform_pipeline",
                (
                    ("Transform Pipeline.joblib" if transform_pipeline_expected else "Transform Pipeline Configuration.txt"),
                    enter_prompt,
                ),
                "",
                timeout_seconds=180,
            )
        )

        base = Path("geopi_output") / request.experiment_name / request.run_name
        model_output = base / "artifacts" / "image" / "model_output"
        expected_outputs = [
            str(base / "artifacts" / "model" / f"{model_display}.joblib"),
            str(base / "artifacts" / "data" / "X Abnormal Detection.xlsx"),
            str(base / "artifacts" / "data" / "X Normal.xlsx"),
            str(base / "artifacts" / "data" / "X Abnormal.xlsx"),
            str(model_output / f"Anomaly Detection Density Estimation - {model_display}.png"),
            str(model_output / f"Anomaly Detection Density Estimation - {model_display}.xlsx"),
            str(base / "artifacts" / "Transform Pipeline Configuration.txt"),
        ]
        if final_feature_count >= 3:
            expected_outputs.extend(
                [
                    str(model_output / f"Anomaly Detection Two-Dimensional Diagram - {model_display}.png"),
                    str(model_output / f"Anomaly Detection Two-Dimensional Diagram - {model_display}.xlsx"),
                    str(model_output / f"Anomaly Detection Three-Dimensional Diagram - {model_display}.png"),
                    str(model_output / f"Anomaly Detection Three-Dimensional Diagram - {model_display}.xlsx"),
                ]
            )
        if request.model.type == "local_outlier_factor":
            expected_outputs.extend(
                [
                    str(model_output / "Lof Score Diagram - Local Outlier Factor.png"),
                    str(model_output / "Lof Score Diagram - Local Outlier Factor.xlsx"),
                ]
            )
        if transform_pipeline_expected:
            expected_outputs.append(str(base / "artifacts" / "model" / "Transform Pipeline.joblib"))
        return InteractionPlan(
            schema_version=INTERACTION_PLAN_VERSION,
            name=f"anomaly-detection-{request.model.type}-v1",
            public_command=_command_with_analysis_options(
                (
                    str(executable),
                    "data-mining",
                    "--data",
                    str(data_path),
                ),
                request.world_map,
                request.existing_experiment_id,
            ),
            steps=tuple(steps),
            expected_output_relative_paths=tuple(expected_outputs),
            requires_source_row_pairing=True,
        )


class TimeSeriesPlanCompiler:
    """Compile the noninteractive, seeded Time Series production command."""

    def compile(
        self,
        request: TimeSeriesRequest,
        cli_executable: Optional[Path] = None,
        *,
        dataset_context: DatasetCompilationContext | None = None,
    ) -> InteractionPlan:
        dataset_context = dataset_context or DatasetCompilationContext()
        data_path = request.training_dataset_path.expanduser().resolve()
        if not data_path.is_file():
            raise PlanCompilationError(f"Training data file does not exist: {data_path}")
        columns = _read_dataset_columns(
            data_path,
            source=dataset_context.training_source,
            sheet=request.sheet,
        )
        if request.mode == "reference_anomaly_series":
            observation_roles = (
                request.time_column,
                *request.signal_columns,
                request.reference_label_column,
                *((request.comparison_label_column,) if request.comparison_label_column is not None else ()),
            )
            required_columns = set(request.resolved_selected_columns)
            missing = sorted(required_columns - set(columns))
            if missing:
                raise PlanCompilationError("Reference anomaly Time Series columns are absent from the observation dataset: " f"{missing}")
            scan_columns = tuple(dict.fromkeys(observation_roles))
            positions = [columns.index(column) for column in scan_columns]
            row_count = 0
            time_values = set()
            for row_number, raw_values in enumerate(
                _iter_selected_rows(data_path, positions, sheet=request.sheet),
                start=2,
            ):
                row = dict(zip(scan_columns, raw_values))
                missing_values = [column for column in scan_columns if _is_missing(row[column])]
                if missing_values:
                    raise PlanCompilationError("Reference anomaly Time Series columns contain missing values at " f"data row {row_number}: {missing_values}.")
                time_identity = str(row[request.time_column]).strip()
                if time_identity in time_values:
                    raise PlanCompilationError("Reference anomaly observation time identifiers must be unique; " f"duplicate value {time_identity!r} occurs at data row {row_number}.")
                time_values.add(time_identity)
                for column in request.signal_columns:
                    try:
                        number = float(row[column])
                    except (TypeError, ValueError) as exc:
                        raise PlanCompilationError(f"Signal column {column!r} contains a non-numeric value at data row {row_number}.") from exc
                    if not math.isfinite(number):
                        raise PlanCompilationError(f"Signal column {column!r} contains a non-finite value at data row {row_number}.")
                row_count += 1
            if row_count == 0:
                raise PlanCompilationError("Reference anomaly Time Series input must contain at least one observation row.")

            event_path = request.event_dataset_path.expanduser().resolve() if request.event_dataset_path is not None else None
            if event_path is not None:
                if not event_path.is_file():
                    raise PlanCompilationError(f"Event data file does not exist: {event_path}")
                event_columns = _read_dataset_columns(
                    event_path,
                    source="path",
                    sheet=request.event_sheet,
                )
                event_roles = (
                    request.event_time_column,
                    *((request.event_identifier_column,) if request.event_identifier_column is not None else ()),
                    *((request.event_filter_column,) if request.event_filter_column is not None else ()),
                )
                missing_event_columns = sorted(set(event_roles) - set(event_columns))
                if missing_event_columns:
                    raise PlanCompilationError("Reference anomaly event columns are absent from the event dataset: " f"{missing_event_columns}")

            executable = Path(cli_executable).expanduser().resolve() if cli_executable is not None else resolve_public_cli_executable()
            command = [
                str(executable),
                "reference-anomaly-time-series",
                "--input",
                str(data_path),
                "--sheet",
                request.sheet,
                "--time-column",
                request.time_column,
                "--reference-label-column",
                request.reference_label_column,
                "--reference-label-provenance",
                request.reference_label_provenance,
                "--experiment-name",
                request.experiment_name,
                "--run-name",
                request.run_name,
            ]
            for column in request.signal_columns:
                command.extend(("--signal-column", column))
            for value in request.reference_positive_values:
                command.extend(("--reference-positive-value", value))
            if request.comparison_label_column is not None:
                command.extend(
                    (
                        "--comparison-label-column",
                        request.comparison_label_column,
                        "--comparison-label-provenance",
                        request.comparison_label_provenance,
                    )
                )
                for value in request.comparison_positive_values:
                    command.extend(("--comparison-positive-value", value))
            if event_path is not None:
                command.extend(
                    (
                        "--event-input",
                        str(event_path),
                        "--event-sheet",
                        request.event_sheet,
                        "--event-time-column",
                        request.event_time_column,
                    )
                )
                if request.event_identifier_column is not None:
                    command.extend(("--event-identifier-column", request.event_identifier_column))
                if request.event_filter_column is not None:
                    command.extend(("--event-filter-column", request.event_filter_column))
                    for value in request.event_filter_values:
                        command.extend(("--event-filter-value", value))
                if request.association_window_days is not None:
                    command.extend(
                        (
                            "--association-window-days",
                            format(request.association_window_days, ".15g"),
                        )
                    )
                command.extend(("--association-direction", request.association_direction))
            base = Path(request.experiment_name) / request.run_name
            return InteractionPlan(
                schema_version=INTERACTION_PLAN_VERSION,
                name="time-series-reference-anomaly-v1",
                public_command=tuple(command),
                steps=(),
                expected_output_relative_paths=(
                    (base / "artifacts" / "data" / "Reference Anomaly Time Series.csv").as_posix(),
                    (base / "artifacts" / "data" / "Reference Anomaly Event Associations.csv").as_posix(),
                    (base / "artifacts" / "image" / "model_output" / "Reference Anomaly Time Series.png").as_posix(),
                    (base / "artifacts" / "image" / "model_output" / "Reference Anomaly Time Series.pdf").as_posix(),
                    (base / "metrics" / "Reference Anomaly Time Series Metrics.json").as_posix(),
                    (base / "parameters" / "Reference Anomaly Time Series Parameters.json").as_posix(),
                    (base / "summary" / "Reference Anomaly Artifact Index.json").as_posix(),
                    (base / "summary" / "Reference Anomaly Time Series Manifest.json").as_posix(),
                ),
                workflow_family="time_series",
                workflow_mode="reference_anomaly_series",
                method="reference_label_event_overlay",
                adapter_id="geochemistrypi-cli.time-series.reference-anomaly-series",
                adapter_version="1",
            )
        if request.mode == "element_mean":
            selected_columns = request.resolved_selected_columns
            required_columns = set(selected_columns)
            if request.identifier_column is not None:
                required_columns.add(request.identifier_column)
            missing = sorted(required_columns - set(columns))
            if missing:
                raise PlanCompilationError(f"Time Series configured columns are absent from the training dataset: {missing}")
            numeric_columns = tuple(
                dict.fromkeys(
                    (
                        request.age_column,
                        *request.element_columns,
                        *((request.filter_column,) if request.filter_column is not None else ()),
                    )
                )
            )
            positions = [columns.index(column) for column in numeric_columns]
            drop_columns = tuple(getattr(request.missing_values, "columns", ()))
            if request.missing_values.method == "drop_rows" and not drop_columns:
                drop_columns = numeric_columns
            retained_rows = 0
            maximum_observed_age = 0.0
            for row_number, raw_values in enumerate(
                _iter_selected_rows(data_path, positions, sheet=request.sheet),
                start=2,
            ):
                row = dict(zip(numeric_columns, raw_values))
                if request.missing_values.method == "drop_rows" and any(_is_missing(row[column]) for column in drop_columns):
                    continue
                missing_selected = [column for column in numeric_columns if _is_missing(row[column])]
                if missing_selected:
                    raise PlanCompilationError(f"Time Series selected columns contain missing values at data row {row_number}: {missing_selected}.")
                values: dict[str, float] = {}
                for column in numeric_columns:
                    try:
                        number = float(row[column])
                    except (TypeError, ValueError) as exc:
                        raise PlanCompilationError(f"Time Series column {column!r} contains a non-numeric value at data row {row_number}.") from exc
                    if not math.isfinite(number):
                        raise PlanCompilationError(f"Time Series column {column!r} contains a non-finite value at data row {row_number}.")
                    values[column] = number
                if values[request.age_column] < 0:
                    raise PlanCompilationError(f"Time Series ages must be non-negative; data row {row_number} contains {values[request.age_column]}.")
                retained_rows += 1
                maximum_observed_age = max(maximum_observed_age, values[request.age_column])
            if retained_rows == 0:
                raise PlanCompilationError("Time Series input must contain at least one retained data row.")
            if maximum_observed_age <= 0:
                raise PlanCompilationError("Time Series input must contain at least one positive age.")
            bin_count = math.ceil(maximum_observed_age / request.bin_width)
            if bin_count > 10_000:
                raise PlanCompilationError(f"bin_width creates {bin_count} bins; the safety limit is 10000.")
            return InteractionPlan(
                schema_version=INTERACTION_PLAN_VERSION,
                name="time-series-element-mean-unbound-v1",
                public_command=(),
                steps=(),
                workflow_family="time_series",
                workflow_mode="element_mean",
                method="binned_arithmetic_mean",
                adapter_id=None,
                adapter_version=None,
                adapter_status="unavailable",
                execution_ready=False,
                blocking_issues=(
                    "The public GeochemistryPi CLI has no element_mean Time Series command; "
                    "the scientific request is valid but cannot be executed without changing "
                    "the prohibited scientific/CLI layer.",
                ),
            )
        roles = (
            request.age_column,
            request.maximum_age_column,
            request.probability_column,
            request.latitude_column,
            request.longitude_column,
        )
        selected_columns = request.resolved_selected_columns
        required_columns = set(selected_columns)
        if request.identifier_column is not None:
            required_columns.add(request.identifier_column)
        missing = sorted(required_columns - set(columns))
        if missing:
            raise PlanCompilationError(f"Time Series configured columns are absent from the training dataset: {missing}")
        scan_columns = tuple(dict.fromkeys(selected_columns))
        positions = [columns.index(column) for column in scan_columns]
        drop_columns = tuple(getattr(request.missing_values, "columns", ()))
        if request.missing_values.method == "drop_rows" and not drop_columns:
            drop_columns = selected_columns
        row_count = 0
        maximum_observed_age = 0.0
        for row_number, raw_values in enumerate(
            _iter_selected_rows(data_path, positions, sheet=request.sheet),
            start=2,
        ):
            row = dict(zip(scan_columns, raw_values))
            if request.missing_values.method == "drop_rows" and any(_is_missing(row[column]) for column in drop_columns):
                continue
            if request.missing_values.method == "error":
                missing_selected = [column for column in selected_columns if _is_missing(row[column])]
                if missing_selected:
                    raise PlanCompilationError(f"Time Series selected columns contain missing values at data row {row_number}: {missing_selected}.")
            row_count += 1
            values = []
            for column in roles:
                raw = row[column]
                try:
                    number = float(raw)
                except (TypeError, ValueError) as exc:
                    raise PlanCompilationError(f"Time Series column {column!r} contains a non-numeric value at data row {row_number}.") from exc
                if not math.isfinite(number):
                    raise PlanCompilationError(f"Time Series column {column!r} contains a missing or non-finite value at data row {row_number}.")
                values.append(number)
            age, age_max, probability, latitude, longitude = values
            if age < 0:
                raise PlanCompilationError(f"Time Series ages must be non-negative; data row {row_number} contains {age}.")
            if age_max < 0:
                raise PlanCompilationError(f"Time Series comparison ages must be non-negative; data row {row_number} contains {age_max}.")
            if probability < 0 or probability > 1:
                raise PlanCompilationError(f"Time Series probability must be between 0 and 1 at data row {row_number}.")
            if latitude < -90 or latitude > 90:
                raise PlanCompilationError(f"Time Series latitude must be between -90 and 90 degrees at data row {row_number}.")
            if longitude < -180 or longitude > 180:
                raise PlanCompilationError(f"Time Series longitude must be between -180 and 180 degrees at data row {row_number}.")
            maximum_observed_age = max(maximum_observed_age, age)
        if row_count == 0:
            raise PlanCompilationError("Time Series input must contain at least one data row.")
        if maximum_observed_age <= 0:
            raise PlanCompilationError("Time Series input must contain at least one positive age.")
        bin_count = math.ceil(maximum_observed_age / request.bin_width)
        if bin_count > 10_000:
            raise PlanCompilationError(f"bin_width creates {bin_count} bins; the safety limit is 10000.")

        executable = Path(cli_executable).expanduser().resolve() if cli_executable is not None else resolve_public_cli_executable()
        command = [
            str(executable),
            "time-series",
            "--input",
            str(data_path),
            "--sheet",
            request.sheet,
            "--bin-width",
            format(request.bin_width, ".15g"),
            "--iterations",
            str(request.iterations),
            "--seed",
            str(request.seed),
            "--experiment-name",
            request.experiment_name,
            "--run-name",
            request.run_name,
            "--age-column",
            request.age_column,
            "--maximum-age-column",
            request.maximum_age_column,
            "--probability-column",
            request.probability_column,
            "--latitude-column",
            request.latitude_column,
            "--longitude-column",
            request.longitude_column,
        ]
        if request.identifier_column is not None:
            command.extend(("--identifier-column", request.identifier_column))
        for column in selected_columns:
            command.extend(("--selected-column", column))
        command.extend(("--missing-values", request.missing_values.method))
        for column in tuple(getattr(request.missing_values, "columns", ())):
            command.extend(("--drop-missing-column", column))
        command.extend(
            (
                "--feature-engineering",
                request.feature_engineering,
                "--age-unit",
                request.age_unit,
                "--fit-curve" if request.fit_curve else "--no-fit-curve",
            )
        )
        base = Path(request.experiment_name) / request.run_name
        return InteractionPlan(
            schema_version=INTERACTION_PLAN_VERSION,
            name="time-series-subaerial-proportion-v1",
            public_command=tuple(command),
            steps=(),
            expected_output_relative_paths=(
                (base / "artifacts" / "data" / "Subaerial Proportion.csv").as_posix(),
                (base / "artifacts" / "image" / "model_output" / "Subaerial Proportion.pdf").as_posix(),
                (base / "metrics" / "Time Series Metrics.json").as_posix(),
                (base / "parameters" / "Time Series Parameters.json").as_posix(),
            ),
            workflow_family="time_series",
            workflow_mode="subaerial_proportion",
            method="subaerial_proportion_bootstrap",
            adapter_id="geochemistrypi-cli.time-series.subaerial-proportion",
            adapter_version="1",
        )


class AnalysisPlanCompiler:
    """Dispatch a validated learning request to its task-specific compiler."""

    def __init__(self) -> None:
        self.classification = ClassificationPlanCompiler()
        self.regression = RegressionPlanCompiler()
        self.clustering = ClusteringPlanCompiler()
        self.decomposition = DecompositionPlanCompiler()
        self.anomaly_detection = AnomalyDetectionPlanCompiler()
        self.time_series = TimeSeriesPlanCompiler()

    @staticmethod
    def _tail_after_common_setup(plan: InteractionPlan, model_name: str) -> Tuple[InteractionStep, ...]:
        model_index = next(index for index, step in enumerate(plan.steps) if step.id == model_name)
        index = model_index + 1
        common_ids = {
            "continue_after_model",
            "enable_automl",
            "disable_automl",
            "continue_after_automl",
            "continue_after_inference_skip",
            "continue_after_application_preparation",
        }
        while index < len(plan.steps) and plan.steps[index].id in common_ids:
            index += 1
        return plan.steps[index:]

    @staticmethod
    def _prefixed_steps(model_name: str, steps: Sequence[InteractionStep]) -> List[InteractionStep]:
        return [
            InteractionStep(
                f"{model_name}.{step.id}",
                step.output_anchors,
                step.response,
                step.timeout_seconds,
            )
            for step in steps
        ]

    @staticmethod
    def bind_scientific_adapter(
        plan: InteractionPlan,
        request: ClassificationRequest | RegressionRequest | ClusteringRequest | DecompositionRequest | AnomalyDetectionRequest | TimeSeriesRequest,
    ) -> InteractionPlan:
        """Attach a paper-agnostic scientific identity to an existing CLI plan."""
        if request.task == "time_series":
            method = {
                "subaerial_proportion": "subaerial_proportion_bootstrap",
                "element_mean": "binned_arithmetic_mean",
                "reference_anomaly_series": "reference_label_event_overlay",
            }[request.mode]
            effective_seeds = (("model", request.seed),) if request.mode == "subaerial_proportion" else ()
            binding_fields = _scientific_binding_fields(
                request,
                "time_series",
                request.mode,
                method,
                plan.expected_output_relative_paths,
                effective_seeds,
            )
            if plan.workflow_family != "legacy":
                return dataclass_replace(
                    plan,
                    scientific_contract_id=f"scientific-contract-v2/time_series/{request.mode}/{method}",
                    seed_binding="request_parameter" if request.mode == "subaerial_proportion" else "not_applicable",
                    required_cli_capabilities=(("command:reference-anomaly-time-series",) if request.mode == "reference_anomaly_series" else plan.required_cli_capabilities),
                    **binding_fields,
                )
            return dataclass_replace(
                plan,
                workflow_family="time_series",
                workflow_mode=request.mode,
                method=method,
                scientific_contract_id=f"scientific-contract-v2/time_series/{request.mode}/{method}",
                adapter_id=f"geochemistrypi-cli.time-series.{request.mode.replace('_', '-')}",
                adapter_version="1",
                seed_binding="request_parameter" if request.mode == "subaerial_proportion" else "not_applicable",
                required_cli_capabilities=(("command:reference-anomaly-time-series",) if request.mode == "reference_anomaly_series" else plan.required_cli_capabilities),
                **binding_fields,
            )
        if request.task == "decomposition" and request.mode == "embedding_label_overlay":
            binding_fields = _scientific_binding_fields(
                request,
                "artifact_composition",
                request.mode,
                "exact_identifier_join",
                plan.expected_output_relative_paths,
                (),
            )
            binding_fields["model_parameter_binding"] = "not_applicable"
            return dataclass_replace(
                plan,
                scientific_contract_id="scientific-contract-v2/artifact_composition/embedding_label_overlay/exact_identifier_join",
                seed_binding="not_applicable",
                required_cli_capabilities=("command:embedding-label-overlay",),
                **binding_fields,
            )
        all_models = request.model_selection.mode == "all"
        model_type = "all_public_methods" if all_models else request.model.type
        if request.task in {"classification", "regression"}:
            family = "supervised_learning"
            mode = request.task
            method = model_type
        elif request.task == "decomposition":
            family = "dimension_reduction"
            mode = "embedding"
            method = {
                "pca": "PCA",
                "tsne": "tSNE",
            }.get(model_type, model_type)
        elif request.task == "clustering":
            family = "clustering"
            mode = "clustering"
            method = {
                "agglomerative": "hierarchical",
                "dbscan": "DBSCAN",
            }.get(model_type, model_type)
        else:
            family = "anomaly_detection"
            mode = "outlier_detection"
            method = {
                "local_outlier_factor": "LOF",
            }.get(model_type, model_type)
        selected_tuning = request.model_selection.tuning if all_models else getattr(request, "tuning", "manual")
        scientific_execution = _scientific_execution_contract_json(request, family, mode)
        if scientific_execution is not None:
            scientific_execution_json, effective_seeds = scientific_execution
            requested_seed_roles = {
                role
                for role, value in (
                    ("split", request.reproducibility.split_seed),
                    ("model", request.reproducibility.model_seed),
                    ("tuning", request.reproducibility.tuning_seed),
                )
                if value is not None
            }
            effective_seed_roles = {role for role, _ in effective_seeds}
            if not effective_seed_roles:
                seed_binding = "not_applicable"
            elif not requested_seed_roles:
                seed_binding = "cli_fixed"
            elif requested_seed_roles == effective_seed_roles:
                seed_binding = "request_parameter"
            else:
                seed_binding = "mixed"
        else:
            scientific_execution_json = None
            effective_seeds = (
                (("split", _CLI_FIXED_RANDOM_SEED), ("model", _CLI_FIXED_RANDOM_SEED), ("tuning", _CLI_FIXED_RANDOM_SEED))
                if request.task in {"classification", "regression"} and selected_tuning == "automl"
                else (("split", _CLI_FIXED_RANDOM_SEED), ("model", _CLI_FIXED_RANDOM_SEED))
                if request.task in {"classification", "regression"}
                else (("model", _CLI_FIXED_RANDOM_SEED),)
            )
            seed_binding = "cli_fixed"
        expected_output_relative_paths = plan.expected_output_relative_paths
        if scientific_execution is not None:
            output_root = Path("geopi_output") / request.experiment_name / request.run_name
            additional_paths = [
                str(output_root / "parameters" / "Scientific Execution Attestation.json"),
            ]
            if request.task in {"classification", "regression"}:
                additional_paths.append(str(output_root / "artifacts" / "data" / "Split Membership.xlsx"))
            if request.task == "regression":
                display_name = REGRESSION_MODEL_DISPLAY_NAMES[model_type]
                additional_paths.extend(
                    (
                        str(output_root / "artifacts" / "data" / "Y Train Predict.xlsx"),
                        str(output_root / "metrics" / f"Training Model Score - {display_name}.txt"),
                        str(output_root / "metrics" / f"Cross Validation - {display_name}.txt"),
                    )
                )
                if request.evaluation.mode != "external_labeled":
                    additional_paths.append(str(output_root / "artifacts" / "image" / "model_output" / f"Predicted vs. Actual Density - {display_name}.png"))
                if request.target_transformations:
                    additional_paths.append(str(output_root / "artifacts" / "data" / "Y Original Target.xlsx"))
                if request.evaluation.mode == "external_labeled":
                    additional_paths.extend(
                        (
                            str(output_root / "metrics" / f"External Evaluation Model Score - {display_name}.txt"),
                            str(output_root / "artifacts" / "data" / f"External Evaluation Predictions - {display_name}.xlsx"),
                            str(output_root / "artifacts" / "data" / f"External Evaluation Residuals - {display_name}.xlsx"),
                            str(output_root / "artifacts" / "image" / "model_output" / f"External Predicted vs. Actual - {display_name}.png"),
                        )
                    )
            elif request.task == "classification":
                display_name = MODEL_DISPLAY_NAMES[model_type]
                additional_paths.append(str(output_root / "metrics" / f"Cross Validation - {display_name}.txt"))
                normalization = request.evaluation.confusion_matrix_normalization
                if normalization is not None:
                    normalized_name = f"Normalized Confusion Matrix ({normalization}) - {display_name}"
                    additional_paths.extend(
                        (
                            str(output_root / "artifacts" / "image" / "model_output" / f"{normalized_name}.png"),
                            str(output_root / "artifacts" / "image" / "model_output" / f"{normalized_name}.xlsx"),
                        )
                    )
            expected_output_relative_paths = tuple(dict.fromkeys((*expected_output_relative_paths, *additional_paths)))
        binding_fields = _scientific_binding_fields(
            request,
            family,
            mode,
            method,
            expected_output_relative_paths,
            effective_seeds,
            (_effective_scientific_model_parameters(request, effective_seeds) if scientific_execution is not None else None),
        )
        return dataclass_replace(
            plan,
            workflow_family=family,
            workflow_mode=mode,
            method=method,
            scientific_contract_id=f"scientific-contract-v{'3' if scientific_execution is not None else '2'}/{family}/{mode}/{method}",
            adapter_id=f"geochemistrypi-cli.data-mining.{request.task}.{model_type}",
            adapter_version="2" if scientific_execution is not None else "1",
            seed_binding=seed_binding,
            scientific_execution_contract_json=scientific_execution_json,
            expected_output_relative_paths=expected_output_relative_paths,
            required_cli_capabilities=(("option:data-mining:--scientific-config",) if scientific_execution is not None else plan.required_cli_capabilities),
            **binding_fields,
        )

    def _compile_all_models(
        self,
        request: ClassificationRequest | RegressionRequest | ClusteringRequest | DecompositionRequest | AnomalyDetectionRequest,
        cli_executable: Optional[Path],
        dataset_context: DatasetCompilationContext,
    ) -> InteractionPlan:
        if getattr(request.missing_values, "method", None) == "keep":
            raise PlanCompilationError("model_selection.mode='all' requires missing values to be rejected, dropped, or imputed so every task model receives compatible data.")
        contracts = {
            "classification": (
                self.classification,
                MODEL_ORDER,
                MODEL_DISPLAY_NAMES,
                TypeAdapter(ClassificationModelSettings),
            ),
            "regression": (
                self.regression,
                REGRESSION_MODEL_ORDER,
                REGRESSION_MODEL_DISPLAY_NAMES,
                TypeAdapter(RegressionModelSettings),
            ),
            "clustering": (
                self.clustering,
                CLUSTERING_MODEL_ORDER,
                CLUSTERING_MODEL_DISPLAY_NAMES,
                TypeAdapter(ClusteringModelSettings),
            ),
            "decomposition": (
                self.decomposition,
                DECOMPOSITION_MODEL_ORDER,
                DECOMPOSITION_MODEL_DISPLAY_NAMES,
                TypeAdapter(DecompositionModelSettings),
            ),
            "anomaly_detection": (
                self.anomaly_detection,
                ANOMALY_DETECTION_MODEL_ORDER,
                ANOMALY_DETECTION_MODEL_DISPLAY_NAMES,
                TypeAdapter(AnomalyDetectionModelSettings),
            ),
        }
        compiler, model_order, display_names, adapter = contracts[request.task]
        tuning = request.model_selection.tuning
        plans = []
        for model_name in model_order:
            child_tuning = tuning
            if request.task == "regression" and model_name in REGRESSION_MODELS_WITHOUT_AUTOML:
                child_tuning = "manual"
            child = request.model_copy(
                update={
                    "model": adapter.validate_python({"type": model_name}),
                    **({"tuning": child_tuning} if request.task in {"classification", "regression"} else {}),
                }
            )
            plans.append(
                (
                    model_name,
                    compiler.compile(
                        child,
                        cli_executable=cli_executable,
                        dataset_context=dataset_context,
                    ),
                )
            )

        first_model, first_plan = plans[0]
        selection_index = next(index for index, step in enumerate(first_plan.steps) if step.id == first_model)
        steps = list(first_plan.steps[:selection_index])
        selection_anchors = first_plan.steps[selection_index].output_anchors
        steps.extend(
            [
                InteractionStep(
                    "all_models",
                    selection_anchors,
                    str(len(model_order) + 1),
                ),
                InteractionStep("continue_after_model", ("(Press Enter",), ""),
            ]
        )
        if request.task in {"classification", "regression"}:
            steps.extend(
                [
                    InteractionStep(
                        "enable_automl" if tuning == "automl" else "disable_automl",
                        ("automated machine learning", "(Model)"),
                        "1" if tuning == "automl" else "2",
                    ),
                    InteractionStep("continue_after_automl", ("(Press Enter",), ""),
                ]
            )
            common_inference = next(
                (
                    step
                    for step in first_plan.steps[selection_index + 1 :]
                    if step.id
                    in {
                        "continue_after_inference_skip",
                        "continue_after_application_preparation",
                    }
                ),
                None,
            )
            if common_inference is not None:
                steps.append(common_inference)

        for model_name, plan in plans:
            aggregate_tail = tuple(
                step
                for step in self._tail_after_common_setup(plan, model_name)
                if step.id
                not in {
                    "continue_after_training",
                    "continue_after_transform_pipeline",
                }
            )
            steps.extend(
                self._prefixed_steps(
                    model_name,
                    aggregate_tail,
                )
            )

        base = Path("geopi_output") / request.experiment_name / request.run_name
        expected_outputs = [(base / "summary" / "Aggregate Model Results.json").as_posix()]
        for (model_name, plan) in plans:
            display = display_names[model_name]
            single_run_base = Path("geopi_output") / request.experiment_name / request.run_name
            for relative in plan.expected_output_relative_paths:
                candidate = Path(relative)
                try:
                    child_tail = candidate.relative_to(single_run_base)
                except ValueError as exc:
                    raise PlanCompilationError("A child model projected an output outside its run directory: " f"{relative!r}.") from exc
                expected_outputs.append((base / display / child_tail).as_posix())
        return InteractionPlan(
            schema_version=INTERACTION_PLAN_VERSION,
            name=(
                f"regression-all-models-{tuning}-multi-output-v1" if isinstance(request, RegressionRequest) and len(request.resolved_target_columns) > 1 else f"{request.task}-all-models-{tuning}-v1"
            ),
            public_command=first_plan.public_command,
            steps=tuple(steps),
            expected_output_relative_paths=tuple(dict.fromkeys(expected_outputs)),
            requires_source_row_pairing=True,
        )

    def compile(
        self,
        request: ClassificationRequest | RegressionRequest | ClusteringRequest | DecompositionRequest | AnomalyDetectionRequest | TimeSeriesRequest,
        cli_executable: Optional[Path] = None,
        *,
        dataset_context: DatasetCompilationContext | None = None,
    ) -> InteractionPlan:
        dataset_context = dataset_context or DatasetCompilationContext()
        if request.task != "time_series" and request.model_selection.mode == "all":
            plan = self._compile_all_models(request, cli_executable, dataset_context)
            return self.bind_scientific_adapter(plan, request)
        if request.task == "classification":
            plan = self.classification.compile(request, cli_executable=cli_executable, dataset_context=dataset_context)
            return self.bind_scientific_adapter(plan, request)
        if request.task == "regression":
            plan = self.regression.compile(request, cli_executable=cli_executable, dataset_context=dataset_context)
            return self.bind_scientific_adapter(plan, request)
        if request.task == "clustering":
            plan = self.clustering.compile(request, cli_executable=cli_executable, dataset_context=dataset_context)
            return self.bind_scientific_adapter(plan, request)
        if request.task == "decomposition":
            plan = self.decomposition.compile(request, cli_executable=cli_executable, dataset_context=dataset_context)
            return self.bind_scientific_adapter(plan, request)
        if request.task == "time_series":
            plan = self.time_series.compile(
                request,
                cli_executable=cli_executable,
                dataset_context=dataset_context,
            )
            return self.bind_scientific_adapter(plan, request)
        plan = self.anomaly_detection.compile(request, cli_executable=cli_executable, dataset_context=dataset_context)
        return self.bind_scientific_adapter(plan, request)
