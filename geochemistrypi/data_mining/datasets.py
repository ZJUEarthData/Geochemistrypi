"""Read-only discovery metadata for built-in and Desktop CLI datasets."""

import csv
import hashlib
import json
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from .constants import BUILT_IN_DATASET_PATH

DATASET_CATALOG_SCHEMA_VERSION = 1
SUPPORTED_DATASET_SUFFIXES = {".csv": "csv", ".xlsx": "xlsx"}
_BUILT_IN_DATASETS = {
    "ApplicationData_Classification.xlsx": (
        "builtin:application_classification",
        "classification",
        "application",
    ),
    "ApplicationData_Regression.xlsx": (
        "builtin:application_regression",
        "regression",
        "application",
    ),
    "Data_AnomalyDetection.xlsx": (
        "builtin:anomaly_detection",
        "anomaly_detection",
        "training",
    ),
    "Data_Classification.xlsx": (
        "builtin:classification",
        "classification",
        "training",
    ),
    "Data_Clustering.xlsx": (
        "builtin:clustering",
        "clustering",
        "training",
    ),
    "Data_Decomposition.xlsx": (
        "builtin:decomposition",
        "decomposition",
        "training",
    ),
    "Data_Regression.xlsx": (
        "builtin:regression",
        "regression",
        "training",
    ),
    "Data_Time_Series.xlsx": (
        "builtin:time_series",
        "time_series",
        "training",
    ),
}


class DatasetCatalogError(RuntimeError):
    """Raised when dataset discovery cannot produce trustworthy metadata."""


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _shape(path: Path) -> Tuple[Optional[int], Optional[int]]:
    if path.suffix.lower() == ".csv":
        try:
            with path.open("r", encoding="utf-8-sig", newline="") as stream:
                reader = csv.reader(stream)
                header = next(reader)
                rows = sum(1 for row in reader if any(value.strip() for value in row))
                return rows, len(header)
        except (OSError, UnicodeError, csv.Error, StopIteration):
            return None, None
    try:
        with path.open("rb") as stream:
            workbook = load_workbook(stream, read_only=True, data_only=True)
            try:
                worksheet = workbook.active
                rows = worksheet.iter_rows(values_only=True)
                header = next(rows, ())
                column_count = max(
                    (index for index, value in enumerate(header, start=1) if value is not None),
                    default=0,
                )
                row_count = sum(1 for row in rows if any(value is not None for value in row))
                return row_count, column_count
            finally:
                workbook.close()
    except (OSError, ValueError):
        return None, None


def _json_value(value: Any) -> Any:
    if isinstance(value, np.generic):
        return _json_value(value.item())
    if isinstance(value, (pd.Timestamp, pd.Timedelta)):
        return value.isoformat()
    if isinstance(value, float) and not np.isfinite(value):
        if np.isnan(value):
            return "NaN"
        return "Infinity" if value > 0 else "-Infinity"
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def _inspect(
    path: Path,
    selected_columns: Sequence[str] = (),
) -> Dict[str, Any]:
    """Return a read-only, JSON-safe scientific inspection of one dataset."""

    try:
        if path.suffix.lower() == ".csv":
            frame = pd.read_csv(path)
        else:
            frame = pd.read_excel(path)
    except (OSError, ValueError) as exc:
        raise DatasetCatalogError(f"Unable to inspect dataset {path.name!r}.") from exc

    columns = [str(column) for column in frame.columns]
    unknown_columns = sorted(set(selected_columns) - set(columns))
    if unknown_columns:
        raise DatasetCatalogError(f"Dataset {path.name!r} does not contain inspection columns: " f"{unknown_columns}.")
    selected_names = list(selected_columns) if selected_columns else columns
    selected_frame = frame.loc[:, selected_names]
    selected_missing_rows = selected_frame.isna().any(axis=1)
    selected_nonfinite_rows = pd.Series(False, index=selected_frame.index)
    for column in selected_frame.select_dtypes(include=[np.number]).columns:
        numeric = selected_frame[column].to_numpy(dtype=float, na_value=np.nan)
        selected_nonfinite_rows |= np.isinf(numeric)
    selected_invalid_rows = selected_missing_rows | selected_nonfinite_rows
    dtypes: Dict[str, str] = {}
    missing_counts: Dict[str, int] = {}
    nonfinite_counts: Dict[str, Optional[int]] = {}
    minimums: Dict[str, Any] = {}
    maximums: Dict[str, Any] = {}
    distinct_non_null_counts: Dict[str, int] = {}
    low_cardinality_value_counts: Dict[str, List[Dict[str, Any]]] = {}

    for position, column in enumerate(frame.columns):
        name = str(column)
        series = frame.iloc[:, position]
        dtypes[name] = str(series.dtype)
        missing_counts[name] = int(series.isna().sum())
        distinct_count = int(series.nunique(dropna=True))
        distinct_non_null_counts[name] = distinct_count
        if pd.api.types.is_numeric_dtype(series.dtype):
            numeric = series.to_numpy(dtype=float, na_value=np.nan)
            nonfinite_counts[name] = int(np.isinf(numeric).sum())
            finite = numeric[np.isfinite(numeric)]
            minimums[name] = None if not finite.size else _json_value(finite.min())
            maximums[name] = None if not finite.size else _json_value(finite.max())
        else:
            nonfinite_counts[name] = None
            minimums[name] = None
            maximums[name] = None
        if distinct_count <= 20:
            counts = series.dropna().value_counts(sort=False)
            low_cardinality_value_counts[name] = [{"value": _json_value(value), "count": int(count)} for value, count in counts.items()]

    return {
        "row_count": int(frame.shape[0]),
        "column_count": int(frame.shape[1]),
        "columns": columns,
        "selected_columns": selected_names,
        "selected_rows_with_any_missing": int(selected_missing_rows.sum()),
        "selected_rows_with_any_nonfinite": int(selected_nonfinite_rows.sum()),
        "selected_rows_with_any_invalid": int(selected_invalid_rows.sum()),
        "selected_complete_row_count": int((~selected_invalid_rows).sum()),
        "dtypes": dtypes,
        "missing_counts": missing_counts,
        "nonfinite_counts": nonfinite_counts,
        "minimums": minimums,
        "maximums": maximums,
        "distinct_non_null_counts": distinct_non_null_counts,
        "low_cardinality_value_counts": low_cardinality_value_counts,
    }


def _entry(
    path: Path,
    source: str,
    dataset_id: str,
    task: Optional[str],
    role: str,
) -> Dict[str, Any]:
    metadata = path.stat()
    rows, columns = _shape(path)
    analysis_blockers = []
    return {
        "dataset_id": dataset_id,
        "source": source,
        "role": role,
        "task": task,
        "file_name": path.name,
        "path": str(path),
        "format": SUPPORTED_DATASET_SUFFIXES[path.suffix.lower()],
        "size_bytes": metadata.st_size,
        "sha256": _sha256(path),
        "row_count": rows,
        "column_count": columns,
        "analysis_blockers": analysis_blockers,
    }


def _built_in_entries(dataset_ids: Sequence[str] = ()) -> List[Dict[str, Any]]:
    root = Path(BUILT_IN_DATASET_PATH).resolve(strict=True)
    requested_ids = tuple(dict.fromkeys(dataset_ids))
    if requested_ids:
        by_id = {dataset_id: (name, task, role) for name, (dataset_id, task, role) in _BUILT_IN_DATASETS.items()}
        unknown = sorted(set(requested_ids) - set(by_id))
        if unknown:
            raise DatasetCatalogError(f"Unknown built-in dataset IDs: {unknown}.")
        selected = [(by_id[dataset_id][0], dataset_id, by_id[dataset_id][1], by_id[dataset_id][2]) for dataset_id in requested_ids]
        return [_entry(root / name, "builtin", dataset_id, task, role) for name, dataset_id, task, role in selected]
    discovered = {path.name for path in root.iterdir() if path.is_file() and path.suffix.lower() in SUPPORTED_DATASET_SUFFIXES}
    declared = set(_BUILT_IN_DATASETS)
    if discovered != declared:
        raise DatasetCatalogError("Built-in dataset declarations are stale. " f"Undeclared files: {sorted(discovered - declared)}; missing files: {sorted(declared - discovered)}")
    return [_entry(root / name, "builtin", dataset_id, task, role) for name, (dataset_id, task, role) in sorted(_BUILT_IN_DATASETS.items())]


def desktop_input_root() -> Path:
    """Return the same Desktop/geopi_input location used by the human CLI."""
    return (Path.home() / "Desktop" / "geopi_input").resolve()


def _desktop_entries(file_names: Sequence[str] = ()) -> Tuple[List[Dict[str, Any]], List[str], Path]:
    root = desktop_input_root()
    if not root.exists():
        return [], ["Desktop/geopi_input does not exist; discovery did not create it."], root
    if not root.is_dir():
        raise DatasetCatalogError(f"Desktop dataset location is not a directory: {root}")
    entries: List[Dict[str, Any]] = []
    warnings: List[str] = []
    requested_names = tuple(dict.fromkeys(file_names))
    if requested_names:
        invalid = [name for name in requested_names if not name or Path(name).name != name or Path(name).is_absolute() or Path(name).suffix.lower() not in SUPPORTED_DATASET_SUFFIXES]
        if invalid:
            raise DatasetCatalogError(f"Desktop exact selectors must be plain supported file names: {invalid}.")
        candidates = [root / name for name in requested_names]
    else:
        try:
            candidates = sorted(root.iterdir(), key=lambda item: item.name.casefold())
        except OSError as exc:
            raise DatasetCatalogError(f"Desktop dataset location could not be read: {root}") from exc
    for candidate in candidates:
        if candidate.suffix.lower() not in SUPPORTED_DATASET_SUFFIXES:
            continue
        if candidate.is_symlink():
            if requested_names:
                raise DatasetCatalogError(f"Desktop exact selector is a symbolic link: {candidate.name}")
            warnings.append(f"Ignored symbolic-link Desktop entry: {candidate.name}")
            continue
        try:
            resolved = candidate.resolve(strict=True)
            resolved.relative_to(root)
        except (OSError, RuntimeError, ValueError):
            if requested_names:
                raise DatasetCatalogError(f"Desktop dataset was not found or is unsafe: {candidate.name}")
            warnings.append(f"Ignored unsafe Desktop entry: {candidate.name}")
            continue
        if not resolved.is_file():
            warnings.append(f"Ignored non-file Desktop entry: {candidate.name}")
            continue
        stable_name = candidate.name.replace("%", "%25").replace(":", "%3A")
        try:
            entries.append(
                _entry(
                    resolved,
                    "desktop",
                    f"desktop:{stable_name}",
                    None,
                    "unspecified",
                )
            )
        except OSError:
            warnings.append(f"Ignored Desktop entry that changed while being read: {candidate.name}")
    return entries, warnings, root


def dataset_catalog(
    source: str = "all",
    *,
    dataset_ids: Sequence[str] = (),
    file_names: Sequence[str] = (),
    detail: str = "compact",
    inspection_columns: Sequence[str] = (),
) -> Dict[str, Any]:
    """Return bounded JSON-ready dataset metadata without changing the filesystem."""
    if source not in {"all", "builtin", "desktop"}:
        raise DatasetCatalogError("source must be one of: all, builtin, desktop")
    if dataset_ids and source == "desktop":
        raise DatasetCatalogError("Built-in dataset IDs require source 'builtin' or 'all'.")
    if file_names and source == "builtin":
        raise DatasetCatalogError("Desktop file names require source 'desktop' or 'all'.")
    if detail not in {"compact", "full"}:
        raise DatasetCatalogError("detail must be one of: compact, full")
    if inspection_columns and detail != "full":
        raise DatasetCatalogError("inspection columns require detail 'full'")
    entries: List[Dict[str, Any]] = []
    warnings: List[str] = []
    desktop_root: Optional[Path] = None
    if source in {"all", "builtin"}:
        entries.extend(_built_in_entries(dataset_ids))
    if source in {"all", "desktop"}:
        desktop, desktop_warnings, desktop_root = _desktop_entries(file_names)
        entries.extend(desktop)
        warnings.extend(desktop_warnings)
    if detail == "full":
        for entry in entries:
            entry["inspection"] = _inspect(
                Path(entry["path"]),
                inspection_columns,
            )
    return {
        "schema_version": DATASET_CATALOG_SCHEMA_VERSION,
        "source_filter": source,
        "supported_formats": ["csv", "xlsx"],
        "desktop_root": None if desktop_root is None else str(desktop_root),
        "datasets": entries,
        "warnings": warnings,
    }


def dataset_catalog_json(
    source: str = "all",
    *,
    dataset_ids: Sequence[str] = (),
    file_names: Sequence[str] = (),
    detail: str = "compact",
    inspection_columns: Sequence[str] = (),
) -> str:
    return json.dumps(
        dataset_catalog(
            source,
            dataset_ids=dataset_ids,
            file_names=file_names,
            detail=detail,
            inspection_columns=inspection_columns,
        ),
        allow_nan=False,
        ensure_ascii=False,
        sort_keys=True,
    )
