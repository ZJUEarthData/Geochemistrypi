import csv
import json
from pathlib import Path

import numpy

# The repository's Python 3.9 lock uses a NumPy version that still exposes
# these aliases.  Keep this test runnable in the host's newer NumPy too.
if not hasattr(numpy, "short"):
    numpy.short = numpy.int16
if not hasattr(numpy, "ushort"):
    numpy.ushort = numpy.uint16
if not hasattr(numpy, "intc"):
    numpy.intc = numpy.int32
if not hasattr(numpy, "uintc"):
    numpy.uintc = numpy.uint32
if not hasattr(numpy, "int0"):
    numpy.int0 = numpy.intp
if not hasattr(numpy, "uint0"):
    numpy.uint0 = numpy.uintp
if not hasattr(numpy, "longlong"):
    numpy.longlong = numpy.int64
if not hasattr(numpy, "ulonglong"):
    numpy.ulonglong = numpy.uint64

from openpyxl import Workbook
from typer.testing import CliRunner

from geochemistrypi.cli import app
from geochemistrypi.data_preparation import prepare_data


def test_prepare_compound_header_projection_and_filter(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Measurements"
    sheet.append(["Sample", "Oxides", "Oxides"])
    sheet.append(["ID", "SiO2", "MgO"])
    sheet.append(["A", 50.0, 8.0])
    sheet.append(["B", 42.0, 12.0])
    workbook.save(source)

    output = tmp_path / "prepared.csv"
    record = prepare_data(
        source,
        output,
        {
            "worksheet": "Measurements",
            "header_row_indices": [0, 1],
            "selected_columns": ["Sample | ID", "Oxides | MgO"],
            "filters": [{"column": "Oxides | SiO2", "operator": "between", "minimum": 43, "maximum": 51}],
        },
    )

    assert record["source_row_count"] == 1
    with output.open(encoding="utf-8", newline="") as stream:
        assert list(csv.reader(stream)) == [["Sample | ID", "Oxides | MgO"], ["A", "8"]]


def test_prepare_worksheet_union_adds_lineage(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "MORB"
    first.append(["ID", "Value"])
    first.append(["M1", 1])
    second = workbook.create_sheet("OIB")
    second.append(["ID", "Value"])
    second.append(["O1", 2])
    workbook.save(source)

    output = tmp_path / "union.csv"
    record = prepare_data(
        source,
        output,
        {
            "worksheets": ["MORB", "OIB"],
            "union_mode": "rows",
            "source_sheet_column": "SourceSheet",
            "source_row_column": "SourceRow",
            "selected_columns": ["ID", "SourceSheet", "SourceRow", "Value"],
        },
    )

    assert record["source_row_count"] == 2
    with output.open(encoding="utf-8", newline="") as stream:
        assert list(csv.reader(stream)) == [
            ["ID", "SourceSheet", "SourceRow", "Value"],
            ["M1", "MORB", "2", "1"],
            ["O1", "OIB", "2", "2"],
        ]


def test_prepare_data_cli_writes_audit_manifest(tmp_path: Path) -> None:
    source = tmp_path / "source.csv"
    source.write_text("ID,Value\nA,1\n", encoding="utf-8")
    config = tmp_path / "config.json"
    config.write_text(json.dumps({"selected_columns": ["ID", "Value"]}), encoding="utf-8")
    output = tmp_path / "prepared.csv"

    result = CliRunner().invoke(
        app,
        ["prepare-data", "--source", str(source), "--config", str(config), "--output", str(output)],
    )

    assert result.exit_code == 0, result.output
    assert output.is_file()
    assert output.with_suffix(".preparation.json").is_file()
