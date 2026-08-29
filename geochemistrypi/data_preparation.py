"""Paper-independent preparation of source tables for the GeochemistryPi CLI."""

import csv
import json
import math
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Dict, Iterable, List, Sequence, Tuple

from openpyxl import load_workbook


class DataPreparationError(ValueError):
    """Raised when a preparation contract cannot be applied deterministically."""


def _cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (date, datetime, time)):
        return value.isoformat()
    return value


def _headers(rows: Sequence[Sequence[Any]], config: Dict[str, Any]) -> List[str]:
    width = max((len(row) for row in rows), default=0)
    separator = str(config.get("header_join_separator", " | "))
    forward_fill = config.get("empty_header_policy", "forward_fill") == "forward_fill"
    layers: List[List[str]] = []
    for row in rows:
        layer: List[str] = []
        previous = ""
        for index in range(width):
            value = "" if index >= len(row) or row[index] is None else str(row[index]).strip()
            if not value and forward_fill:
                value = previous
            if value:
                previous = value
            layer.append(value)
        layers.append(layer)
    result = []
    for index in range(width):
        tokens: List[str] = []
        for layer in layers:
            token = layer[index]
            if token and (not tokens or tokens[-1] != token):
                tokens.append(token)
        result.append(separator.join(tokens))
    if config.get("header_whitespace_policy") == "strip":
        result = [value.strip() for value in result]
    if config.get("header_bom_policy") == "strip" and result:
        result[0] = result[0].lstrip("\ufeff")
    duplicates = {name for name in result if name and result.count(name) > 1}
    if duplicates and config.get("duplicate_header_policy", "reject") != "suffix":
        raise DataPreparationError(f"Duplicate column names: {sorted(duplicates)}")
    counts: Dict[str, int] = {}
    for index, name in enumerate(result):
        counts[name] = counts.get(name, 0) + 1
        if name in duplicates:
            result[index] = f"{name}{separator}{counts[name]}"
    if any(not name for name in result):
        raise DataPreparationError("The selected header contains an empty column name.")
    return result


def _is_null(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip()) or (isinstance(value, float) and math.isnan(value))


def _compare(value: Any, operand: Any) -> Any:
    if isinstance(operand, bool):
        normalized = str(value).strip().lower()
        if normalized in {"true", "1"}:
            return True
        if normalized in {"false", "0"}:
            return False
        raise DataPreparationError(f"Value {value!r} is not boolean.")
    if isinstance(operand, (int, float)) and not isinstance(operand, bool):
        try:
            return float(value)
        except (TypeError, ValueError) as exc:
            raise DataPreparationError(f"Value {value!r} is not numeric.") from exc
    return str(value)


def _keep(row: Sequence[Any], positions: Dict[str, int], filters: Sequence[Dict[str, Any]]) -> bool:
    for rule in filters:
        column = rule["column"]
        if column not in positions:
            raise DataPreparationError(f"Filter column is absent: {column!r}.")
        value = row[positions[column]]
        operator = rule["operator"]
        if operator == "not_null":
            if _is_null(value):
                return False
            continue
        if _is_null(value):
            return False
        if operator == "between":
            comparable = _compare(value, rule["minimum"])
            low, high = float(rule["minimum"]), float(rule["maximum"])
            inside = low <= comparable <= high if rule.get("inclusive", True) else low < comparable < high
            if not inside:
                return False
            continue
        if operator == "in":
            if not any(_compare(value, item) == _compare(item, item) for item in rule["values"]):
                return False
            continue
        expected = rule["value"]
        left, right = _compare(value, expected), _compare(expected, expected)
        outcomes = {
            "equal": left == right,
            "not_equal": left != right,
            "greater_than": left > right,
            "greater_than_or_equal": left >= right,
            "less_than": left < right,
            "less_than_or_equal": left <= right,
        }
        if operator not in outcomes:
            raise DataPreparationError(f"Unsupported filter operator: {operator!r}.")
        if not outcomes[operator]:
            return False
    return True


def _table(path: Path, sheet: Any, config: Dict[str, Any]) -> Tuple[List[str], Iterable[Tuple[Any, ...]], str, int, Any]:
    header_indices = config.get("header_row_indices")
    indices = list(header_indices) if header_indices is not None else [int(config.get("header_row_index", 0))]
    if indices != sorted(set(indices)) or not indices or indices[0] < 0:
        raise DataPreparationError("Header row indices must be unique, increasing, and zero-based.")
    if path.suffix.lower() == ".csv":
        if sheet is not None:
            raise DataPreparationError("CSV input does not accept a worksheet.")
        stream = path.open("r", encoding="utf-8-sig", newline="")
        rows = iter(csv.reader(stream))
        owner, sheet_name = stream, "csv"
    elif path.suffix.lower() == ".xlsx":
        owner = load_workbook(path, read_only=True, data_only=True)
        if sheet is None:
            if len(owner.sheetnames) != 1:
                owner.close()
                raise DataPreparationError(f"Select one worksheet from {owner.sheetnames}.")
            worksheet = owner[owner.sheetnames[0]]
        else:
            try:
                worksheet = owner[str(sheet)]
            except KeyError as exc:
                owner.close()
                raise DataPreparationError(f"Worksheet does not exist: {sheet!r}.") from exc
        rows, sheet_name = worksheet.iter_rows(values_only=True), worksheet.title
    else:
        raise DataPreparationError("Input must be .csv or .xlsx.")
    selected = []
    try:
        for index in range(indices[-1] + 1):
            row = tuple(next(rows))
            if index in indices:
                selected.append(row)
    except StopIteration as exc:
        owner.close()
        raise DataPreparationError("Header rows exceed the available table rows.") from exc
    return _headers(selected, config), rows, sheet_name, indices[-1] + 2, owner


def prepare_data(source: Path, destination: Path, config: Dict[str, Any]) -> Dict[str, Any]:
    """Apply a JSON preparation contract and write a stable CLI-ready CSV."""
    source = source.expanduser().resolve()
    destination = destination.expanduser().resolve()
    if not source.is_file():
        raise DataPreparationError(f"Source dataset does not exist: {source}")
    sheets = config.get("worksheets") or [config.get("worksheet")]
    if config.get("worksheets") and config.get("union_mode") != "rows":
        raise DataPreparationError("Multiple worksheets require union_mode='rows'.")
    generated_sheet = config.get("source_sheet_column")
    generated_row = config.get("source_row_column")
    if len(sheets) > 1 and (not generated_sheet or not generated_row):
        raise DataPreparationError("Worksheet unions require source_sheet_column and source_row_column.")
    selected = list(config.get("selected_columns") or [])
    excluded = set(config.get("excluded_columns") or [])
    if selected and excluded:
        raise DataPreparationError("selected_columns and excluded_columns are mutually exclusive.")
    filters = list(config.get("filters") or [])
    output_rows: List[List[Any]] = []
    output_header: List[str] = selected
    input_count = 0
    per_sheet: Dict[str, Dict[str, int]] = {}
    for requested_sheet in sheets:
        header, rows, sheet_name, first_row, owner = _table(source, requested_sheet, config)
        positions = {name: index for index, name in enumerate(header)}
        source_columns = [name for name in selected if name not in {generated_sheet, generated_row}] if selected else [name for name in header if name not in excluded]
        missing = [name for name in source_columns if name not in positions]
        if missing:
            owner.close()
            raise DataPreparationError(f"Selected columns are absent: {missing}")
        if not output_header:
            output_header = source_columns
        kept = 0
        try:
            for row_number, raw in enumerate(rows, start=first_row):
                input_count += 1
                row = list(raw[: len(header)]) + [""] * max(0, len(header) - len(raw))
                if not _keep(row, positions, filters):
                    continue
                values = {name: row[positions[name]] for name in source_columns}
                if generated_sheet:
                    values[generated_sheet] = sheet_name
                if generated_row:
                    values[generated_row] = row_number
                output_rows.append([_cell(values[name]) for name in output_header])
                kept += 1
        finally:
            owner.close()
        per_sheet[sheet_name] = {"input_row_count": input_count - sum(item["input_row_count"] for item in per_sheet.values()), "source_row_count": kept}
    if not output_rows:
        raise DataPreparationError("No rows remain after preparation.")
    destination.parent.mkdir(parents=True, exist_ok=True)
    with destination.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.writer(stream, lineterminator="\n")
        writer.writerow(output_header)
        writer.writerows(output_rows)
    return {
        "schema_version": 1,
        "source": str(source),
        "output": str(destination),
        "columns": output_header,
        "input_row_count": input_count,
        "source_row_count": len(output_rows),
        "filtered_row_count": input_count - len(output_rows),
        "per_sheet": per_sheet,
        "contract": config,
    }


def load_preparation_config(path: Path) -> Dict[str, Any]:
    try:
        value = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise DataPreparationError(f"Cannot read preparation config: {path}") from exc
    if not isinstance(value, dict):
        raise DataPreparationError("Preparation config must be a JSON object.")
    return value
